from __future__ import annotations

from collections import Counter
from copy import copy
import csv
import hashlib
import io
import json
import logging
import os
import re
import time
import uuid
from datetime import date, datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.cell.cell import MergedCell
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ...deps import get_db_conn, get_current_user, apply_rate_limit
from ...schemas import (
    AppleDaytimeShiftOut,
    ImportPreviewIssueLocationOut,
    ImportPreviewIssueOut,
    ScheduleBulkCreateIn,
    ScheduleBulkCreateOut,
    AppleOvertimeCreate,
    AppleOvertimeOut,
    DailyEventCreate,
    DailyEventOut,
    DutyLogOut,
    DutyLogRowOut,
    FinanceSubmissionPreviewOut,
    FinanceSubmissionStatusOut,
    ImportApplyOut,
    ImportApplyRowOut,
    ImportPreviewOut,
    ImportPreviewRowOut,
    ScheduleImportMappingEntryOut,
    ScheduleImportMappingProfileOut,
    LateShiftCreate,
    LateShiftOut,
    ScheduleCloserUpdate,
    ScheduleLeaderCandidateOut,
    ScheduleLeaderCandidatesOut,
    ScheduleCreateRow,
    SupportRoundtripApplyOut,
    SupportRoundtripApplyRowOut,
    SupportRoundtripPreviewOut,
    SupportRoundtripPreviewRowOut,
    SupportRoundtripPreviewMetadataOut,
    SupportRosterHqReviewIssueOut,
    SupportRosterHqReviewRowOut,
    SupportRosterHqApplyOut,
    SupportRosterHqApplyScopeOut,
    SupportRosterHqScopeSummaryOut,
    SupportRosterHqUploadInspectOut,
    SupportRosterHqUploadMetaOut,
    SupportRosterHqWorkspaceOut,
    SupportRosterHqWorkspaceSiteOut,
    SupportRoundtripStatusOut,
    ScheduleTemplateBulkCreateIn,
    ScheduleTemplateCreate,
    ScheduleTemplateOut,
    ScheduleTemplateSingleCreateIn,
    ScheduleTemplateUpdate,
    ScheduleUpdate,
    SiteShiftPolicyOut,
    SiteShiftPolicyUpdate,
    SupportAssignmentCreate,
    SupportAssignmentOut,
    SupportStatusAssignmentOut,
    SupportStatusWorkspaceOut,
    SupportStatusWorkspaceRowOut,
)
from ...services.p1_schedule import (
    build_duty_log,
    create_apple_overtime_log,
    create_daily_event_log,
    create_late_shift_log,
    delete_daily_event_log,
    delete_late_shift_log,
    delete_support_assignment,
    generate_apple_daytime_shift,
    get_or_create_site_shift_policy,
    list_apple_report_overnight_records,
    list_apple_overtime_logs,
    list_daily_event_logs,
    list_late_shift_logs,
    list_support_assignments,
    parse_support_entry_text,
    resolve_employee,
    resolve_site,
    resolve_support_entries_to_assignments,
    resolve_tenant,
    upsert_site_shift_policy,
    upsert_support_assignment,
)
from ...services.push_notifications import send_push_notification_to_users
from ...utils.permissions import can_manage_schedule, is_super_admin, normalize_role, normalize_user_role
from ...utils.tenant_context import enforce_staff_site_scope, resolve_scoped_tenant

router = APIRouter(prefix="/schedules", tags=["schedules"], dependencies=[Depends(apply_rate_limit)])
logger = logging.getLogger(__name__)


SHIFT_TYPE_ALIASES = {
    "leave": "off",
}
ALLOWED_SHIFT_TYPES = {"day", "overtime", "night", "off", "holiday"}
NON_WORKING_SHIFT_TYPES = {"off", "holiday"}
TEAM_MANAGER_DUTY_ROLE = "TEAM_MANAGER"
VICE_SUPERVISOR_DUTY_ROLE = "VICE_SUPERVISOR"
GUARD_DUTY_ROLE = "GUARD"
IMPORT_FORMATS = {"csv", "xlsx"}
IMPORT_PREVIEW_LIMIT = 300
IMPORT_REPORT_LIMIT = 300
SHIFT_TYPE_LABELS = {
    "day": "09:00-18:00",
    "overtime": "초과근무",
    "night": "18:00-09:00",
    "off": "휴무",
    "holiday": "공휴일",
}
SUPERVISOR_DAY_SHIFT_START = "08:00:00"
SUPERVISOR_DAY_SHIFT_END = "18:00:00"
SUPERVISOR_DAY_SHIFT_HOURS = 10.0
GUARD_DAY_SHIFT_START = "10:00:00"
GUARD_DAY_SHIFT_END = "22:00:00"
GUARD_DAY_SHIFT_HOURS = 12.0
DEFAULT_NIGHT_SHIFT_START = "22:00:00"
DEFAULT_NIGHT_SHIFT_END = "08:00:00"
DEFAULT_NIGHT_SHIFT_HOURS = 10.0
VALIDATION_MESSAGES = {
    "tenant_code_mismatch": "요청 계정 테넌트와 일치하지 않습니다.",
    "tenant_match_failed": "테넌트 코드 매칭 실패",
    "company_match_failed": "조직(회사) 코드 매칭 실패",
    "site_match_failed": "사이트 코드 매칭 실패",
    "employee_match_failed": "직원 코드 매칭 실패",
    "required_column_missing": "필수 컬럼 누락",
    "invalid_shift_type": "shift_type 값이 유효하지 않습니다.",
    "invalid_schedule_date": "schedule_date 형식이 올바르지 않습니다. (YYYY-MM-DD)",
    "time_conflict": "같은 직원/날짜 스케줄과 충돌합니다.",
}

SCHEDULE_TEMPLATE_DUTY_ALIASES = {
    "day": "day",
    "daytime": "day",
    "주간근무": "day",
    "주간": "day",
    "overtime": "overtime",
    "ot": "overtime",
    "초과근무": "overtime",
    "night": "night",
    "nighttime": "night",
    "야간근무": "night",
    "야간": "night",
}
SCHEDULE_TEMPLATE_DUTY_LABELS = {
    "day": "주간근무",
    "overtime": "초과근무",
    "night": "야간근무",
}
SHIFT_TYPE_BY_DUTY_TYPE = {
    "day": "day",
    "overtime": "overtime",
    "night": "night",
}
ARLS_MONTHLY_BASE_UPLOAD_SOURCE = "arls_monthly_base_upload"
ARLS_MONTHLY_BASE_SOURCE_ALIASES = {
    ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
    "workbook_import",
}
ARLS_DAYTIME_NEED_SOURCE_ALIASES = {
    "monthly_workbook",
    ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
}
SENTRIX_SUPPORT_REQUEST_WORKFLOW = "arls_monthly_base_upload"
SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS = "active"
SENTRIX_SUPPORT_REQUEST_RETRACTED_STATUS = "retracted"
ARLS_SHEET_NAME = "본사 스케쥴 양식"
ARLS_DATE_START_COL = 4   # D
ARLS_DATE_END_COL = 34    # AH
ARLS_MIN_DATA_ROW = 5
ARLS_SUMMARY_START_COL = 35  # AI
ARLS_SUMMARY_END_COL = 37    # AK
ARLS_METADATA_SHEET_NAME = "_ARLS_EXPORT_META"
ARLS_SUPPORT_METADATA_SHEET_NAME = "_ARLS_SUPPORT_META"
SENTRIX_SUPPORT_HQ_METADATA_SHEET_NAME = "_SENTRIX_SUPPORT_HQ_META"
ARLS_EXPORT_TEMPLATE_VERSION = "monthly_schedule_template.v5"
ARLS_EXPORT_SOURCE_VERSION = "schedule_export.phase2.roundtrip"
ARLS_SUPPORT_FORM_VERSION = "support_roundtrip.phase3.v1"
SENTRIX_HQ_ROSTER_ASSIGNMENT_SOURCE = "HQ_ROUNDTRIP"
SENTRIX_HQ_ROSTER_AUTO_APPROVED_STATUS = "auto_approved"
SENTRIX_HQ_ROSTER_PENDING_STATUS = "approval_pending"
SENTRIX_HQ_ROSTER_FINAL_APPROVED_STATE = "approved"
SENTRIX_HQ_ROSTER_NOTIFICATION_MESSAGE = "[HQ] 지원근무자 업데이트 발생"
SENTRIX_ARLS_BRIDGE_SOURCE = "sentrix_support_ticket"
SENTRIX_ARLS_BRIDGE_ACTION_UPSERT = "UPSERT"
SENTRIX_ARLS_BRIDGE_ACTION_RETRACT = "RETRACT"
SENTRIX_SUPPORT_MATERIALIZATION_STATUS_ACTIVE = "active"
SENTRIX_SUPPORT_MATERIALIZATION_STATUS_RETRACTED = "retracted"
SENTRIX_SUPPORT_MATERIALIZATION_MODE_OWNED = "owned_schedule"
SENTRIX_SUPPORT_MATERIALIZATION_MODE_LINKED = "linked_existing_schedule"
SENTRIX_HQ_ROSTER_MULTI_PERSON_PATTERN = re.compile(r"[\r\n,;/|&]|(?:\s+\+\s+)")
SENTRIX_HQ_ROSTER_SELF_STAFF_PATTERN = re.compile(r"^자체\s+(\S+)$")
ARLS_SUPPORTED_IMPORT_SOURCE_VERSIONS = {
    "schedule_export.phase2.roundtrip",
}
ARLS_TEMPLATE_SEARCH_PATHS = (
    Path(__file__).resolve().parents[2] / "templates" / "monthly_schedule_template.xlsx",
    Path(__file__).resolve().parents[2] / "templates" / "월간 근무표 템플릿 예시.xlsx",
)
ARLS_SUPPORT_WEEKLY_LABEL = "주간\n지원 근무자"
ARLS_SUPPORT_NIGHT_LABEL = "야간 \n지원 근무자\n"
ARLS_TEMPLATE_FAMILY_LABEL = "ARLS 월간 근무표 workbook"
ARLS_ADDITIONAL_DAY_KEYWORDS = {
    "주간 추가 근무자",
    "주간추가근무자",
    "주간 지원 근무자",
    "주간지원근무자",
}
ARLS_ADDITIONAL_NIGHT_KEYWORDS = {
    "야간 근무자",
    "야간근무자",
    "야간 지원 근무자",
    "야간지원근무자",
}
ARLS_VENDOR_COUNT_KEYWORDS = {
    "업체투입 수",
    "외부인원 투입 수",
    "외부인원투입수",
    "외부 인원 투입 수",
}
ARLS_NEED_COUNT_KEYWORDS = {"필요인원 수"}
ARLS_WORK_NOTE_KEYWORDS = {"작업 내용", "작업 목적"}
ARLS_DAY_SUPPORT_BLOCK_ALIASES = {
    "주간 추가 근무자",
    "주간 지원 근무자",
    "주간 근무자(직원) 수",
}
ARLS_NIGHT_SUPPORT_BLOCK_ALIASES = {
    "야간 추가 근무자",
    "야간 지원 근무자",
    "야간 근무자",
    "야간 근무자 총 수",
}
ARLS_IGNORE_VALUE_TOKENS = {
    "업체투입 수",
    "외부인원 투입 수",
    "외부 인원 투입 수",
    "필요인원 수",
    "총 수",
    "작업 내용",
    "작업 목적",
    "critical incidents",
}
ARLS_LEAVE_MARKERS = {"연차", "휴가", "반차", "휴무", "공휴일"}
ARLS_WORKER_ROW_PATTERN = re.compile(r"^근무자\s*(\d+)$")
ARLS_MULTI_PERSON_SPLIT_PATTERN = re.compile(r"[\r\n]+|[;,/&]|(?:\s+\+\s+)|(?:\s{2,})")
ARLS_REQUIRED_SECTION_ROWS = (
    ("day_support", "주간 지원 블록"),
    ("night_support", "야간 지원 블록"),
)
SCHEDULE_IMPORT_MAPPING_ROW_TYPES = {"day", "overtime", "night"}
SCHEDULE_IMPORT_ISSUE_CATALOG: dict[str, dict[str, str]] = {
    "TEMPLATE_FAMILY_MISMATCH": {
        "severity": "blocking",
        "message": "지원하지 않는 월간 근무표 템플릿입니다.",
        "guidance": "현재 ARLS 빈 양식 또는 최신 기준본으로 다시 시작하세요.",
    },
    "TEMPLATE_REVISION_STALE": {
        "severity": "blocking",
        "message": "업로드한 파일이 현재 ARLS 기준보다 오래되었습니다.",
        "guidance": "최신 기준본을 다시 다운로드한 뒤 수정본을 업로드하세요.",
    },
    "SECTION_NOT_FOUND": {
        "severity": "blocking",
        "message": "필수 섹션을 찾지 못했습니다.",
        "guidance": "현재 지원 템플릿 family의 workbook인지 확인하세요.",
    },
    "DATE_HEADER_PARSE_FAILED": {
        "severity": "blocking",
        "message": "날짜 헤더를 읽지 못했습니다.",
        "guidance": "월간 근무표의 날짜 헤더가 손상되지 않았는지 확인하세요.",
    },
    "EMPLOYEE_ROW_GROUP_INVALID": {
        "severity": "blocking",
        "message": "직원 3행 묶음 구조가 올바르지 않습니다.",
        "guidance": "주간근무/초과근무/야간근무 3행 구조를 유지하세요.",
    },
    "TEMPLATE_MAPPING_MISSING": {
        "severity": "blocking",
        "message": "시간값에 대응하는 근무 템플릿 매핑이 없습니다.",
        "guidance": "테넌트 import mapping profile에 (행 유형, 시간) 매핑을 준비하세요.",
    },
    "MULTI_PERSON_CELL": {
        "severity": "blocking",
        "message": "한 셀에 2명 이상이 입력되어 있습니다.",
        "guidance": "근무자 1셀에는 1명만 입력하세요.",
    },
    "WORKER_CELL_INVALID": {
        "severity": "blocking",
        "message": "근무자 셀 형식이 올바르지 않습니다.",
        "guidance": "지원근무자 셀에는 1명의 이름 또는 업체/이름 한 건만 입력하세요.",
    },
    "SUPPORT_BLOCK_REQUIRED_COUNT_INVALID": {
        "severity": "blocking",
        "message": "필요 인원 수를 숫자로 해석할 수 없습니다.",
        "guidance": "필요인원 수는 숫자 또는 n인 형식으로 입력하세요.",
    },
    "PROTECTED_FIELD_IGNORED": {
        "severity": "warning",
        "message": "시스템 관리 영역은 이번 단계에서 반영되지 않습니다.",
        "guidance": "해당 영역은 검토만 가능하며 실제 연동은 후속 단계에서 처리됩니다.",
    },
    "UNSUPPORTED_CELL_FORMAT": {
        "severity": "blocking",
        "message": "지원하지 않는 셀 형식입니다.",
        "guidance": "숫자, 공란, 휴가 마커, 단일 텍스트만 사용하세요.",
    },
    "CANNOT_RESOLVE_TEMPLATE": {
        "severity": "blocking",
        "message": "매핑된 템플릿을 해석할 수 없습니다.",
        "guidance": "mapping profile 대상 템플릿이 현재 tenant/site에서 유효한지 확인하세요.",
    },
    "EMPLOYEE_MATCH_FAILED": {
        "severity": "blocking",
        "message": "직원을 현재 지점 active 직원과 매칭할 수 없습니다.",
        "guidance": "직원명 오탈자 또는 현장 소속을 확인하세요.",
    },
    "EMPLOYEE_MATCH_AMBIGUOUS": {
        "severity": "blocking",
        "message": "동일 이름의 active 직원이 여러 명 있어 자동 매칭할 수 없습니다.",
        "guidance": "동일 지점 내 동명이인 정리 후 다시 업로드하세요.",
    },
    "NON_BASE_LINEAGE_CONFLICT": {
        "severity": "blocking",
        "message": "같은 슬롯에 다른 lineage의 스케줄이 있어 base upload로 덮어쓸 수 없습니다.",
        "guidance": "수동/다른 source 스케줄을 먼저 정리한 뒤 다시 분석하세요.",
    },
    "TEMPLATE_PROFILE_NOT_PREPARED": {
        "severity": "blocking",
        "message": "테넌트 import mapping profile이 준비되지 않았습니다.",
        "guidance": "매핑 프로필과 필수 entry를 먼저 준비하세요.",
    },
}

ARLS_THIN_SIDE = Side(style="thin", color="D0D7E2")
ARLS_HEADER_FILL = PatternFill("solid", fgColor="F3F5F8")
ARLS_SUBHEADER_FILL = PatternFill("solid", fgColor="FAFBFC")
ARLS_DAY_ROW_FILL = PatternFill("solid", fgColor="F7FBF8")
ARLS_OT_ROW_FILL = PatternFill("solid", fgColor="FFF8EF")
ARLS_NIGHT_ROW_FILL = PatternFill("solid", fgColor="F9F5FF")
ARLS_DAY_VALUE_FILL = PatternFill("solid", fgColor="C2D6EC")
ARLS_OT_VALUE_FILL = PatternFill("solid", fgColor="FDF4D0")
ARLS_NIGHT_VALUE_FILL = PatternFill("solid", fgColor="F0C9AF")
ARLS_ANNUAL_LEAVE_FILL = PatternFill("solid", fgColor="4FAEEA")
ARLS_HIRE_MARKER_FILL = PatternFill("solid", fgColor="4FAD5B")
ARLS_LEAVE_MARKER_FILL = PatternFill("solid", fgColor="EA3323")
ARLS_HIRE_MARKER_FONT_COLOR = "000000"
ARLS_LEAVE_MARKER_FONT_COLOR = "FFFFFF"
ARLS_SUPPORT_WEEKLY_MAX_SLOTS = 6
ARLS_SUPPORT_NIGHT_MAX_SLOTS = 5


def _normalize_schedule_template_duty_type(value: str | None) -> str:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return ""
    return SCHEDULE_TEMPLATE_DUTY_ALIASES.get(normalized, normalized)


def _schedule_template_duty_label(value: str | None) -> str:
    duty_type = _normalize_schedule_template_duty_type(value)
    return SCHEDULE_TEMPLATE_DUTY_LABELS.get(duty_type, duty_type or "-")


def _normalize_template_label_token(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def _label_contains_any(value: object, keywords: set[str]) -> bool:
    token = _normalize_template_label_token(value)
    return any(_normalize_template_label_token(keyword) in token for keyword in keywords)


def _normalize_import_issue_code(value: str | None) -> str:
    return str(value or "").strip().upper()


def _get_import_issue_meta(code: str | None) -> dict[str, str]:
    normalized = _normalize_import_issue_code(code)
    return dict(SCHEDULE_IMPORT_ISSUE_CATALOG.get(normalized) or {})


def _issue_is_blocking(code: str | None) -> bool:
    meta = _get_import_issue_meta(code)
    return str(meta.get("severity") or "blocking").strip().lower() == "blocking"


def _excel_col_label(col_idx: int | None) -> str | None:
    if not col_idx:
        return None
    try:
        return get_column_letter(int(col_idx))
    except Exception:
        return None


def _build_import_issue(
    code: str,
    *,
    message: str | None = None,
    severity: str | None = None,
    guidance: str | None = None,
    sheet_name: str | None = None,
    row_no: int | None = None,
    col_no: int | None = None,
    section: str | None = None,
) -> dict[str, Any]:
    normalized = _normalize_import_issue_code(code)
    meta = _get_import_issue_meta(normalized)
    return {
        "code": normalized,
        "severity": str(severity or meta.get("severity") or "blocking").strip().lower(),
        "message": str(message or meta.get("message") or normalized).strip(),
        "guidance": str(guidance or meta.get("guidance") or "").strip() or None,
        "count": 1,
        "example_rows": [int(row_no)] if row_no else [],
        "location": {
            "sheet": str(sheet_name or "").strip() or None,
            "row": int(row_no) if row_no else None,
            "col": int(col_no) if col_no else None,
            "col_label": _excel_col_label(col_no),
            "section": str(section or "").strip() or None,
        },
    }


def _normalize_hours_key(value: float | int | str | None) -> str:
    try:
        return f"{float(value):.2f}"
    except (TypeError, ValueError):
        return ""


def _row_type_hours_mapping_key(row_type: str | None, numeric_hours: float | int | str | None) -> tuple[str, str]:
    return (_normalize_schedule_template_duty_type(row_type), _normalize_hours_key(numeric_hours))


def _is_worker_slot_label(value: object) -> bool:
    return bool(ARLS_WORKER_ROW_PATTERN.match(str(value or "").strip()))


def _parse_worker_slot_number(value: object) -> int | None:
    match = ARLS_WORKER_ROW_PATTERN.match(str(value or "").strip())
    if not match:
        return None
    try:
        return int(match.group(1))
    except Exception:
        return None


def _support_cell_has_multiple_people(text: str) -> bool:
    parts = [item.strip() for item in re.split(ARLS_MULTI_PERSON_SPLIT_PATTERN, str(text or "").strip()) if item and item.strip()]
    return len(parts) > 1


def _parse_support_worker_cell(value: object) -> dict[str, Any]:
    raw_text = _normalize_workbook_display_value(value)
    if not raw_text:
        return {
            "raw_value": "",
            "semantic_type": "blank",
            "is_filled": False,
            "issue_code": None,
            "issue_message": None,
        }
    if raw_text.startswith("#"):
        return {
            "raw_value": raw_text,
            "semantic_type": "invalid",
            "is_filled": True,
            "issue_code": "WORKER_CELL_INVALID",
            "issue_message": "수식 오류 값은 지원근무자 셀에 사용할 수 없습니다.",
        }
    if _support_cell_has_multiple_people(raw_text):
        return {
            "raw_value": raw_text,
            "semantic_type": "multi_person",
            "is_filled": True,
            "issue_code": "MULTI_PERSON_CELL",
            "issue_message": "한 셀에 2명 이상이 입력되어 있습니다.",
        }
    if raw_text in ARLS_IGNORE_VALUE_TOKENS:
        return {
            "raw_value": raw_text,
            "semantic_type": "invalid",
            "is_filled": True,
            "issue_code": "WORKER_CELL_INVALID",
            "issue_message": "근무자 셀에 라벨 값이 입력되어 있습니다.",
        }
    return {
        "raw_value": raw_text,
        "semantic_type": "single_person",
        "is_filled": True,
        "issue_code": None,
        "issue_message": None,
    }


def _parse_support_count_value(value: object) -> tuple[int | None, str]:
    text = _normalize_workbook_display_value(value)
    if not text:
        return 0, ""
    numeric = _parse_numeric_hours(text)
    if numeric is not None:
        return max(0, int(numeric)), text
    match = re.search(r"(\d+)\s*(?:인|명|건)?", text)
    if match and match.group(1):
        return max(0, int(match.group(1))), text
    if text in {"-", "없음", "0건", "0", "없음(0)"}:
        return 0, text
    return None, text


def _classify_import_body_semantic_type(value: object) -> tuple[str, float | None]:
    text = _normalize_workbook_display_value(value)
    if not text:
        return "blank", None
    if text in ARLS_LEAVE_MARKERS:
        return "leave_marker", None
    numeric = _parse_numeric_hours(text)
    if numeric is not None:
        return "numeric_hours", float(numeric)
    return "invalid", None


def _normalize_time_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d{1,2}:\d{2}", text):
        hour, minute = text.split(":")
        hh = int(hour)
        mm = int(minute)
        if hh == 24 and mm == 0:
            return "24:00:00"
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}:00"
    if re.fullmatch(r"\d{2}:\d{2}:\d{2}", text):
        hh, mm, ss = [int(part) for part in text.split(":")]
        if hh == 24 and mm == 0 and ss == 0:
            return "24:00:00"
        if 0 <= hh <= 23 and 0 <= mm <= 59 and 0 <= ss <= 59:
            return f"{hh:02d}:{mm:02d}:{ss:02d}"
    return None


def _format_time_range(start_time: object, end_time: object) -> str:
    start_text = _normalize_time_text(start_time)
    end_text = _normalize_time_text(end_time)
    if start_text and end_text:
        return f"{start_text[:5]}-{end_text[:5]}"
    return ""


def _time_text_to_minutes(value: object) -> int | None:
    normalized = _normalize_time_text(value)
    if not normalized:
        return None
    try:
        hour_text, minute_text, _ = normalized.split(":")
        hours = int(hour_text)
        minutes = int(minute_text)
    except (TypeError, ValueError):
        return None
    if hours == 24 and minutes == 0:
        return 1440
    if not (0 <= hours <= 23 and 0 <= minutes <= 59):
        return None
    return (hours * 60) + minutes


def _coerce_float_or_none(value: object) -> float | None:
    try:
        if value in (None, ""):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _hours_match(left: float | None, right: float | None, *, tolerance: float = 0.001) -> bool:
    if left is None or right is None:
        return False
    return abs(float(left) - float(right)) <= tolerance


def _infer_canonical_shift_hours(start_time: object, end_time: object) -> float | None:
    start_minutes = _time_text_to_minutes(start_time)
    end_minutes = _time_text_to_minutes(end_time)
    if start_minutes is None or end_minutes is None:
        return None
    normalized_end = end_minutes
    if normalized_end <= start_minutes:
        normalized_end += 1440
    duration_hours = (normalized_end - start_minutes) / 60.0
    return duration_hours if duration_hours > 0 else None


def _derive_time_range_from_start_hours(start_time: object, hours: object) -> tuple[str | None, str | None, float | None]:
    normalized_start = _normalize_time_text(start_time)
    normalized_hours = _coerce_float_or_none(hours)
    start_minutes = _time_text_to_minutes(normalized_start)
    if not normalized_start or start_minutes is None or normalized_hours is None or normalized_hours <= 0:
        return None, None, None
    rounded_minutes = int(round(normalized_hours * 60))
    if rounded_minutes <= 0:
        return None, None, None
    absolute_end_minutes = start_minutes + rounded_minutes
    wrapped_end_minutes = absolute_end_minutes % 1440
    if wrapped_end_minutes == 0 and absolute_end_minutes > start_minutes:
        normalized_end = "24:00:00"
    else:
        normalized_end = f"{wrapped_end_minutes // 60:02d}:{wrapped_end_minutes % 60:02d}:00"
    return normalized_start, normalized_end, normalized_hours


def _resolve_day_role_default_range(row: dict) -> tuple[str, str, float]:
    soc_role = str(row.get("soc_role") or "").strip().lower()
    duty_role = str(row.get("duty_role") or "").strip().upper()
    if soc_role in {"supervisor", "vice_supervisor", "hq_admin"}:
        return SUPERVISOR_DAY_SHIFT_START, SUPERVISOR_DAY_SHIFT_END, SUPERVISOR_DAY_SHIFT_HOURS
    if duty_role in {TEAM_MANAGER_DUTY_ROLE, VICE_SUPERVISOR_DUTY_ROLE}:
        return SUPERVISOR_DAY_SHIFT_START, SUPERVISOR_DAY_SHIFT_END, SUPERVISOR_DAY_SHIFT_HOURS
    return GUARD_DAY_SHIFT_START, GUARD_DAY_SHIFT_END, GUARD_DAY_SHIFT_HOURS


def _resolve_canonical_schedule_time(row: dict) -> dict[str, Any]:
    shift_type = _normalize_shift_type(row.get("shift_type"))
    duty_type = _normalize_schedule_template_duty_type(row.get("duty_type"))
    if duty_type not in {"day", "night", "overtime"}:
        if shift_type == "night":
            duty_type = "night"
        elif shift_type == "overtime":
            duty_type = "overtime"
        else:
            duty_type = "day"

    row_start = _normalize_time_text(row.get("shift_start_time"))
    row_end = _normalize_time_text(row.get("shift_end_time"))
    template_start = _normalize_time_text(row.get("template_start_time") or row.get("start_time"))
    template_end = _normalize_time_text(row.get("template_end_time") or row.get("end_time"))

    row_hours = _infer_canonical_shift_hours(row_start, row_end) if row_start and row_end else None
    template_hours = _infer_canonical_shift_hours(template_start, template_end) if template_start and template_end else None
    explicit_hours = _coerce_float_or_none(row.get("paid_hours"))
    template_paid_hours = _coerce_float_or_none(row.get("template_paid_hours"))

    def _valid_day_range(start_value: str | None, end_value: str | None, hours_value: float | None) -> bool:
        if not start_value or not end_value or hours_value is None:
            return False
        start_minutes = _time_text_to_minutes(start_value)
        end_minutes = _time_text_to_minutes(end_value)
        if start_minutes is None or end_minutes is None:
            return False
        return end_minutes > start_minutes and 0 < hours_value <= 16

    def _valid_night_or_ot_range(start_value: str | None, end_value: str | None, hours_value: float | None) -> bool:
        return bool(start_value and end_value and hours_value is not None and 0 < hours_value <= 16)

    start_time = None
    end_time = None
    hours = None
    source = "none"

    if duty_type == "day":
        if _valid_day_range(row_start, row_end, row_hours):
            start_time, end_time, hours, source = row_start, row_end, row_hours, "row"
        elif row_start and explicit_hours is not None and 0 < explicit_hours <= 16:
            start_time, end_time, hours = _derive_time_range_from_start_hours(row_start, explicit_hours)
            source = "row_hours" if start_time and end_time else "none"
        elif _valid_day_range(template_start, template_end, template_hours):
            start_time, end_time, hours, source = template_start, template_end, template_hours, "template"
        elif template_start and template_paid_hours is not None and 0 < template_paid_hours <= 16:
            start_time, end_time, hours = _derive_time_range_from_start_hours(template_start, template_paid_hours)
            source = "template_hours" if start_time and end_time else "none"
        else:
            default_start, default_end, default_hours = _resolve_day_role_default_range(row)
            if row_start or row_end or template_start or template_end or explicit_hours is not None or template_paid_hours is not None:
                start_time, end_time, hours, source = default_start, default_end, default_hours, "role_default"
    elif duty_type in {"night", "overtime"} or shift_type == "night":
        if _valid_night_or_ot_range(row_start, row_end, row_hours):
            start_time, end_time, hours, source = row_start, row_end, row_hours, "row"
        elif row_start and explicit_hours is not None and 0 < explicit_hours <= 16:
            start_time, end_time, hours = _derive_time_range_from_start_hours(row_start, explicit_hours)
            source = "row_hours" if start_time and end_time else "none"
        elif _valid_night_or_ot_range(template_start, template_end, template_hours):
            start_time, end_time, hours, source = template_start, template_end, template_hours, "template"
        elif template_start and template_paid_hours is not None and 0 < template_paid_hours <= 16:
            start_time, end_time, hours = _derive_time_range_from_start_hours(template_start, template_paid_hours)
            source = "template_hours" if start_time and end_time else "none"
        elif row_start:
            start_time, end_time, hours = _derive_time_range_from_start_hours(row_start, DEFAULT_NIGHT_SHIFT_HOURS)
            source = "row_default" if start_time and end_time else "none"
        elif template_start:
            start_time, end_time, hours = _derive_time_range_from_start_hours(template_start, DEFAULT_NIGHT_SHIFT_HOURS)
            source = "template_default" if start_time and end_time else "none"
        elif shift_type == "night":
            start_time, end_time, hours, source = (
                DEFAULT_NIGHT_SHIFT_START,
                DEFAULT_NIGHT_SHIFT_END,
                DEFAULT_NIGHT_SHIFT_HOURS,
                "night_default",
            )

    if hours is None:
        hours = _infer_canonical_shift_hours(start_time, end_time) if start_time and end_time else None
    if hours is None:
        if start_time and end_time:
            hours = _infer_export_hours_from_range(start_time, end_time)
        elif explicit_hours is not None:
            hours = explicit_hours
        elif template_paid_hours is not None:
            hours = template_paid_hours

    if source == "row" and template_start and template_end and explicit_hours is not None:
        if not _hours_match(hours, explicit_hours) and _hours_match(template_hours, explicit_hours):
            start_time, end_time, hours, source = template_start, template_end, template_hours, "template"

    if not start_time or not end_time:
        label = _shift_label(str(row.get("shift_type") or ""))
    else:
        label = _format_time_range(start_time, end_time)

    return {
        "start_time": start_time,
        "end_time": end_time,
        "hours": hours,
        "label": label,
        "source": source,
        "duty_type": duty_type,
        "shift_type": shift_type,
    }


def _support_roundtrip_normalize_role(user: dict | None) -> str:
    return normalize_user_role((user or {}).get("role"))


def _can_use_support_roundtrip_source(user: dict | None) -> bool:
    return _support_roundtrip_normalize_role(user) in {"developer", "hq_admin", "supervisor", "vice_supervisor"}


def _can_use_support_roundtrip_hq(user: dict | None) -> bool:
    return _support_roundtrip_normalize_role(user) in {"developer", "hq_admin"}


def _can_use_support_roundtrip_final_download(user: dict | None) -> bool:
    return _support_roundtrip_normalize_role(user) in {"developer", "hq_admin", "supervisor", "vice_supervisor"}


def _finance_submission_normalize_role(user: dict | None) -> str:
    return normalize_user_role((user or {}).get("role"))


def _can_view_finance_submission(user: dict | None) -> bool:
    return _finance_submission_normalize_role(user) in {"developer", "hq_admin", "supervisor", "vice_supervisor"}


def _can_download_finance_review(user: dict | None) -> bool:
    return _finance_submission_normalize_role(user) in {"developer", "hq_admin"}


def _can_upload_finance_final(user: dict | None) -> bool:
    return _finance_submission_normalize_role(user) in {"developer", "supervisor"}


def _can_download_finance_final(user: dict | None) -> bool:
    return _finance_submission_normalize_role(user) in {"developer", "hq_admin"}


def _build_schedule_overlap_range(start_time: object, end_time: object) -> tuple[int, int] | None:
    start_minutes = _time_text_to_minutes(start_time)
    end_minutes = _time_text_to_minutes(end_time)
    if start_minutes is None or end_minutes is None:
        return None
    normalized_end = end_minutes
    if normalized_end <= start_minutes:
        normalized_end += 1440
    return start_minutes, normalized_end


def _schedule_time_ranges_overlap(
    left_start: object,
    left_end: object,
    right_start: object,
    right_end: object,
) -> bool:
    left_range = _build_schedule_overlap_range(left_start, left_end)
    right_range = _build_schedule_overlap_range(right_start, right_end)
    if not left_range or not right_range:
        return True
    return left_range[0] < right_range[1] and right_range[0] < left_range[1]


def _schedule_time_ranges_do_not_overlap(
    left_start: object,
    left_end: object,
    right_start: object,
    right_end: object,
) -> bool:
    left_range = _build_schedule_overlap_range(left_start, left_end)
    right_range = _build_schedule_overlap_range(right_start, right_end)
    if not left_range or not right_range:
        return False
    return left_range[1] <= right_range[0] or right_range[1] <= left_range[0]


def _merge_board_items_for_calendar(items: list[dict]) -> list[dict]:
    rows = [dict(item) for item in (items or []) if isinstance(item, dict)]
    if len(rows) < 2:
        return rows

    def _minutes_to_time_text(value: int | None) -> str | None:
        if value is None:
            return None
        normalized = int(value) % 1440
        return f"{normalized // 60:02d}:{normalized % 60:02d}:00"

    def _employee_key(item: dict) -> str:
        return str(
            item.get("employee_id")
            or item.get("employee_code")
            or item.get("employee_name")
            or ""
        ).strip().upper()

    def _site_key(item: dict) -> str:
        return str(item.get("site_code") or item.get("site_name") or "").strip().upper()

    def _start_minutes(item: dict) -> int:
        value = _time_text_to_minutes(item.get("start_time"))
        return value if value is not None else 10**9

    rows.sort(key=lambda item: (_site_key(item), _employee_key(item), _start_minutes(item)))

    merged: list[dict] = []
    index = 0
    while index < len(rows):
        current = rows[index]
        next_item = rows[index + 1] if index + 1 < len(rows) else None
        current_shift = _normalize_shift_type(str(current.get("shift_type") or ""))
        if (
            next_item
            and str(current.get("status") or "").strip().lower() != "leave"
            and str(next_item.get("status") or "").strip().lower() != "leave"
            and _employee_key(current)
            and _employee_key(current) == _employee_key(next_item)
            and _site_key(current)
            and _site_key(current) == _site_key(next_item)
        ):
            next_shift = _normalize_shift_type(str(next_item.get("shift_type") or ""))
            shift_pair = {current_shift, next_shift}
            if shift_pair == {"day", "night"} and _schedule_time_ranges_do_not_overlap(
                current.get("start_time"),
                current.get("end_time"),
                next_item.get("start_time"),
                next_item.get("end_time"),
            ):
                current_start = _time_text_to_minutes(current.get("start_time"))
                next_start = _time_text_to_minutes(next_item.get("start_time"))
                current_end_range = _build_schedule_overlap_range(current.get("start_time"), current.get("end_time"))
                next_end_range = _build_schedule_overlap_range(next_item.get("start_time"), next_item.get("end_time"))
                merged_start = min(
                    value for value in [current_start, next_start] if value is not None
                )
                merged_end = max(
                    value for value in [
                        current_end_range[1] if current_end_range else None,
                        next_end_range[1] if next_end_range else None,
                    ] if value is not None
                )
                base = current if current_shift == "day" else next_item
                merged.append(
                    {
                        **base,
                        "schedule_id": "",
                        "shift_type": "day",
                        "display_variant": "combined",
                        "display_shift_types": ["day", "night"],
                        "combined_schedule_ids": [
                            str(current.get("schedule_id") or "").strip(),
                            str(next_item.get("schedule_id") or "").strip(),
                        ],
                        "shift_label": "주간+야간",
                        "start_time": _format_time_for_response(
                            _minutes_to_time_text(merged_start)
                        ),
                        "end_time": _format_time_for_response(
                            _minutes_to_time_text(merged_end)
                        ),
                    }
                )
                index += 2
                continue
        merged.append(current)
        index += 1
    return merged


def _resolve_schedule_row_conflict_range(row: dict | None) -> tuple[str | None, str | None]:
    payload = dict(row or {})
    canonical = _resolve_canonical_schedule_time(payload)
    start_time = _normalize_time_text(canonical.get("start_time")) or _normalize_time_text(payload.get("shift_start_time"))
    end_time = _normalize_time_text(canonical.get("end_time")) or _normalize_time_text(payload.get("shift_end_time"))
    return start_time, end_time


def _has_existing_schedule_time_conflict(
    cur,
    *,
    tenant_id: str,
    employee_id: str,
    schedule_date: date,
    shift_start_time: str | None,
    shift_end_time: str | None,
) -> bool:
    cur.execute(
        """
        SELECT ms.shift_start_time,
               ms.shift_end_time,
               ms.shift_type,
               ms.paid_hours,
               ms.template_id,
               st.duty_type,
               st.start_time AS template_start_time,
               st.end_time AS template_end_time,
               st.paid_hours AS template_paid_hours
        FROM monthly_schedules ms
        LEFT JOIN schedule_templates st ON st.id = ms.template_id
        WHERE ms.tenant_id = %s
          AND ms.employee_id = %s
          AND ms.schedule_date = %s
        """,
        (tenant_id, employee_id, schedule_date),
    )
    rows = cur.fetchall()
    if not rows:
        return False
    if not shift_start_time or not shift_end_time:
        return True
    for row in rows:
        row_start_time, row_end_time = _resolve_schedule_row_conflict_range(row)
        if _schedule_time_ranges_overlap(
            shift_start_time,
            shift_end_time,
            row_start_time,
            row_end_time,
        ):
            return True
    return False


def _load_existing_schedule_rows_for_dates(
    cur,
    *,
    tenant_id: str,
    employee_id: str,
    schedule_dates: list[date],
) -> dict[str, list[dict]]:
    normalized_dates = sorted({item for item in schedule_dates if isinstance(item, date)})
    if not normalized_dates:
        return {}
    cur.execute(
        """
        SELECT ms.schedule_date,
               ms.shift_start_time,
               ms.shift_end_time,
               ms.shift_type,
               ms.paid_hours,
               ms.template_id,
               st.duty_type,
               st.start_time AS template_start_time,
               st.end_time AS template_end_time,
               st.paid_hours AS template_paid_hours
        FROM monthly_schedules ms
        LEFT JOIN schedule_templates st ON st.id = ms.template_id
        WHERE ms.tenant_id = %s
          AND ms.employee_id = %s
          AND ms.schedule_date = ANY(%s::date[])
        """,
        (tenant_id, employee_id, normalized_dates),
    )
    grouped: dict[str, list[dict]] = {}
    for row in cur.fetchall():
        schedule_date_raw = row.get("schedule_date")
        schedule_date_key = schedule_date_raw.isoformat() if isinstance(schedule_date_raw, date) else str(schedule_date_raw or "").strip()
        if not schedule_date_key:
            continue
        row_payload = dict(row)
        row_start_time, row_end_time = _resolve_schedule_row_conflict_range(row_payload)
        row_payload["shift_start_time"] = row_start_time
        row_payload["shift_end_time"] = row_end_time
        grouped.setdefault(schedule_date_key, []).append(row_payload)
    return grouped


def _parse_numeric_hours(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if number < 0:
            return None
        return number
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    if re.fullmatch(r"\d+(\.\d+)?", text):
        return float(text)
    return None


def _normalize_name_token(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    return re.sub(r"\s+", "", text)


def _parse_month_text(value: object) -> tuple[int, int] | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month
    normalized = text.replace(".", "-").replace("/", "-").replace(" ", "")
    if re.fullmatch(r"\d{4}-\d{1,2}", normalized):
        year_text, month_text = normalized.split("-")
        year = int(year_text)
        month = int(month_text)
        if 1 <= month <= 12:
            return year, month
    return None


def _import_status_label(validation_code: str | None) -> str:
    if not validation_code:
        return "정상"
    code = _normalize_import_issue_code(validation_code)
    if code in {"EMPLOYEE_MATCH_FAILED", "EMPLOYEE_MATCH_FAILED"}:
        return "직원 미존재"
    if code in {"TEMPLATE_MAPPING_MISSING", "CANNOT_RESOLVE_TEMPLATE"}:
        return "템플릿 없음"
    if code == "TIME_CONFLICT":
        return "중복 일정"
    if code in {"UNSUPPORTED_CELL_FORMAT", "WORKER_CELL_INVALID", "SUPPORT_BLOCK_REQUIRED_COUNT_INVALID"}:
        return "잘못된 값"
    if code == "MULTI_PERSON_CELL":
        return "셀 다중 인원"
    if code == "TEMPLATE_REVISION_STALE":
        return "구버전 파일"
    return "오류"


def _import_diff_status_label(*, diff_category: str | None, validation_code: str | None, is_blocking: bool = False) -> str:
    if validation_code:
        normalized_code = _normalize_import_issue_code(validation_code)
        if normalized_code == "TEMPLATE_FAMILY_MISMATCH":
            return "메타데이터 불일치"
        if normalized_code == "TEMPLATE_REVISION_STALE":
            return "구버전 파일"
        if normalized_code == "PROTECTED_FIELD_IGNORED":
            return "보호영역 수정"
        if normalized_code == "MULTI_ROW_CONFLICT":
            return "다중 일정 충돌"
        return _import_status_label(validation_code)
    category = str(diff_category or "").strip().lower()
    if category == "create":
        return "추가"
    if category == "update":
        return "수정"
    if category == "delete":
        return "삭제"
    if category == "ignored_protected":
        return "보호영역 무시"
    if category == "unchanged":
        return "변경 없음"
    if is_blocking:
        return "검토 필요"
    return "검토"


def _normalize_workbook_display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return _format_export_hours_value(value)
    return str(value).strip()


def _metadata_sheet_to_dict(sheet) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row_idx in range(1, min(sheet.max_row, 64) + 1):
        key = str(sheet.cell(row=row_idx, column=1).value or "").strip()
        if not key:
            continue
        value = sheet.cell(row=row_idx, column=2).value
        metadata[key] = _normalize_workbook_display_value(value)
    return metadata


def _read_arls_export_metadata(workbook: Workbook) -> dict[str, str]:
    if ARLS_METADATA_SHEET_NAME not in workbook.sheetnames:
        return {}
    return _metadata_sheet_to_dict(workbook[ARLS_METADATA_SHEET_NAME])


def _is_supported_import_template_version(value: str | None) -> bool:
    text = str(value or "").strip()
    return bool(text) and text.startswith(ARLS_EXPORT_TEMPLATE_VERSION)


def _is_supported_import_source_version(value: str | None) -> bool:
    text = str(value or "").strip()
    return text in ARLS_SUPPORTED_IMPORT_SOURCE_VERSIONS


def _validate_arls_import_metadata(
    metadata: dict[str, str],
    *,
    expected_tenant_code: str,
    expected_site_code: str,
    expected_month: str,
) -> list[str]:
    errors: list[str] = []
    required_fields = (
        "tenant_code",
        "site_code",
        "month",
        "export_revision",
        "template_version",
        "export_source_version",
    )
    for field_name in required_fields:
        if not str(metadata.get(field_name) or "").strip():
            errors.append(f"metadata_missing:{field_name}")
    if errors:
        return errors
    if str(metadata.get("tenant_code") or "").strip() != expected_tenant_code:
        errors.append("metadata_mismatch:tenant_code")
    if str(metadata.get("site_code") or "").strip() != expected_site_code:
        errors.append("metadata_mismatch:site_code")
    if str(metadata.get("month") or "").strip() != expected_month:
        errors.append("metadata_mismatch:month")
    if not _is_supported_import_template_version(metadata.get("template_version")):
        errors.append("metadata_mismatch:template_version")
    if not _is_supported_import_source_version(metadata.get("export_source_version")):
        errors.append("metadata_mismatch:export_source_version")
    return errors


def _append_blocked_reason(blocked_reasons: list[str], reason: str | None) -> None:
    text = str(reason or "").strip()
    if text and text not in blocked_reasons:
        blocked_reasons.append(text)


def _summarize_import_issues(issues: list[dict[str, Any]]) -> tuple[Counter[str], list[dict[str, Any]]]:
    issue_counts: Counter[str] = Counter()
    grouped: dict[str, dict[str, Any]] = {}
    for item in issues:
        code = _normalize_import_issue_code(item.get("code"))
        if not code:
            continue
        issue_counts[code] += 1
        group = grouped.get(code)
        if not group:
            group = {
                "code": code,
                "severity": str(item.get("severity") or "blocking").strip().lower(),
                "message": str(item.get("message") or code).strip(),
                "guidance": str(item.get("guidance") or "").strip() or None,
                "count": 0,
                "example_rows": [],
                "location": dict(item.get("location") or {}),
            }
            grouped[code] = group
        group["count"] += int(item.get("count") or 1)
        for row_no in item.get("example_rows") or []:
            if row_no and row_no not in group["example_rows"] and len(group["example_rows"]) < 5:
                group["example_rows"].append(int(row_no))
    ordered_groups = sorted(
        grouped.values(),
        key=lambda item: (
            0 if str(item.get("severity") or "").lower() == "blocking" else 1,
            str(item.get("code") or ""),
        ),
    )
    return issue_counts, ordered_groups


def _detect_arls_import_workbook_context(
    workbook: Workbook,
    *,
    selected_month: str,
    expected_tenant_code: str,
    expected_site_code: str,
    current_revision: str | None,
) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    blocked_reasons: list[str] = []
    metadata = _read_arls_export_metadata(workbook)
    visible_sheet = workbook[ARLS_SHEET_NAME] if ARLS_SHEET_NAME in workbook.sheetnames else None
    if visible_sheet is None:
        issues.append(_build_import_issue("TEMPLATE_FAMILY_MISMATCH"))
        blocked_reasons.append("지원하지 않는 월간 근무표 파일입니다. ARLS 호환 workbook을 사용하세요.")
        return {
            "metadata": metadata,
            "visible_sheet": None,
            "parsed_sheet": {
                "month_ctx": None,
                "body_cells": [],
                "need_cells": [],
                "support_cells": [],
                "support_blocks": [],
                "issues": issues,
                "section_rows": {},
                "summary_start_row": None,
            },
            "workbook_kind": "unsupported",
            "workbook_valid": False,
            "is_stale": False,
            "blocked_reasons": blocked_reasons,
            "issues": issues,
            "template_family": ARLS_TEMPLATE_FAMILY_LABEL,
            "revision_status": "invalid",
        }

    parsed_sheet = _parse_arls_canonical_import_sheet(visible_sheet)
    issues.extend(parsed_sheet.get("issues") or [])
    month_ctx = parsed_sheet.get("month_ctx")
    parsed_month = f"{month_ctx[0]:04d}-{month_ctx[1]:02d}" if month_ctx else None
    workbook_kind = "template_only"
    is_stale = False

    if metadata:
        workbook_kind = "latest_base"
        if str(metadata.get("tenant_code") or "").strip() != expected_tenant_code:
            issues.append(
                _build_import_issue(
                    "TEMPLATE_FAMILY_MISMATCH",
                    message="업로드한 파일의 tenant 정보가 현재 선택값과 일치하지 않습니다.",
                    sheet_name=visible_sheet.title,
                    section="metadata",
                )
            )
            _append_blocked_reason(blocked_reasons, "업로드 파일의 tenant 정보가 현재 계정 범위와 맞지 않습니다.")
        if str(metadata.get("site_code") or "").strip() != expected_site_code:
            issues.append(
                _build_import_issue(
                    "TEMPLATE_FAMILY_MISMATCH",
                    message="업로드한 파일의 지점 정보가 현재 선택값과 일치하지 않습니다.",
                    sheet_name=visible_sheet.title,
                    section="metadata",
                )
            )
            _append_blocked_reason(blocked_reasons, "업로드 파일의 지점 정보가 현재 선택 지점과 일치하지 않습니다.")
        if str(metadata.get("month") or "").strip() != selected_month:
            issues.append(
                _build_import_issue(
                    "TEMPLATE_FAMILY_MISMATCH",
                    message="업로드한 파일의 대상월이 현재 선택값과 일치하지 않습니다.",
                    sheet_name=visible_sheet.title,
                    section="metadata",
                )
            )
            _append_blocked_reason(blocked_reasons, "업로드 파일의 대상월이 현재 선택한 월과 일치하지 않습니다.")
        if not _is_supported_import_template_version(metadata.get("template_version")):
            issues.append(
                _build_import_issue(
                    "TEMPLATE_FAMILY_MISMATCH",
                    message="지원하지 않는 템플릿 revision입니다.",
                    sheet_name=visible_sheet.title,
                    section="metadata",
                )
            )
            _append_blocked_reason(blocked_reasons, "현재 지원하지 않는 월간 근무표 템플릿 revision입니다.")
        if not _is_supported_import_source_version(metadata.get("export_source_version")):
            issues.append(
                _build_import_issue(
                    "TEMPLATE_FAMILY_MISMATCH",
                    message="지원하지 않는 workbook source version입니다.",
                    sheet_name=visible_sheet.title,
                    section="metadata",
                )
            )
            _append_blocked_reason(blocked_reasons, "현재 업로드에서 지원하지 않는 workbook source version입니다.")
        export_revision = str(metadata.get("export_revision") or "").strip()
        if current_revision and export_revision and export_revision != str(current_revision).strip():
            is_stale = True
            issues.append(
                _build_import_issue(
                    "TEMPLATE_REVISION_STALE",
                    sheet_name=visible_sheet.title,
                    section="metadata",
                )
            )
            _append_blocked_reason(blocked_reasons, "업로드한 파일이 현재 ARLS 기준보다 오래되었습니다. 최신 기준본을 다시 사용하세요.")
    if not parsed_sheet.get("body_cells"):
        issues.append(_build_import_issue("SECTION_NOT_FOUND", sheet_name=visible_sheet.title, section="base_schedule"))
        _append_blocked_reason(blocked_reasons, "기본 근무표 영역을 찾지 못했습니다.")
    if parsed_month and parsed_month != selected_month:
        issues.append(
            _build_import_issue(
                "DATE_HEADER_PARSE_FAILED",
                message="워크북 날짜 헤더의 월 정보가 현재 선택월과 다릅니다.",
                sheet_name=visible_sheet.title,
                section="date_header",
            )
        )
        _append_blocked_reason(blocked_reasons, "워크북 날짜 헤더의 월 정보가 현재 선택월과 다릅니다.")

    workbook_valid = not any(str(item.get("severity") or "").lower() == "blocking" for item in issues)
    revision_status = "stale" if is_stale else ("latest" if metadata else "template_only")
    return {
        "metadata": metadata,
        "visible_sheet": visible_sheet,
        "parsed_sheet": parsed_sheet,
        "workbook_kind": workbook_kind,
        "workbook_valid": workbook_valid,
        "is_stale": is_stale,
        "blocked_reasons": blocked_reasons,
        "issues": issues,
        "template_family": ARLS_TEMPLATE_FAMILY_LABEL,
        "revision_status": revision_status,
    }


def _parse_daytime_need_value(value: object) -> tuple[int | None, str]:
    text = _normalize_workbook_display_value(value)
    if not text:
        return 0, ""
    numeric = _parse_numeric_hours(text)
    if numeric is not None:
        return max(0, int(numeric)), text
    match = re.search(r"(\d+)\s*인", text)
    if match:
        return max(0, int(match.group(1))), text
    if text in {"-", "없음", "0건", "0"}:
        return 0, text
    return None, text


def _extract_arls_date_columns(sheet) -> tuple[dict[int, date], tuple[int, int] | None]:
    month_ctx = _parse_month_text(sheet.cell(row=2, column=2).value)
    date_map: dict[int, date] = {}
    for col in range(ARLS_DATE_START_COL, ARLS_DATE_END_COL + 1):
        header = sheet.cell(row=2, column=col).value
        resolved_date: date | None = None
        if isinstance(header, datetime):
            resolved_date = header.date()
        elif isinstance(header, date):
            resolved_date = header
        else:
            day_number: int | None = None
            if isinstance(header, (int, float)) and not isinstance(header, bool):
                day_number = int(header)
            else:
                text = str(header or "").strip()
                if text and re.fullmatch(r"\d{1,2}", text):
                    day_number = int(text)
            if day_number and month_ctx:
                year, month = month_ctx
                try:
                    resolved_date = date(year, month, day_number)
                except ValueError:
                    resolved_date = None
        if resolved_date:
            date_map[col] = resolved_date
    return date_map, month_ctx


def _normalize_shift_type(value: str | None) -> str:
    normalized = (value or "").strip().lower()
    if not normalized:
        return ""
    return SHIFT_TYPE_ALIASES.get(normalized, normalized)


def _schedule_import_headers() -> list[str]:
    return [
        "tenant_code",
        "company_code",
        "site_code",
        "employee_code",
        "schedule_date",
        "shift_type",
    ]


def _month_bounds(value: str) -> tuple[date, date]:
    try:
        start = datetime.strptime(value, "%Y-%m").date()
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail={"error": "INVALID_MONTH", "message": "month must be YYYY-MM"},
        ) from exc

    if start.month == 12:
        end = datetime(start.year + 1, 1, 1).date()
    else:
        end = datetime(start.year, start.month + 1, 1).date()
    return start, end


def _month_day_keys(start: date, end_exclusive: date) -> list[str]:
    keys: list[str] = []
    cursor = start
    while cursor < end_exclusive:
        keys.append(cursor.isoformat())
        cursor += timedelta(days=1)
    return keys


def _calendar_grid_bounds(month_key: str) -> tuple[date, date]:
    start, end = _month_bounds(month_key)
    start_js_weekday = (start.weekday() + 1) % 7
    grid_start = start - timedelta(days=start_js_weekday)
    total_days = (end - start).days
    slot_count = start_js_weekday + total_days
    trailing_days = (7 - (slot_count % 7)) % 7
    grid_end = end + timedelta(days=trailing_days)
    return grid_start, grid_end


def _shift_label(shift_type: str | None) -> str:
    normalized = _normalize_shift_type(shift_type)
    return SHIFT_TYPE_LABELS.get(normalized, normalized or "-")


def _row_shift_label(row: dict) -> str:
    canonical = _resolve_canonical_schedule_time(row)
    custom_label = str(canonical.get("label") or "").strip()
    if custom_label:
        return custom_label
    return _shift_label(str(row.get("shift_type") or ""))


def _normalize_header(value: str | None) -> str:
    return (value or "").strip().strip("\ufeff").lower()


def _normalize_import_row(row: dict[str, object]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        header = _normalize_header(str(key))
        if value is None:
            normalized[header] = ""
            continue
        if isinstance(value, datetime):
            normalized[header] = value.date().isoformat()
            continue
        if isinstance(value, date):
            normalized[header] = value.isoformat()
            continue
        if isinstance(value, float) and value.is_integer():
            normalized[header] = str(int(value))
            continue
        normalized[header] = str(value).strip()
    return normalized


def _parse_date_or_none(value: str) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _read_import_rows(file: UploadFile, raw_bytes: bytes) -> tuple[str, list[str], list[dict[str, str]]]:
    filename = (file.filename or "").lower()
    detected = "xlsx" if filename.endswith(".xlsx") else "csv"
    if detected not in IMPORT_FORMATS:
        detected = "csv"

    if detected == "xlsx":
        workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            header_cells = next(rows_iter, None)
            if not header_cells:
                return detected, [], []
            headers = [_normalize_header(str(cell) if cell is not None else "") for cell in header_cells]
            rows: list[dict[str, str]] = []
            for values in rows_iter:
                row_map: dict[str, str] = {}
                for idx, header in enumerate(headers):
                    cell_value = values[idx] if idx < len(values) else None
                    if cell_value is None:
                        row_map[header] = ""
                    elif isinstance(cell_value, datetime):
                        row_map[header] = cell_value.date().isoformat()
                    elif isinstance(cell_value, date):
                        row_map[header] = cell_value.isoformat()
                    elif isinstance(cell_value, float) and cell_value.is_integer():
                        row_map[header] = str(int(cell_value))
                    else:
                        row_map[header] = str(cell_value).strip()
                rows.append(row_map)
            return detected, headers, rows
        finally:
            workbook.close()

    try:
        raw = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        raw = raw_bytes.decode("cp949")
    reader = csv.DictReader(io.StringIO(raw))
    raw_headers = [_normalize_header(item) for item in (reader.fieldnames or [])]
    rows = [_normalize_import_row(row) for row in reader]
    return detected, raw_headers, rows


def _resolve_import_refs(
    cur,
    user,
    tenant_code: str,
    company_code: str,
    site_code: str,
    employee_code: str,
) -> tuple[dict | None, str | None, str | None]:
    cur.execute("SELECT id FROM tenants WHERE tenant_code = %s", (tenant_code,))
    tenant = cur.fetchone()
    if not tenant:
        return None, "tenant_match_failed", VALIDATION_MESSAGES["tenant_match_failed"]

    if not is_super_admin(user["role"]) and str(tenant["id"]) != str(user["tenant_id"]):
        return None, "tenant_code_mismatch", VALIDATION_MESSAGES["tenant_code_mismatch"]

    cur.execute(
        "SELECT id FROM companies WHERE tenant_id = %s AND company_code = %s",
        (tenant["id"], company_code),
    )
    company = cur.fetchone()
    if not company:
        return None, "company_match_failed", VALIDATION_MESSAGES["company_match_failed"]

    cur.execute(
        "SELECT id FROM sites WHERE tenant_id = %s AND company_id = %s AND site_code = %s",
        (tenant["id"], company["id"], site_code),
    )
    site = cur.fetchone()
    if not site:
        return None, "site_match_failed", VALIDATION_MESSAGES["site_match_failed"]

    cur.execute(
        "SELECT id, site_id FROM employees WHERE tenant_id = %s AND employee_code = %s",
        (tenant["id"], employee_code),
    )
    employee = cur.fetchone()
    if not employee:
        return None, "employee_match_failed", VALIDATION_MESSAGES["employee_match_failed"]

    if str(employee["site_id"]) != str(site["id"]):
        return None, "site_match_failed", "직원이 해당 사이트에 배정되어 있지 않습니다."

    refs = {
        "tenant_id": tenant["id"],
        "company_id": company["id"],
        "site_id": site["id"],
        "employee_id": employee["id"],
    }
    return refs, None, None


def _parse_export_period(
    month: str | None,
    start_date: str | None,
    end_date: str | None,
) -> tuple[date, date, str]:
    if start_date or end_date:
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="start_date and end_date must be provided together")
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="start_date/end_date must be YYYY-MM-DD") from exc
        if end < start:
            raise HTTPException(status_code=400, detail="end_date must be greater than or equal to start_date")
        return start, end + timedelta(days=1), f"{start.isoformat()}_{end.isoformat()}"

    if not month:
        raise HTTPException(status_code=400, detail="month or start_date/end_date is required")

    start, next_month_start = _month_bounds(month)
    return start, next_month_start, month


def _fetch_export_rows(
    conn,
    tenant_id: str,
    period_start: date,
    period_end_exclusive: date,
    company_code: str,
    site_code: str,
) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT t.tenant_code,
                   c.company_code,
                   s.site_code,
                   e.employee_code,
                   ms.schedule_date,
                   ms.shift_type
            FROM monthly_schedules ms
            JOIN tenants t ON t.id = ms.tenant_id
            JOIN companies c ON c.id = ms.company_id
            JOIN sites s ON s.id = ms.site_id
            JOIN employees e ON e.id = ms.employee_id
            WHERE ms.tenant_id = %s
              AND ms.schedule_date >= %s
              AND ms.schedule_date < %s
              AND (%s = '' OR c.company_code = %s)
              AND (%s = '' OR s.site_code = %s)
            ORDER BY e.employee_code, ms.schedule_date
            """,
            (
                tenant_id,
                period_start,
                period_end_exclusive,
                company_code,
                company_code,
                site_code,
                site_code,
            ),
        )
        return [dict(row) for row in cur.fetchall()]


@router.get("/import/template")
def download_schedule_template(
    file_format: str = Query("xlsx", alias="format"),
    month: str | None = Query(default=None, description="YYYY-MM"),
    tenant_code: str | None = Query(default=None, max_length=64),
    site_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not (can_manage_schedule(user["role"]) or _can_use_support_roundtrip_source(user)):
        raise HTTPException(status_code=403, detail="forbidden")

    normalized_format = (file_format or "xlsx").strip().lower()
    if normalized_format != "xlsx":
        raise HTTPException(status_code=400, detail="format must be xlsx")

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    month_key = (month or datetime.utcnow().strftime("%Y-%m")).strip()
    _month_bounds(month_key)

    effective_site_code = _resolve_scoped_schedule_site_code(user, request_site_code=site_code)
    resolved_site_name = ""
    if effective_site_code:
        site_row = _resolve_site_context_by_code(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_code=effective_site_code,
        )
        if site_row:
            resolved_site_name = str(site_row.get("site_name") or "").strip()

    workbook, _template_path = _load_required_arls_month_workbook()
    _prepare_blank_arls_template_workbook(
        workbook,
        month_key=month_key,
        site_name=resolved_site_name,
    )

    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    safe_month = month_key.replace("-", "")
    filename = f"schedule_monthly_template_{safe_month}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/import/latest-base")
def download_schedule_latest_base(
    month: str = Query(..., description="YYYY-MM"),
    tenant_code: str | None = Query(default=None, max_length=64),
    site_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not (can_manage_schedule(user["role"]) or _can_use_support_roundtrip_source(user)):
        raise HTTPException(status_code=403, detail="forbidden")

    month_key = str(month or "").strip()
    _month_bounds(month_key)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    effective_site_code = _resolve_scoped_schedule_site_code(user, request_site_code=site_code)
    if not effective_site_code:
        raise HTTPException(status_code=400, detail="site_code is required for latest base download")

    scope_site = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=effective_site_code,
    )
    if not scope_site:
        raise HTTPException(status_code=404, detail="site not found")

    export_ctx = _collect_monthly_export_context(
        conn,
        target_tenant=target_tenant,
        site_row=scope_site,
        month_key=month_key,
        user=user,
    )
    workbook = export_ctx["workbook"]
    out = BytesIO()
    try:
        workbook.save(out)
    finally:
        workbook.close()
    out.seek(0)

    safe_month = month_key.replace("-", "")
    filename = f"schedule_monthly_latest_base_{effective_site_code}_{safe_month}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _lookup_refs(conn, tenant_id: str, tenant_code: str, company_code: str, site_code: str, employee_code: str):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT t.id AS tenant_id, c.id AS company_id, s.id AS site_id, e.id AS employee_id
            FROM tenants t
            JOIN companies c ON c.tenant_id = t.id
            JOIN sites s ON s.company_id = c.id
            JOIN employees e ON e.site_id = s.id
            WHERE t.id = %s
              AND t.tenant_code = %s
              AND c.company_code = %s
              AND s.site_code = %s
              AND e.employee_code = %s
            """,
            (tenant_id, tenant_code, company_code, site_code, employee_code),
        )
        return cur.fetchone()


def _resolve_target_tenant(conn, user, tenant_code: str | None):
    return resolve_scoped_tenant(
        conn,
        user,
        query_tenant_code=tenant_code,
        header_tenant_id=user.get("active_tenant_id"),
        require_dev_context=True,
    )


def _format_time_for_response(value: object) -> str | None:
    return _normalize_time_text(value)


def _format_schedule_template_row(row: dict) -> ScheduleTemplateOut:
    return ScheduleTemplateOut(
        id=row["id"],
        tenant_id=row["tenant_id"],
        template_name=str(row.get("template_name") or "").strip(),
        duty_type=_normalize_schedule_template_duty_type(row.get("duty_type")),
        start_time=_format_time_for_response(row.get("start_time")),
        end_time=_format_time_for_response(row.get("end_time")),
        paid_hours=float(row["paid_hours"]) if row.get("paid_hours") is not None else None,
        break_minutes=int(row["break_minutes"]) if row.get("break_minutes") is not None else None,
        site_id=row.get("site_id"),
        site_code=str(row.get("site_code") or "").strip() or None,
        site_name=str(row.get("site_name") or "").strip() or None,
        is_default=bool(row.get("is_default")),
        is_active=bool(row.get("is_active")),
        created_at=row.get("created_at"),
        updated_at=row.get("updated_at"),
    )


def _fetch_schedule_templates(conn, *, tenant_id: str, site_id: str | None = None, include_inactive: bool = False) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT st.id,
                   st.tenant_id,
                   st.template_name,
                   st.duty_type,
                   st.start_time,
                   st.end_time,
                   st.paid_hours,
                   st.break_minutes,
                   st.site_id,
                   s.site_code,
                   s.site_name,
                   st.is_default,
                   st.is_active,
                   st.created_at,
                   st.updated_at
            FROM schedule_templates st
            LEFT JOIN sites s ON s.id = st.site_id
            WHERE st.tenant_id = %s
              AND (%s = '' OR st.site_id::text = %s OR st.site_id IS NULL)
              AND (%s OR st.is_active = TRUE)
            ORDER BY st.is_default DESC, st.is_active DESC, st.updated_at DESC, st.template_name ASC
            """,
            (tenant_id, str(site_id or ""), str(site_id or ""), bool(include_inactive)),
        )
        return [dict(row) for row in cur.fetchall()]


def _fetch_template_by_id_for_scope(conn, *, tenant_id: str, template_id: str) -> dict | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, tenant_id, template_name, duty_type, start_time, end_time, paid_hours, break_minutes,
                   site_id, is_default, is_active
            FROM schedule_templates
            WHERE tenant_id = %s
              AND id = %s
            LIMIT 1
            """,
            (tenant_id, str(template_id)),
        )
        row = cur.fetchone()
    return dict(row) if row else None


def _fetch_active_schedule_import_mapping_profile(conn, *, tenant_id: str) -> dict[str, Any] | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT p.id,
                   p.profile_name,
                   p.is_active,
                   p.created_at,
                   p.updated_at
            FROM schedule_import_mapping_profiles p
            WHERE p.tenant_id = %s
              AND p.is_active = TRUE
            ORDER BY p.updated_at DESC, p.created_at DESC
            LIMIT 1
            """,
            (tenant_id,),
        )
        profile_row = cur.fetchone()
        if not profile_row:
            return None
        cur.execute(
            """
            SELECT e.id,
                   e.row_type,
                   e.numeric_hours,
                   e.template_id,
                   st.template_name,
                   st.site_id,
                   st.is_active AS template_is_active,
                   s.site_code AS template_site_code
            FROM schedule_import_mapping_entries e
            LEFT JOIN schedule_templates st ON st.id = e.template_id
            LEFT JOIN sites s ON s.id = st.site_id
            WHERE e.profile_id = %s
            ORDER BY
              CASE e.row_type WHEN 'day' THEN 1 WHEN 'overtime' THEN 2 WHEN 'night' THEN 3 ELSE 9 END,
              e.numeric_hours ASC
            """,
            (profile_row["id"],),
        )
        entries = [dict(row) for row in cur.fetchall()]
    return {
        "profile_id": str(profile_row["id"]),
        "profile_name": str(profile_row.get("profile_name") or "").strip(),
        "is_active": bool(profile_row.get("is_active")),
        "entries": entries,
        "created_at": profile_row.get("created_at"),
        "updated_at": profile_row.get("updated_at"),
    }


def _build_schedule_import_mapping_summary(profile: dict[str, Any] | None) -> dict[str, Any]:
    if not profile:
        return {
            "profile_id": None,
            "profile_name": None,
            "is_active": False,
            "entry_count": 0,
            "missing_required_entries": [],
            "entries": [],
        }
    entries_out: list[dict[str, Any]] = []
    for entry in profile.get("entries") or []:
        template_name = str(entry.get("template_name") or "").strip()
        template_is_active = bool(entry.get("template_is_active"))
        issue_code = None
        issue_message = None
        status = "ready"
        if not str(entry.get("template_id") or "").strip():
            issue_code = "CANNOT_RESOLVE_TEMPLATE"
            issue_message = "매핑된 템플릿이 존재하지 않습니다."
            status = "invalid"
        elif not template_name or not template_is_active:
            issue_code = "CANNOT_RESOLVE_TEMPLATE"
            issue_message = "매핑된 템플릿이 비활성 또는 조회 불가 상태입니다."
            status = "invalid"
        entries_out.append(
            {
                "row_type": _normalize_schedule_template_duty_type(entry.get("row_type")),
                "numeric_hours": float(entry["numeric_hours"]) if entry.get("numeric_hours") is not None else None,
                "template_id": str(entry.get("template_id") or "").strip() or None,
                "template_name": template_name or None,
                "template_site_code": str(entry.get("template_site_code") or "").strip() or None,
                "status": status,
                "issue_code": issue_code,
                "issue_message": issue_message,
            }
        )
    return {
        "profile_id": str(profile.get("profile_id") or "").strip() or None,
        "profile_name": str(profile.get("profile_name") or "").strip() or None,
        "is_active": bool(profile.get("is_active")),
        "entry_count": len(entries_out),
        "updated_at": profile.get("updated_at"),
        "missing_required_entries": [],
        "entries": entries_out,
    }


def _build_schedule_import_mapping_lookup(profile: dict[str, Any] | None) -> dict[tuple[str, str], dict[str, Any]]:
    lookup: dict[tuple[str, str], dict[str, Any]] = {}
    if not profile:
        return lookup
    for entry in profile.get("entries") or []:
        key = _row_type_hours_mapping_key(entry.get("row_type"), entry.get("numeric_hours"))
        if not key[0] or not key[1]:
            continue
        lookup[key] = dict(entry)
    return lookup


def _resolve_shift_type_from_duty_type(duty_type: str | None) -> str:
    normalized = _normalize_schedule_template_duty_type(duty_type)
    mapped = SHIFT_TYPE_BY_DUTY_TYPE.get(normalized, "day")
    return _normalize_shift_type(mapped)


def _template_matches_hours(template_row: dict, paid_hours: float | None) -> bool:
    if paid_hours is None:
        return False
    template_hours = template_row.get("paid_hours")
    if template_hours is None:
        return False
    try:
        return abs(float(template_hours) - float(paid_hours)) < 0.001
    except (TypeError, ValueError):
        return False


def _choose_template_by_duty_hours(
    templates: list[dict],
    *,
    duty_type: str,
    paid_hours: float | None = None,
    require_default: bool = False,
    strict_hours: bool = False,
) -> dict | None:
    normalized_duty = _normalize_schedule_template_duty_type(duty_type)
    candidates = [row for row in templates if _normalize_schedule_template_duty_type(row.get("duty_type")) == normalized_duty]
    if require_default:
        candidates = [row for row in candidates if bool(row.get("is_default"))]
    if paid_hours is not None:
        exact = [row for row in candidates if _template_matches_hours(row, paid_hours)]
        if exact:
            exact.sort(key=lambda row: (not bool(row.get("is_default")), str(row.get("template_name") or "")))
            return exact[0]
        if strict_hours:
            return None
    if not candidates:
        return None
    candidates.sort(key=lambda row: (not bool(row.get("is_default")), str(row.get("template_name") or "")))
    return candidates[0]


def _resolve_site_context_by_code(conn, *, tenant_id: str, site_code: str) -> dict | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT s.id, s.site_code, s.site_name, COALESCE(s.address, '') AS address, s.company_id, c.company_code
            FROM sites s
            JOIN companies c ON c.id = s.company_id
            WHERE s.tenant_id = %s
              AND s.site_code = %s
            LIMIT 1
            """,
            (tenant_id, site_code),
        )
        row = cur.fetchone()
    return dict(row) if row else None


def _resolve_site_context_by_id(conn, *, tenant_id: str, site_id: str) -> dict | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT s.id, s.site_code, s.site_name, COALESCE(s.address, '') AS address, s.company_id, c.company_code
            FROM sites s
            JOIN companies c ON c.id = s.company_id
            WHERE s.tenant_id = %s
              AND s.id = %s
            LIMIT 1
            """,
            (tenant_id, site_id),
        )
        row = cur.fetchone()
    return dict(row) if row else None


def _resolve_scoped_schedule_site_code(user: dict, *, request_site_code: str | None = None) -> str:
    requested_site_code = str(request_site_code or "").strip().upper()
    if requested_site_code.lower() == "all":
        requested_site_code = ""
    staff_scope = enforce_staff_site_scope(user, request_site_code=requested_site_code or None)
    if staff_scope:
        return str(staff_scope.get("site_code") or "").strip().upper()
    return requested_site_code


def _list_site_contexts_for_export(conn, *, tenant_id: str) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT s.id, s.site_code, s.site_name, COALESCE(s.address, '') AS address, s.company_id, c.company_code
            FROM sites s
            JOIN companies c ON c.id = s.company_id
            WHERE s.tenant_id = %s
            ORDER BY COALESCE(s.is_active, TRUE) DESC, s.site_code ASC, s.site_name ASC
            """,
            (tenant_id,),
        )
        return [dict(row) for row in cur.fetchall()]


def _resolve_employee_by_code(conn, *, tenant_id: str, employee_code: str) -> dict | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, tenant_id, company_id, site_id, employee_code, full_name
            FROM employees
            WHERE tenant_id = %s
              AND employee_code = %s
            LIMIT 1
            """,
            (tenant_id, employee_code),
        )
        row = cur.fetchone()
    return dict(row) if row else None


def _resolve_employee_by_id(conn, *, tenant_id: str, employee_id: str) -> dict | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, tenant_id, company_id, site_id, employee_code, full_name
            FROM employees
            WHERE tenant_id = %s
              AND id = %s
            LIMIT 1
            """,
            (tenant_id, employee_id),
        )
        row = cur.fetchone()
    return dict(row) if row else None


def _insert_monthly_schedule_row(
    cur,
    *,
    tenant_id: str,
    company_id: str,
    site_id: str,
    employee_id: str,
    schedule_date: date,
    shift_type: str,
    template_id: str | None = None,
    shift_start_time: str | None = None,
    shift_end_time: str | None = None,
    paid_hours: float | None = None,
    schedule_note: str | None = None,
    source: str | None = None,
    source_batch_id: str | None = None,
    source_revision: str | None = None,
    source_ticket_uuid: str | None = None,
    source_ticket_state: str | None = None,
    source_action: str | None = None,
    source_self_staff: bool = False,
) -> str:
    schedule_id = str(uuid.uuid4())
    cur.execute(
        """
        INSERT INTO monthly_schedules (
            id,
            tenant_id,
            company_id,
            site_id,
            employee_id,
            schedule_date,
            shift_type,
            leader_user_id,
            template_id,
            shift_start_time,
            shift_end_time,
            paid_hours,
            schedule_note,
            source,
            source_batch_id,
            source_revision,
            source_ticket_uuid,
            source_ticket_state,
            source_action,
            source_self_staff
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, NULL, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            schedule_id,
            tenant_id,
            company_id,
            site_id,
            employee_id,
            schedule_date,
            shift_type,
            template_id,
            shift_start_time,
            shift_end_time,
            paid_hours,
            schedule_note,
            source,
            source_batch_id,
            source_revision,
            source_ticket_uuid,
            source_ticket_state,
            source_action,
            bool(source_self_staff),
        ),
    )
    return schedule_id


def _update_monthly_schedule_row(
    cur,
    *,
    schedule_id: str,
    shift_type: str,
    template_id: str | None = None,
    shift_start_time: str | None = None,
    shift_end_time: str | None = None,
    paid_hours: float | None = None,
    schedule_note: str | None = None,
    source: str | None = None,
    source_batch_id: str | None = None,
    source_revision: str | None = None,
    source_ticket_uuid: str | None = None,
    source_ticket_state: str | None = None,
    source_action: str | None = None,
    source_self_staff: bool | None = None,
) -> None:
    cur.execute(
        """
        UPDATE monthly_schedules
        SET shift_type = %s,
            template_id = %s,
            shift_start_time = %s,
            shift_end_time = %s,
            paid_hours = %s,
            schedule_note = %s,
            source = %s,
            source_batch_id = %s,
            source_revision = %s,
            source_ticket_uuid = %s,
            source_ticket_state = %s,
            source_action = %s,
            source_self_staff = COALESCE(%s, source_self_staff)
        WHERE id = %s
        """,
        (
            shift_type,
            template_id,
            shift_start_time,
            shift_end_time,
            paid_hours,
            schedule_note,
            source or ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
            source_batch_id,
            source_revision,
            source_ticket_uuid,
            source_ticket_state,
            source_action,
            source_self_staff,
            schedule_id,
        ),
    )


def _delete_monthly_schedule_row(cur, *, schedule_id: str) -> None:
    cur.execute("DELETE FROM monthly_schedules WHERE id = %s", (schedule_id,))


def _upsert_daytime_need_count_row(
    cur,
    *,
    tenant_id: str,
    site_id: str,
    work_date: date,
    required_count: int,
    raw_text: str | None,
    updated_by: str,
    source: str = "monthly_workbook",
    source_batch_id: str | None = None,
    source_revision: str | None = None,
) -> None:
    cur.execute(
        """
        INSERT INTO site_daytime_need_counts (
            id, tenant_id, site_id, work_date, required_count, raw_text, source, updated_by,
            source_batch_id, source_revision
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (tenant_id, site_id, work_date)
        DO UPDATE SET
            required_count = EXCLUDED.required_count,
            raw_text = EXCLUDED.raw_text,
            source = EXCLUDED.source,
            updated_by = EXCLUDED.updated_by,
            source_batch_id = EXCLUDED.source_batch_id,
            source_revision = EXCLUDED.source_revision,
            updated_at = timezone('utc', now())
        """,
        (
            uuid.uuid4(),
            tenant_id,
            site_id,
            work_date,
            max(0, int(required_count)),
            str(raw_text or "").strip() or None,
            source,
            updated_by,
            source_batch_id,
            source_revision,
        ),
    )


def _delete_daytime_need_count_row(cur, *, tenant_id: str, site_id: str, work_date: date) -> None:
    cur.execute(
        """
        DELETE FROM site_daytime_need_counts
        WHERE tenant_id = %s
          AND site_id = %s
          AND work_date = %s
        """,
        (tenant_id, site_id, work_date),
    )


def _split_name_cells(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    lowered = text.lower()
    if lowered in ARLS_IGNORE_VALUE_TOKENS:
        return []
    normalized = text.replace("\n", ",")
    normalized = normalized.replace(";", ",")
    normalized = normalized.replace("/", ",")
    normalized = normalized.replace("|", ",")
    candidates = [item.strip() for item in normalized.split(",")]
    return [item for item in candidates if item]


def _build_employee_name_index(employee_rows: list[dict]) -> dict[str, list[dict]]:
    index: dict[str, list[dict]] = {}
    for row in employee_rows:
        token = _normalize_name_token(row.get("full_name"))
        if not token:
            continue
        index.setdefault(token, []).append(row)
    return index


def _load_site_employees(conn, *, tenant_id: str, site_id: str) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, tenant_id, company_id, site_id, employee_code, full_name,
                   sequence_no, hire_date, leave_date
            FROM employees
            WHERE tenant_id = %s
              AND site_id = %s
            ORDER BY employee_code ASC
            """,
            (tenant_id, site_id),
        )
        return [dict(row) for row in cur.fetchall()]


def _employee_is_active_for_schedule_date(row: dict[str, Any], schedule_date: date | None) -> bool:
    if not isinstance(schedule_date, date):
        return False
    hire_date = row.get("hire_date")
    leave_date = row.get("leave_date")
    if isinstance(hire_date, date) and schedule_date < hire_date:
        return False
    if isinstance(leave_date, date) and schedule_date > leave_date:
        return False
    return True


def _resolve_import_employee_match(
    employee_index: dict[str, list[dict[str, Any]]],
    *,
    employee_name: str,
    schedule_date: date | None,
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    token = _normalize_name_token(employee_name)
    matches = [dict(item) for item in (employee_index.get(token) or [])]
    if not matches:
        return None, "EMPLOYEE_MATCH_FAILED", "직원을 찾을 수 없습니다."
    active_matches = [
        item for item in matches
        if _employee_is_active_for_schedule_date(item, schedule_date)
    ]
    if len(active_matches) == 1:
        return active_matches[0], None, None
    if len(active_matches) > 1:
        return None, "EMPLOYEE_MATCH_AMBIGUOUS", "동일 이름의 active 직원이 2명 이상입니다."
    return None, "EMPLOYEE_MATCH_FAILED", "현재 날짜 기준 active 직원으로 매칭되지 않습니다."


def _load_existing_schedule_keys(
    conn,
    *,
    tenant_id: str,
    site_id: str,
    start_date: date,
    end_date: date,
) -> set[tuple[str, str]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT employee_id, schedule_date
            FROM monthly_schedules
            WHERE tenant_id = %s
              AND site_id = %s
              AND schedule_date >= %s
              AND schedule_date <= %s
            """,
            (tenant_id, site_id, start_date, end_date),
        )
        rows = cur.fetchall()
    keys: set[tuple[str, str]] = set()
    for row in rows:
        employee_id = str(row.get("employee_id") or "").strip()
        schedule_date = row.get("schedule_date")
        if not employee_id or not isinstance(schedule_date, date):
            continue
        keys.add((employee_id, schedule_date.isoformat()))
    return keys


def _parse_arls_candidate_rows(sheet) -> tuple[list[dict], tuple[int, int] | None]:
    date_columns, month_ctx = _extract_arls_date_columns(sheet)
    if not date_columns:
        raise HTTPException(status_code=400, detail="ARLS 템플릿 날짜 헤더를 읽지 못했습니다.")

    rows: list[dict] = []
    mode = "base"

    for row_idx in range(ARLS_MIN_DATA_ROW, sheet.max_row + 1):
        employee_cell = sheet.cell(row=row_idx, column=2).value
        duty_cell = sheet.cell(row=row_idx, column=3).value
        employee_name = str(employee_cell or "").strip()
        duty_text = str(duty_cell or "").strip()
        employee_name_no_space = employee_name.replace(" ", "")
        duty_type = _normalize_schedule_template_duty_type(duty_text)

        if employee_name_no_space in ARLS_ADDITIONAL_DAY_KEYWORDS:
            mode = "additional_day"
            continue
        if employee_name_no_space in ARLS_ADDITIONAL_NIGHT_KEYWORDS:
            mode = "additional_night"
            continue

        if mode == "base":
            if duty_type not in {"day", "overtime", "night"}:
                continue
            if not employee_name:
                continue
            for col_idx, schedule_date in date_columns.items():
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                value_text = str(value).strip()
                if not value_text:
                    continue
                if value_text in {"연차", "휴가"}:
                    rows.append(
                        {
                            "row_no": row_idx,
                            "employee_name": employee_name,
                            "duty_type": duty_type,
                            "schedule_date": schedule_date,
                            "work_value": value_text,
                            "is_leave": True,
                            "paid_hours": None,
                            "source_block": "base",
                        }
                    )
                    continue
                paid_hours = _parse_numeric_hours(value)
                if paid_hours is None:
                    rows.append(
                        {
                            "row_no": row_idx,
                            "employee_name": employee_name,
                            "duty_type": duty_type,
                            "schedule_date": schedule_date,
                            "work_value": value_text,
                            "is_leave": False,
                            "paid_hours": None,
                            "source_block": "base",
                            "hard_error": "잘못된 값",
                        }
                    )
                    continue
                rows.append(
                    {
                        "row_no": row_idx,
                        "employee_name": employee_name,
                        "duty_type": duty_type,
                        "schedule_date": schedule_date,
                        "work_value": value_text,
                        "is_leave": False,
                        "paid_hours": paid_hours,
                        "source_block": "base",
                    }
                )
            continue

        if mode in {"additional_day", "additional_night"}:
            duty_for_block = "day" if mode == "additional_day" else "night"
            for col_idx, schedule_date in date_columns.items():
                value = sheet.cell(row=row_idx, column=col_idx).value
                for parsed_name in _split_name_cells(value):
                    rows.append(
                        {
                            "row_no": row_idx,
                            "employee_name": parsed_name,
                            "duty_type": duty_for_block,
                            "schedule_date": schedule_date,
                            "work_value": parsed_name,
                            "is_leave": False,
                            "paid_hours": None,
                            "source_block": mode,
                            "force_default_template": True,
                        }
                    )
            continue

    return rows, month_ctx


def _resolve_arls_preview_rows(
    conn,
    *,
    tenant_id: str,
    tenant_code: str,
    company_code: str,
    site_id: str,
    site_code: str,
    parsed_rows: list[dict],
) -> tuple[list[dict], Counter[str]]:
    templates = _fetch_schedule_templates(conn, tenant_id=tenant_id, site_id=site_id, include_inactive=False)
    employees = _load_site_employees(conn, tenant_id=tenant_id, site_id=site_id)
    employee_index = _build_employee_name_index(employees)
    error_counts: Counter[str] = Counter()
    resolved: list[dict] = []

    if parsed_rows:
        min_date = min(item["schedule_date"] for item in parsed_rows if isinstance(item.get("schedule_date"), date))
        max_date = max(item["schedule_date"] for item in parsed_rows if isinstance(item.get("schedule_date"), date))
        existing_keys = _load_existing_schedule_keys(
            conn,
            tenant_id=tenant_id,
            site_id=site_id,
            start_date=min_date,
            end_date=max_date,
        )
    else:
        existing_keys = set()

    seen_keys: set[tuple[str, str]] = set()
    for row in parsed_rows:
        employee_name = str(row.get("employee_name") or "").strip()
        duty_type = _normalize_schedule_template_duty_type(row.get("duty_type"))
        schedule_date = row.get("schedule_date")
        work_value = str(row.get("work_value") or "").strip()
        validation_code: str | None = None
        validation_error: str | None = None
        template_row: dict | None = None
        employee_row: dict | None = None
        shift_type = "off" if bool(row.get("is_leave")) else _resolve_shift_type_from_duty_type(duty_type)
        shift_start_time: str | None = None
        shift_end_time: str | None = None
        paid_hours = row.get("paid_hours")

        if not isinstance(schedule_date, date):
            validation_code = "invalid_value"
            validation_error = "잘못된 날짜 값입니다."
        elif row.get("hard_error"):
            validation_code = "invalid_value"
            validation_error = str(row.get("hard_error") or "잘못된 값").strip() or "잘못된 값"

        if not validation_code:
            token = _normalize_name_token(employee_name)
            matches = employee_index.get(token, [])
            if len(matches) != 1:
                validation_code = "employee_match_failed"
                validation_error = "직원을 찾을 수 없습니다."
            else:
                employee_row = matches[0]

        if not validation_code and not bool(row.get("is_leave")):
            template_row = _choose_template_by_duty_hours(
                templates,
                duty_type=duty_type,
                paid_hours=float(paid_hours) if paid_hours is not None else None,
                require_default=bool(row.get("force_default_template")),
                strict_hours=bool(row.get("source_block") == "base" and paid_hours is not None),
            )
            if not template_row:
                validation_code = "template_match_failed"
                validation_error = "매칭 가능한 근무 템플릿이 없습니다."
            else:
                shift_start_time = _normalize_time_text(template_row.get("start_time"))
                shift_end_time = _normalize_time_text(template_row.get("end_time"))
                if paid_hours is None:
                    paid_hours = template_row.get("paid_hours")

        if not validation_code and employee_row:
            key = (str(employee_row.get("id") or ""), schedule_date.isoformat())
            if key in seen_keys or key in existing_keys:
                validation_code = "time_conflict"
                validation_error = "이미 등록된 일정과 충돌합니다."
            else:
                seen_keys.add(key)

        if validation_code:
            error_counts[validation_code] += 1

        if employee_row:
            company_id = str(employee_row.get("company_id") or "").strip()
            employee_id = str(employee_row.get("id") or "").strip()
            employee_code = str(employee_row.get("employee_code") or "").strip()
        else:
            company_id = ""
            employee_id = ""
            employee_code = ""

        resolved.append(
            {
                "row_no": int(row.get("row_no") or 0),
                "tenant_id": tenant_id,
                "tenant_code": tenant_code,
                "company_id": company_id or None,
                "company_code": company_code,
                "site_id": site_id,
                "site_code": site_code,
                "employee_id": employee_id or None,
                "employee_code": employee_code,
                "employee_name": employee_name,
                "schedule_date": schedule_date,
                "shift_type": shift_type,
                "duty_type": duty_type,
                "template_id": str(template_row.get("id") or "").strip() if template_row else None,
                "template_name": str(template_row.get("template_name") or "").strip() if template_row else None,
                "work_value": work_value,
                "shift_start_time": shift_start_time,
                "shift_end_time": shift_end_time,
                "paid_hours": float(paid_hours) if paid_hours is not None else None,
                "is_valid": validation_code is None,
                "validation_code": validation_code,
                "validation_error": validation_error,
            }
        )

    return resolved, error_counts


def _resolve_arls_template_path() -> Path | None:
    env_path = str(os.getenv("ARLS_MONTHLY_SCHEDULE_TEMPLATE_PATH") or "").strip()
    candidates: list[Path] = []
    if env_path:
        candidates.append(Path(env_path).expanduser())
    candidates.extend(ARLS_TEMPLATE_SEARCH_PATHS)
    for path in candidates:
        if path.exists():
            return path
    return None


def _load_required_arls_month_workbook() -> tuple[Workbook, Path]:
    template_path = _resolve_arls_template_path()
    if not template_path:
        raise HTTPException(status_code=500, detail="monthly schedule template unavailable")
    return load_workbook(filename=template_path), template_path


def _describe_arls_template_version(template_path: Path) -> str:
    try:
        stat = template_path.stat()
    except OSError:
        return ARLS_EXPORT_TEMPLATE_VERSION
    return f"{ARLS_EXPORT_TEMPLATE_VERSION}:{template_path.name}:{int(stat.st_mtime)}"


def _read_monthly_board_rows_for_export(conn, *, tenant_id: str, month_key: str) -> list[dict]:
    start_date, end_date = _month_bounds(month_key)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT ms.employee_id,
                   ms.schedule_date,
                   ms.shift_type,
                   ms.paid_hours,
                   ms.shift_start_time,
                   ms.shift_end_time,
                   ms.template_id,
                   ms.source,
                   ms.schedule_note,
                   st.duty_type,
                   st.start_time AS template_start_time,
                   st.end_time AS template_end_time,
                   st.paid_hours AS template_paid_hours,
                   e.employee_code,
                   e.sequence_no,
                   e.full_name AS employee_name,
                   e.hire_date,
                   e.leave_date,
                   COALESCE(e.soc_role, '') AS soc_role,
                   COALESCE(e.duty_role, '') AS duty_role,
                   s.site_code
            FROM monthly_schedules ms
            JOIN sites s ON s.id = ms.site_id
            JOIN employees e ON e.id = ms.employee_id
            LEFT JOIN schedule_templates st ON st.id = ms.template_id
            WHERE ms.tenant_id = %s
              AND ms.schedule_date >= %s
              AND ms.schedule_date < %s
            ORDER BY ms.schedule_date ASC, s.site_code ASC, e.employee_code ASC
            """,
            (tenant_id, start_date, end_date),
        )
        return [dict(row) for row in cur.fetchall()]


def _read_monthly_support_assignment_rows_for_export(conn, *, tenant_id: str, site_code: str, month_key: str) -> list[dict]:
    start_date, end_date = _month_bounds(month_key)
    rows = list_support_assignments(conn, tenant_id=tenant_id, work_date=None, site_id=None)
    filtered: list[dict] = []
    for row in rows:
        work_date = row.get("work_date")
        if not isinstance(work_date, date):
            continue
        if str(row.get("site_code") or "").strip() != site_code:
            continue
        if start_date <= work_date < end_date:
            filtered.append(dict(row))
    return filtered


def _read_monthly_overnight_rows_for_export(conn, *, tenant_id: str, site_code: str, month_key: str) -> list[dict]:
    start_date, end_date = _month_bounds(month_key)
    rows = list_apple_report_overnight_records(conn, tenant_id=tenant_id, work_date=None, site_id=None)
    filtered: list[dict] = []
    for row in rows:
        work_date = row.get("work_date")
        if not isinstance(work_date, date):
            continue
        if str(row.get("site_code") or "").strip() != site_code:
            continue
        if start_date <= work_date < end_date:
            filtered.append(dict(row))
    return filtered


def _read_monthly_employee_overnight_rows_for_export(
    conn,
    *,
    tenant_id: str,
    site_code: str,
    month_key: str,
) -> list[dict]:
    start_date, end_date = _month_bounds(month_key)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT ao.employee_id,
                   ao.work_date,
                   COALESCE(aor.hours, 10.0) AS hours,
                   COALESCE(aor.time_range, '22:00-08:00') AS time_range,
                   e.employee_code,
                   e.full_name AS employee_name,
                   e.sequence_no,
                   e.hire_date,
                   e.leave_date,
                   COALESCE(e.soc_role, '') AS soc_role,
                   COALESCE(e.duty_role, '') AS duty_role,
                   s.site_code
            FROM apple_overnight_reports ao
            JOIN employees e ON e.id = ao.employee_id
            LEFT JOIN sites s ON s.id = ao.site_id
            LEFT JOIN apple_report_overnight_records aor
              ON aor.tenant_id = ao.tenant_id
             AND aor.site_id = ao.site_id
             AND aor.work_date = ao.work_date
            WHERE ao.tenant_id = %s
              AND ao.overnight_approved = TRUE
              AND ao.work_date >= %s
              AND ao.work_date < %s
              AND COALESCE(s.site_code, '') = %s
            ORDER BY ao.work_date ASC, e.employee_code ASC
            """,
            (tenant_id, start_date, end_date, site_code),
        )
        return [dict(row) for row in cur.fetchall()]


def _read_monthly_daytime_need_rows_for_export(
    conn,
    *,
    tenant_id: str,
    site_id: str,
    month_key: str,
) -> list[dict]:
    start_date, end_date = _month_bounds(month_key)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT work_date, required_count, COALESCE(raw_text, '') AS raw_text
            FROM site_daytime_need_counts
            WHERE tenant_id = %s
              AND site_id = %s
              AND work_date >= %s
              AND work_date < %s
            ORDER BY work_date ASC
            """,
            (tenant_id, site_id, start_date, end_date),
        )
        return [dict(row) for row in cur.fetchall()]


def _read_monthly_support_request_rows_for_export(
    conn,
    *,
    tenant_id: str,
    site_id: str,
    month_key: str,
) -> list[dict]:
    start_date, end_date = _month_bounds(month_key)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT work_date,
                   shift_kind,
                   request_count,
                   work_purpose,
                   status,
                   detail_json
            FROM sentrix_support_request_tickets
            WHERE tenant_id = %s
              AND site_id = %s
              AND work_date >= %s
              AND work_date < %s
              AND status = %s
              AND source_workflow = %s
            ORDER BY work_date ASC, shift_kind ASC
            """,
            (
                tenant_id,
                site_id,
                start_date,
                end_date,
                SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS,
                SENTRIX_SUPPORT_REQUEST_WORKFLOW,
            ),
        )
        return [dict(row) for row in cur.fetchall()]


def _build_support_request_ticket_index(rows: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    index: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        work_date = row.get("work_date")
        if not isinstance(work_date, date):
            continue
        shift_kind = "night" if str(row.get("shift_kind") or "day").strip().lower() == "night" else "day"
        index[(work_date.isoformat(), shift_kind)] = dict(row)
    return index


def _build_schedule_export_revision(
    conn,
    *,
    tenant_id: str,
    site_id: str,
    site_code: str,
    month_key: str,
) -> str:
    board_rows = [
        row
        for row in _read_monthly_board_rows_for_export(conn, tenant_id=tenant_id, month_key=month_key)
        if str(row.get("site_code") or "").strip() == site_code
    ]
    support_rows = _read_monthly_support_assignment_rows_for_export(
        conn,
        tenant_id=tenant_id,
        site_code=site_code,
        month_key=month_key,
    )
    overnight_rows = _read_monthly_overnight_rows_for_export(
        conn,
        tenant_id=tenant_id,
        site_code=site_code,
        month_key=month_key,
    )
    employee_overnight_rows = _read_monthly_employee_overnight_rows_for_export(
        conn,
        tenant_id=tenant_id,
        site_code=site_code,
        month_key=month_key,
    )
    daytime_need_rows = _read_monthly_daytime_need_rows_for_export(
        conn,
        tenant_id=tenant_id,
        site_id=site_id,
        month_key=month_key,
    )
    support_request_rows = _read_monthly_support_request_rows_for_export(
        conn,
        tenant_id=tenant_id,
        site_id=site_id,
        month_key=month_key,
    )
    payload = {
        "month": month_key,
        "site_code": site_code,
        "board_rows": [
            {
                "employee_id": str(row.get("employee_id") or ""),
                "schedule_date": str(row.get("schedule_date") or ""),
                "shift_type": str(row.get("shift_type") or ""),
                "duty_type": str(row.get("duty_type") or ""),
                "template_id": str(row.get("template_id") or ""),
                "paid_hours": row.get("paid_hours"),
                "shift_start_time": _normalize_time_text(row.get("shift_start_time")),
                "shift_end_time": _normalize_time_text(row.get("shift_end_time")),
                "schedule_note": str(row.get("schedule_note") or ""),
                "source": str(row.get("source") or ""),
                "source_ticket_id": str(row.get("source_ticket_id") or ""),
            }
            for row in board_rows
        ],
        "support_rows": [
            {
                "employee_id": str(row.get("employee_id") or ""),
                "work_date": str(row.get("work_date") or ""),
                "site_code": str(row.get("site_code") or ""),
                "support_period": str(row.get("support_period") or "day"),
                "slot_index": int(row.get("slot_index") or 1),
                "worker_type": str(row.get("worker_type") or ""),
                "name": str(row.get("name") or ""),
                "affiliation": str(row.get("affiliation") or ""),
                "source": str(row.get("source") or ""),
                "source_ticket_id": str(row.get("source_ticket_id") or ""),
            }
            for row in support_rows
        ],
        "overnight_rows": [
            {
                "work_date": str(row.get("work_date") or ""),
                "site_code": str(row.get("site_code") or ""),
                "hours": row.get("hours"),
                "time_range": str(row.get("time_range") or ""),
                "headcount": row.get("headcount"),
            }
            for row in overnight_rows
        ],
        "employee_overnight_rows": [
            {
                "employee_id": str(row.get("employee_id") or ""),
                "work_date": str(row.get("work_date") or ""),
                "site_code": str(row.get("site_code") or ""),
                "hours": row.get("hours"),
                "time_range": str(row.get("time_range") or ""),
            }
            for row in employee_overnight_rows
        ],
        "daytime_need_rows": [
            {
                "work_date": str(row.get("work_date") or ""),
                "required_count": int(row.get("required_count") or 0),
                "raw_text": str(row.get("raw_text") or ""),
            }
            for row in daytime_need_rows
        ],
        "support_request_rows": [
            {
                "work_date": str(row.get("work_date") or ""),
                "shift_kind": str(row.get("shift_kind") or "day"),
                "request_count": int(row.get("request_count") or 0),
                "work_purpose": str(row.get("work_purpose") or ""),
                "status": str(row.get("status") or ""),
                "detail_json": dict(row.get("detail_json") or {}),
            }
            for row in support_request_rows
        ],
    }
    payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(payload_json.encode("utf-8")).hexdigest()[:16]


def _format_export_hours_value(value: object) -> str:
    if value is None or value == "":
        return ""
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.1f}".rstrip("0").rstrip(".")


def _export_value_to_hours(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _coerce_export_workbook_cell_value(value: object) -> object:
    text = str(value or "").strip()
    if not text:
        return None
    numeric = _parse_numeric_hours(text)
    if numeric is None:
        return text
    if float(numeric).is_integer():
        return int(numeric)
    return float(numeric)


def _apply_export_numeric_cell_number_format(cell, *, value: object) -> None:
    numeric = _parse_numeric_hours(value)
    if numeric is None:
        return
    if float(numeric).is_integer():
        cell.number_format = "0"
    else:
        cell.number_format = "0.0"


def _infer_export_hours_from_range(start_time: object, end_time: object) -> float | None:
    start_minutes = _time_text_to_minutes(start_time)
    end_minutes = _time_text_to_minutes(end_time)
    if start_minutes is None or end_minutes is None:
        return None
    if end_minutes <= start_minutes:
        overnight_hours = (end_minutes + 1440 - start_minutes) / 60.0
        if 0 < overnight_hours <= 16:
            return overnight_hours
        half_day_hours = ((end_minutes + 720) - start_minutes) / 60.0
        if 0 < half_day_hours <= 16:
            return half_day_hours
        return overnight_hours
    duration_hours = (end_minutes - start_minutes) / 60.0
    return duration_hours if duration_hours > 0 else None


def _resolve_export_row_hours(row: dict) -> object:
    canonical = _resolve_canonical_schedule_time(row)
    if canonical.get("hours") not in (None, ""):
        return canonical.get("hours")
    explicit_hours = row.get("paid_hours")
    if explicit_hours not in (None, ""):
        return explicit_hours
    template_hours = row.get("template_paid_hours")
    if template_hours not in (None, ""):
        return template_hours
    return None


def _resolve_export_overnight_value(hours: object, time_range: object = None) -> str:
    if hours not in (None, ""):
        return _format_export_hours_value(hours)
    start_time, end_time = _parse_export_shift_label_range(time_range)
    inferred = _infer_canonical_shift_hours(start_time, end_time)
    if inferred is None:
        return ""
    return _format_export_hours_value(inferred)


def _export_employee_role_rank(row: dict) -> int:
    soc_role = str(row.get("soc_role") or "").strip().lower()
    if soc_role in {"supervisor", "hq_admin"}:
        return 0
    if soc_role == "vice_supervisor":
        return 1
    if soc_role == "officer":
        return 2

    duty_role = str(row.get("duty_role") or "").strip().upper()
    if duty_role == TEAM_MANAGER_DUTY_ROLE:
        return 0
    if duty_role == VICE_SUPERVISOR_DUTY_ROLE:
        return 1
    if duty_role == GUARD_DUTY_ROLE:
        return 2
    return 9


def _resolve_internal_support_default_hours(row: dict, *, support_period: str) -> str:
    normalized_period = str(support_period or "day").strip().lower() or "day"
    if normalized_period == "night":
        return "10"
    soc_role = str(row.get("soc_role") or "").strip().lower()
    if soc_role in {"supervisor", "vice_supervisor", "hq_admin"}:
        return "10"
    duty_role = str(row.get("duty_role") or "").strip().upper()
    if duty_role in {TEAM_MANAGER_DUTY_ROLE, VICE_SUPERVISOR_DUTY_ROLE}:
        return "10"
    return "12"


def _parse_export_shift_label_range(label: object) -> tuple[str | None, str | None]:
    text = str(label or "").strip()
    match = re.fullmatch(r"(\d{2}:\d{2})-(\d{2}:\d{2})", text)
    if not match:
        return None, None
    return _normalize_time_text(match.group(1)), _normalize_time_text(match.group(2))


def _load_employee_export_meta(conn, *, tenant_id: str, employee_ids: list[str]) -> dict[str, dict[str, Any]]:
    unique_ids = [str(item).strip() for item in employee_ids if str(item).strip()]
    if not unique_ids:
        return {}
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id,
                   employee_code,
                   sequence_no,
                   COALESCE(soc_role, '') AS soc_role,
                   COALESCE(duty_role, '') AS duty_role
            FROM employees
            WHERE tenant_id = %s
              AND id = ANY(%s::uuid[])
            """,
            (tenant_id, unique_ids),
        )
        return {str(row["id"]): dict(row) for row in cur.fetchall()}


def _build_export_rows_from_board_payload(
    conn,
    *,
    tenant_id: str,
    board_payload: dict,
    site_code: str,
) -> list[dict]:
    employee_ids = [str(item.get("id") or "").strip() for item in board_payload.get("employees") or []]
    employee_meta = _load_employee_export_meta(conn, tenant_id=tenant_id, employee_ids=employee_ids)
    flattened: list[dict] = []
    for day in board_payload.get("days") or []:
        date_key = str(day.get("date") or "").strip()
        if not date_key:
            continue
        for item in day.get("items") or []:
            if str(item.get("site_code") or "").strip() != site_code:
                continue
            employee_id = str(item.get("employee_id") or "").strip()
            meta = employee_meta.get(employee_id, {})
            shift_start_time = _normalize_time_text(item.get("start_time") or item.get("shift_start_time"))
            shift_end_time = _normalize_time_text(item.get("end_time") or item.get("shift_end_time"))
            if not shift_start_time or not shift_end_time:
                shift_start_time, shift_end_time = _parse_export_shift_label_range(item.get("shift_label"))
            flattened.append(
                {
                    "employee_id": employee_id or None,
                    "employee_code": str(item.get("employee_code") or meta.get("employee_code") or "").strip(),
                    "employee_name": str(item.get("employee_name") or "").strip(),
                    "sequence_no": meta.get("sequence_no"),
                    "soc_role": meta.get("soc_role"),
                    "duty_role": meta.get("duty_role"),
                    "schedule_date": date_key,
                    "shift_type": item.get("shift_type"),
                    "duty_type": item.get("duty_type"),
                    "paid_hours": item.get("paid_hours"),
                    "shift_start_time": shift_start_time,
                    "shift_end_time": shift_end_time,
                    "template_id": item.get("template_id"),
                    "template_start_time": None,
                    "template_end_time": None,
                    "template_paid_hours": None,
                    "source": None,
                    "schedule_note": None,
                    "site_code": site_code,
                }
            )
    return flattened


def _merge_export_rows_with_board_fallback(rows: list[dict], fallback_rows: list[dict]) -> list[dict]:
    merged: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    order: list[tuple[str, str, str, str]] = []

    def row_key(row: dict) -> tuple[str, str, str, str]:
        return (
            str(row.get("employee_id") or row.get("employee_code") or "").strip(),
            str(row.get("schedule_date") or "").strip(),
            _normalize_schedule_template_duty_type(row.get("duty_type")),
            _normalize_shift_type(row.get("shift_type")),
        )

    for row in rows:
        key = row_key(row)
        merged[key] = dict(row)
        order.append(key)

    for row in fallback_rows:
        key = row_key(row)
        existing = merged.get(key)
        if not existing:
            merged[key] = dict(row)
            order.append(key)
            continue
        for field in (
            "paid_hours",
            "shift_start_time",
            "shift_end_time",
            "soc_role",
            "duty_role",
            "sequence_no",
            "employee_name",
            "employee_code",
        ):
            if existing.get(field) in (None, "", 0) and row.get(field) not in (None, ""):
                existing[field] = row.get(field)

    return [merged[key] for key in order]


def _merge_export_cell_value(existing: object, incoming: object) -> str:
    existing_text = str(existing or "").strip()
    incoming_text = str(incoming or "").strip()
    if not existing_text:
        return incoming_text
    if not incoming_text or existing_text == incoming_text:
        return existing_text
    try:
        merged = float(existing_text) + float(incoming_text)
    except (TypeError, ValueError):
        return f"{existing_text}/{incoming_text}"
    return _format_export_hours_value(merged)


def _get_export_non_working_marker(row: dict) -> str:
    shift_type = _normalize_shift_type(row.get("shift_type"))
    if shift_type == "holiday":
        return "공휴일"
    note_text = str(row.get("schedule_note") or "").strip()
    source_text = str(row.get("source") or "").strip().lower()
    note_lower = note_text.lower()
    if any(token in note_lower for token in ("연차", "leave", "vacation")) or source_text in {"leave", "annual_leave"}:
        return "연차"
    if any(token in note_lower for token in ("반차", "half")):
        return "반차"
    return "휴무"


def _find_template_data_start_row(sheet) -> int:
    for row_idx in range(1, min(sheet.max_row, 120) + 1):
        if str(sheet.cell(row=row_idx, column=3).value or "").strip() == "주간근무":
            return row_idx
    return 5


def _find_template_summary_start_row(sheet, *, fallback: int) -> int:
    start_row = max(1, fallback)
    for row_idx in range(start_row, min(sheet.max_row, 320) + 1):
        left = sheet.cell(row=row_idx, column=2).value
        right = sheet.cell(row=row_idx, column=3).value
        if (
            _label_contains_any(left, ARLS_DAY_SUPPORT_BLOCK_ALIASES)
            or _label_contains_any(left, ARLS_NIGHT_SUPPORT_BLOCK_ALIASES)
            or _label_contains_any(right, ARLS_NEED_COUNT_KEYWORDS)
            or _label_contains_any(right, ARLS_VENDOR_COUNT_KEYWORDS)
            or _label_contains_any(right, ARLS_WORK_NOTE_KEYWORDS)
        ):
            return row_idx
    return min(sheet.max_row + 1, fallback + 42)


def _copy_template_block_rows(sheet, *, src_start: int, dest_start: int) -> None:
    for offset in range(3):
        src_row = src_start + offset
        dest_row = dest_start + offset
        sheet.row_dimensions[dest_row].height = sheet.row_dimensions[src_row].height
        for col_idx in range(1, ARLS_SUMMARY_END_COL + 1):
            source = sheet.cell(row=src_row, column=col_idx)
            target = sheet.cell(row=dest_row, column=col_idx)
            target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.font:
                target.font = copy(source.font)
            if source.fill:
                target.fill = copy(source.fill)
            if source.border:
                target.border = copy(source.border)
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.protection:
                target.protection = copy(source.protection)
            target.value = None
    sheet.merge_cells(start_row=dest_start, start_column=1, end_row=dest_start + 2, end_column=1)
    sheet.merge_cells(start_row=dest_start, start_column=2, end_row=dest_start + 2, end_column=2)
    sheet.merge_cells(start_row=dest_start, start_column=36, end_row=dest_start + 2, end_column=36)
    sheet.merge_cells(start_row=dest_start, start_column=37, end_row=dest_start + 2, end_column=37)


def _ensure_template_employee_capacity(sheet, *, employee_count: int, data_start_row: int, summary_start_row: int) -> int:
    existing_groups = max(0, (summary_start_row - data_start_row) // 3)
    if employee_count <= existing_groups:
        return existing_groups
    missing_groups = employee_count - existing_groups
    insert_at = summary_start_row
    sheet.insert_rows(insert_at, amount=missing_groups * 3)
    src_start = data_start_row
    for idx in range(missing_groups):
        _copy_template_block_rows(sheet, src_start=src_start, dest_start=insert_at + (idx * 3))
    return employee_count


def _clear_template_hidden_sheet(sheet) -> None:
    for row in range(2, min(sheet.max_row, 64) + 1):
        for col in range(2, min(sheet.max_column, 26) + 1):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                continue
            cell.value = None


def _set_sheet_cell_value_if_writable(sheet, *, row: int, column: int, value: Any) -> bool:
    cell = sheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def _reset_arls_month_template_sheet(sheet, *, employee_count: int) -> tuple[object, int, int]:
    workbook = sheet.parent
    data_start_row = _find_template_data_start_row(sheet)
    summary_start_row = _find_template_summary_start_row(sheet, fallback=data_start_row + 42)
    slot_count = _ensure_template_employee_capacity(
        sheet,
        employee_count=max(1, employee_count),
        data_start_row=data_start_row,
        summary_start_row=summary_start_row,
    )
    summary_start_row = _find_template_summary_start_row(sheet, fallback=data_start_row + (slot_count * 3))

    for row_idx in range(data_start_row, summary_start_row):
        group_offset = (row_idx - data_start_row) % 3
        for col_idx in range(1, ARLS_SUMMARY_END_COL + 1):
            if col_idx == 3:
                continue
            if group_offset != 0 and col_idx in {1, 2, 36, 37}:
                continue
            _set_sheet_cell_value_if_writable(sheet, row=row_idx, column=col_idx, value=None)
        _set_sheet_cell_value_if_writable(
            sheet,
            row=row_idx,
            column=3,
            value=_schedule_template_duty_label(("day", "overtime", "night")[group_offset]),
        )
        if group_offset == 0:
            _set_sheet_cell_value_if_writable(sheet, row=row_idx, column=36, value=None)
            _set_sheet_cell_value_if_writable(sheet, row=row_idx, column=37, value=None)

    for col_idx in range(ARLS_DATE_START_COL, ARLS_DATE_END_COL + 1):
        _set_sheet_cell_value_if_writable(sheet, row=4, column=col_idx, value=None)

    for row_idx in range(summary_start_row, min(sheet.max_row, 84) + 1):
        for col_idx in range(1, ARLS_SUMMARY_END_COL + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            if row_idx == summary_start_row and col_idx in {2, 3}:
                continue
            cell.value = None if col_idx >= ARLS_DATE_START_COL or col_idx in {1, 35, 36, 37} else cell.value

    hidden_sheet_name = "출동.잔업 초과수당(2)"
    if hidden_sheet_name in workbook.sheetnames:
        _clear_template_hidden_sheet(workbook[hidden_sheet_name])

    return sheet, data_start_row, summary_start_row


def _reset_arls_month_template_workbook(workbook: Workbook, *, employee_count: int) -> tuple[object, int, int]:
    if ARLS_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[ARLS_SHEET_NAME]
    else:
        sheet = workbook.active
        sheet.title = ARLS_SHEET_NAME
    return _reset_arls_month_template_sheet(sheet, employee_count=employee_count)


def _prepare_blank_arls_template_workbook(
    workbook: Workbook,
    *,
    month_key: str,
    site_name: str = "",
    site_address: str = "",
) -> None:
    sheet, _data_start_row, _summary_start_row = _reset_arls_month_template_workbook(
        workbook,
        employee_count=1,
    )
    day_keys = _populate_arls_month_template_headers(
        sheet,
        month_key=month_key,
        site_name=site_name,
        site_address=site_address,
    )
    # Template downloads must not contain sample date/month/site payload.
    sheet["B1"] = None
    sheet["B2"] = None
    sheet["B3"] = None
    for idx, _date_key in enumerate(day_keys):
        col_idx = ARLS_DATE_START_COL + idx
        _set_sheet_cell_value_if_writable(sheet, row=2, column=col_idx, value=None)
        _set_sheet_cell_value_if_writable(sheet, row=3, column=col_idx, value=None)
        _set_sheet_cell_value_if_writable(sheet, row=4, column=col_idx, value=None)
    if ARLS_METADATA_SHEET_NAME in workbook.sheetnames:
        workbook.remove(workbook[ARLS_METADATA_SHEET_NAME])


def _build_export_employee_blocks(
    rows: list[dict],
    *,
    support_rows: list[dict] | None = None,
    overnight_rows: list[dict] | None = None,
    employee_overnight_rows: list[dict] | None = None,
) -> list[dict]:
    support_rows = support_rows or []
    overnight_rows = overnight_rows or []
    employee_overnight_rows = employee_overnight_rows or []
    grouped: dict[str, dict[str, Any]] = {}
    ordering: list[str] = []

    def ensure_employee(row: dict) -> dict[str, Any] | None:
        employee_id = str(row.get("employee_id") or "").strip()
        employee_code = str(row.get("employee_code") or "").strip()
        employee_name = str(row.get("employee_name") or "").strip()
        employee_key = employee_id or employee_code or employee_name
        if not employee_key:
            return None
        if employee_key not in grouped:
            grouped[employee_key] = {
                "employee_id": employee_id or None,
                "employee_code": employee_code,
                "employee_name": employee_name or employee_code or employee_key,
                "sequence_no": row.get("sequence_no"),
                "role_rank": _export_employee_role_rank(row),
                "day": {},
                "overtime": {},
                "night": {},
                "notes": [],
                "work_dates": set(),
                "hire_date": row.get("hire_date") if isinstance(row.get("hire_date"), date) else None,
                "leave_date": row.get("leave_date") if isinstance(row.get("leave_date"), date) else None,
            }
            ordering.append(employee_key)
        block = grouped[employee_key]
        if not block.get("employee_id") and employee_id:
            block["employee_id"] = employee_id
        if not str(block.get("employee_code") or "").strip() and employee_code:
            block["employee_code"] = employee_code
        if not str(block.get("employee_name") or "").strip() and employee_name:
            block["employee_name"] = employee_name
        if block.get("sequence_no") in (None, "") and row.get("sequence_no") not in (None, ""):
            block["sequence_no"] = row.get("sequence_no")
        current_role_rank = block.get("role_rank")
        next_role_rank = _export_employee_role_rank(row)
        if current_role_rank is None or next_role_rank < current_role_rank:
            block["role_rank"] = next_role_rank
        if block.get("hire_date") is None and isinstance(row.get("hire_date"), date):
            block["hire_date"] = row.get("hire_date")
        if block.get("leave_date") is None and isinstance(row.get("leave_date"), date):
            block["leave_date"] = row.get("leave_date")
        return block

    for row in rows:
        block = ensure_employee(row)
        if not block:
            continue
        schedule_date = row.get("schedule_date")
        if isinstance(schedule_date, date):
            date_key = schedule_date.isoformat()
        else:
            date_key = str(schedule_date or "").strip()
        if not date_key:
            continue
        duty_type = _normalize_schedule_template_duty_type(row.get("duty_type"))
        shift_type = _normalize_shift_type(row.get("shift_type"))
        if duty_type not in {"day", "overtime", "night"}:
            if shift_type == "night":
                duty_type = "night"
            elif shift_type == "overtime":
                duty_type = "overtime"
            else:
                duty_type = "day"
        if shift_type in NON_WORKING_SHIFT_TYPES:
            duty_type = "overtime"
            value = _get_export_non_working_marker(row)
        else:
            value = _format_export_hours_value(_resolve_export_row_hours(row))
        block[duty_type][date_key] = _merge_export_cell_value(block[duty_type].get(date_key), value)
        if shift_type not in NON_WORKING_SHIFT_TYPES and duty_type in {"day", "night"} and _export_value_to_hours(value) > 0:
            block["work_dates"].add(date_key)
        note_text = str(row.get("schedule_note") or "").strip()
        if note_text and note_text not in block["notes"]:
            block["notes"].append(note_text)

    for support in support_rows:
        support_period = str(support.get("support_period") or "day").strip().lower() or "day"
        if support_period not in {"day", "night"}:
            support_period = "day"
        employee_id = str(support.get("employee_id") or "").strip()
        work_date = support.get("work_date")
        if not employee_id or not isinstance(work_date, date):
            continue
        date_key = work_date.isoformat()
        block = ensure_employee(
            {
                "employee_id": employee_id,
                "employee_code": support.get("employee_code"),
                "employee_name": support.get("employee_name") or support.get("worker_name") or support.get("name"),
                "sequence_no": None,
                "soc_role": support.get("soc_role"),
                "duty_role": support.get("duty_role"),
            }
        )
        if not block:
            continue
        if not bool(support.get("is_internal")):
            continue
        if str(support.get("source") or "").strip().upper() == SENTRIX_HQ_ROSTER_ASSIGNMENT_SOURCE:
            # Sentrix HQ roster rows must reach employee-visible export cells only
            # through canonical monthly schedule truth after ARLS materialization.
            continue
        target_bucket = "night" if support_period == "night" else "day"
        target_value = _resolve_internal_support_default_hours(support, support_period=support_period)
        current_value = str(block[target_bucket].get(date_key) or "").strip()
        if current_value:
            if current_value != target_value:
                note_text = f"지원근무 충돌 검토 필요 {date_key} ({'야간' if target_bucket == 'night' else '주간'})"
                if note_text not in block["notes"]:
                    block["notes"].append(note_text)
            continue
        block[target_bucket][date_key] = target_value
        if _export_value_to_hours(target_value) > 0:
            block["work_dates"].add(date_key)

    # Employee-visible night values must come from canonical monthly schedule truth
    # (or support-worker overlays already merged into that truth contract). Raw
    # apple_overnight_reports rows are kept for reconciliation, but they must not
    # inject extra night cells that the calendar does not show.

    for block in grouped.values():
        work_dates = sorted(str(item).strip() for item in block.get("work_dates") or set() if str(item).strip())
        hire_date = block.get("hire_date")
        leave_date = block.get("leave_date")
        if not work_dates:
            continue
        if isinstance(hire_date, date):
            hire_key = hire_date.isoformat()
            if hire_key == work_dates[0]:
                block["night"][hire_key] = "입사"
        if isinstance(leave_date, date):
            leave_key = leave_date.isoformat()
            if leave_key == work_dates[-1]:
                block["night"][leave_key] = "퇴사"

    def sort_key(employee_key: str) -> tuple[Any, str, str]:
        block = grouped[employee_key]
        role_rank_value = block.get("role_rank")
        role_rank = int(role_rank_value) if role_rank_value is not None else 9
        sequence_no = block.get("sequence_no")
        try:
            sequence_value = int(sequence_no)
        except (TypeError, ValueError):
            sequence_value = 999999
        return (
            role_rank,
            sequence_value,
            str(block.get("employee_code") or ""),
            str(block.get("employee_name") or ""),
        )

    return [grouped[key] for key in sorted(ordering, key=sort_key)]


def _build_template_site_display(site_name: str, site_address: str) -> str:
    name_text = str(site_name or "").strip()
    address_text = str(site_address or "").strip()
    if name_text and address_text:
        return f"{name_text}\n{address_text}"
    return name_text or address_text


def _populate_arls_month_template_headers(sheet, *, month_key: str, site_name: str, site_address: str) -> list[str]:
    start_date, end_date = _month_bounds(month_key)
    sheet["B1"] = f"{start_date.month}월"
    sheet["B2"] = f"{start_date.year}년 {start_date.month}월"
    sheet["B3"] = _build_template_site_display(site_name, site_address)
    day_keys = _month_day_keys(start_date, end_date)
    weekday_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for idx, date_key in enumerate(day_keys):
        if idx >= (ARLS_DATE_END_COL - ARLS_DATE_START_COL + 1):
            break
        col_idx = ARLS_DATE_START_COL + idx
        current_date = datetime.strptime(date_key, "%Y-%m-%d").date()
        sheet.cell(row=2, column=col_idx, value=current_date)
        sheet.cell(row=3, column=col_idx, value=weekday_labels[current_date.weekday()])
        sheet.cell(row=4, column=col_idx, value=None)
    for idx in range(len(day_keys), ARLS_DATE_END_COL - ARLS_DATE_START_COL + 1):
        col_idx = ARLS_DATE_START_COL + idx
        sheet.cell(row=2, column=col_idx, value=None)
        sheet.cell(row=3, column=col_idx, value=None)
        sheet.cell(row=4, column=col_idx, value=None)
    return day_keys


def _populate_arls_month_template_summary_sections(
    sheet,
    *,
    summary_start_row: int,
    day_keys: list[str],
    employee_blocks: list[dict],
    daytime_need_rows: list[dict] | None = None,
    support_request_rows: list[dict] | None = None,
) -> None:
    day_row_label = "주간 근무자(직원) 수"
    night_row_label = "야간 근무자 총 수"
    day_count_row = None
    night_count_row = None
    day_need_row = None
    night_need_row = None
    day_vendor_count_row = None
    night_vendor_count_row = None
    work_note_row = None
    daytime_need_rows = daytime_need_rows or []
    support_request_rows = support_request_rows or []
    need_by_day = {
        row.get("work_date").isoformat(): dict(row)
        for row in daytime_need_rows
        if isinstance(row.get("work_date"), date)
    }
    request_by_key = {
        (
            row.get("work_date").isoformat(),
            "night" if str(row.get("shift_kind") or "day").strip().lower() == "night" else "day",
        ): dict(row)
        for row in support_request_rows
        if isinstance(row.get("work_date"), date)
    }
    rows_meta = _locate_support_section_rows(sheet)
    day_count_row = rows_meta.get("weekly_count_row")
    night_count_row = rows_meta.get("night_count_row")
    day_need_row = rows_meta.get("day_need_row")
    night_need_row = rows_meta.get("night_need_row")
    day_vendor_count_row = rows_meta.get("day_vendor_count_row")
    night_vendor_count_row = rows_meta.get("night_vendor_count_row")
    work_note_row = rows_meta.get("work_note_row")
    if day_count_row:
        total = 0
        for idx, date_key in enumerate(day_keys):
            col_idx = ARLS_DATE_START_COL + idx
            count = 0
            for block in employee_blocks:
                value = str(block["day"].get(date_key) or "").strip()
                if _export_value_to_hours(value) > 0:
                    count += 1
            _set_sheet_cell_value_if_writable(sheet, row=day_count_row, column=col_idx, value=count or 0)
            total += count
        _set_sheet_cell_value_if_writable(
            sheet,
            row=day_count_row,
            column=ARLS_SUMMARY_START_COL,
            value=total or 0,
        )
    if night_count_row:
        total = 0
        for idx, date_key in enumerate(day_keys):
            col_idx = ARLS_DATE_START_COL + idx
            count = 0
            for block in employee_blocks:
                value = str(block["night"].get(date_key) or "").strip()
                if _export_value_to_hours(value) > 0:
                    count += 1
            _set_sheet_cell_value_if_writable(sheet, row=night_count_row, column=col_idx, value=count or 0)
            total += count
        _set_sheet_cell_value_if_writable(
            sheet,
            row=night_count_row,
            column=ARLS_SUMMARY_START_COL,
            value=total or 0,
        )
    if day_need_row:
        total = 0
        for idx, date_key in enumerate(day_keys):
            col_idx = ARLS_DATE_START_COL + idx
            row = need_by_day.get(date_key) or {}
            raw_text = str(row.get("raw_text") or "").strip()
            required_count = int(row.get("required_count") or 0)
            if raw_text:
                value = raw_text
            elif required_count > 0:
                value = f"섭외 {required_count}인 요청"
            else:
                value = ""
            _set_sheet_cell_value_if_writable(sheet, row=day_need_row, column=col_idx, value=value)
            total += required_count
        _set_sheet_cell_value_if_writable(
            sheet,
            row=day_need_row,
            column=ARLS_SUMMARY_START_COL,
            value=total or 0,
        )
    if night_need_row or day_vendor_count_row or night_vendor_count_row or work_note_row:
        for idx, date_key in enumerate(day_keys):
            col_idx = ARLS_DATE_START_COL + idx
            day_request = request_by_key.get((date_key, "day")) or {}
            night_request = request_by_key.get((date_key, "night")) or {}
            day_detail = dict(day_request.get("detail_json") or {})
            night_detail = dict(night_request.get("detail_json") or {})
            if day_vendor_count_row:
                day_vendor_value = str(day_detail.get("external_count_raw") or "").strip()
                _set_sheet_cell_value_if_writable(sheet, row=day_vendor_count_row, column=col_idx, value=day_vendor_value or None)
            if night_vendor_count_row:
                night_vendor_value = str(night_detail.get("external_count_raw") or "").strip()
                _set_sheet_cell_value_if_writable(sheet, row=night_vendor_count_row, column=col_idx, value=night_vendor_value or None)
            if night_need_row:
                night_required_value = (
                    str(night_detail.get("required_count_raw") or "").strip()
                    or (str(int(night_request.get("request_count") or 0)) if int(night_request.get("request_count") or 0) > 0 else "")
                )
                _set_sheet_cell_value_if_writable(sheet, row=night_need_row, column=col_idx, value=night_required_value or None)
            if work_note_row:
                work_purpose = str(night_request.get("work_purpose") or "").strip()
                _set_sheet_cell_value_if_writable(sheet, row=work_note_row, column=col_idx, value=work_purpose or None)


def _upsert_arls_export_metadata_sheet(
    workbook: Workbook,
    *,
    tenant_code: str,
    site_code: str,
    site_name: str,
    month_key: str,
    export_revision: str,
    template_version: str,
    source_version: str,
    employee_count: int,
    row_count: int,
    support_row_count: int,
    overnight_row_count: int,
    employee_overnight_row_count: int,
) -> None:
    if ARLS_METADATA_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[ARLS_METADATA_SHEET_NAME]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(ARLS_METADATA_SHEET_NAME)
    sheet.sheet_state = "hidden"
    exported_at = datetime.now(timezone(timedelta(hours=9))).isoformat()
    entries = [
        ("tenant_code", tenant_code),
        ("site_code", site_code),
        ("site_name", site_name),
        ("month", month_key),
        ("exported_at_kst", exported_at),
        ("export_revision", export_revision),
        ("template_version", template_version),
        ("export_source_version", source_version),
        ("employee_count", employee_count),
        ("schedule_row_count", row_count),
        ("support_assignment_row_count", support_row_count),
        ("overnight_site_row_count", overnight_row_count),
        ("overnight_employee_row_count", employee_overnight_row_count),
    ]
    for index, (key, value) in enumerate(entries, start=1):
        sheet.cell(row=index, column=1, value=key)
        sheet.cell(row=index, column=2, value=value)


def _build_all_sites_export_revision(contexts: list[dict[str, Any]]) -> str:
    payload = [
        {
            "site_code": str(item.get("site_code") or "").strip(),
            "site_name": str(item.get("site_name") or "").strip(),
            "export_revision": str(item.get("export_revision") or "").strip(),
        }
        for item in contexts
    ]
    payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(payload_json.encode("utf-8")).hexdigest()[:16]


def _normalize_excel_sheet_title(raw: object, *, fallback: str) -> str:
    text = str(raw or "").strip() or fallback
    text = re.sub(r"[:\\\\/?*\\[\\]]", "_", text)
    text = text.strip().strip("'") or fallback
    return text[:31] or fallback


def _apply_export_value_cell_style(cell, *, duty_type: str, value: object) -> None:
    value_text = str(value or "").strip()
    if not value_text:
        return
    duty = _normalize_schedule_template_duty_type(duty_type)
    if duty == "day":
        if _parse_numeric_hours(value_text) is not None:
            _apply_export_numeric_cell_number_format(cell, value=value_text)
            cell.fill = copy(ARLS_DAY_VALUE_FILL)
        return
    if duty == "overtime":
        if value_text == "연차":
            cell.fill = copy(ARLS_ANNUAL_LEAVE_FILL)
            return
        if _parse_numeric_hours(value_text) is not None:
            _apply_export_numeric_cell_number_format(cell, value=value_text)
            cell.fill = copy(ARLS_OT_VALUE_FILL)
        return
    if duty == "night":
        if value_text == "입사":
            cell.fill = copy(ARLS_HIRE_MARKER_FILL)
            next_font = copy(cell.font)
            next_font.color = ARLS_HIRE_MARKER_FONT_COLOR
            cell.font = next_font
            return
        if value_text == "퇴사":
            cell.fill = copy(ARLS_LEAVE_MARKER_FILL)
            next_font = copy(cell.font)
            next_font.color = ARLS_LEAVE_MARKER_FONT_COLOR
            cell.font = next_font
            return
        if _parse_numeric_hours(value_text) is not None:
            _apply_export_numeric_cell_number_format(cell, value=value_text)
            cell.fill = copy(ARLS_NIGHT_VALUE_FILL)


def _build_arls_month_sheet(
    workbook: Workbook,
    *,
    month_key: str,
    rows: list[dict],
    site_name: str = "",
    site_address: str = "",
    tenant_code: str = "",
    site_code: str = "",
    support_rows: list[dict] | None = None,
    support_request_rows: list[dict] | None = None,
    overnight_rows: list[dict] | None = None,
    employee_overnight_rows: list[dict] | None = None,
    daytime_need_rows: list[dict] | None = None,
    export_revision: str = "",
    template_version: str = ARLS_EXPORT_TEMPLATE_VERSION,
    source_version: str = ARLS_EXPORT_SOURCE_VERSION,
    sheet_name: str | None = None,
    write_metadata: bool = True,
) -> None:
    employee_blocks = _build_export_employee_blocks(
        rows,
        support_rows=support_rows,
        overnight_rows=overnight_rows,
        employee_overnight_rows=employee_overnight_rows,
    )
    if sheet_name and sheet_name in workbook.sheetnames:
        target_sheet = workbook[sheet_name]
    elif sheet_name and workbook.active.title == sheet_name:
        target_sheet = workbook.active
    elif sheet_name and ARLS_SHEET_NAME in workbook.sheetnames:
        target_sheet = workbook[ARLS_SHEET_NAME]
        target_sheet.title = sheet_name
    elif ARLS_SHEET_NAME in workbook.sheetnames:
        target_sheet = workbook[ARLS_SHEET_NAME]
    else:
        target_sheet = workbook.active
        target_sheet.title = sheet_name or ARLS_SHEET_NAME

    sheet, data_start_row, summary_start_row = _reset_arls_month_template_sheet(
        target_sheet,
        employee_count=len(employee_blocks),
    )
    day_keys = _populate_arls_month_template_headers(
        sheet,
        month_key=month_key,
        site_name=site_name,
        site_address=site_address,
    )

    current_row = data_start_row
    for index, block in enumerate(employee_blocks, start=1):
        employee_name = str(block.get("employee_name") or "").strip()
        if not employee_name:
            employee_name = str(block.get("employee_code") or "").strip()
        notes = [f"* {item}" for item in block.get("notes") or [] if str(item).strip()]
        day_total = 0.0
        overtime_total = 0.0
        night_total = 0.0
        day_count = 0

        for duty_offset, duty_type in enumerate(("day", "overtime", "night")):
            row_idx = current_row + duty_offset
            if duty_offset == 0:
                sheet.cell(row=row_idx, column=1, value=index)
                sheet.cell(row=row_idx, column=2, value=employee_name)
            sheet.cell(row=row_idx, column=3, value=_schedule_template_duty_label(duty_type))
            for idx, date_key in enumerate(day_keys):
                col_idx = ARLS_DATE_START_COL + idx
                raw_value = str(block[duty_type].get(date_key) or "").strip()
                cell_value = _coerce_export_workbook_cell_value(raw_value)
                cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                _apply_export_value_cell_style(cell, duty_type=duty_type, value=raw_value)
                numeric_hours = _export_value_to_hours(raw_value)
                if duty_type == "day":
                    day_total += numeric_hours
                    if numeric_hours > 0:
                        day_count += 1
                elif duty_type == "overtime":
                    overtime_total += numeric_hours
                else:
                    night_total += numeric_hours
        sheet.cell(row=current_row, column=ARLS_SUMMARY_START_COL, value=day_total or 0)
        sheet.cell(row=current_row + 1, column=ARLS_SUMMARY_START_COL, value=overtime_total or 0)
        sheet.cell(row=current_row + 2, column=ARLS_SUMMARY_START_COL, value=night_total or 0)
        sheet.cell(row=current_row, column=ARLS_SUMMARY_START_COL + 1, value=day_count or 0)
        sheet.cell(row=current_row, column=ARLS_SUMMARY_START_COL + 2, value="\n".join(notes) if notes else None)
        current_row += 3

    _populate_arls_month_template_summary_sections(
        sheet,
        summary_start_row=summary_start_row,
        day_keys=day_keys,
        employee_blocks=employee_blocks,
        daytime_need_rows=daytime_need_rows,
        support_request_rows=support_request_rows,
    )
    _populate_support_assignment_sections(
        sheet,
        day_keys=day_keys,
        assignment_rows=support_rows or [],
        include_internal=False,
    )
    if write_metadata:
        _upsert_arls_export_metadata_sheet(
            workbook,
            tenant_code=tenant_code,
            site_code=site_code,
            site_name=site_name,
            month_key=month_key,
            export_revision=export_revision or str(uuid.uuid4()),
            template_version=template_version,
            source_version=source_version,
            employee_count=len(employee_blocks),
            row_count=len(rows),
            support_row_count=len(support_rows or []),
            overnight_row_count=len(overnight_rows or []),
            employee_overnight_row_count=len(employee_overnight_rows or []),
        )


def _classify_arls_summary_row(label_col_b: object, label_col_c: object) -> tuple[str | None, str | None]:
    left = str(label_col_b or "").strip().replace("\n", " ")
    right = str(label_col_c or "").strip().replace("\n", " ")
    if _label_contains_any(label_col_c, ARLS_NEED_COUNT_KEYWORDS):
        return "daytime_need", "필요인원 수"
    if left == "주간 근무자(직원) 수":
        return "protected_summary", "주간 근무자(직원) 수"
    if left == "주간 추가 근무자 수":
        return "protected_summary", "주간 추가 근무자 수"
    if left == "주간 출근자 총 수":
        return "protected_summary", "주간 출근자 총 수"
    if _label_contains_any(label_col_b, ARLS_ADDITIONAL_DAY_KEYWORDS):
        return "protected_support_names", "주간 지원 근무자"
    if _label_contains_any(label_col_b, ARLS_ADDITIONAL_NIGHT_KEYWORDS):
        return "protected_night_names", "야간 지원 근무자"
    if left == "야간 근무자 총 수":
        return "protected_summary", "야간 근무자 총 수"
    if _label_contains_any(label_col_c, ARLS_VENDOR_COUNT_KEYWORDS):
        return "protected_vendor_count", "외부인원 투입 수"
    if _label_contains_any(label_col_c, ARLS_WORK_NOTE_KEYWORDS):
        return "protected_work_note", "작업 목적"
    return None, None


def _parse_arls_canonical_import_sheet(sheet) -> dict[str, Any]:
    date_columns, month_ctx = _extract_arls_date_columns(sheet)
    data_start_row = _find_template_data_start_row(sheet)
    rows_meta = _locate_support_section_rows(sheet)
    fallback_summary_start = _find_template_summary_start_row(sheet, fallback=data_start_row + 42)
    anchor_rows = [
        int(value)
        for key, value in rows_meta.items()
        if key.endswith("_row") and isinstance(value, int) and int(value) >= data_start_row
    ]
    anchor_rows.extend([int(item) for item in rows_meta.get("weekly_rows") or [] if int(item) >= data_start_row])
    anchor_rows.extend([int(item) for item in rows_meta.get("night_rows") or [] if int(item) >= data_start_row])
    summary_start_row = min([fallback_summary_start, *anchor_rows]) if anchor_rows else fallback_summary_start
    body_cells: list[dict[str, Any]] = []
    support_cells: list[dict[str, Any]] = []
    need_cells: list[dict[str, Any]] = []
    support_blocks: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []

    if not date_columns:
        issues.append(
            _build_import_issue(
                "DATE_HEADER_PARSE_FAILED",
                sheet_name=sheet.title,
                section="date_header",
            )
        )
        return {
            "month_ctx": month_ctx,
            "body_cells": body_cells,
            "need_cells": need_cells,
            "support_cells": support_cells,
            "support_blocks": support_blocks,
            "issues": issues,
            "section_rows": rows_meta,
            "summary_start_row": summary_start_row,
        }

    body_end_row = max(data_start_row - 1, summary_start_row - 1)
    row_idx = data_start_row
    while row_idx <= body_end_row:
        if not any(
            _normalize_workbook_display_value(sheet.cell(row=row_idx + offset, column=col_idx).value)
            for offset in range(0, min(3, max(body_end_row - row_idx + 1, 0)))
            for col_idx in range(1, 4)
        ):
            row_idx += 1
            continue
        expected = ("day", "overtime", "night")
        labels = tuple(
            _normalize_schedule_template_duty_type(sheet.cell(row=row_idx + offset, column=3).value)
            for offset in range(0, min(3, max(body_end_row - row_idx + 1, 0)))
        )
        if row_idx + 2 > body_end_row or labels != expected:
            issues.append(
                _build_import_issue(
                    "EMPLOYEE_ROW_GROUP_INVALID",
                    sheet_name=sheet.title,
                    row_no=row_idx,
                    col_no=3,
                    section="base_schedule",
                )
            )
            row_idx += 1
            continue
        employee_name = _normalize_workbook_display_value(sheet.cell(row=row_idx, column=2).value)
        employee_sequence = sheet.cell(row=row_idx, column=1).value
        if not employee_name or employee_name in {"0", "-"}:
            row_idx += 3
            continue
        for duty_offset, duty_type in enumerate(expected):
            duty_row = row_idx + duty_offset
            for col_idx, schedule_date in date_columns.items():
                raw_value = sheet.cell(row=duty_row, column=col_idx).value
                work_value = _normalize_workbook_display_value(raw_value)
                semantic_type, numeric_hours = _classify_import_body_semantic_type(raw_value)
                issue_code = None
                issue_message = None
                if semantic_type == "invalid" and work_value:
                    issue_code = "UNSUPPORTED_CELL_FORMAT"
                    issue_message = "지원하지 않는 셀 값 형식입니다."
                    issues.append(
                        _build_import_issue(
                            issue_code,
                            message=issue_message,
                            sheet_name=sheet.title,
                            row_no=duty_row,
                            col_no=col_idx,
                            section="base_schedule",
                        )
                    )
                body_cells.append(
                    {
                        "row_no": duty_row,
                        "col_no": col_idx,
                        "source_sheet": sheet.title,
                        "employee_name": employee_name,
                        "employee_sequence_no": int(employee_sequence) if isinstance(employee_sequence, (int, float)) else None,
                        "duty_type": duty_type,
                        "schedule_date": schedule_date,
                        "work_value": work_value,
                        "source_block": "body",
                        "section_label": _schedule_template_duty_label(duty_type),
                        "parsed_semantic_type": semantic_type,
                        "numeric_hours": numeric_hours,
                        "issue_code": issue_code,
                        "issue_message": issue_message,
                        "month_ctx": month_ctx,
                    }
                )
        row_idx += 3

    if not rows_meta.get("weekly_rows"):
        issues.append(_build_import_issue("SECTION_NOT_FOUND", sheet_name=sheet.title, section="day_support"))
    if not rows_meta.get("night_rows"):
        issues.append(_build_import_issue("SECTION_NOT_FOUND", sheet_name=sheet.title, section="night_support"))

    def append_support_value_row(
        *,
        row_no: int,
        col_no: int,
        schedule_date: date,
        source_block: str,
        section_label: str,
        raw_value: object,
        parsed_semantic_type: str,
        issue_code: str | None = None,
        issue_message: str | None = None,
        extra: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        payload = {
            "row_no": row_no,
            "col_no": col_no,
            "source_sheet": sheet.title,
            "schedule_date": schedule_date,
            "work_value": _normalize_workbook_display_value(raw_value),
            "source_block": source_block,
            "section_label": section_label,
            "parsed_semantic_type": parsed_semantic_type,
            "issue_code": issue_code,
            "issue_message": issue_message,
        }
        if extra:
            payload.update(extra)
        return payload

    for block_type, block_rows, section_label in (
        ("day_support", list(rows_meta.get("weekly_rows") or []), "주간 지원 근무자"),
        ("night_support", list(rows_meta.get("night_rows") or []), "야간 지원 근무자"),
    ):
        for col_idx, schedule_date in date_columns.items():
            block_issues: list[str] = []
            worker_slots: list[dict[str, Any]] = []
            valid_filled_count = 0
            invalid_filled_count = 0
            for worker_row in block_rows:
                raw_value = sheet.cell(row=worker_row, column=col_idx).value
                parsed_worker = _parse_support_worker_cell(raw_value)
                slot_index = _parse_worker_slot_number(sheet.cell(row=worker_row, column=3).value) or (len(worker_slots) + 1)
                issue_code = parsed_worker.get("issue_code")
                issue_message = parsed_worker.get("issue_message")
                if issue_code:
                    invalid_filled_count += 1
                    block_issues.append(_normalize_import_issue_code(issue_code))
                    issues.append(
                        _build_import_issue(
                            issue_code,
                            message=issue_message,
                            sheet_name=sheet.title,
                            row_no=worker_row,
                            col_no=col_idx,
                            section=block_type,
                        )
                    )
                elif parsed_worker.get("is_filled"):
                    valid_filled_count += 1
                slot_payload = append_support_value_row(
                    row_no=worker_row,
                    col_no=col_idx,
                    schedule_date=schedule_date,
                    source_block=f"{block_type}_worker",
                    section_label=section_label,
                    raw_value=raw_value,
                    parsed_semantic_type=str(parsed_worker.get("semantic_type") or ""),
                    issue_code=issue_code,
                    issue_message=issue_message,
                    extra={"slot_index": slot_index},
                )
                worker_slots.append(slot_payload)
                support_cells.append(slot_payload)

            required_count_raw = None
            required_count_numeric = None
            required_row_key = "day_need_row" if block_type == "day_support" else "night_need_row"
            required_row = rows_meta.get(required_row_key)
            if required_row:
                required_row = int(required_row)
                raw_value = sheet.cell(row=required_row, column=col_idx).value
                required_count_numeric, required_count_raw = _parse_daytime_need_value(raw_value)
                issue_code = None
                issue_message = None
                if required_count_numeric is None:
                    issue_code = "SUPPORT_BLOCK_REQUIRED_COUNT_INVALID"
                    issue_message = "필요 인원 수를 해석할 수 없습니다."
                    block_issues.append(issue_code)
                    issues.append(
                        _build_import_issue(
                            issue_code,
                            message=issue_message,
                            sheet_name=sheet.title,
                            row_no=required_row,
                            col_no=col_idx,
                            section=block_type,
                        )
                    )
                row_payload = append_support_value_row(
                    row_no=required_row,
                    col_no=col_idx,
                    schedule_date=schedule_date,
                    source_block=f"{block_type}_required_count",
                    section_label="필요인원 수",
                    raw_value=raw_value,
                    parsed_semantic_type="numeric_count" if required_count_numeric is not None else "invalid",
                    issue_code=issue_code,
                    issue_message=issue_message,
                )
                need_cells.append({**row_payload, "required_count_numeric": required_count_numeric})

            external_count_raw = None
            external_count_numeric = None
            vendor_row_key = "day_vendor_count_row" if block_type == "day_support" else "night_vendor_count_row"
            vendor_row = rows_meta.get(vendor_row_key)
            if vendor_row:
                vendor_row = int(vendor_row)
                raw_value = sheet.cell(row=vendor_row, column=col_idx).value
                external_count_numeric, external_count_raw = _parse_support_count_value(raw_value)
                issue_code = None
                issue_message = None
                if external_count_numeric is None:
                    issue_code = "UNSUPPORTED_CELL_FORMAT"
                    issue_message = "외부인원 투입 수를 숫자로 해석할 수 없습니다."
                    block_issues.append(issue_code)
                    issues.append(
                        _build_import_issue(
                            issue_code,
                            message=issue_message,
                            sheet_name=sheet.title,
                            row_no=vendor_row,
                            col_no=col_idx,
                            section=block_type,
                        )
                    )
                support_cells.append(
                    append_support_value_row(
                        row_no=vendor_row,
                        col_no=col_idx,
                        schedule_date=schedule_date,
                        source_block=f"{block_type}_external_count",
                        section_label="외부인원 투입 수",
                        raw_value=raw_value,
                        parsed_semantic_type="numeric_count" if external_count_numeric is not None else "invalid",
                        issue_code=issue_code,
                        issue_message=issue_message,
                    )
                )

            purpose_text = None
            if block_type == "night_support" and rows_meta.get("work_note_row"):
                work_note_row = int(rows_meta["work_note_row"])
                raw_value = sheet.cell(row=work_note_row, column=col_idx).value
                purpose_text = _normalize_workbook_display_value(raw_value) or None
                support_cells.append(
                    append_support_value_row(
                        row_no=work_note_row,
                        col_no=col_idx,
                        schedule_date=schedule_date,
                        source_block="night_support_purpose",
                        section_label="작업 목적",
                        raw_value=raw_value,
                        parsed_semantic_type="text" if purpose_text else "blank",
                    )
                )

            summary_row = rows_meta.get("weekly_count_row") if block_type == "day_support" else rows_meta.get("night_count_row")
            if summary_row:
                support_cells.append(
                    append_support_value_row(
                        row_no=int(summary_row),
                        col_no=col_idx,
                        schedule_date=schedule_date,
                        source_block=f"{block_type}_summary_count",
                        section_label="주간 추가 근무자 수" if block_type == "day_support" else "야간 근무자 총 수",
                        raw_value=sheet.cell(row=int(summary_row), column=col_idx).value,
                        parsed_semantic_type="protected_summary",
                        issue_code="PROTECTED_FIELD_IGNORED",
                        issue_message="요약 행은 검토만 가능하며 직접 반영되지 않습니다.",
                    )
                )

            support_blocks.append(
                {
                    "site": None,
                    "target_month": f"{month_ctx[0]:04d}-{month_ctx[1]:02d}" if month_ctx else None,
                    "target_date": schedule_date,
                    "block_type": block_type,
                    "required_count_raw": required_count_raw,
                    "required_count_numeric": required_count_numeric,
                    "external_count_raw": external_count_raw,
                    "external_count_numeric": external_count_numeric,
                    "purpose_text": purpose_text,
                    "worker_slots": worker_slots,
                    "worker_slot_count": len(block_rows),
                    "valid_filled_count": valid_filled_count,
                    "invalid_filled_count": invalid_filled_count,
                    "required_row_no": required_row,
                    "vendor_row_no": vendor_row,
                    "purpose_row_no": int(rows_meta["work_note_row"]) if block_type == "night_support" and rows_meta.get("work_note_row") else None,
                    "source_sheet": sheet.title,
                    "source_col": _excel_col_label(col_idx),
                    "issues": sorted(set(_normalize_import_issue_code(code) for code in block_issues if code)),
                }
            )

    return {
        "month_ctx": month_ctx,
        "body_cells": body_cells,
        "need_cells": need_cells,
        "support_cells": support_cells,
        "support_blocks": support_blocks,
        "issues": issues,
        "section_rows": rows_meta,
        "summary_start_row": summary_start_row,
    }


def _build_visible_value_index(rows: list[dict], *, include_employee: bool = True) -> dict[tuple[Any, ...], dict[str, Any]]:
    index: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        if include_employee:
            key = (
                _normalize_name_token(row.get("employee_name")),
                _normalize_schedule_template_duty_type(row.get("duty_type")),
                row.get("schedule_date").isoformat() if isinstance(row.get("schedule_date"), date) else str(row.get("schedule_date") or "").strip(),
            )
        else:
            key = (
                str(row.get("source_block") or "").strip(),
                row.get("schedule_date").isoformat() if isinstance(row.get("schedule_date"), date) else str(row.get("schedule_date") or "").strip(),
            )
        index[key] = dict(row)
    return index


def _build_support_value_index(rows: list[dict]) -> dict[tuple[str, str, int, int], dict[str, Any]]:
    index: dict[tuple[str, str, int, int], dict[str, Any]] = {}
    for row in rows:
        schedule_date = row.get("schedule_date")
        date_key = schedule_date.isoformat() if isinstance(schedule_date, date) else str(schedule_date or "").strip()
        source_block = str(row.get("source_block") or "").strip()
        row_no = int(row.get("row_no") or 0)
        col_no = int(row.get("col_no") or 0)
        if not source_block or not date_key or row_no <= 0:
            continue
        index[(source_block, date_key, row_no, col_no)] = dict(row)
    return index


def _collect_monthly_export_context(
    conn,
    *,
    target_tenant: dict,
    site_row: dict,
    month_key: str,
    user: dict,
    build_workbook: bool = True,
) -> dict[str, Any]:
    rows = [
        row
        for row in _read_monthly_board_rows_for_export(
            conn,
            tenant_id=str(target_tenant["id"]),
            month_key=month_key,
        )
        if str(row.get("site_code") or "").strip() == str(site_row.get("site_code") or "").strip()
    ]
    board_payload = monthly_board_lite(
        month=month_key,
        tenant_code=str(target_tenant.get("tenant_code") or "").strip(),
        conn=conn,
        user=user,
    )
    rows = _merge_export_rows_with_board_fallback(
        rows,
        _build_export_rows_from_board_payload(
            conn,
            tenant_id=str(target_tenant["id"]),
            board_payload=board_payload,
            site_code=str(site_row.get("site_code") or "").strip(),
        ),
    )
    support_rows = _read_monthly_support_assignment_rows_for_export(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_row["site_code"]),
        month_key=month_key,
    )
    overnight_rows = _read_monthly_overnight_rows_for_export(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_row["site_code"]),
        month_key=month_key,
    )
    employee_overnight_rows = _read_monthly_employee_overnight_rows_for_export(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_row["site_code"]),
        month_key=month_key,
    )
    daytime_need_rows = _read_monthly_daytime_need_rows_for_export(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month_key,
    )
    support_request_rows = _read_monthly_support_request_rows_for_export(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month_key,
    )
    if not rows and not support_rows and not overnight_rows and not employee_overnight_rows and not daytime_need_rows and not support_request_rows:
        raise HTTPException(status_code=404, detail="monthly schedule export data not found")
    employee_blocks = _build_export_employee_blocks(
        rows,
        support_rows=support_rows,
        overnight_rows=overnight_rows,
        employee_overnight_rows=employee_overnight_rows,
    )
    if not employee_blocks:
        raise HTTPException(status_code=409, detail="employee mapping unavailable for monthly export")
    workbook = None
    template_path = _resolve_arls_template_path()
    template_version = _describe_arls_template_version(template_path) if template_path else ARLS_EXPORT_TEMPLATE_VERSION
    export_revision = _build_schedule_export_revision(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        site_code=str(site_row["site_code"]),
        month_key=month_key,
    )
    parsed_sheet = None
    metadata = None
    if build_workbook:
        workbook, template_path = _load_required_arls_month_workbook()
        template_version = _describe_arls_template_version(template_path)
        _build_arls_month_sheet(
            workbook,
            month_key=month_key,
            rows=rows,
            tenant_code=str(target_tenant.get("tenant_code") or "").strip(),
            site_code=str(site_row.get("site_code") or "").strip(),
            site_name=str(site_row.get("site_name") or "").strip(),
            site_address=str(site_row.get("address") or "").strip(),
            support_rows=support_rows,
            support_request_rows=support_request_rows,
            overnight_rows=overnight_rows,
            employee_overnight_rows=employee_overnight_rows,
            daytime_need_rows=daytime_need_rows,
            export_revision=export_revision,
            template_version=template_version,
            source_version=ARLS_EXPORT_SOURCE_VERSION,
        )
        parsed_sheet = _parse_arls_canonical_import_sheet(workbook[ARLS_SHEET_NAME])
        metadata = _read_arls_export_metadata(workbook)
    return {
        "site_id": str(site_row.get("id") or "").strip(),
        "site_code": str(site_row.get("site_code") or "").strip(),
        "site_name": str(site_row.get("site_name") or "").strip(),
        "site_address": str(site_row.get("address") or "").strip(),
        "rows": rows,
        "support_rows": support_rows,
        "support_request_rows": support_request_rows,
        "overnight_rows": overnight_rows,
        "employee_overnight_rows": employee_overnight_rows,
        "daytime_need_rows": daytime_need_rows,
        "employee_blocks": employee_blocks,
        "workbook": workbook,
        "template_path": template_path,
        "template_version": template_version,
        "export_revision": export_revision,
        "parsed_sheet": parsed_sheet,
        "metadata": metadata,
    }


def _resolve_employee_by_full_name(conn, *, tenant_id: str, site_id: str, full_name: str) -> dict[str, Any] | None:
    normalized = str(full_name or "").strip()
    if not normalized:
        return None
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id,
                   tenant_id,
                   company_id,
                   site_id,
                   employee_code,
                   full_name,
                   sequence_no,
                   COALESCE(soc_role, '') AS soc_role,
                   COALESCE(duty_role, '') AS duty_role
            FROM employees
            WHERE tenant_id = %s
              AND site_id = %s
              AND lower(full_name) = lower(%s)
            ORDER BY sequence_no NULLS LAST, employee_code ASC
            LIMIT 1
            """,
            (tenant_id, site_id, normalized),
        )
        return cur.fetchone()


def _fetch_default_schedule_template_map(conn, *, tenant_id: str, site_id: str) -> dict[str, dict[str, Any]]:
    rows = _fetch_schedule_templates(conn, tenant_id=tenant_id, site_id=site_id, include_inactive=False)
    template_map: dict[str, dict[str, Any]] = {}
    for row in rows:
        duty_type = _normalize_schedule_template_duty_type(row.get("duty_type"))
        if duty_type not in {"day", "night", "overtime"}:
            continue
        current = template_map.get(duty_type)
        if row.get("is_default") and not current:
            template_map[duty_type] = dict(row)
            continue
        if current is None:
            template_map[duty_type] = dict(row)
    return template_map


def _build_support_form_sheet_name(site_name: str, site_code: str) -> str:
    preferred = str(site_name or "").strip() or str(site_code or "").strip() or "지원근무"
    cleaned = re.sub(r"[:\\\\/*?\\[\\]]", "_", preferred).strip()
    if len(cleaned) <= 31:
        return cleaned
    return cleaned[:31]


def _read_support_roundtrip_metadata(workbook: Workbook) -> dict[str, str]:
    if ARLS_SUPPORT_METADATA_SHEET_NAME not in workbook.sheetnames:
        return {}
    return _metadata_sheet_to_dict(workbook[ARLS_SUPPORT_METADATA_SHEET_NAME])


def _write_support_roundtrip_metadata_sheet(
    workbook: Workbook,
    *,
    tenant_code: str,
    site_code: str,
    site_name: str,
    month_key: str,
    source_revision: str,
    template_version: str,
    support_form_version: str,
) -> None:
    if ARLS_SUPPORT_METADATA_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[ARLS_SUPPORT_METADATA_SHEET_NAME]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(ARLS_SUPPORT_METADATA_SHEET_NAME)
    sheet.sheet_state = "hidden"
    entries = [
        ("tenant_code", tenant_code),
        ("site_code", site_code),
        ("site_name", site_name),
        ("month", month_key),
        ("source_revision", source_revision),
        ("template_version", template_version),
        ("support_form_version", support_form_version),
        ("extracted_at_kst", datetime.now(timezone(timedelta(hours=9))).isoformat()),
    ]
    for index, (key, value) in enumerate(entries, start=1):
        sheet.cell(row=index, column=1, value=key)
        sheet.cell(row=index, column=2, value=value)


def _validate_support_roundtrip_metadata(
    metadata: dict[str, str],
    *,
    expected_tenant_code: str,
    expected_site_code: str,
    expected_month: str,
    expected_source_revision: str,
) -> list[str]:
    errors: list[str] = []
    required_fields = (
        "tenant_code",
        "site_code",
        "month",
        "source_revision",
        "template_version",
        "support_form_version",
    )
    for field_name in required_fields:
        if not str(metadata.get(field_name) or "").strip():
            errors.append(f"metadata_missing:{field_name}")
    if errors:
        return errors
    if str(metadata.get("tenant_code") or "").strip() != expected_tenant_code:
        errors.append("metadata_mismatch:tenant_code")
    if str(metadata.get("site_code") or "").strip() != expected_site_code:
        errors.append("metadata_mismatch:site_code")
    if str(metadata.get("month") or "").strip() != expected_month:
        errors.append("metadata_mismatch:month")
    if str(metadata.get("source_revision") or "").strip() != expected_source_revision:
        errors.append("stale_source_revision")
    if not str(metadata.get("template_version") or "").strip().startswith(ARLS_EXPORT_TEMPLATE_VERSION):
        errors.append("metadata_mismatch:template_version")
    if str(metadata.get("support_form_version") or "").strip() != ARLS_SUPPORT_FORM_VERSION:
        errors.append("metadata_mismatch:support_form_version")
    return errors


def _build_sentrix_support_hq_sheet_name(site_name: str) -> str | None:
    normalized = str(site_name or "").strip()
    if not normalized:
        return None
    if len(normalized) > 31:
        return None
    if re.search(r"[:\\\\/*?\\[\\]]", normalized):
        return None
    return normalized


def _build_sentrix_support_hq_bundle_revision(site_entries: list[dict[str, Any]], *, month_key: str, scope: str) -> str:
    payload = {
        "month": month_key,
        "scope": scope,
        "sites": [
            {
                "site_code": str(entry.get("site_code") or "").strip(),
                "site_name": str(entry.get("site_name") or "").strip(),
                "source_revision": str(entry.get("source_revision") or "").strip(),
            }
            for entry in site_entries
        ],
    }
    digest = hashlib.sha1(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
    return digest[:16]


def _read_sentrix_support_hq_metadata(workbook: Workbook) -> dict[str, str]:
    if SENTRIX_SUPPORT_HQ_METADATA_SHEET_NAME not in workbook.sheetnames:
        return {}
    return _metadata_sheet_to_dict(workbook[SENTRIX_SUPPORT_HQ_METADATA_SHEET_NAME])


def _write_sentrix_support_hq_metadata_sheet(
    workbook: Workbook,
    *,
    tenant_code: str,
    month_key: str,
    download_scope: str,
    template_version: str,
    site_entries: list[dict[str, Any]],
    selected_site_code: str | None = None,
    selected_site_name: str | None = None,
) -> None:
    if SENTRIX_SUPPORT_HQ_METADATA_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SENTRIX_SUPPORT_HQ_METADATA_SHEET_NAME]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(SENTRIX_SUPPORT_HQ_METADATA_SHEET_NAME)
    sheet.sheet_state = "hidden"
    bundle_revision = _build_sentrix_support_hq_bundle_revision(site_entries, month_key=month_key, scope=download_scope)
    site_codes = [str(entry.get("site_code") or "").strip() for entry in site_entries if str(entry.get("site_code") or "").strip()]
    site_names = [str(entry.get("site_name") or "").strip() for entry in site_entries if str(entry.get("site_name") or "").strip()]
    site_name_code_map = {
        str(entry.get("site_name") or "").strip(): str(entry.get("site_code") or "").strip()
        for entry in site_entries
        if str(entry.get("site_name") or "").strip() and str(entry.get("site_code") or "").strip()
    }
    site_revision_map = {
        str(entry.get("site_code") or "").strip(): str(entry.get("source_revision") or "").strip()
        for entry in site_entries
        if str(entry.get("site_code") or "").strip()
    }
    entries = [
        ("tenant_code", tenant_code),
        ("month", month_key),
        ("download_scope", download_scope),
        ("workbook_family", ARLS_SUPPORT_FORM_VERSION),
        ("template_version", template_version),
        ("bundle_revision", bundle_revision),
        ("site_count", len(site_entries)),
        ("site_codes_json", json.dumps(site_codes, ensure_ascii=False)),
        ("site_names_json", json.dumps(site_names, ensure_ascii=False)),
        ("site_name_code_map_json", json.dumps(site_name_code_map, ensure_ascii=False)),
        ("site_revision_map_json", json.dumps(site_revision_map, ensure_ascii=False)),
        ("selected_site_code", str(selected_site_code or "").strip()),
        ("selected_site_name", str(selected_site_name or "").strip()),
        ("extracted_at_kst", datetime.now(timezone(timedelta(hours=9))).isoformat()),
    ]
    for index, (key, value) in enumerate(entries, start=1):
        sheet.cell(row=index, column=1, value=key)
        sheet.cell(row=index, column=2, value=value)


def _safe_load_metadata_json_list(value: object) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except Exception:
        return []
    if not isinstance(parsed, list):
        return []
    return [str(item or "").strip() for item in parsed if str(item or "").strip()]


def _safe_load_metadata_json_dict(value: object) -> dict[str, str]:
    raw = str(value or "").strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(parsed, dict):
        return {}
    result: dict[str, str] = {}
    for key, item in parsed.items():
        normalized_key = str(key or "").strip()
        normalized_value = str(item or "").strip()
        if normalized_key:
            result[normalized_key] = normalized_value
    return result


def _list_support_roundtrip_workspace_sites(
    conn,
    *,
    tenant_id: str,
    month_key: str,
) -> list[dict[str, Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT s.id,
                   s.site_code,
                   s.site_name,
                   src.id AS source_id,
                   COALESCE(src.state, 'source_missing') AS source_state,
                   src.source_revision,
                   src.latest_hq_revision,
                   COALESCE(src.hq_merge_stale, FALSE) AS hq_merge_stale
            FROM sites s
            LEFT JOIN schedule_support_roundtrip_sources src
                   ON src.tenant_id = s.tenant_id
                  AND src.site_id = s.id
                  AND src.month_key = %s
            WHERE s.tenant_id = %s
              AND COALESCE(s.is_active, TRUE) = TRUE
            ORDER BY s.site_name ASC, s.site_code ASC
            """,
            (month_key, tenant_id),
        )
        rows = [dict(row) for row in (cur.fetchall() or [])]
    results: list[dict[str, Any]] = []
    for row in rows:
        exact_sheet_name = _build_sentrix_support_hq_sheet_name(str(row.get("site_name") or "").strip())
        source_available = bool(row.get("source_id"))
        latest_status = "latest" if source_available and exact_sheet_name else ("sheet_name_invalid" if not exact_sheet_name else "source_missing")
        results.append(
            {
                "site_id": str(row.get("id") or "").strip(),
                "site_code": str(row.get("site_code") or "").strip(),
                "site_name": str(row.get("site_name") or "").strip(),
                "sheet_name": exact_sheet_name or str(row.get("site_name") or "").strip() or str(row.get("site_code") or "").strip(),
                "sheet_name_valid": bool(exact_sheet_name),
                "download_ready": bool(source_available and exact_sheet_name),
                "source_state": str(row.get("source_state") or "source_missing").strip() or "source_missing",
                "source_revision": str(row.get("source_revision") or "").strip() or None,
                "latest_hq_revision": str(row.get("latest_hq_revision") or "").strip() or None,
                "latest_status": latest_status,
            }
        )
    return results


def _clone_support_hq_sheet_to_workbook(source_sheet, *, target_workbook: Workbook, title: str) -> None:
    target_sheet = target_workbook.create_sheet(title=title)
    target_sheet.freeze_panes = source_sheet.freeze_panes
    target_sheet.sheet_format.defaultColWidth = source_sheet.sheet_format.defaultColWidth
    target_sheet.sheet_format.defaultRowHeight = source_sheet.sheet_format.defaultRowHeight
    target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
    target_sheet.sheet_view.zoomScale = source_sheet.sheet_view.zoomScale
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.print_options = copy(source_sheet.print_options)
    target_sheet.auto_filter.ref = source_sheet.auto_filter.ref
    for col_key, source_dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[col_key]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.bestFit = source_dimension.bestFit
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed
        target_dimension.style = source_dimension.style
    for row_key, source_dimension in source_sheet.row_dimensions.items():
        target_dimension = target_sheet.row_dimensions[row_key]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed
    for row in source_sheet.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target_sheet.cell(row=source_cell.row, column=source_cell.column, value=source_cell.value)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))


def _build_support_roundtrip_workspace_issue(
    code: str,
    *,
    count: int = 1,
    message: str | None = None,
) -> SupportRosterHqReviewIssueOut:
    catalog = {
        "WORKBOOK_METADATA_MISSING": ("blocking", "메타데이터 누락", "지원근무 workbook 메타데이터를 읽지 못했습니다.", "현재 Sentrix HQ 다운로드본으로 다시 시작하세요."),
        "WORKBOOK_FAMILY_MISMATCH": ("blocking", "workbook family 불일치", "지원하지 않는 지원근무 workbook family 입니다.", "현재 Sentrix HQ 다운로드본을 다시 내려받아 사용하세요."),
        "WORKBOOK_MONTH_MISMATCH": ("blocking", "대상월 불일치", "선택한 대상월과 업로드 파일의 대상월이 일치하지 않습니다.", "대상월을 맞추거나 올바른 workbook을 선택하세요."),
        "WORKBOOK_SCOPE_MISMATCH": ("blocking", "다운로드 범위 불일치", "업로드 파일의 범위와 현재 검토 컨텍스트가 맞지 않습니다.", "전체/지점별 범위를 다시 확인하세요."),
        "WORKBOOK_REVISION_STALE": ("blocking", "구버전 workbook", "현재 source revision 보다 오래된 workbook 입니다.", "최신 전체/지점별 다운로드본을 다시 사용하세요."),
        "WORKBOOK_SITE_SHEET_MISMATCH": ("blocking", "시트 구성 불일치", "workbook metadata와 실제 시트 구성이 일치하지 않습니다.", "시트명을 수정하지 말고 새 다운로드본으로 다시 작업하세요."),
        "WORKBOOK_SITE_NOT_FOUND": ("blocking", "지점 식별 실패", "시트명 또는 metadata 기준으로 지점을 찾지 못했습니다.", "지점명 변경 없이 최신 workbook을 사용하세요."),
    }
    severity, title, default_message, guidance = catalog.get(
        code,
        ("warning", code, message or code, None),
    )
    return SupportRosterHqReviewIssueOut(
        code=code,
        severity=severity,
        title=title,
        message=message or default_message,
        guidance=guidance,
        count=max(int(count or 1), 1),
    )


def _list_sentrix_hq_assignment_rows_for_download(
    conn,
    *,
    tenant_id: str,
    site_code: str,
    month_key: str,
    source_id: str | None = None,
    source_revision: str | None = None,
) -> list[dict[str, Any]]:
    month_rows = _read_monthly_support_assignment_rows_for_export(
        conn,
        tenant_id=tenant_id,
        site_code=site_code,
        month_key=month_key,
    )
    current_rows = [
        dict(row)
        for row in month_rows
        if str(row.get("source") or "").strip().upper() == SENTRIX_HQ_ROSTER_ASSIGNMENT_SOURCE
    ]
    if current_rows:
        return current_rows
    if source_id and source_revision:
        return _list_support_roundtrip_assignments(
            conn,
            source_id=source_id,
            source_revision=source_revision,
        )
    return []


def _sentrix_hq_issue_template(code: str) -> tuple[str, str, str, str | None]:
    catalog = {
        "WORKBOOK_METADATA_MISSING": ("blocking", "메타데이터 누락", "Sentrix HQ 다운로드 메타데이터를 읽지 못했습니다.", "현재 Sentrix HQ 다운로드본으로 다시 시작하세요."),
        "WORKBOOK_FAMILY_MISMATCH": ("blocking", "workbook family 불일치", "지원하지 않는 HQ roster workbook family 입니다.", "현재 Sentrix HQ 다운로드본을 다시 내려받아 사용하세요."),
        "WORKBOOK_MONTH_MISMATCH": ("blocking", "대상월 불일치", "선택한 대상월과 업로드 파일의 대상월이 일치하지 않습니다.", "대상월을 다시 확인하고 올바른 workbook을 선택하세요."),
        "WORKBOOK_SCOPE_MISMATCH": ("blocking", "다운로드 범위 불일치", "업로드 파일의 다운로드 범위가 현재 검토 컨텍스트와 맞지 않습니다.", "전체/지점별 범위를 다시 확인하세요."),
        "OUTDATED_WORKBOOK": ("blocking", "구버전 workbook", "현재 Supervisor 기준본보다 오래된 workbook 입니다.", "최신 전체/지점별 다운로드본을 다시 사용하세요."),
        "SITE_SHEET_NOT_FOUND": ("blocking", "시트 지점 식별 실패", "시트명으로 지점을 정확히 찾지 못했습니다.", "시트명을 수정하지 말고 최신 workbook을 다시 사용하세요."),
        "BLOCK_SECTION_NOT_FOUND": ("blocking", "지원 블록 누락", "주간 또는 야간 지원 블록을 찾지 못했습니다.", "다운로드한 workbook 구조를 변경하지 않았는지 확인하세요."),
        "DATE_SCOPE_NOT_RESOLVED": ("blocking", "날짜 범위 해석 실패", "시트의 날짜 헤더 또는 월 범위를 해석하지 못했습니다.", "월 헤더와 날짜 컬럼을 수정하지 않았는지 확인하세요."),
        "TICKET_SCOPE_NOT_FOUND": ("blocking", "지원요청 ticket 없음", "해당 지점/날짜/주야간 범위의 기존 Sentrix ticket을 찾지 못했습니다.", "ARLS 원본 업로드로 생성된 ticket scope를 먼저 확인하세요."),
        "MULTI_PERSON_CELL": ("blocking", "한 셀 다중 인원", "한 셀에는 1명만 입력할 수 있습니다.", "여러 명을 각각 다른 근무자 셀에 입력하세요."),
        "SELF_STAFF_FORMAT_INVALID": ("blocking", "자체 인원 표기 오류", "자체 인원은 정확히 '자체 {이름}' 형식만 허용됩니다.", "예: '자체 조태환' 형식으로 수정하세요."),
        "SELF_STAFF_EMPLOYEE_NOT_FOUND": ("blocking", "자체 인원 매칭 실패", "자체 인원을 해당 지점 active employee master에서 찾지 못했습니다.", "이름 또는 지점 소속 정보를 다시 확인하세요."),
        "SELF_STAFF_EMPLOYEE_AMBIGUOUS": ("blocking", "자체 인원 중복 매칭", "동일 이름의 active 직원이 2명 이상 있어 자동 매칭할 수 없습니다.", "employee master에서 중복 이름을 해소하세요."),
        "WORKER_CELL_INVALID": ("blocking", "근무자 셀 형식 오류", "근무자 셀 값을 해석할 수 없습니다.", "허용된 지원근무자 표기 형식으로 수정하세요."),
        "REPLACE_SCOPE_CONFLICT": ("blocking", "replace 범위 충돌", "동일 지점/날짜/주야간/슬롯 범위가 workbook 안에서 중복되었습니다.", "중복된 슬롯 입력을 제거하세요."),
        "REQUEST_COUNT_MISMATCH_UNDER": ("warning", "필요 인원 미달", "유효 입력 인원이 기존 ticket 요청 수보다 적습니다.", "현재 업로드를 적용하면 ticket 상태가 승인대기로 계산됩니다."),
        "REQUEST_COUNT_MISMATCH_OVER": ("warning", "필요 인원 초과", "유효 입력 인원이 기존 ticket 요청 수보다 많습니다.", "현재 업로드를 적용하면 ticket 상태가 승인대기로 계산됩니다."),
        "PURPOSE_FIELD_PARSE_WARNING": ("warning", "작업 목적 확인", "야간 작업 목적 셀을 참고 정보로만 유지합니다.", "기존 ticket 작업 목적은 유지되고 workbook 값은 정보성으로만 남습니다."),
    }
    return catalog.get(code, ("warning", code, code, None))


def _build_sentrix_hq_roster_issue(
    code: str,
    *,
    message: str | None = None,
    guidance: str | None = None,
    sheet_name: str | None = None,
    site_code: str | None = None,
    site_name: str | None = None,
    work_date: date | None = None,
    shift_kind: str | None = None,
) -> dict[str, Any]:
    severity, title, default_message, default_guidance = _sentrix_hq_issue_template(code)
    return {
        "code": code,
        "severity": severity,
        "title": title,
        "message": message or default_message,
        "guidance": guidance if guidance is not None else default_guidance,
        "sheet_name": sheet_name,
        "site_code": site_code,
        "site_name": site_name,
        "work_date": work_date,
        "shift_kind": shift_kind,
    }


def _group_sentrix_hq_roster_issues(raw_issues: list[dict[str, Any]]) -> list[SupportRosterHqReviewIssueOut]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for issue in raw_issues:
        code = str(issue.get("code") or "").strip() or "UNKNOWN"
        message = str(issue.get("message") or "").strip() or code
        key = (code, message)
        bucket = grouped.get(key)
        if bucket is None:
            bucket = dict(issue)
            bucket["count"] = 0
            grouped[key] = bucket
        bucket["count"] = int(bucket.get("count") or 0) + 1
    ordered = sorted(
        grouped.values(),
        key=lambda item: (
            0 if str(item.get("severity") or "").strip().lower() == "blocking" else 1,
            str(item.get("code") or "").strip(),
            str(item.get("message") or "").strip(),
        ),
    )
    return [
        SupportRosterHqReviewIssueOut(
            code=str(item.get("code") or "").strip() or "UNKNOWN",
            severity=str(item.get("severity") or "warning").strip() or "warning",
            title=str(item.get("title") or item.get("code") or "").strip() or "검토 이슈",
            message=str(item.get("message") or item.get("code") or "").strip() or "검토 이슈",
            guidance=str(item.get("guidance") or "").strip() or None,
            count=max(int(item.get("count") or 1), 1),
            sheet_name=str(item.get("sheet_name") or "").strip() or None,
            site_code=str(item.get("site_code") or "").strip() or None,
            site_name=str(item.get("site_name") or "").strip() or None,
            work_date=item.get("work_date") if isinstance(item.get("work_date"), date) else None,
            shift_kind=str(item.get("shift_kind") or "").strip() or None,
        )
        for item in ordered
    ]


def _extract_sentrix_ticket_hq_roster_status(ticket: dict[str, Any] | None) -> str | None:
    if not ticket:
        return None
    detail_json = ticket.get("detail_json")
    if not isinstance(detail_json, dict):
        return None
    hq_roster = detail_json.get("hq_roster")
    if not isinstance(hq_roster, dict):
        return None
    final_state = _normalize_sentrix_hq_roster_final_state(
        hq_roster.get("final_state") or hq_roster.get("state")
    )
    if final_state:
        return final_state
    normalized = str(hq_roster.get("status") or "").strip().lower()
    if normalized in {SENTRIX_HQ_ROSTER_AUTO_APPROVED_STATUS, SENTRIX_HQ_ROSTER_PENDING_STATUS}:
        return normalized
    return None


def _load_sentrix_support_ticket_scope_map(
    conn,
    *,
    tenant_id: str,
    month_key: str,
    site_codes: list[str],
) -> dict[tuple[str, str, str], dict[str, Any]]:
    start_date, end_date = _month_bounds(month_key)
    normalized_codes = [str(code or "").strip().upper() for code in site_codes if str(code or "").strip()]
    clauses = [
        "sr.tenant_id = %s",
        "sr.work_date >= %s",
        "sr.work_date < %s",
        "sr.source_workflow = %s",
        "sr.status = %s",
    ]
    params: list[Any] = [
        tenant_id,
        start_date,
        end_date,
        SENTRIX_SUPPORT_REQUEST_WORKFLOW,
        SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS,
    ]
    if normalized_codes:
        clauses.append("upper(sr.site_code) = ANY(%s)")
        params.append(normalized_codes)
    with conn.cursor() as cur:
        cur.execute(
            f"""
            SELECT sr.id,
                   sr.site_id,
                   sr.site_code,
                   s.site_name,
                   sr.work_date,
                   sr.shift_kind,
                   sr.request_count,
                   sr.work_purpose,
                   sr.status,
                   sr.source_revision,
                   sr.detail_json,
                   sr.updated_at
            FROM sentrix_support_request_tickets sr
            JOIN sites s ON s.id = sr.site_id
            WHERE {' AND '.join(clauses)}
            """,
            tuple(params),
        )
        rows = [dict(row) for row in (cur.fetchall() or [])]
    result: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        work_date = row.get("work_date")
        if not isinstance(work_date, date):
            continue
        site_code = str(row.get("site_code") or "").strip().upper()
        shift_kind = "night" if str(row.get("shift_kind") or "").strip().lower() == "night" else "day"
        if not site_code:
            continue
        result[(site_code, work_date.isoformat(), shift_kind)] = row
    return result


def _normalize_sentrix_hq_worker_affiliation(value: str | None) -> str:
    return str(value or "").strip().upper()


def _parse_sentrix_hq_worker_cell(
    raw_value: object,
    *,
    schedule_date: date,
    employee_index: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    raw_text = _normalize_workbook_display_value(raw_value)
    compact_text = re.sub(r"\s+", " ", raw_text).strip() if raw_text else ""
    if not compact_text or compact_text in {"-", "없음"}:
        return {
            "raw_text": raw_text,
            "compact_text": compact_text,
            "is_filled": False,
            "is_valid": True,
            "countable": False,
            "self_staff": False,
            "affiliation": "",
            "name": "",
            "display_value": None,
            "employee_id": None,
            "employee_code": None,
            "employee_name": None,
            "issue_code": None,
            "issue_message": None,
            "worker_type": None,
        }

    if compact_text.startswith("자체"):
        match = SENTRIX_HQ_ROSTER_SELF_STAFF_PATTERN.fullmatch(compact_text)
        if not match:
            return {
                "raw_text": raw_text,
                "compact_text": compact_text,
                "is_filled": True,
                "is_valid": False,
                "countable": False,
                "self_staff": True,
                "affiliation": "",
                "name": "",
                "display_value": compact_text,
                "employee_id": None,
                "employee_code": None,
                "employee_name": None,
                "issue_code": "SELF_STAFF_FORMAT_INVALID",
                "issue_message": _sentrix_hq_issue_template("SELF_STAFF_FORMAT_INVALID")[2],
                "worker_type": "INTERNAL",
            }
        employee_name = str(match.group(1) or "").strip()
        match_row, issue_code, issue_message = _resolve_import_employee_match(
            employee_index,
            employee_name=employee_name,
            schedule_date=schedule_date,
        )
        if issue_code == "EMPLOYEE_MATCH_FAILED":
            issue_code = "SELF_STAFF_EMPLOYEE_NOT_FOUND"
            issue_message = _sentrix_hq_issue_template(issue_code)[2]
        elif issue_code == "EMPLOYEE_MATCH_AMBIGUOUS":
            issue_code = "SELF_STAFF_EMPLOYEE_AMBIGUOUS"
            issue_message = _sentrix_hq_issue_template(issue_code)[2]
        return {
            "raw_text": raw_text,
            "compact_text": compact_text,
            "is_filled": True,
            "is_valid": not issue_code,
            "countable": not issue_code,
            "self_staff": True,
            "affiliation": "",
            "name": employee_name,
            "display_value": f"자체 {employee_name}".strip(),
            "employee_id": str((match_row or {}).get("id") or "").strip() or None,
            "employee_code": str((match_row or {}).get("employee_code") or "").strip() or None,
            "employee_name": str((match_row or {}).get("full_name") or employee_name).strip() or employee_name,
            "issue_code": issue_code,
            "issue_message": issue_message,
            "worker_type": "INTERNAL",
        }

    if SENTRIX_HQ_ROSTER_MULTI_PERSON_PATTERN.search(compact_text):
        return {
            "raw_text": raw_text,
            "compact_text": compact_text,
            "is_filled": True,
            "is_valid": False,
            "countable": False,
            "self_staff": False,
            "affiliation": "",
            "name": compact_text,
            "display_value": compact_text,
            "employee_id": None,
            "employee_code": None,
            "employee_name": None,
            "issue_code": "MULTI_PERSON_CELL",
            "issue_message": _sentrix_hq_issue_template("MULTI_PERSON_CELL")[2],
            "worker_type": None,
        }

    parts = compact_text.split(None, 1)
    if len(parts) > 1:
        affiliation = _normalize_sentrix_hq_worker_affiliation(parts[0])
        worker_name = str(parts[1] or "").strip()
    else:
        affiliation = ""
        worker_name = compact_text
    if not worker_name:
        return {
            "raw_text": raw_text,
            "compact_text": compact_text,
            "is_filled": True,
            "is_valid": False,
            "countable": False,
            "self_staff": False,
            "affiliation": affiliation,
            "name": "",
            "display_value": compact_text,
            "employee_id": None,
            "employee_code": None,
            "employee_name": None,
            "issue_code": "WORKER_CELL_INVALID",
            "issue_message": _sentrix_hq_issue_template("WORKER_CELL_INVALID")[2],
            "worker_type": None,
        }
    display_value = f"{affiliation} {worker_name}".strip() if affiliation else worker_name
    return {
        "raw_text": raw_text,
        "compact_text": compact_text,
        "is_filled": True,
        "is_valid": True,
        "countable": True,
        "self_staff": False,
        "affiliation": affiliation,
        "name": worker_name,
        "display_value": display_value,
        "employee_id": None,
        "employee_code": None,
        "employee_name": None,
        "issue_code": None,
        "issue_message": None,
        "worker_type": "BK" if affiliation == "BK" else "F",
    }


def _build_support_roster_hq_workspace_payload(
    conn,
    *,
    tenant_id: str,
    tenant_code: str,
    month_key: str,
) -> SupportRosterHqWorkspaceOut:
    sites = _list_support_roundtrip_workspace_sites(conn, tenant_id=tenant_id, month_key=month_key)
    return SupportRosterHqWorkspaceOut(
        tenant_code=tenant_code,
        month=month_key,
        default_scope="all",
        workbook_family=ARLS_SUPPORT_FORM_VERSION,
        template_version=ARLS_EXPORT_TEMPLATE_VERSION,
        latest_status="latest",
        total_site_count=len(sites),
        ready_site_count=sum(1 for site in sites if bool(site.get("download_ready"))),
        sites=[
            SupportRosterHqWorkspaceSiteOut(
                site_code=str(site.get("site_code") or "").strip(),
                site_name=str(site.get("site_name") or "").strip(),
                sheet_name=str(site.get("sheet_name") or "").strip(),
                download_ready=bool(site.get("download_ready")),
                source_state=str(site.get("source_state") or "source_missing").strip() or "source_missing",
                source_revision=str(site.get("source_revision") or "").strip() or None,
                latest_hq_revision=str(site.get("latest_hq_revision") or "").strip() or None,
                latest_status=str(site.get("latest_status") or "source_missing").strip() or "source_missing",
            )
            for site in sites
        ],
    )


def _resolve_support_roster_hq_download_sites(
    conn,
    *,
    tenant_id: str,
    month_key: str,
    scope: str,
    site_code: str | None = None,
) -> list[dict[str, Any]]:
    sites = _list_support_roundtrip_workspace_sites(conn, tenant_id=tenant_id, month_key=month_key)
    normalized_scope = "site" if str(scope or "").strip().lower() == "site" else "all"
    if normalized_scope == "site":
        normalized_site_code = str(site_code or "").strip()
        selected = next((site for site in sites if str(site.get("site_code") or "").strip() == normalized_site_code), None)
        if not selected:
            raise HTTPException(status_code=404, detail="site not found")
        if not bool(selected.get("download_ready")):
            raise HTTPException(status_code=409, detail="selected site workbook is not ready")
        return [selected]
    ready_sites = [site for site in sites if bool(site.get("download_ready"))]
    if not ready_sites:
        raise HTTPException(status_code=409, detail="no support workbook source is ready for the selected month")
    return ready_sites


def _build_support_roster_hq_source_map(
    conn,
    *,
    tenant_id: str,
    month_key: str,
    site_codes: list[str],
) -> dict[str, dict[str, Any]]:
    normalized_codes = [str(code or "").strip() for code in site_codes if str(code or "").strip()]
    if not normalized_codes:
        return {}
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT s.site_code,
                   s.site_name,
                   src.source_revision,
                   COALESCE(src.state, 'source_missing') AS source_state
            FROM sites s
            LEFT JOIN schedule_support_roundtrip_sources src
                   ON src.tenant_id = s.tenant_id
                  AND src.site_id = s.id
                  AND src.month_key = %s
            WHERE s.tenant_id = %s
              AND s.site_code = ANY(%s)
            """,
            (month_key, tenant_id, normalized_codes),
        )
        rows = [dict(row) for row in (cur.fetchall() or [])]
    return {
        str(row.get("site_code") or "").strip(): {
            "site_name": str(row.get("site_name") or "").strip(),
            "source_revision": str(row.get("source_revision") or "").strip() or None,
            "source_state": str(row.get("source_state") or "source_missing").strip() or "source_missing",
        }
        for row in rows
        if str(row.get("site_code") or "").strip()
    }


def _build_support_roster_hq_download_workbook(
    conn,
    *,
    target_tenant: dict,
    month_key: str,
    scope: str,
    selected_site_code: str | None,
    user: dict,
) -> tuple[Workbook, list[dict[str, Any]]]:
    selected_sites = _resolve_support_roster_hq_download_sites(
        conn,
        tenant_id=str(target_tenant["id"]),
        month_key=month_key,
        scope=scope,
        site_code=selected_site_code,
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    written_sites: list[dict[str, Any]] = []
    template_version = ARLS_EXPORT_TEMPLATE_VERSION
    for site_entry in selected_sites:
        exact_sheet_name = _build_sentrix_support_hq_sheet_name(str(site_entry.get("site_name") or "").strip())
        if not exact_sheet_name:
            raise HTTPException(status_code=409, detail=f"site name cannot be used as exact sheet name: {site_entry.get('site_name')}")
        site_row = _resolve_site_context_by_code(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_code=str(site_entry.get("site_code") or "").strip(),
        )
        if not site_row:
            raise HTTPException(status_code=404, detail="site not found")
        source_row = _get_support_roundtrip_source(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_id=str(site_row["id"]),
            month_key=month_key,
        )
        if not source_row:
            raise HTTPException(status_code=409, detail="support roundtrip source missing")
        export_ctx = _collect_monthly_export_context(
            conn,
            target_tenant=target_tenant,
            site_row=site_row,
            month_key=month_key,
            user=user,
        )
        template_version = str(export_ctx.get("template_version") or template_version)
        active_assignments = _list_sentrix_hq_assignment_rows_for_download(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_code=str(site_row.get("site_code") or "").strip(),
            month_key=month_key,
            source_id=str(source_row["id"]),
            source_revision=str(source_row.get("source_revision") or "").strip(),
        )
        site_workbook = _build_support_only_workbook(
            export_ctx=export_ctx,
            target_tenant=target_tenant,
            site_row=site_row,
            month_key=month_key,
            source_revision=str(source_row.get("source_revision") or "").strip(),
            active_assignments=active_assignments,
        )
        visible_sheet_name = next(
            (
                name
                for name in site_workbook.sheetnames
                if name not in {ARLS_SUPPORT_METADATA_SHEET_NAME, SENTRIX_SUPPORT_HQ_METADATA_SHEET_NAME}
            ),
            None,
        )
        if not visible_sheet_name:
            raise HTTPException(status_code=409, detail="support workbook sheet missing")
        _clone_support_hq_sheet_to_workbook(
            site_workbook[visible_sheet_name],
            target_workbook=workbook,
            title=exact_sheet_name,
        )
        written_sites.append(
            {
                "site_code": str(site_row.get("site_code") or "").strip(),
                "site_name": str(site_row.get("site_name") or "").strip(),
                "source_revision": str(source_row.get("source_revision") or "").strip(),
            }
        )
    if not workbook.worksheets:
        raise HTTPException(status_code=409, detail="support workbook generation failed")
    workbook.active = 0
    _write_sentrix_support_hq_metadata_sheet(
        workbook,
        tenant_code=str(target_tenant.get("tenant_code") or "").strip(),
        month_key=month_key,
        download_scope="site" if str(scope or "").strip().lower() == "site" else "all",
        template_version=template_version,
        site_entries=written_sites,
        selected_site_code=str(selected_site_code or "").strip() or None,
        selected_site_name=written_sites[0]["site_name"] if len(written_sites) == 1 else None,
    )
    return workbook, written_sites


def _build_support_roster_hq_upload_inspect_result(
    conn,
    *,
    workbook: Workbook,
    target_tenant: dict,
    selected_month: str,
    filename: str,
    user: dict,
) -> SupportRosterHqUploadInspectOut:
    raw_issues: list[dict[str, Any]] = []
    sentrix_metadata = _read_sentrix_support_hq_metadata(workbook)
    legacy_metadata = _read_support_roundtrip_metadata(workbook) if not sentrix_metadata else {}
    meta_sheet_names = {ARLS_SUPPORT_METADATA_SHEET_NAME, SENTRIX_SUPPORT_HQ_METADATA_SHEET_NAME}
    visible_sheet_names = [
        name
        for name in workbook.sheetnames
        if name not in meta_sheet_names and workbook[name].sheet_state == "visible"
    ]
    workspace_sites = _list_support_roundtrip_workspace_sites(
        conn,
        tenant_id=str(target_tenant["id"]),
        month_key=selected_month,
    )
    workspace_sites_by_name = {str(site.get("site_name") or "").strip(): dict(site) for site in workspace_sites}
    workspace_sites_by_code = {
        str(site.get("site_code") or "").strip().upper(): dict(site)
        for site in workspace_sites
        if str(site.get("site_code") or "").strip()
    }

    def add_issue(
        code: str,
        *,
        message: str | None = None,
        guidance: str | None = None,
        sheet_name: str | None = None,
        site_code: str | None = None,
        site_name: str | None = None,
        work_date: date | None = None,
        shift_kind: str | None = None,
    ) -> None:
        raw_issues.append(
            _build_sentrix_hq_roster_issue(
                code,
                message=message,
                guidance=guidance,
                sheet_name=sheet_name,
                site_code=site_code,
                site_name=site_name,
                work_date=work_date,
                shift_kind=shift_kind,
            )
        )

    download_scope = "all"
    workbook_family = None
    template_version = None
    revision = None
    file_month = None
    site_names: list[str] = []
    site_codes: list[str] = []
    selected_site_code = None
    selected_site_name = None
    site_revision_map: dict[str, str] = {}

    if sentrix_metadata:
        download_scope = str(sentrix_metadata.get("download_scope") or "all").strip().lower() or "all"
        workbook_family = str(sentrix_metadata.get("workbook_family") or "").strip() or None
        template_version = str(sentrix_metadata.get("template_version") or "").strip() or None
        revision = str(sentrix_metadata.get("bundle_revision") or "").strip() or None
        file_month = str(sentrix_metadata.get("month") or "").strip() or None
        site_names = _safe_load_metadata_json_list(sentrix_metadata.get("site_names_json"))
        site_codes = _safe_load_metadata_json_list(sentrix_metadata.get("site_codes_json"))
        site_revision_map = _safe_load_metadata_json_dict(sentrix_metadata.get("site_revision_map_json"))
        selected_site_code = str(sentrix_metadata.get("selected_site_code") or "").strip() or None
        selected_site_name = str(sentrix_metadata.get("selected_site_name") or "").strip() or None
    elif legacy_metadata:
        download_scope = "site"
        workbook_family = str(legacy_metadata.get("support_form_version") or "").strip() or None
        template_version = str(legacy_metadata.get("template_version") or "").strip() or None
        revision = str(legacy_metadata.get("source_revision") or "").strip() or None
        file_month = str(legacy_metadata.get("month") or "").strip() or None
        selected_site_code = str(legacy_metadata.get("site_code") or "").strip() or None
        selected_site_name = str(legacy_metadata.get("site_name") or "").strip() or None
        site_names = [selected_site_name] if selected_site_name else []
        site_codes = [selected_site_code] if selected_site_code else []
        if selected_site_code and revision:
            site_revision_map[selected_site_code] = revision
    else:
        add_issue("WORKBOOK_METADATA_MISSING")

    if workbook_family and workbook_family != ARLS_SUPPORT_FORM_VERSION:
        add_issue("WORKBOOK_FAMILY_MISMATCH")
    elif not workbook_family and sentrix_metadata:
        add_issue("WORKBOOK_FAMILY_MISMATCH")

    if file_month and file_month != selected_month:
        add_issue("WORKBOOK_MONTH_MISMATCH")

    if download_scope not in {"all", "site"}:
        add_issue("WORKBOOK_SCOPE_MISMATCH")

    expected_sheet_name_set = set(site_names or [])
    actual_sheet_name_set = set(visible_sheet_names)
    if expected_sheet_name_set and expected_sheet_name_set != actual_sheet_name_set:
        for missing_sheet in sorted(expected_sheet_name_set - actual_sheet_name_set):
            add_issue(
                "SITE_SHEET_NOT_FOUND",
                message=f"{missing_sheet} 시트가 workbook에서 누락되었습니다.",
                sheet_name=missing_sheet,
                site_name=missing_sheet,
            )
        for extra_sheet in sorted(actual_sheet_name_set - expected_sheet_name_set):
            add_issue(
                "SITE_SHEET_NOT_FOUND",
                message=f"{extra_sheet} 시트가 metadata 범위에 없습니다.",
                sheet_name=extra_sheet,
                site_name=extra_sheet,
            )

    inferred_site_codes: list[str] = []
    for sheet_name in visible_sheet_names:
        matched_site = workspace_sites_by_name.get(sheet_name)
        if not matched_site:
            add_issue(
                "SITE_SHEET_NOT_FOUND",
                message=f"{sheet_name} 시트와 일치하는 active 지점을 찾지 못했습니다.",
                sheet_name=sheet_name,
                site_name=sheet_name,
            )
            continue
        inferred_site_codes.append(str(matched_site.get("site_code") or "").strip().upper())
    if inferred_site_codes:
        site_codes = list(
            dict.fromkeys(
                [
                    *[str(code or "").strip().upper() for code in site_codes if str(code or "").strip()],
                    *inferred_site_codes,
                ]
            )
        )

    current_source_map = _build_support_roster_hq_source_map(
        conn,
        tenant_id=str(target_tenant["id"]),
        month_key=file_month or selected_month,
        site_codes=site_codes,
    )
    ticket_scope_map = _load_sentrix_support_ticket_scope_map(
        conn,
        tenant_id=str(target_tenant["id"]),
        month_key=selected_month,
        site_codes=site_codes,
    )

    employee_index_cache: dict[str, dict[str, list[dict[str, Any]]]] = {}
    raw_review_rows: list[dict[str, Any]] = []
    scope_summaries: list[SupportRosterHqScopeSummaryOut] = []
    valid_sheet_count = 0
    total_scope_count = 0
    valid_scope_count = 0

    for sheet_name in visible_sheet_names:
        sheet = workbook[sheet_name]
        site_entry = workspace_sites_by_name.get(sheet_name)
        site_code = str((site_entry or {}).get("site_code") or "").strip().upper()
        site_name = str((site_entry or {}).get("site_name") or sheet_name).strip() or sheet_name
        site_blocking_count_before = len(
            [issue for issue in raw_issues if str(issue.get("sheet_name") or "").strip() == sheet_name and str(issue.get("severity") or "").strip() == "blocking"]
        )

        if not site_entry or not site_code:
            raw_review_rows.append(
                {
                    "row_kind": "scope_summary",
                    "sheet_name": sheet_name,
                    "site_name": site_name,
                    "site_code": None,
                    "work_date": None,
                    "shift_kind": None,
                    "slot_index": 0,
                    "raw_cell_text": None,
                    "parsed_display_value": None,
                    "ticket_id": None,
                    "request_count": 0,
                    "valid_filled_count": 0,
                    "target_status": None,
                    "status": "blocking",
                    "reason": "시트명을 기준으로 지점을 정확히 찾지 못했습니다.",
                    "issue_code": "SITE_SHEET_NOT_FOUND",
                }
            )
            continue

        expected_revision = str(site_revision_map.get(site_code) or "").strip() or None
        current_source = current_source_map.get(site_code) or {}
        current_source_revision = str(current_source.get("source_revision") or "").strip() or None
        if expected_revision and expected_revision != current_source_revision:
            add_issue(
                "OUTDATED_WORKBOOK",
                message=f"{sheet_name} 시트의 기준 revision이 현재 source revision과 다릅니다.",
                sheet_name=sheet_name,
                site_code=site_code,
                site_name=site_name,
            )
        elif expected_revision and not current_source_revision:
            add_issue(
                "OUTDATED_WORKBOOK",
                message=f"{sheet_name} 시트의 현재 source revision을 찾지 못했습니다.",
                sheet_name=sheet_name,
                site_code=site_code,
                site_name=site_name,
            )

        date_columns, month_ctx = _extract_arls_date_columns(sheet)
        parsed_month = f"{month_ctx[0]:04d}-{month_ctx[1]:02d}" if month_ctx else None
        if not date_columns or parsed_month != selected_month:
            add_issue(
                "DATE_SCOPE_NOT_RESOLVED",
                message="시트의 날짜 헤더 또는 대상월을 현재 업로드 컨텍스트로 해석하지 못했습니다.",
                sheet_name=sheet_name,
                site_code=site_code,
                site_name=site_name,
            )
            raw_review_rows.append(
                {
                    "row_kind": "scope_summary",
                    "sheet_name": sheet_name,
                    "site_name": site_name,
                    "site_code": site_code,
                    "work_date": None,
                    "shift_kind": None,
                    "slot_index": 0,
                    "raw_cell_text": None,
                    "parsed_display_value": None,
                    "ticket_id": None,
                    "request_count": 0,
                    "valid_filled_count": 0,
                    "target_status": None,
                    "status": "blocking",
                    "reason": "날짜 범위를 해석하지 못했습니다.",
                    "issue_code": "DATE_SCOPE_NOT_RESOLVED",
                }
            )
            continue

        rows_meta = _locate_support_section_rows(sheet)
        if not list(rows_meta.get("weekly_rows") or []):
            add_issue(
                "BLOCK_SECTION_NOT_FOUND",
                message="주간 지원 블록을 찾지 못했습니다.",
                sheet_name=sheet_name,
                site_code=site_code,
                site_name=site_name,
                shift_kind="day",
            )
        if not list(rows_meta.get("night_rows") or []):
            add_issue(
                "BLOCK_SECTION_NOT_FOUND",
                message="야간 지원 블록을 찾지 못했습니다.",
                sheet_name=sheet_name,
                site_code=site_code,
                site_name=site_name,
                shift_kind="night",
            )

        employee_index = employee_index_cache.get(site_code)
        if employee_index is None:
            site_employee_rows = _load_site_employees(
                conn,
                tenant_id=str(target_tenant["id"]),
                site_id=str(site_entry.get("site_id") or "").strip(),
            )
            employee_index = _build_employee_name_index(site_employee_rows)
            employee_index_cache[site_code] = employee_index

        seen_scope_keys: set[str] = set()
        for shift_kind, block_rows, required_row_key, vendor_row_key, section_label in (
            ("day", list(rows_meta.get("weekly_rows") or []), "day_need_row", "day_vendor_count_row", "주간 지원"),
            ("night", list(rows_meta.get("night_rows") or []), "night_need_row", "night_vendor_count_row", "야간 지원"),
        ):
            if not block_rows:
                continue
            required_row_no = rows_meta.get(required_row_key)
            vendor_row_no = rows_meta.get(vendor_row_key)
            purpose_row_no = rows_meta.get("work_note_row") if shift_kind == "night" else None
            for col_idx, schedule_date in sorted(date_columns.items(), key=lambda item: item[1]):
                scope_key = f"{site_code}:{schedule_date.isoformat()}:{shift_kind}"
                if scope_key in seen_scope_keys:
                    add_issue(
                        "REPLACE_SCOPE_CONFLICT",
                        message="동일 지점/날짜/주야간 범위가 workbook 안에서 중복되었습니다.",
                        sheet_name=sheet_name,
                        site_code=site_code,
                        site_name=site_name,
                        work_date=schedule_date,
                        shift_kind=shift_kind,
                    )
                    continue
                seen_scope_keys.add(scope_key)

                scope_issue_start = len(raw_issues)
                workbook_required_raw = _normalize_workbook_display_value(sheet.cell(row=int(required_row_no), column=col_idx).value) if required_row_no else None
                workbook_required_count = None
                if workbook_required_raw:
                    workbook_required_count, _raw_text = _parse_daytime_need_value(workbook_required_raw)
                external_count_raw = _normalize_workbook_display_value(sheet.cell(row=int(vendor_row_no), column=col_idx).value) if vendor_row_no else None
                purpose_text = _normalize_workbook_display_value(sheet.cell(row=int(purpose_row_no), column=col_idx).value) if purpose_row_no else None

                ticket = ticket_scope_map.get((site_code, schedule_date.isoformat(), shift_kind))
                request_count = max(0, int((ticket or {}).get("request_count") or 0))
                current_status = _extract_sentrix_ticket_hq_roster_status(ticket) if ticket else None
                valid_filled_count = 0
                invalid_filled_count = 0
                meaningful_scope = False
                worker_rows: list[dict[str, Any]] = []
                seen_slot_indexes: set[int] = set()

                for position, row_no in enumerate(block_rows, start=1):
                    slot_index = _parse_worker_slot_number(sheet.cell(row=row_no, column=3).value) or position
                    if slot_index in seen_slot_indexes:
                        add_issue(
                            "REPLACE_SCOPE_CONFLICT",
                            message="동일 scope 안에 중복된 slot index가 있습니다.",
                            sheet_name=sheet_name,
                            site_code=site_code,
                            site_name=site_name,
                            work_date=schedule_date,
                            shift_kind=shift_kind,
                        )
                        continue
                    seen_slot_indexes.add(slot_index)
                    parsed_worker = _parse_sentrix_hq_worker_cell(
                        sheet.cell(row=row_no, column=col_idx).value,
                        schedule_date=schedule_date,
                        employee_index=employee_index,
                    )
                    if parsed_worker.get("is_filled"):
                        meaningful_scope = True
                    issue_code = str(parsed_worker.get("issue_code") or "").strip() or None
                    issue_message = str(parsed_worker.get("issue_message") or "").strip() or None
                    if issue_code:
                        invalid_filled_count += 1
                        add_issue(
                            issue_code,
                            message=issue_message,
                            sheet_name=sheet_name,
                            site_code=site_code,
                            site_name=site_name,
                            work_date=schedule_date,
                            shift_kind=shift_kind,
                        )
                    elif parsed_worker.get("is_filled"):
                        valid_filled_count += 1
                    if issue_code or parsed_worker.get("is_filled"):
                        worker_rows.append(
                            {
                                "row_kind": "worker",
                                "sheet_name": sheet_name,
                                "site_name": site_name,
                                "site_code": site_code,
                                "work_date": schedule_date,
                                "shift_kind": shift_kind,
                                "slot_index": slot_index,
                                "raw_cell_text": str(parsed_worker.get("raw_text") or "").strip() or None,
                                "parsed_display_value": str(parsed_worker.get("display_value") or "").strip() or None,
                                "ticket_id": ticket.get("id") if ticket else None,
                                "request_count": request_count,
                                "valid_filled_count": valid_filled_count,
                                "target_status": None,
                                "status": "blocking" if issue_code else "parsed",
                                "reason": issue_message if issue_code else "유효 입력",
                                "issue_code": issue_code,
                                "payload": {
                                    "scope_key": scope_key,
                                    "sheet_name": sheet_name,
                                    "site_id": str(site_entry.get("site_id") or "").strip() or None,
                                    "site_code": site_code,
                                    "site_name": site_name,
                                    "work_date": schedule_date.isoformat(),
                                    "shift_kind": shift_kind,
                                    "slot_index": slot_index,
                                    "ticket_id": str(ticket.get("id") or "").strip() if ticket else None,
                                    "request_count": request_count,
                                    "raw_cell_text": str(parsed_worker.get("raw_text") or "").strip() or None,
                                    "parsed_display_value": str(parsed_worker.get("display_value") or "").strip() or None,
                                    "self_staff": bool(parsed_worker.get("self_staff")),
                                    "affiliation": str(parsed_worker.get("affiliation") or "").strip() or None,
                                    "worker_name": str(parsed_worker.get("name") or "").strip() or None,
                                    "worker_type": str(parsed_worker.get("worker_type") or "").strip() or None,
                                    "employee_id": str(parsed_worker.get("employee_id") or "").strip() or None,
                                    "employee_code": str(parsed_worker.get("employee_code") or "").strip() or None,
                                    "employee_name": str(parsed_worker.get("employee_name") or "").strip() or None,
                                    "countable": bool(parsed_worker.get("countable")),
                                    "issue_code": issue_code,
                                    "issue_message": issue_message,
                                },
                            }
                        )

                if purpose_text and ticket and str(ticket.get("work_purpose") or "").strip() and str(ticket.get("work_purpose") or "").strip() != purpose_text:
                    add_issue(
                        "PURPOSE_FIELD_PARSE_WARNING",
                        message="야간 작업 목적은 현재 ticket 값을 유지하고 workbook 값은 참고 정보로만 보존합니다.",
                        sheet_name=sheet_name,
                        site_code=site_code,
                        site_name=site_name,
                        work_date=schedule_date,
                        shift_kind=shift_kind,
                    )

                if not ticket and meaningful_scope:
                    add_issue(
                        "TICKET_SCOPE_NOT_FOUND",
                        message="입력된 지원근무 범위와 매칭되는 기존 Sentrix ticket을 찾지 못했습니다.",
                        sheet_name=sheet_name,
                        site_code=site_code,
                        site_name=site_name,
                        work_date=schedule_date,
                        shift_kind=shift_kind,
                    )

                target_status = None
                scope_status = "empty"
                scope_reason = "입력 없음"
                if ticket:
                    if valid_filled_count == request_count:
                        target_status = SENTRIX_HQ_ROSTER_AUTO_APPROVED_STATUS
                        scope_status = SENTRIX_HQ_ROSTER_AUTO_APPROVED_STATUS
                        scope_reason = f"유효 {valid_filled_count}명 / 요청 {request_count}명"
                    elif valid_filled_count < request_count:
                        target_status = SENTRIX_HQ_ROSTER_PENDING_STATUS
                        scope_status = SENTRIX_HQ_ROSTER_PENDING_STATUS
                        scope_reason = f"유효 {valid_filled_count}명 / 요청 {request_count}명"
                        add_issue(
                            "REQUEST_COUNT_MISMATCH_UNDER",
                            sheet_name=sheet_name,
                            site_code=site_code,
                            site_name=site_name,
                            work_date=schedule_date,
                            shift_kind=shift_kind,
                        )
                    else:
                        target_status = SENTRIX_HQ_ROSTER_PENDING_STATUS
                        scope_status = SENTRIX_HQ_ROSTER_PENDING_STATUS
                        scope_reason = f"유효 {valid_filled_count}명 / 요청 {request_count}명"
                        add_issue(
                            "REQUEST_COUNT_MISMATCH_OVER",
                            sheet_name=sheet_name,
                            site_code=site_code,
                            site_name=site_name,
                            work_date=schedule_date,
                            shift_kind=shift_kind,
                        )
                elif meaningful_scope:
                    scope_status = "blocking"
                    scope_reason = "매칭되는 Sentrix ticket이 없습니다."

                scope_issues = raw_issues[scope_issue_start:]
                blocking_issue_count = sum(
                    1 for issue in scope_issues if str(issue.get("severity") or "").strip().lower() == "blocking"
                )
                warning_issue_count = sum(
                    1 for issue in scope_issues if str(issue.get("severity") or "").strip().lower() == "warning"
                )
                if ticket or meaningful_scope:
                    total_scope_count += 1
                    if ticket and blocking_issue_count == 0:
                        valid_scope_count += 1
                    scope_summary = SupportRosterHqScopeSummaryOut(
                        scope_key=scope_key,
                        sheet_name=sheet_name,
                        site_name=site_name,
                        site_code=site_code,
                        work_date=schedule_date,
                        shift_kind=shift_kind,
                        ticket_id=ticket.get("id") if ticket else None,
                        request_count=request_count,
                        valid_filled_count=valid_filled_count,
                        invalid_filled_count=invalid_filled_count,
                        target_status=target_status,
                        current_status=current_status,
                        workbook_required_count=workbook_required_count,
                        workbook_required_raw=workbook_required_raw,
                        external_count_raw=external_count_raw,
                        purpose_text=purpose_text,
                        matched_ticket=bool(ticket),
                        blocking_issue_count=blocking_issue_count,
                        warning_issue_count=warning_issue_count,
                    )
                    scope_summaries.append(scope_summary)
                    raw_review_rows.append(
                        {
                            "row_kind": "scope_summary",
                            "sheet_name": sheet_name,
                            "site_name": site_name,
                            "site_code": site_code,
                            "work_date": schedule_date,
                            "shift_kind": shift_kind,
                            "slot_index": 0,
                            "raw_cell_text": None,
                            "parsed_display_value": f"유효 {valid_filled_count}명 / 요청 {request_count}명" if ticket else f"유효 {valid_filled_count}명",
                            "ticket_id": ticket.get("id") if ticket else None,
                            "request_count": request_count,
                            "valid_filled_count": valid_filled_count,
                            "target_status": target_status,
                            "status": "blocking" if blocking_issue_count > 0 else scope_status,
                            "reason": scope_reason,
                            "issue_code": None,
                            "payload": {
                                "scope_key": scope_key,
                                "row_kind": "scope_summary",
                                "sheet_name": sheet_name,
                                "site_id": str(site_entry.get("site_id") or "").strip() or None,
                                "site_code": site_code,
                                "site_name": site_name,
                                "work_date": schedule_date.isoformat(),
                                "shift_kind": shift_kind,
                                "ticket_id": str(ticket.get("id") or "").strip() if ticket else None,
                                "request_count": request_count,
                                "valid_filled_count": valid_filled_count,
                                "invalid_filled_count": invalid_filled_count,
                                "target_status": target_status,
                                "current_status": current_status,
                                "workbook_required_count": workbook_required_count,
                                "workbook_required_raw": workbook_required_raw,
                                "external_count_raw": external_count_raw,
                                "purpose_text": purpose_text,
                                "matched_ticket": bool(ticket),
                                "blocking_issue_count": blocking_issue_count,
                                "warning_issue_count": warning_issue_count,
                            },
                        }
                    )
                    raw_review_rows.extend(worker_rows)

        site_blocking_count_after = len(
            [issue for issue in raw_issues if str(issue.get("sheet_name") or "").strip() == sheet_name and str(issue.get("severity") or "").strip() == "blocking"]
        )
        if site_blocking_count_after == site_blocking_count_before:
            valid_sheet_count += 1

    issue_list = _group_sentrix_hq_roster_issues(raw_issues)
    blocking_issue_count = sum(issue.count for issue in issue_list if issue.severity == "blocking")
    latest_status = "latest"
    if blocking_issue_count > 0:
        latest_status = "outdated" if any(issue.code == "OUTDATED_WORKBOOK" for issue in issue_list) else "invalid"

    raw_review_rows.sort(
        key=lambda item: (
            str(item.get("sheet_name") or ""),
            str(item.get("work_date") or ""),
            1 if str(item.get("shift_kind") or "").strip() == "night" else 0,
            0 if str(item.get("row_kind") or "").strip() == "scope_summary" else 1,
            int(item.get("slot_index") or 0),
        )
    )
    review_rows = [
        SupportRosterHqReviewRowOut(
            row_kind=str(item.get("row_kind") or "worker").strip() or "worker",
            sheet_name=str(item.get("sheet_name") or "").strip() or "-",
            site_name=str(item.get("site_name") or "").strip() or None,
            site_code=str(item.get("site_code") or "").strip() or None,
            work_date=item.get("work_date") if isinstance(item.get("work_date"), date) else None,
            shift_kind=str(item.get("shift_kind") or "").strip() or None,
            slot_index=max(int(item.get("slot_index") or 0), 0),
            raw_cell_text=str(item.get("raw_cell_text") or "").strip() or None,
            parsed_display_value=str(item.get("parsed_display_value") or "").strip() or None,
            ticket_id=item.get("ticket_id"),
            request_count=max(int(item.get("request_count") or 0), 0),
            valid_filled_count=max(int(item.get("valid_filled_count") or 0), 0),
            target_status=str(item.get("target_status") or "").strip() or None,
            status=str(item.get("status") or "pending").strip() or "pending",
            reason=str(item.get("reason") or "").strip() or None,
            issue_code=str(item.get("issue_code") or "").strip() or None,
        )
        for item in raw_review_rows
    ]

    summary = {
        "scope_total": total_scope_count,
        "scope_valid": valid_scope_count,
        "auto_approved": sum(1 for item in scope_summaries if item.target_status == SENTRIX_HQ_ROSTER_AUTO_APPROVED_STATUS and item.blocking_issue_count == 0),
        "approval_pending": sum(1 for item in scope_summaries if item.target_status == SENTRIX_HQ_ROSTER_PENDING_STATUS and item.blocking_issue_count == 0),
        "blocking": sum(1 for item in scope_summaries if item.blocking_issue_count > 0),
        "worker_rows": sum(1 for item in review_rows if item.row_kind == "worker"),
        "valid_workers": sum(int(item.valid_filled_count or 0) for item in scope_summaries),
    }

    upload_meta = SupportRosterHqUploadMetaOut(
        file_name=str(filename or "support_roster.xlsx").strip() or "support_roster.xlsx",
        month=file_month or selected_month,
        download_scope=download_scope,
        workbook_family=workbook_family,
        template_version=template_version,
        revision=revision,
        latest_status=latest_status,
        latest=latest_status == "latest",
        site_count=len(site_names or visible_sheet_names),
        site_names=list(site_names or visible_sheet_names),
        site_codes=list(site_codes),
        selected_site_code=selected_site_code,
        selected_site_name=selected_site_name,
    )
    can_apply = blocking_issue_count == 0 and valid_scope_count > 0
    next_step_message = (
        "차단 이슈를 모두 해결한 뒤 다시 검토하세요."
        if blocking_issue_count > 0
        else ("검토 완료. 적용하면 최신 roster snapshot으로 ticket 상태를 다시 계산합니다." if can_apply else "적용 가능한 ticket scope가 없습니다.")
    )
    batch_id = _persist_sentrix_hq_roster_preview_batch(
        conn,
        tenant_id=str(target_tenant["id"]),
        selected_month=selected_month,
        upload_meta=upload_meta,
        grouped_issues=issue_list,
        scope_summaries=scope_summaries,
        review_rows=raw_review_rows,
        summary=summary,
        site_revision_map=site_revision_map,
        user=user,
    )
    return SupportRosterHqUploadInspectOut(
        batch_id=batch_id,
        workbook_valid=blocking_issue_count == 0,
        can_apply=can_apply,
        upload_meta=upload_meta,
        total_sheet_count=len(visible_sheet_names),
        valid_sheet_count=valid_sheet_count,
        total_scope_count=total_scope_count,
        valid_scope_count=valid_scope_count,
        issue_count=sum(issue.count for issue in issue_list),
        summary=summary,
        issues=issue_list,
        scope_summaries=scope_summaries,
        review_rows=review_rows,
        next_step_message=next_step_message,
    )


def _persist_sentrix_hq_roster_preview_batch(
    conn,
    *,
    tenant_id: str,
    selected_month: str,
    upload_meta: SupportRosterHqUploadMetaOut,
    grouped_issues: list[SupportRosterHqReviewIssueOut],
    scope_summaries: list[SupportRosterHqScopeSummaryOut],
    review_rows: list[dict[str, Any]],
    summary: dict[str, int],
    site_revision_map: dict[str, str],
    user: dict,
) -> uuid.UUID:
    batch_id = uuid.uuid4()
    blocking_issue_count = sum(issue.count for issue in grouped_issues if issue.severity == "blocking")
    upload_meta_payload = upload_meta.model_dump(mode="json")
    upload_meta_payload["site_revision_map"] = dict(site_revision_map or {})
    summary_payload = {
        "summary": dict(summary or {}),
        "issues": [issue.model_dump(mode="json") for issue in grouped_issues],
        "scope_count": len(scope_summaries),
    }
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO sentrix_support_hq_roster_batches (
                id, tenant_id, month_key, download_scope, selected_site_code, filename,
                workbook_family, template_version, bundle_revision, latest_status, status,
                uploaded_by, uploaded_role, issue_count, blocking_issue_count,
                total_scope_count, valid_scope_count, upload_meta_json, summary_json,
                created_at
            )
            VALUES (
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s::jsonb, %s::jsonb,
                timezone('utc', now())
            )
            """,
            (
                batch_id,
                tenant_id,
                selected_month,
                str(upload_meta.download_scope or "all").strip() or "all",
                str(upload_meta.selected_site_code or "").strip() or None,
                str(upload_meta.file_name or "support_roster.xlsx").strip() or "support_roster.xlsx",
                str(upload_meta.workbook_family or "").strip() or None,
                str(upload_meta.template_version or "").strip() or None,
                str(upload_meta.revision or "").strip() or None,
                str(upload_meta.latest_status or "unknown").strip() or "unknown",
                "blocked" if blocking_issue_count > 0 else "previewed",
                user["id"],
                normalize_role(user.get("role")),
                sum(issue.count for issue in grouped_issues),
                blocking_issue_count,
                len(scope_summaries),
                max(0, int(summary.get("scope_valid") or 0)),
                json.dumps(upload_meta_payload, ensure_ascii=False, default=str),
                json.dumps(summary_payload, ensure_ascii=False, default=str),
            ),
        )
        for row in review_rows:
            payload = dict(row.get("payload") or {})
            work_date_raw = payload.get("work_date") or row.get("work_date")
            work_date_value = None
            if isinstance(work_date_raw, date):
                work_date_value = work_date_raw
            elif str(work_date_raw or "").strip():
                try:
                    work_date_value = date.fromisoformat(str(work_date_raw).strip())
                except Exception:
                    work_date_value = None
            cur.execute(
                """
                INSERT INTO sentrix_support_hq_roster_rows (
                    id, batch_id, tenant_id, sheet_name, site_id, site_code, site_name,
                    work_date, shift_kind, slot_index, row_kind, status, severity,
                    issue_code, ticket_id, raw_cell_text, parsed_display_value, effect_text,
                    payload_json, created_at
                )
                VALUES (
                    %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s::jsonb, timezone('utc', now())
                )
                """,
                (
                    uuid.uuid4(),
                    batch_id,
                    tenant_id,
                    str(row.get("sheet_name") or "").strip() or "-",
                    str(payload.get("site_id") or "").strip() or None,
                    str(row.get("site_code") or payload.get("site_code") or "").strip() or None,
                    str(row.get("site_name") or payload.get("site_name") or "").strip() or None,
                    work_date_value,
                    str(row.get("shift_kind") or payload.get("shift_kind") or "").strip() or None,
                    max(int(row.get("slot_index") or payload.get("slot_index") or 0), 0),
                    str(row.get("row_kind") or "worker").strip() or "worker",
                    str(row.get("status") or "pending").strip() or "pending",
                    "blocking" if str(row.get("status") or "").strip() == "blocking" else None,
                    str(row.get("issue_code") or payload.get("issue_code") or "").strip() or None,
                    str(row.get("ticket_id") or payload.get("ticket_id") or "").strip() or None,
                    str(row.get("raw_cell_text") or payload.get("raw_cell_text") or "").strip() or None,
                    str(row.get("parsed_display_value") or payload.get("parsed_display_value") or "").strip() or None,
                    str(row.get("reason") or "").strip() or None,
                    json.dumps(payload or row, ensure_ascii=False, default=str),
                ),
            )
    return batch_id


def _get_sentrix_hq_roster_batch(conn, *, batch_id: uuid.UUID, tenant_id: str) -> dict[str, Any] | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT *
            FROM sentrix_support_hq_roster_batches
            WHERE id = %s
              AND tenant_id = %s
            LIMIT 1
            """,
            (batch_id, tenant_id),
        )
        row = cur.fetchone()
    return dict(row) if row else None


def _load_sentrix_hq_roster_batch_rows(conn, *, batch_id: uuid.UUID) -> list[dict[str, Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT *
            FROM sentrix_support_hq_roster_rows
            WHERE batch_id = %s
            ORDER BY site_code ASC NULLS LAST, work_date ASC NULLS LAST, shift_kind ASC NULLS LAST, row_kind ASC, slot_index ASC
            """,
            (batch_id,),
        )
        return [dict(row) for row in (cur.fetchall() or [])]


def _validate_sentrix_hq_roster_batch_freshness(
    conn,
    *,
    tenant_id: str,
    batch: dict[str, Any],
    month_key: str,
) -> list[str]:
    upload_meta_json = batch.get("upload_meta_json") or {}
    if not isinstance(upload_meta_json, dict):
        return ["batch upload metadata is invalid"]
    expected_revision_map = {
        str(key or "").strip().upper(): str(value or "").strip()
        for key, value in dict(upload_meta_json.get("site_revision_map") or {}).items()
        if str(key or "").strip()
    }
    if not expected_revision_map:
        return []
    current_source_map = _build_support_roster_hq_source_map(
        conn,
        tenant_id=tenant_id,
        month_key=month_key,
        site_codes=list(expected_revision_map.keys()),
    )
    blocked_reasons: list[str] = []
    for site_code, expected_revision in expected_revision_map.items():
        current_revision = str((current_source_map.get(site_code) or {}).get("source_revision") or "").strip()
        if not current_revision or current_revision != expected_revision:
            blocked_reasons.append(f"{site_code} source revision이 변경되어 현재 upload batch는 stale 상태입니다.")
    return blocked_reasons


def _restore_sentrix_hq_roster_apply_result(batch: dict[str, Any] | None) -> SupportRosterHqApplyOut | None:
    if not batch:
        return None
    summary_json = batch.get("summary_json")
    if not isinstance(summary_json, dict):
        return None
    payload = summary_json.get("apply_result")
    if not isinstance(payload, dict):
        return None
    try:
        return SupportRosterHqApplyOut(**payload)
    except Exception:
        return None


def _write_sentrix_hq_roster_batch_apply_audit(
    cur,
    *,
    batch: dict[str, Any],
    batch_id: uuid.UUID,
    status: str,
    user: dict,
    result_json: dict[str, Any],
    error_text: str | None = None,
) -> None:
    summary_payload = dict(batch.get("summary_json") or {})
    summary_payload["apply_result"] = dict(result_json or {})
    summary_payload["apply_audit"] = {
        "status": status,
        "applied_by": str(user.get("id") or "").strip() or None,
        "applied_role": normalize_role(user.get("role")),
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "error_text": str(error_text or "").strip() or None,
    }
    upload_meta_payload = dict(batch.get("upload_meta_json") or {})
    upload_meta_payload["last_applied_by"] = str(user.get("id") or "").strip() or None
    upload_meta_payload["last_applied_role"] = normalize_role(user.get("role"))
    cur.execute(
        """
        UPDATE sentrix_support_hq_roster_batches
        SET status = %s,
            completed_at = timezone('utc', now()),
            upload_meta_json = %s::jsonb,
            summary_json = %s::jsonb
        WHERE id = %s
        """,
        (
            str(status or "failed").strip() or "failed",
            json.dumps(upload_meta_payload, ensure_ascii=False, default=str),
            json.dumps(summary_payload, ensure_ascii=False, default=str),
            batch_id,
        ),
    )


def _normalize_sentrix_hq_roster_final_state(value: object) -> str | None:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return None
    if normalized in {SENTRIX_HQ_ROSTER_AUTO_APPROVED_STATUS, SENTRIX_HQ_ROSTER_FINAL_APPROVED_STATE, "active"}:
        return SENTRIX_HQ_ROSTER_FINAL_APPROVED_STATE
    if normalized in {SENTRIX_HQ_ROSTER_PENDING_STATUS, "pending", "requested", "waiting"}:
        return SENTRIX_HQ_ROSTER_PENDING_STATUS
    if normalized in {"cancelled", "rejected", "deleted", "unavailable", "retracted"}:
        return normalized
    return normalized


def _extract_sentrix_hq_roster_final_state(ticket: dict[str, Any] | None) -> str | None:
    if not ticket:
        return None
    detail_json = ticket.get("detail_json")
    if not isinstance(detail_json, dict):
        return None
    hq_roster = detail_json.get("hq_roster")
    if not isinstance(hq_roster, dict):
        return None
    return _normalize_sentrix_hq_roster_final_state(
        hq_roster.get("final_state") or hq_roster.get("state") or hq_roster.get("status")
    )


def _sentrix_hq_final_state_is_approved(value: object) -> bool:
    return _normalize_sentrix_hq_roster_final_state(value) == SENTRIX_HQ_ROSTER_FINAL_APPROVED_STATE


def _build_sentrix_hq_roster_ticket_detail_json(
    ticket: dict[str, Any],
    *,
    batch_id: uuid.UUID,
    batch: dict[str, Any],
    user: dict,
    scope_payload: dict[str, Any],
    valid_worker_payloads: list[dict[str, Any]],
    target_status: str,
) -> dict[str, Any]:
    detail_json = dict(ticket.get("detail_json") or {})
    worker_values = [
        str(
            payload.get("parsed_display_value")
            or payload.get("display_value")
            or payload.get("raw_cell_text")
            or payload.get("worker_name")
            or ""
        ).strip()
        for payload in valid_worker_payloads
        if str(
            payload.get("parsed_display_value")
            or payload.get("display_value")
            or payload.get("raw_cell_text")
            or payload.get("worker_name")
            or ""
        ).strip()
    ]
    hq_roster_detail = dict(detail_json.get("hq_roster") or {})
    previous_status = str(hq_roster_detail.get("status") or "").strip().lower() or None
    previous_final_state = _normalize_sentrix_hq_roster_final_state(
        hq_roster_detail.get("final_state") or hq_roster_detail.get("state") or previous_status
    )
    final_state = _normalize_sentrix_hq_roster_final_state(target_status) or SENTRIX_HQ_ROSTER_PENDING_STATUS
    hq_roster_detail.update(
        {
            "status": target_status,
            "previous_status": previous_status,
            "previous_final_state": previous_final_state,
            "final_state": final_state,
            "state_transition": {
                "from": previous_final_state,
                "to": final_state,
            },
            "source": "sentrix_hq_roster_upload",
            "batch_id": str(batch_id),
            "month": str(batch.get("month_key") or "").strip() or None,
            "download_scope": str(batch.get("download_scope") or "").strip() or None,
            "workbook_family": str(batch.get("workbook_family") or "").strip() or None,
            "template_version": str(batch.get("template_version") or "").strip() or None,
            "bundle_revision": str(batch.get("bundle_revision") or "").strip() or None,
            "sheet_name": str(scope_payload.get("sheet_name") or "").strip() or None,
            "site_code": str(scope_payload.get("site_code") or "").strip() or None,
            "site_name": str(scope_payload.get("site_name") or "").strip() or None,
            "work_date": str(scope_payload.get("work_date") or "").strip() or None,
            "shift_kind": str(scope_payload.get("shift_kind") or "").strip() or None,
            "request_count": max(int(scope_payload.get("request_count") or 0), 0),
            "valid_filled_count": max(int(scope_payload.get("valid_filled_count") or 0), 0),
            "invalid_filled_count": max(int(scope_payload.get("invalid_filled_count") or 0), 0),
            "assignment_count": len(valid_worker_payloads),
            "worker_values": worker_values,
            "workbook_required_count": _coerce_int_or_none(scope_payload.get("workbook_required_count")),
            "workbook_required_raw": str(scope_payload.get("workbook_required_raw") or "").strip() or None,
            "external_count_raw": str(scope_payload.get("external_count_raw") or "").strip() or None,
            "purpose_text": str(scope_payload.get("purpose_text") or "").strip() or None,
            "applied_by": str(user.get("id") or "").strip() or None,
            "applied_role": normalize_role(user.get("role")),
            "applied_at": datetime.now(timezone.utc).isoformat(),
        }
    )
    detail_json["hq_roster"] = hq_roster_detail
    return detail_json


def _build_sentrix_hq_snapshot_entries(
    *,
    valid_worker_payloads: list[dict[str, Any]],
    spec: dict[str, Any],
    batch_id: uuid.UUID,
) -> list[dict[str, Any]]:
    snapshot_entries: list[dict[str, Any]] = []
    for worker_payload in sorted(
        list(valid_worker_payloads or []),
        key=lambda item: max(int(item.get("slot_index") or 0), 0),
    ):
        slot_index = max(int(worker_payload.get("slot_index") or 0), 1)
        raw_cell_text = str(worker_payload.get("raw_cell_text") or "").strip() or None
        normalized_affiliation = str(worker_payload.get("affiliation") or "").strip().upper() or None
        normalized_name = str(
            worker_payload.get("worker_name")
            or worker_payload.get("employee_name")
            or worker_payload.get("name")
            or ""
        ).strip() or None
        employee_id = str(worker_payload.get("employee_id") or "").strip() or None
        employee_code = str(worker_payload.get("employee_code") or "").strip() or None
        employee_name = str(worker_payload.get("employee_name") or normalized_name or "").strip() or None
        display_value = str(
            worker_payload.get("parsed_display_value")
            or worker_payload.get("display_value")
            or raw_cell_text
            or employee_name
            or ""
        ).strip() or None
        snapshot_entries.append(
            {
                "slot_index": slot_index,
                "raw_cell_text": raw_cell_text,
                "normalized_affiliation": normalized_affiliation,
                "normalized_name": normalized_name,
                "display_value": display_value,
                "self_staff": bool(worker_payload.get("self_staff")),
                "employee_id": employee_id,
                "employee_code": employee_code,
                "employee_name": employee_name,
                "worker_type": str(worker_payload.get("worker_type") or "").strip().upper() or None,
                "validity_state": "valid",
                "issue_code": str(worker_payload.get("issue_code") or "").strip() or None,
                "sheet_name": str(spec.get("sheet_name") or "").strip() or None,
                "site_id": str(spec.get("ticket", {}).get("site_id") or "").strip() or None,
                "site_code": str(spec.get("site_code") or "").strip() or None,
                "site_name": str(spec.get("site_name") or "").strip() or None,
                "work_date": spec["work_date"].isoformat(),
                "shift_kind": str(spec.get("shift_kind") or "").strip() or None,
                "ticket_id": str(spec.get("ticket_id") or "").strip() or None,
                "upload_batch_id": str(batch_id),
            }
        )
    return snapshot_entries


def _build_sentrix_hq_snapshot_signature(
    *,
    entries: list[dict[str, Any]],
    request_count: int,
    valid_filled_count: int,
    invalid_filled_count: int,
    ticket_state: str | None,
) -> str:
    normalized_entries = [
        {
            "slot_index": max(int(entry.get("slot_index") or 0), 0),
            "raw_cell_text": str(entry.get("raw_cell_text") or "").strip() or None,
            "normalized_affiliation": str(entry.get("normalized_affiliation") or "").strip().upper() or None,
            "normalized_name": str(entry.get("normalized_name") or "").strip() or None,
            "display_value": str(entry.get("display_value") or "").strip() or None,
            "self_staff": bool(entry.get("self_staff")),
            "employee_id": str(entry.get("employee_id") or "").strip() or None,
            "employee_code": str(entry.get("employee_code") or "").strip() or None,
            "employee_name": str(entry.get("employee_name") or "").strip() or None,
            "worker_type": str(entry.get("worker_type") or "").strip().upper() or None,
            "validity_state": str(entry.get("validity_state") or "").strip().lower() or None,
            "issue_code": str(entry.get("issue_code") or "").strip() or None,
        }
        for entry in sorted(
            list(entries or []),
            key=lambda item: (
                max(int(item.get("slot_index") or 0), 0),
                str(item.get("display_value") or ""),
                str(item.get("employee_id") or ""),
            ),
        )
    ]
    payload = {
        "request_count": max(int(request_count or 0), 0),
        "valid_filled_count": max(int(valid_filled_count or 0), 0),
        "invalid_filled_count": max(int(invalid_filled_count or 0), 0),
        "ticket_state": _normalize_sentrix_hq_roster_final_state(ticket_state),
        "entries": normalized_entries,
    }
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()[:24]


def _load_current_sentrix_hq_snapshot_map(
    conn,
    *,
    tenant_id: str,
    ticket_ids: list[str],
) -> dict[str, dict[str, Any]]:
    normalized_ticket_ids = sorted({str(ticket_id or "").strip() for ticket_id in ticket_ids if str(ticket_id or "").strip()})
    if not normalized_ticket_ids:
        return {}
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id,
                   batch_id,
                   ticket_id,
                   site_id,
                   site_code,
                   site_name,
                   month_key,
                   work_date,
                   shift_kind,
                   sheet_name,
                   previous_snapshot_id,
                   previous_ticket_state,
                   ticket_state,
                   request_count,
                   valid_filled_count,
                   invalid_filled_count,
                   changed,
                   is_current,
                   created_at
            FROM sentrix_support_roster_snapshots
            WHERE tenant_id = %s
              AND is_current = TRUE
              AND ticket_id::text = ANY(%s)
            """,
            (tenant_id, normalized_ticket_ids),
        )
        snapshot_rows = [dict(row) for row in (cur.fetchall() or [])]
    if not snapshot_rows:
        return {}

    snapshot_id_map = {
        str(row.get("id") or "").strip(): row
        for row in snapshot_rows
        if str(row.get("id") or "").strip()
    }
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT snapshot_id,
                   slot_index,
                   raw_cell_text,
                   normalized_affiliation,
                   normalized_name,
                   display_value,
                   self_staff,
                   employee_id,
                   employee_code,
                   employee_name,
                   worker_type,
                   validity_state,
                   issue_code
            FROM sentrix_support_roster_snapshot_entries
            WHERE snapshot_id::text = ANY(%s)
            ORDER BY slot_index ASC, id ASC
            """,
            (list(snapshot_id_map.keys()),),
        )
        entry_rows = [dict(row) for row in (cur.fetchall() or [])]

    bundle_map: dict[str, dict[str, Any]] = {}
    for row in snapshot_rows:
        ticket_id = str(row.get("ticket_id") or "").strip()
        if not ticket_id:
            continue
        bundle_map[ticket_id] = {
            "snapshot": row,
            "entries": [],
        }
    for row in entry_rows:
        snapshot_id = str(row.get("snapshot_id") or "").strip()
        snapshot = snapshot_id_map.get(snapshot_id)
        if not snapshot:
            continue
        ticket_id = str(snapshot.get("ticket_id") or "").strip()
        if not ticket_id or ticket_id not in bundle_map:
            continue
        bundle_map[ticket_id]["entries"].append(row)
    return bundle_map


def _persist_sentrix_hq_roster_snapshot(
    cur,
    *,
    tenant_id: str,
    batch_id: uuid.UUID,
    spec: dict[str, Any],
    ticket_state: str,
    previous_bundle: dict[str, Any] | None,
) -> tuple[uuid.UUID, bool, list[dict[str, Any]], str | None]:
    previous_snapshot = dict((previous_bundle or {}).get("snapshot") or {})
    previous_entries = list((previous_bundle or {}).get("entries") or [])
    snapshot_entries = _build_sentrix_hq_snapshot_entries(
        valid_worker_payloads=spec["valid_worker_payloads"],
        spec=spec,
        batch_id=batch_id,
    )
    previous_signature = _build_sentrix_hq_snapshot_signature(
        entries=previous_entries,
        request_count=max(int(previous_snapshot.get("request_count") or 0), 0),
        valid_filled_count=max(int(previous_snapshot.get("valid_filled_count") or 0), 0),
        invalid_filled_count=max(int(previous_snapshot.get("invalid_filled_count") or 0), 0),
        ticket_state=str(previous_snapshot.get("ticket_state") or "").strip() or None,
    ) if previous_snapshot else None
    current_signature = _build_sentrix_hq_snapshot_signature(
        entries=snapshot_entries,
        request_count=spec["request_count"],
        valid_filled_count=spec["valid_filled_count"],
        invalid_filled_count=spec["invalid_filled_count"],
        ticket_state=ticket_state,
    )
    changed = previous_signature != current_signature
    previous_snapshot_id = str(previous_snapshot.get("id") or "").strip() or None
    previous_ticket_state = _normalize_sentrix_hq_roster_final_state(previous_snapshot.get("ticket_state"))

    if previous_snapshot_id:
        cur.execute(
            """
            UPDATE sentrix_support_roster_snapshots
            SET is_current = FALSE
            WHERE id = %s
            """,
            (previous_snapshot_id,),
        )

    snapshot_id = uuid.uuid4()
    cur.execute(
        """
        INSERT INTO sentrix_support_roster_snapshots (
            id, tenant_id, batch_id, ticket_id, site_id, site_code, site_name, month_key,
            work_date, shift_kind, sheet_name, previous_snapshot_id, previous_ticket_state,
            ticket_state, request_count, valid_filled_count, invalid_filled_count,
            changed, is_current, created_at, updated_at
        )
        VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, TRUE, timezone('utc', now()), timezone('utc', now())
        )
        """,
        (
            snapshot_id,
            tenant_id,
            batch_id,
            spec["ticket_id"],
            spec["ticket"]["site_id"],
            str(spec.get("site_code") or "").strip() or None,
            str(spec.get("site_name") or "").strip() or None,
            str(spec.get("month_key") or "").strip() or None,
            spec["work_date"],
            str(spec.get("shift_kind") or "").strip() or None,
            str(spec.get("sheet_name") or "").strip() or None,
            previous_snapshot_id,
            previous_ticket_state,
            ticket_state,
            spec["request_count"],
            spec["valid_filled_count"],
            spec["invalid_filled_count"],
            changed,
        ),
    )

    for entry in snapshot_entries:
        cur.execute(
            """
            INSERT INTO sentrix_support_roster_snapshot_entries (
                id, snapshot_id, tenant_id, batch_id, ticket_id, sheet_name, site_id, site_code, site_name, work_date,
                shift_kind, slot_index, raw_cell_text, normalized_affiliation, normalized_name,
                display_value, self_staff, employee_id, employee_code, employee_name,
                worker_type, validity_state, issue_code, upload_scope_key, payload_json,
                created_at, updated_at
            )
            VALUES (
                arls_random_uuid(), %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s::jsonb,
                timezone('utc', now()), timezone('utc', now())
            )
            """,
            (
                snapshot_id,
                tenant_id,
                batch_id,
                spec["ticket_id"],
                str(spec.get("sheet_name") or "").strip() or None,
                spec["ticket"]["site_id"],
                str(entry.get("site_code") or "").strip() or None,
                str(entry.get("site_name") or "").strip() or None,
                spec["work_date"],
                str(entry.get("shift_kind") or "").strip() or None,
                max(int(entry.get("slot_index") or 0), 1),
                entry.get("raw_cell_text"),
                entry.get("normalized_affiliation"),
                entry.get("normalized_name"),
                entry.get("display_value"),
                bool(entry.get("self_staff")),
                entry.get("employee_id"),
                entry.get("employee_code"),
                entry.get("employee_name"),
                entry.get("worker_type"),
                entry.get("validity_state"),
                entry.get("issue_code"),
                str(spec.get("scope_key") or "").strip() or None,
                json.dumps(entry, ensure_ascii=False, default=str),
            ),
        )

    return snapshot_id, changed, snapshot_entries, previous_ticket_state


def _build_sentrix_hq_bridge_candidates(entries: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    candidates: dict[str, dict[str, Any]] = {}
    for entry in entries or []:
        if not bool(entry.get("self_staff")):
            continue
        employee_id = str(entry.get("employee_id") or "").strip()
        if not employee_id:
            continue
        if str(entry.get("validity_state") or "").strip().lower() != "valid":
            continue
        if employee_id in candidates:
            continue
        candidates[employee_id] = {
            "employee_id": employee_id,
            "employee_code": str(entry.get("employee_code") or "").strip() or None,
            "employee_name": str(entry.get("employee_name") or entry.get("normalized_name") or "").strip() or None,
            "display_value": str(entry.get("display_value") or "").strip() or None,
        }
    return candidates


def _queue_sentrix_hq_arls_bridge_actions(
    cur,
    *,
    batch_id: uuid.UUID,
    snapshot_id: uuid.UUID,
    tenant_id: str,
    spec: dict[str, Any],
    ticket_state: str,
    previous_ticket_state: str | None,
    previous_entries: list[dict[str, Any]],
    current_entries: list[dict[str, Any]],
) -> dict[str, int]:
    previous_candidates = _build_sentrix_hq_bridge_candidates(previous_entries)
    current_candidates = _build_sentrix_hq_bridge_candidates(current_entries)
    previous_approved = _sentrix_hq_final_state_is_approved(previous_ticket_state)
    current_approved = _sentrix_hq_final_state_is_approved(ticket_state)

    upsert_ids: set[str] = set()
    retract_ids: set[str] = set()
    previous_ids = set(previous_candidates.keys())
    current_ids = set(current_candidates.keys())
    if current_approved:
        if previous_approved:
            upsert_ids = current_ids - previous_ids
            retract_ids = previous_ids - current_ids
        else:
            upsert_ids = current_ids
    elif previous_approved:
        retract_ids = previous_ids

    inserted_total = 0
    upsert_total = 0
    retract_total = 0
    for employee_id in sorted(upsert_ids):
        candidate = current_candidates.get(employee_id) or {}
        payload = {
            "source": SENTRIX_ARLS_BRIDGE_SOURCE,
            "source_ticket_id": str(spec["ticket_id"]),
            "ticket_state": ticket_state,
            "upload_batch_id": str(batch_id),
            "tenant_id": tenant_id,
            "site_id": str(spec["ticket"].get("site_id") or "").strip() or None,
            "site_code": str(spec.get("site_code") or "").strip() or None,
            "site_name": str(spec.get("site_name") or "").strip() or None,
            "work_date": spec["work_date"].isoformat(),
            "shift_kind": str(spec.get("shift_kind") or "").strip() or None,
            "employee_id": employee_id,
            "employee_code": candidate.get("employee_code"),
            "employee_display_name": candidate.get("employee_name") or candidate.get("display_value"),
            "self_staff": True,
            "action": SENTRIX_ARLS_BRIDGE_ACTION_UPSERT,
            "snapshot_id": str(snapshot_id),
            "scope_key": str(spec.get("scope_key") or "").strip() or None,
        }
        idempotency_key = (
            f"sentrix-hq-bridge:{spec['ticket_id']}:{spec['work_date'].isoformat()}:"
            f"{spec['shift_kind']}:{employee_id}:{SENTRIX_ARLS_BRIDGE_ACTION_UPSERT}:{batch_id}"
        )
        cur.execute(
            """
            INSERT INTO sentrix_support_arls_bridge_actions (
                id, tenant_id, batch_id, snapshot_id, ticket_id, site_id, site_code, work_date, shift_kind,
                employee_id, employee_code, employee_name, action, ticket_state, self_staff,
                source, idempotency_key, status, payload_json, created_at, updated_at
            )
            VALUES (
                arls_random_uuid(), %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, TRUE,
                %s, %s, 'pending', %s::jsonb, timezone('utc', now()), timezone('utc', now())
            )
            ON CONFLICT (idempotency_key) DO NOTHING
            """,
            (
                tenant_id,
                batch_id,
                snapshot_id,
                spec["ticket_id"],
                spec["ticket"]["site_id"],
                str(spec.get("site_code") or "").strip() or None,
                spec["work_date"],
                str(spec.get("shift_kind") or "").strip() or None,
                employee_id,
                candidate.get("employee_code"),
                candidate.get("employee_name"),
                SENTRIX_ARLS_BRIDGE_ACTION_UPSERT,
                ticket_state,
                SENTRIX_ARLS_BRIDGE_SOURCE,
                idempotency_key,
                json.dumps(payload, ensure_ascii=False, default=str),
            ),
        )
        inserted = max(int(cur.rowcount or 0), 0)
        inserted_total += inserted
        upsert_total += inserted

    retract_source = previous_candidates if previous_approved else current_candidates
    for employee_id in sorted(retract_ids):
        candidate = retract_source.get(employee_id) or previous_candidates.get(employee_id) or current_candidates.get(employee_id) or {}
        payload = {
            "source": SENTRIX_ARLS_BRIDGE_SOURCE,
            "source_ticket_id": str(spec["ticket_id"]),
            "ticket_state": ticket_state,
            "upload_batch_id": str(batch_id),
            "tenant_id": tenant_id,
            "site_id": str(spec["ticket"].get("site_id") or "").strip() or None,
            "site_code": str(spec.get("site_code") or "").strip() or None,
            "site_name": str(spec.get("site_name") or "").strip() or None,
            "work_date": spec["work_date"].isoformat(),
            "shift_kind": str(spec.get("shift_kind") or "").strip() or None,
            "employee_id": employee_id,
            "employee_code": candidate.get("employee_code"),
            "employee_display_name": candidate.get("employee_name") or candidate.get("display_value"),
            "self_staff": True,
            "action": SENTRIX_ARLS_BRIDGE_ACTION_RETRACT,
            "snapshot_id": str(snapshot_id),
            "scope_key": str(spec.get("scope_key") or "").strip() or None,
            "previous_ticket_state": previous_ticket_state,
        }
        idempotency_key = (
            f"sentrix-hq-bridge:{spec['ticket_id']}:{spec['work_date'].isoformat()}:"
            f"{spec['shift_kind']}:{employee_id}:{SENTRIX_ARLS_BRIDGE_ACTION_RETRACT}:{batch_id}"
        )
        cur.execute(
            """
            INSERT INTO sentrix_support_arls_bridge_actions (
                id, tenant_id, batch_id, snapshot_id, ticket_id, site_id, site_code, work_date, shift_kind,
                employee_id, employee_code, employee_name, action, ticket_state, self_staff,
                source, idempotency_key, status, payload_json, created_at, updated_at
            )
            VALUES (
                arls_random_uuid(), %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, TRUE,
                %s, %s, 'pending', %s::jsonb, timezone('utc', now()), timezone('utc', now())
            )
            ON CONFLICT (idempotency_key) DO NOTHING
            """,
            (
                tenant_id,
                batch_id,
                snapshot_id,
                spec["ticket_id"],
                spec["ticket"]["site_id"],
                str(spec.get("site_code") or "").strip() or None,
                spec["work_date"],
                str(spec.get("shift_kind") or "").strip() or None,
                employee_id,
                candidate.get("employee_code"),
                candidate.get("employee_name"),
                SENTRIX_ARLS_BRIDGE_ACTION_RETRACT,
                ticket_state,
                SENTRIX_ARLS_BRIDGE_SOURCE,
                idempotency_key,
                json.dumps(payload, ensure_ascii=False, default=str),
            ),
        )
        inserted = max(int(cur.rowcount or 0), 0)
        inserted_total += inserted
        retract_total += inserted

    return {
        "created": inserted_total,
        "upserts": upsert_total,
        "retracts": retract_total,
    }


def _resolve_sentrix_hq_notification_user_ids(
    conn,
    *,
    tenant_id: str,
    site_id: str,
) -> list[str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT au.id,
                   au.role,
                   COALESCE(au.site_id, e.site_id) AS scoped_site_id
            FROM arls_users au
            LEFT JOIN employees e ON e.id = au.employee_id
            WHERE au.tenant_id = %s
              AND au.is_active = TRUE
              AND COALESCE(au.is_deleted, FALSE) = FALSE
            """,
            (tenant_id,),
        )
        rows = [dict(row) for row in (cur.fetchall() or [])]
    recipients: list[str] = []
    seen: set[str] = set()
    for row in rows:
        user_id = str(row.get("id") or "").strip()
        if not user_id or user_id in seen:
            continue
        role = normalize_user_role(row.get("role"))
        scoped_site_id = str(row.get("scoped_site_id") or "").strip()
        if role in {"hq_admin", "developer"}:
            recipients.append(user_id)
            seen.add(user_id)
            continue
        if role in {"supervisor", "vice_supervisor"} and scoped_site_id == site_id:
            recipients.append(user_id)
            seen.add(user_id)
    return recipients


def _insert_sentrix_hq_notification_audit(
    cur,
    *,
    tenant_id: str,
    batch_id: uuid.UUID,
    site_id: str,
    site_code: str,
    message: str,
    dedupe_key: str,
    recipient_user_ids: list[str],
    payload: dict[str, Any],
) -> uuid.UUID:
    audit_id = uuid.uuid4()
    cur.execute(
        """
        INSERT INTO sentrix_support_notification_audit (
            id, tenant_id, batch_id, site_id, site_code, message, dedupe_key, toast_recipient_count,
            push_target_count, push_sent_count, push_failed_count, recipient_user_ids_json,
            payload_json, status, created_at, updated_at
        )
        VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s,
            0, 0, 0, %s::jsonb,
            %s::jsonb, 'pending', timezone('utc', now()), timezone('utc', now())
        )
        """,
        (
            audit_id,
            tenant_id,
            batch_id,
            site_id,
            site_code or None,
            message,
            dedupe_key,
            len(recipient_user_ids),
            json.dumps(list(recipient_user_ids or []), ensure_ascii=False, default=str),
            json.dumps(payload or {}, ensure_ascii=False, default=str),
        ),
    )
    return audit_id


def _update_sentrix_hq_notification_audit_after_push(
    conn,
    *,
    audit_id: uuid.UUID,
    push_result: dict[str, int],
    error_text: str | None = None,
) -> None:
    target_count = max(int((push_result or {}).get("target_count") or 0), 0)
    sent_count = max(int((push_result or {}).get("sent_count") or 0), 0)
    failed_count = max(int((push_result or {}).get("failed_count") or 0), 0)
    status = "sent"
    if failed_count > 0 and sent_count > 0:
        status = "partial_failed"
    elif failed_count > 0:
        status = "failed"
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE sentrix_support_notification_audit
            SET push_target_count = %s,
                push_sent_count = %s,
                push_failed_count = %s,
                status = %s,
                error_text = %s,
                updated_at = timezone('utc', now())
            WHERE id = %s
            """,
            (
                target_count,
                sent_count,
                failed_count,
                status,
                str(error_text or "").strip() or None,
                audit_id,
            ),
        )


def _resolve_sentrix_support_materialized_shift_defaults(
    employee: dict[str, Any],
    *,
    shift_kind: str,
) -> tuple[str, str, float]:
    normalized_shift_kind = str(shift_kind or "").strip().lower()
    if normalized_shift_kind == "night":
        return DEFAULT_NIGHT_SHIFT_START, DEFAULT_NIGHT_SHIFT_END, DEFAULT_NIGHT_SHIFT_HOURS
    soc_role = str(employee.get("soc_role") or "").strip().lower()
    duty_role = str(employee.get("duty_role") or "").strip().upper()
    if soc_role in {"supervisor", "vice_supervisor", "hq_admin"} or duty_role in {TEAM_MANAGER_DUTY_ROLE, VICE_SUPERVISOR_DUTY_ROLE}:
        return SUPERVISOR_DAY_SHIFT_START, SUPERVISOR_DAY_SHIFT_END, SUPERVISOR_DAY_SHIFT_HOURS
    return GUARD_DAY_SHIFT_START, GUARD_DAY_SHIFT_END, GUARD_DAY_SHIFT_HOURS


def _resolve_sentrix_support_bridge_site(
    conn,
    *,
    tenant_id: str,
    site_id: str | None,
    site_code: str | None,
) -> dict[str, Any] | None:
    normalized_site_id = str(site_id or "").strip()
    normalized_site_code = str(site_code or "").strip()
    query = [
        """
        SELECT s.id,
               s.site_code,
               s.site_name,
               s.company_id,
               c.company_code
        FROM sites s
        JOIN companies c ON c.id = s.company_id
        WHERE s.tenant_id = %s
          AND COALESCE(s.is_active, TRUE) = TRUE
        """
    ]
    params: list[Any] = [tenant_id]
    if normalized_site_id:
        query.append("AND s.id = %s")
        params.append(normalized_site_id)
    elif normalized_site_code:
        query.append("AND s.site_code = %s")
        params.append(normalized_site_code)
    else:
        return None
    query.append("LIMIT 1")
    with conn.cursor() as cur:
        cur.execute("\n".join(query), tuple(params))
        row = cur.fetchone()
    return dict(row) if row else None


def _resolve_sentrix_support_bridge_employee(
    conn,
    *,
    tenant_id: str,
    site_id: str,
    employee_id: str,
    work_date: date,
) -> dict[str, Any] | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id,
                   tenant_id,
                   company_id,
                   site_id,
                   employee_code,
                   full_name,
                   COALESCE(soc_role, '') AS soc_role,
                   COALESCE(duty_role, '') AS duty_role,
                   hire_date,
                   leave_date
            FROM employees
            WHERE tenant_id = %s
              AND id = %s
              AND site_id = %s
            LIMIT 1
            """,
            (tenant_id, employee_id, site_id),
        )
        row = cur.fetchone()
    if not row:
        return None
    employee = dict(row)
    if not _employee_is_active_for_schedule_date(employee, work_date):
        return None
    return employee


def _load_sentrix_support_materialization_row(
    cur,
    *,
    tenant_id: str,
    ticket_id: str,
    site_id: str,
    work_date: date,
    shift_kind: str,
    employee_id: str,
) -> dict[str, Any] | None:
    cur.execute(
        """
        SELECT id,
               bridge_action_id,
               batch_id,
               snapshot_id,
               monthly_schedule_id,
               coexistence_mode,
               status,
               ticket_state,
               source_action,
               result_json,
               payload_json
        FROM sentrix_support_schedule_materializations
        WHERE tenant_id = %s
          AND ticket_id = %s
          AND site_id = %s
          AND work_date = %s
          AND shift_kind = %s
          AND employee_id = %s
        LIMIT 1
        """,
        (tenant_id, ticket_id, site_id, work_date, shift_kind, employee_id),
    )
    row = cur.fetchone()
    return dict(row) if row else None


def _load_monthly_schedule_row_for_shift(
    cur,
    *,
    tenant_id: str,
    employee_id: str,
    work_date: date,
    shift_kind: str,
) -> dict[str, Any] | None:
    cur.execute(
        """
        SELECT id,
               tenant_id,
               company_id,
               site_id,
               employee_id,
               schedule_date,
               shift_type,
               template_id,
               shift_start_time,
               shift_end_time,
               paid_hours,
               schedule_note,
               source,
               source_batch_id,
               source_revision,
               source_ticket_uuid,
               source_ticket_state,
               source_action,
               source_self_staff
        FROM monthly_schedules
        WHERE tenant_id = %s
          AND employee_id = %s
          AND schedule_date = %s
          AND lower(COALESCE(NULLIF(trim(shift_type), ''), 'day')) = %s
        LIMIT 1
        """,
        (tenant_id, employee_id, work_date, shift_kind),
    )
    row = cur.fetchone()
    return dict(row) if row else None


def _upsert_sentrix_support_materialization_row(
    cur,
    *,
    materialization_id: str | None,
    tenant_id: str,
    ticket_id: str,
    bridge_action_id: str,
    batch_id: str | None,
    snapshot_id: str | None,
    site_id: str,
    site_code: str,
    work_date: date,
    shift_kind: str,
    employee_id: str,
    employee_code: str | None,
    employee_name: str | None,
    monthly_schedule_id: str | None,
    coexistence_mode: str,
    status: str,
    ticket_state: str,
    source_action: str,
    payload: dict[str, Any],
    result_json: dict[str, Any],
    error_text: str | None = None,
) -> str:
    next_id = materialization_id or str(uuid.uuid4())
    cur.execute(
        """
        INSERT INTO sentrix_support_schedule_materializations (
            id, tenant_id, ticket_id, bridge_action_id, batch_id, snapshot_id, site_id, site_code,
            work_date, shift_kind, employee_id, employee_code, employee_name, self_staff,
            monthly_schedule_id, coexistence_mode, status, ticket_state, source,
            source_action, payload_json, result_json, error_text, created_at, updated_at, retracted_at
        )
        VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, TRUE,
            %s, %s, %s, %s, %s,
            %s, %s::jsonb, %s::jsonb, %s, timezone('utc', now()), timezone('utc', now()),
            CASE WHEN %s = %s THEN timezone('utc', now()) ELSE NULL END
        )
        ON CONFLICT (tenant_id, ticket_id, site_id, work_date, shift_kind, employee_id)
        DO UPDATE SET
            bridge_action_id = EXCLUDED.bridge_action_id,
            batch_id = EXCLUDED.batch_id,
            snapshot_id = EXCLUDED.snapshot_id,
            site_code = EXCLUDED.site_code,
            employee_code = EXCLUDED.employee_code,
            employee_name = EXCLUDED.employee_name,
            monthly_schedule_id = EXCLUDED.monthly_schedule_id,
            coexistence_mode = EXCLUDED.coexistence_mode,
            status = EXCLUDED.status,
            ticket_state = EXCLUDED.ticket_state,
            source_action = EXCLUDED.source_action,
            payload_json = EXCLUDED.payload_json,
            result_json = EXCLUDED.result_json,
            error_text = EXCLUDED.error_text,
            updated_at = timezone('utc', now()),
            retracted_at = CASE
                WHEN EXCLUDED.status = %s THEN timezone('utc', now())
                ELSE NULL
            END
        RETURNING id
        """,
        (
            next_id,
            tenant_id,
            ticket_id,
            bridge_action_id,
            batch_id,
            snapshot_id,
            site_id,
            site_code,
            work_date,
            shift_kind,
            employee_id,
            employee_code,
            employee_name,
            monthly_schedule_id,
            coexistence_mode,
            status,
            ticket_state,
            SENTRIX_ARLS_BRIDGE_SOURCE,
            source_action,
            json.dumps(payload or {}, ensure_ascii=False, default=str),
            json.dumps(result_json or {}, ensure_ascii=False, default=str),
            str(error_text or "").strip() or None,
            status,
            SENTRIX_SUPPORT_MATERIALIZATION_STATUS_RETRACTED,
            SENTRIX_SUPPORT_MATERIALIZATION_STATUS_RETRACTED,
        ),
    )
    row = cur.fetchone()
    return str((row or {}).get("id") or next_id)


def _mark_sentrix_support_bridge_action(
    cur,
    *,
    action_id: str,
    status: str,
    result_json: dict[str, Any] | None = None,
    error_text: str | None = None,
) -> None:
    cur.execute(
        """
        UPDATE sentrix_support_arls_bridge_actions
        SET status = %s,
            error_text = %s,
            result_json = %s::jsonb,
            processed_at = timezone('utc', now()),
            updated_at = timezone('utc', now())
        WHERE id = %s
        """,
        (
            status,
            str(error_text or "").strip() or None,
            json.dumps(result_json or {}, ensure_ascii=False, default=str),
            action_id,
        ),
    )


def _apply_sentrix_support_bridge_action(
    cur,
    *,
    tenant_id: str,
    action_row: dict[str, Any],
) -> dict[str, Any]:
    bridge_action_id = str(action_row.get("id") or "").strip()
    ticket_id = str(action_row.get("ticket_id") or "").strip()
    batch_id = str(action_row.get("batch_id") or "").strip() or None
    snapshot_id = str(action_row.get("snapshot_id") or "").strip() or None
    payload = dict(action_row.get("payload_json") or {})
    shift_kind = _normalize_support_status_shift_kind(action_row.get("shift_kind") or payload.get("shift_kind"))
    bridge_action = str(action_row.get("action") or payload.get("action") or "").strip().upper()
    ticket_state = _normalize_sentrix_hq_roster_final_state(action_row.get("ticket_state") or payload.get("ticket_state"))
    work_date = action_row.get("work_date")
    if not bridge_action or bridge_action not in {SENTRIX_ARLS_BRIDGE_ACTION_UPSERT, SENTRIX_ARLS_BRIDGE_ACTION_RETRACT}:
        raise RuntimeError("unsupported Sentrix bridge action")
    if not isinstance(work_date, date):
        raise RuntimeError("work date is missing for Sentrix bridge action")

    site = _resolve_sentrix_support_bridge_site(
        cur.connection,
        tenant_id=tenant_id,
        site_id=str(action_row.get("site_id") or payload.get("site_id") or "").strip() or None,
        site_code=str(action_row.get("site_code") or payload.get("site_code") or "").strip() or None,
    )
    if not site:
        raise RuntimeError("Sentrix bridge site could not be resolved")

    employee_id = str(action_row.get("employee_id") or payload.get("employee_id") or "").strip()
    if not employee_id:
        raise RuntimeError("Sentrix bridge employee id is missing")
    employee = _resolve_sentrix_support_bridge_employee(
        cur.connection,
        tenant_id=tenant_id,
        site_id=str(site["id"]),
        employee_id=employee_id,
        work_date=work_date,
    )
    if not employee:
        raise RuntimeError("Sentrix bridge employee could not be resolved as same-site active employee")

    existing_materialization = _load_sentrix_support_materialization_row(
        cur,
        tenant_id=tenant_id,
        ticket_id=ticket_id,
        site_id=str(site["id"]),
        work_date=work_date,
        shift_kind=shift_kind,
        employee_id=employee_id,
    )
    existing_schedule = _load_monthly_schedule_row_for_shift(
        cur,
        tenant_id=tenant_id,
        employee_id=employee_id,
        work_date=work_date,
        shift_kind=shift_kind,
    )

    if bridge_action == SENTRIX_ARLS_BRIDGE_ACTION_UPSERT and not _sentrix_hq_final_state_is_approved(ticket_state):
        raise RuntimeError("non-approved ticket state cannot materialize into ARLS schedule truth")

    materialization_result: dict[str, Any] = {
        "bridge_action_id": bridge_action_id,
        "ticket_id": ticket_id,
        "site_code": str(site.get("site_code") or "").strip() or None,
        "work_date": work_date.isoformat(),
        "shift_kind": shift_kind,
        "employee_id": employee_id,
        "employee_code": str(employee.get("employee_code") or "").strip() or None,
        "employee_name": str(employee.get("full_name") or action_row.get("employee_name") or payload.get("employee_display_name") or "").strip() or None,
        "ticket_state": ticket_state,
        "action": bridge_action,
    }

    if bridge_action == SENTRIX_ARLS_BRIDGE_ACTION_UPSERT:
        shift_start_time, shift_end_time, paid_hours = _resolve_sentrix_support_materialized_shift_defaults(
            employee,
            shift_kind=shift_kind,
        )
        schedule_row_id: str | None = None
        coexistence_mode = SENTRIX_SUPPORT_MATERIALIZATION_MODE_OWNED
        schedule_effect = "noop"
        if existing_schedule:
            existing_source = str(existing_schedule.get("source") or "").strip()
            existing_ticket_uuid = str(existing_schedule.get("source_ticket_uuid") or "").strip()
            if existing_source == SENTRIX_ARLS_BRIDGE_SOURCE and existing_ticket_uuid and existing_ticket_uuid != ticket_id:
                raise RuntimeError("another Sentrix support lineage already owns this employee/date/shift row")
            if existing_source == SENTRIX_ARLS_BRIDGE_SOURCE:
                _update_monthly_schedule_row(
                    cur,
                    schedule_id=str(existing_schedule["id"]),
                    shift_type=shift_kind,
                    template_id=None,
                    shift_start_time=shift_start_time,
                    shift_end_time=shift_end_time,
                    paid_hours=paid_hours,
                    schedule_note=str(existing_schedule.get("schedule_note") or "").strip() or None,
                    source=SENTRIX_ARLS_BRIDGE_SOURCE,
                    source_batch_id=batch_id,
                    source_revision=snapshot_id,
                    source_ticket_uuid=ticket_id,
                    source_ticket_state=ticket_state,
                    source_action=SENTRIX_ARLS_BRIDGE_ACTION_UPSERT,
                    source_self_staff=True,
                )
                schedule_row_id = str(existing_schedule["id"])
                schedule_effect = "updated"
            else:
                schedule_row_id = str(existing_schedule["id"])
                coexistence_mode = SENTRIX_SUPPORT_MATERIALIZATION_MODE_LINKED
                schedule_effect = "linked_existing"
        else:
            schedule_row_id = _insert_monthly_schedule_row(
                cur,
                tenant_id=tenant_id,
                company_id=str(site["company_id"]),
                site_id=str(site["id"]),
                employee_id=employee_id,
                schedule_date=work_date,
                shift_type=shift_kind,
                template_id=None,
                shift_start_time=shift_start_time,
                shift_end_time=shift_end_time,
                paid_hours=paid_hours,
                schedule_note=None,
                source=SENTRIX_ARLS_BRIDGE_SOURCE,
                source_batch_id=batch_id,
                source_revision=snapshot_id,
                source_ticket_uuid=ticket_id,
                source_ticket_state=ticket_state,
                source_action=SENTRIX_ARLS_BRIDGE_ACTION_UPSERT,
                source_self_staff=True,
            )
            schedule_effect = "created"

        materialization_result.update(
            {
                "schedule_effect": schedule_effect,
                "monthly_schedule_id": schedule_row_id,
                "coexistence_mode": coexistence_mode,
                "shift_start_time": shift_start_time,
                "shift_end_time": shift_end_time,
                "paid_hours": paid_hours,
            }
        )
        materialization_id = _upsert_sentrix_support_materialization_row(
            cur,
            materialization_id=str(existing_materialization["id"]) if existing_materialization else None,
            tenant_id=tenant_id,
            ticket_id=ticket_id,
            bridge_action_id=bridge_action_id,
            batch_id=batch_id,
            snapshot_id=snapshot_id,
            site_id=str(site["id"]),
            site_code=str(site.get("site_code") or "").strip(),
            work_date=work_date,
            shift_kind=shift_kind,
            employee_id=employee_id,
            employee_code=str(employee.get("employee_code") or "").strip() or None,
            employee_name=materialization_result["employee_name"],
            monthly_schedule_id=schedule_row_id,
            coexistence_mode=coexistence_mode,
            status=SENTRIX_SUPPORT_MATERIALIZATION_STATUS_ACTIVE,
            ticket_state=ticket_state or SENTRIX_HQ_ROSTER_FINAL_APPROVED_STATE,
            source_action=SENTRIX_ARLS_BRIDGE_ACTION_UPSERT,
            payload=payload,
            result_json=materialization_result,
        )
        materialization_result["materialization_id"] = materialization_id
        materialization_result["status"] = SENTRIX_SUPPORT_MATERIALIZATION_STATUS_ACTIVE
        return materialization_result

    schedule_deleted = 0
    deleted_schedule_id: str | None = None
    coexistence_mode = str((existing_materialization or {}).get("coexistence_mode") or "").strip() or SENTRIX_SUPPORT_MATERIALIZATION_MODE_OWNED
    current_schedule_row_id = str((existing_materialization or {}).get("monthly_schedule_id") or "").strip() or None
    if current_schedule_row_id and coexistence_mode == SENTRIX_SUPPORT_MATERIALIZATION_MODE_OWNED:
        cur.execute(
            """
            DELETE FROM monthly_schedules
            WHERE id = %s
              AND tenant_id = %s
              AND COALESCE(source, '') = %s
            RETURNING id
            """,
            (current_schedule_row_id, tenant_id, SENTRIX_ARLS_BRIDGE_SOURCE),
        )
        deleted = cur.fetchone()
        if deleted:
            schedule_deleted = 1
            deleted_schedule_id = str(deleted.get("id") or current_schedule_row_id)
    elif existing_schedule:
        existing_source = str(existing_schedule.get("source") or "").strip()
        existing_ticket_uuid = str(existing_schedule.get("source_ticket_uuid") or "").strip()
        if existing_source == SENTRIX_ARLS_BRIDGE_SOURCE and existing_ticket_uuid == ticket_id:
            cur.execute(
                """
                DELETE FROM monthly_schedules
                WHERE id = %s
                  AND tenant_id = %s
                  AND COALESCE(source, '') = %s
                RETURNING id
                """,
                (str(existing_schedule["id"]), tenant_id, SENTRIX_ARLS_BRIDGE_SOURCE),
            )
            deleted = cur.fetchone()
            if deleted:
                schedule_deleted = 1
                deleted_schedule_id = str(deleted.get("id") or existing_schedule["id"])

    materialization_result.update(
        {
            "schedule_effect": "retracted" if schedule_deleted else "noop_already_retracted",
            "deleted_schedule_id": deleted_schedule_id,
            "coexistence_mode": coexistence_mode,
            "monthly_schedule_id": current_schedule_row_id,
        }
    )
    materialization_id = _upsert_sentrix_support_materialization_row(
        cur,
        materialization_id=str(existing_materialization["id"]) if existing_materialization else None,
        tenant_id=tenant_id,
        ticket_id=ticket_id,
        bridge_action_id=bridge_action_id,
        batch_id=batch_id,
        snapshot_id=snapshot_id,
        site_id=str(site["id"]),
        site_code=str(site.get("site_code") or "").strip(),
        work_date=work_date,
        shift_kind=shift_kind,
        employee_id=employee_id,
        employee_code=str(employee.get("employee_code") or "").strip() or None,
        employee_name=materialization_result["employee_name"],
        monthly_schedule_id=None,
        coexistence_mode=coexistence_mode,
        status=SENTRIX_SUPPORT_MATERIALIZATION_STATUS_RETRACTED,
        ticket_state=ticket_state or SENTRIX_HQ_ROSTER_PENDING_STATUS,
        source_action=SENTRIX_ARLS_BRIDGE_ACTION_RETRACT,
        payload=payload,
        result_json=materialization_result,
    )
    materialization_result["materialization_id"] = materialization_id
    materialization_result["status"] = SENTRIX_SUPPORT_MATERIALIZATION_STATUS_RETRACTED
    return materialization_result


def _process_sentrix_support_arls_bridge_actions(
    conn,
    *,
    tenant_id: str,
    batch_id: str | None = None,
    include_failed: bool = False,
) -> dict[str, Any]:
    statuses = ["pending"]
    if include_failed:
        statuses.append("failed")
    clauses = [
        "tenant_id = %s",
        "status = ANY(%s::text[])",
    ]
    params: list[Any] = [tenant_id, statuses]
    if str(batch_id or "").strip():
        clauses.append("batch_id = %s")
        params.append(str(batch_id).strip())
    with conn.cursor() as cur:
        cur.execute(
            f"""
            SELECT id,
                   tenant_id,
                   batch_id,
                   snapshot_id,
                   ticket_id,
                   site_id,
                   site_code,
                   work_date,
                   shift_kind,
                   employee_id,
                   employee_code,
                   employee_name,
                   action,
                   ticket_state,
                   payload_json
            FROM sentrix_support_arls_bridge_actions
            WHERE {' AND '.join(clauses)}
            ORDER BY created_at ASC, id ASC
            """,
            tuple(params),
        )
        action_rows = [dict(row) for row in (cur.fetchall() or [])]

    result = {
        "processed": 0,
        "failed": 0,
        "upserts": 0,
        "retracts": 0,
        "materialized_created": 0,
        "materialized_updated": 0,
        "materialized_linked": 0,
        "materialized_retracted": 0,
        "materialized_noop": 0,
        "failures": [],
    }
    if not action_rows:
        return result

    affected_dates_by_site: dict[str, set[date]] = {}
    for action_row in action_rows:
        try:
            with conn.cursor() as cur:
                action_result = _apply_sentrix_support_bridge_action(
                    cur,
                    tenant_id=tenant_id,
                    action_row=action_row,
                )
                _mark_sentrix_support_bridge_action(
                    cur,
                    action_id=str(action_row["id"]),
                    status="success",
                    result_json=action_result,
                )
            conn.commit()
            result["processed"] += 1
            if str(action_row.get("action") or "").strip().upper() == SENTRIX_ARLS_BRIDGE_ACTION_UPSERT:
                result["upserts"] += 1
                schedule_effect = str(action_result.get("schedule_effect") or "").strip()
                if schedule_effect == "created":
                    result["materialized_created"] += 1
                elif schedule_effect == "updated":
                    result["materialized_updated"] += 1
                elif schedule_effect == "linked_existing":
                    result["materialized_linked"] += 1
                else:
                    result["materialized_noop"] += 1
            else:
                result["retracts"] += 1
                schedule_effect = str(action_result.get("schedule_effect") or "").strip()
                if schedule_effect == "retracted":
                    result["materialized_retracted"] += 1
                else:
                    result["materialized_noop"] += 1
            site_id = str(action_row.get("site_id") or "").strip()
            work_date = action_row.get("work_date")
            if site_id and isinstance(work_date, date):
                affected_dates_by_site.setdefault(site_id, set()).add(work_date)
        except Exception as exc:
            conn.rollback()
            failure_text = str(exc).strip() or "unknown bridge processing failure"
            with conn.cursor() as cur:
                _mark_sentrix_support_bridge_action(
                    cur,
                    action_id=str(action_row["id"]),
                    status="failed",
                    error_text=failure_text,
                    result_json={
                        "bridge_action_id": str(action_row.get("id") or "").strip(),
                        "ticket_id": str(action_row.get("ticket_id") or "").strip(),
                        "site_code": str(action_row.get("site_code") or "").strip() or None,
                        "work_date": action_row.get("work_date").isoformat() if isinstance(action_row.get("work_date"), date) else None,
                        "shift_kind": _normalize_support_status_shift_kind(action_row.get("shift_kind")),
                        "employee_id": str(action_row.get("employee_id") or "").strip() or None,
                        "action": str(action_row.get("action") or "").strip() or None,
                        "error": failure_text,
                    },
                )
            conn.commit()
            result["failed"] += 1
            result["failures"].append(
                {
                    "bridge_action_id": str(action_row.get("id") or "").strip(),
                    "ticket_id": str(action_row.get("ticket_id") or "").strip(),
                    "site_code": str(action_row.get("site_code") or "").strip() or None,
                    "work_date": action_row.get("work_date").isoformat() if isinstance(action_row.get("work_date"), date) else None,
                    "shift_kind": _normalize_support_status_shift_kind(action_row.get("shift_kind")),
                    "employee_id": str(action_row.get("employee_id") or "").strip() or None,
                    "error": failure_text,
                }
            )

    for site_id, schedule_dates in affected_dates_by_site.items():
        if not schedule_dates:
            continue
        try:
            _refresh_daily_leader_defaults_for_dates(
                conn,
                tenant_id=tenant_id,
                site_id=site_id,
                schedule_dates=sorted(schedule_dates),
            )
            conn.commit()
        except Exception as exc:
            conn.rollback()
            result["failures"].append(
                {
                    "site_id": site_id,
                    "error": f"leader refresh failed: {str(exc).strip() or 'unknown error'}",
                }
            )
    return result


def _apply_sentrix_hq_roster_batch(
    conn,
    *,
    batch_id: uuid.UUID,
    batch: dict[str, Any],
    target_tenant: dict,
    user: dict,
) -> SupportRosterHqApplyOut:
    restored_result = _restore_sentrix_hq_roster_apply_result(batch)
    if restored_result and str(batch.get("status") or "").strip() in {"applied", "blocked"}:
        return restored_result

    month_key = str(batch.get("month_key") or "").strip()
    blocked_reasons: list[str] = []
    if not month_key:
        blocked_reasons.append("upload batch 대상월이 없어 적용할 수 없습니다.")
    if max(int(batch.get("blocking_issue_count") or 0), 0) > 0:
        blocked_reasons.append("검토 단계의 차단 이슈가 남아 있어 적용할 수 없습니다.")
    if max(int(batch.get("valid_scope_count") or 0), 0) <= 0:
        blocked_reasons.append("적용 가능한 ticket scope가 없습니다.")
    blocked_reasons.extend(
        _validate_sentrix_hq_roster_batch_freshness(
            conn,
            tenant_id=str(target_tenant["id"]),
            batch=batch,
            month_key=month_key,
        )
    )

    batch_rows = _load_sentrix_hq_roster_batch_rows(conn, batch_id=batch_id)
    if not batch_rows:
        blocked_reasons.append("upload batch preview 데이터가 없어 다시 검토해야 합니다.")

    scope_payloads: dict[str, dict[str, Any]] = {}
    worker_payloads_by_scope: dict[str, list[dict[str, Any]]] = {}
    site_codes: set[str] = set()
    site_ids: set[str] = set()

    for row in batch_rows:
        payload = dict(row.get("payload_json") or {})
        scope_key = str(payload.get("scope_key") or "").strip()
        row_kind = str(row.get("row_kind") or payload.get("row_kind") or "").strip() or "worker"
        site_code = str(payload.get("site_code") or row.get("site_code") or "").strip().upper()
        site_id_raw = str(payload.get("site_id") or row.get("site_id") or "").strip()
        if site_code:
            site_codes.add(site_code)
        if site_id_raw:
            try:
                site_ids.add(str(uuid.UUID(site_id_raw)))
            except Exception:
                blocked_reasons.append(f"{site_code or site_id_raw} site_id를 해석할 수 없습니다.")
        if row_kind == "scope_summary":
            if not scope_key:
                blocked_reasons.append("scope summary key가 없어 적용할 수 없습니다.")
                continue
            if scope_key in scope_payloads:
                blocked_reasons.append(f"{scope_key} scope summary가 batch 안에서 중복되었습니다.")
                continue
            scope_payloads[scope_key] = payload
            continue
        if row_kind == "worker":
            if not scope_key:
                blocked_reasons.append("worker row scope key가 없어 적용할 수 없습니다.")
                continue
            if str(row.get("severity") or "").strip().lower() == "blocking" or str(payload.get("issue_code") or "").strip():
                blocked_reasons.append(f"{scope_key} 범위에 차단 근무자 셀이 남아 있어 적용할 수 없습니다.")
            worker_payloads_by_scope.setdefault(scope_key, []).append(payload)

    orphan_scope_keys = sorted(set(worker_payloads_by_scope.keys()) - set(scope_payloads.keys()))
    if orphan_scope_keys:
        blocked_reasons.append("scope summary 없이 남아 있는 worker row가 있어 다시 검토해야 합니다.")

    if not scope_payloads:
        blocked_reasons.append("적용할 support roster scope summary가 없습니다.")

    current_ticket_scope_map = _load_sentrix_support_ticket_scope_map(
        conn,
        tenant_id=str(target_tenant["id"]),
        month_key=month_key,
        site_codes=sorted(site_codes),
    )
    scope_apply_specs: list[dict[str, Any]] = []

    for scope_key, payload in sorted(scope_payloads.items(), key=lambda item: item[0]):
        site_code = str(payload.get("site_code") or "").strip().upper()
        site_name = str(payload.get("site_name") or "").strip() or None
        sheet_name = str(payload.get("sheet_name") or "").strip() or site_name or site_code or "-"
        work_date_raw = str(payload.get("work_date") or "").strip()
        shift_kind = "night" if str(payload.get("shift_kind") or "").strip().lower() == "night" else "day"
        ticket_id_raw = str(payload.get("ticket_id") or "").strip()
        target_status = str(payload.get("target_status") or "").strip().lower()
        request_count = max(int(payload.get("request_count") or 0), 0)
        valid_filled_count = max(int(payload.get("valid_filled_count") or 0), 0)
        invalid_filled_count = max(int(payload.get("invalid_filled_count") or 0), 0)
        matched_ticket = bool(payload.get("matched_ticket"))

        try:
            work_date = date.fromisoformat(work_date_raw)
        except Exception:
            blocked_reasons.append(f"{sheet_name} 시트의 날짜 범위를 적용 단계에서 해석하지 못했습니다.")
            continue

        current_ticket = current_ticket_scope_map.get((site_code, work_date.isoformat(), shift_kind))
        if not current_ticket or not matched_ticket or not ticket_id_raw:
            blocked_reasons.append(f"{sheet_name} {work_date.isoformat()} {shift_kind} 범위의 기존 ticket을 찾을 수 없습니다.")
            continue
        if str(current_ticket.get("id") or "").strip() != ticket_id_raw:
            blocked_reasons.append(f"{sheet_name} {work_date.isoformat()} {shift_kind} ticket이 검토 이후 변경되었습니다.")
            continue
        current_request_count = max(int(current_ticket.get("request_count") or 0), 0)
        if current_request_count != request_count:
            blocked_reasons.append(f"{sheet_name} {work_date.isoformat()} {shift_kind} 요청 인원이 검토 이후 변경되었습니다.")
            continue
        if target_status not in {SENTRIX_HQ_ROSTER_AUTO_APPROVED_STATUS, SENTRIX_HQ_ROSTER_PENDING_STATUS}:
            blocked_reasons.append(f"{sheet_name} {work_date.isoformat()} {shift_kind} target status를 확정하지 못했습니다.")
            continue

        worker_payloads = sorted(
            list(worker_payloads_by_scope.get(scope_key) or []),
            key=lambda item: max(int(item.get("slot_index") or 0), 0),
        )
        valid_worker_payloads = [
            payload_row
            for payload_row in worker_payloads
            if not str(payload_row.get("issue_code") or "").strip()
            and bool(payload_row.get("countable"))
        ]
        if len(valid_worker_payloads) != valid_filled_count:
            blocked_reasons.append(f"{sheet_name} {work_date.isoformat()} {shift_kind} 유효 인원 수가 preview 결과와 다릅니다.")
            continue

        try:
            ticket_uuid = uuid.UUID(ticket_id_raw)
        except Exception:
            blocked_reasons.append(f"{sheet_name} {work_date.isoformat()} {shift_kind} ticket id를 해석하지 못했습니다.")
            continue

        scope_apply_specs.append(
            {
                "scope_key": scope_key,
                "sheet_name": sheet_name,
                "site_code": site_code,
                "site_name": site_name,
                "month_key": month_key,
                "work_date": work_date,
                "shift_kind": shift_kind,
                "ticket": current_ticket,
                "ticket_id": ticket_uuid,
                "request_count": request_count,
                "valid_filled_count": valid_filled_count,
                "invalid_filled_count": invalid_filled_count,
                "target_status": target_status,
                "scope_payload": payload,
                "worker_payloads": worker_payloads,
                "valid_worker_payloads": valid_worker_payloads,
            }
        )

    if blocked_reasons:
        blocked_reasons = list(dict.fromkeys(item for item in blocked_reasons if str(item).strip()))
        result = SupportRosterHqApplyOut(
            batch_id=batch_id,
            applied=False,
            blocked=True,
            blocked_reasons=blocked_reasons,
            issue_count=max(int(batch.get("issue_count") or 0), 0),
            assignments_created=0,
            assignments_removed=0,
            tickets_updated=0,
            tickets_auto_approved=0,
            tickets_pending=0,
            applied_scope_count=0,
            failed_scope_count=len(scope_payloads),
            audit_timestamp=datetime.now(timezone.utc),
            scope_results=[],
        )
        with conn.cursor() as cur:
            _write_sentrix_hq_roster_batch_apply_audit(
                cur,
                batch=batch,
                batch_id=batch_id,
                status="blocked",
                user=user,
                result_json=result.model_dump(mode="json"),
                error_text="; ".join(blocked_reasons[:5]) or None,
            )
        return result

    assignments_removed = 0
    assignments_created = 0
    tickets_updated = 0
    tickets_auto_approved = 0
    tickets_pending = 0
    snapshots_created = 0
    notifications_created = 0
    notification_sites = 0
    push_sent = 0
    push_failed = 0
    bridge_actions_created = 0
    bridge_upserts = 0
    bridge_retracts = 0
    bridge_processed = 0
    bridge_failed = 0
    arls_materialized_created = 0
    arls_materialized_updated = 0
    arls_materialized_linked = 0
    arls_materialized_retracted = 0
    arls_materialized_noop = 0
    scope_results: list[SupportRosterHqApplyScopeOut] = []
    start_date, end_date = _month_bounds(month_key)
    tenant_id = str(target_tenant["id"])
    current_snapshot_map = _load_current_sentrix_hq_snapshot_map(
        conn,
        tenant_id=tenant_id,
        ticket_ids=[str(spec["ticket_id"]) for spec in scope_apply_specs],
    )
    notification_jobs: list[dict[str, Any]] = []
    partial_failures: list[str] = []

    try:
        with conn.cursor() as cur:
            if site_ids:
                cur.execute(
                    """
                    DELETE FROM support_assignment
                    WHERE tenant_id = %s
                      AND source = %s
                      AND work_date >= %s
                      AND work_date < %s
                      AND site_id::text = ANY(%s)
                    """,
                    (
                        target_tenant["id"],
                        SENTRIX_HQ_ROSTER_ASSIGNMENT_SOURCE,
                        start_date,
                        end_date,
                        sorted(site_ids),
                    ),
                )
                assignments_removed = max(int(cur.rowcount or 0), 0)

            for spec in scope_apply_specs:
                for worker_payload in spec["valid_worker_payloads"]:
                    worker_name = str(
                        worker_payload.get("employee_name")
                        or worker_payload.get("worker_name")
                        or worker_payload.get("name")
                        or ""
                    ).strip()
                    if not worker_name:
                        raise RuntimeError(f"{spec['scope_key']} scope worker_name missing")
                    upsert_support_assignment(
                        conn,
                        tenant_id=tenant_id,
                        site_id=uuid.UUID(str(spec["ticket"].get("site_id"))),
                        work_date=spec["work_date"],
                        worker_type=str(worker_payload.get("worker_type") or "F").strip().upper() or "F",
                        name=worker_name,
                        support_period=spec["shift_kind"],
                        slot_index=max(int(worker_payload.get("slot_index") or 0), 1),
                        source=SENTRIX_HQ_ROSTER_ASSIGNMENT_SOURCE,
                        employee_id=(uuid.UUID(str(worker_payload["employee_id"])) if str(worker_payload.get("employee_id") or "").strip() else None),
                        affiliation=str(worker_payload.get("affiliation") or "").strip() or None,
                        source_event_uid=(
                            f"sentrix:hq-roster:{batch_id}:{spec['site_code']}:{spec['work_date'].isoformat()}:{spec['shift_kind']}:{max(int(worker_payload.get('slot_index') or 0), 1)}"
                        ),
                    )
                    assignments_created += 1

            site_notification_specs: dict[str, dict[str, Any]] = {}
            for spec in scope_apply_specs:
                target_status = spec["target_status"]
                final_ticket_state = _normalize_sentrix_hq_roster_final_state(target_status) or SENTRIX_HQ_ROSTER_PENDING_STATUS
                previous_bundle = current_snapshot_map.get(str(spec["ticket_id"]))
                previous_snapshot = dict((previous_bundle or {}).get("snapshot") or {})
                previous_entries = list((previous_bundle or {}).get("entries") or [])
                previous_ticket_state = (
                    _normalize_sentrix_hq_roster_final_state(previous_snapshot.get("ticket_state"))
                    or _extract_sentrix_hq_roster_final_state(spec["ticket"])
                )
                snapshot_id, snapshot_changed, snapshot_entries, persisted_previous_ticket_state = _persist_sentrix_hq_roster_snapshot(
                    cur,
                    tenant_id=tenant_id,
                    batch_id=batch_id,
                    spec=spec,
                    ticket_state=final_ticket_state,
                    previous_bundle=previous_bundle,
                )
                snapshots_created += 1
                bridge_counts = _queue_sentrix_hq_arls_bridge_actions(
                    cur,
                    batch_id=batch_id,
                    snapshot_id=snapshot_id,
                    tenant_id=tenant_id,
                    spec=spec,
                    ticket_state=final_ticket_state,
                    previous_ticket_state=persisted_previous_ticket_state or previous_ticket_state,
                    previous_entries=previous_entries,
                    current_entries=snapshot_entries,
                )
                bridge_actions_created += bridge_counts["created"]
                bridge_upserts += bridge_counts["upserts"]
                bridge_retracts += bridge_counts["retracts"]
                detail_json = _build_sentrix_hq_roster_ticket_detail_json(
                    spec["ticket"],
                    batch_id=batch_id,
                    batch=batch,
                    user=user,
                    scope_payload={
                        **dict(spec["scope_payload"] or {}),
                        "valid_filled_count": spec["valid_filled_count"],
                        "invalid_filled_count": spec["invalid_filled_count"],
                    },
                    valid_worker_payloads=spec["valid_worker_payloads"],
                    target_status=target_status,
                )
                cur.execute(
                    """
                    UPDATE sentrix_support_request_tickets
                    SET detail_json = %s::jsonb,
                        updated_at = timezone('utc', now())
                    WHERE id = %s
                      AND tenant_id = %s
                    """,
                    (
                        json.dumps(detail_json, ensure_ascii=False, default=str),
                        spec["ticket_id"],
                        tenant_id,
                    ),
                )
                if int(cur.rowcount or 0) != 1:
                    raise RuntimeError(f"{spec['scope_key']} ticket update failed")
                tickets_updated += 1
                if target_status == SENTRIX_HQ_ROSTER_AUTO_APPROVED_STATUS:
                    tickets_auto_approved += 1
                else:
                    tickets_pending += 1
                scope_results.append(
                    SupportRosterHqApplyScopeOut(
                        scope_key=spec["scope_key"],
                        sheet_name=spec["sheet_name"],
                        site_name=spec["site_name"],
                        site_code=spec["site_code"] or None,
                        work_date=spec["work_date"],
                        shift_kind=spec["shift_kind"],
                        ticket_id=spec["ticket_id"],
                        request_count=spec["request_count"],
                        valid_filled_count=spec["valid_filled_count"],
                        previous_status=persisted_previous_ticket_state or previous_ticket_state,
                        target_status=target_status,
                        assignment_count=len(spec["valid_worker_payloads"]),
                        bridge_action_count=bridge_counts["created"],
                        snapshot_changed=snapshot_changed,
                    )
                )
                site_id = str(spec["ticket"].get("site_id") or "").strip()
                if site_id:
                    site_entry = site_notification_specs.setdefault(
                        site_id,
                        {
                            "site_id": site_id,
                            "site_code": str(spec.get("site_code") or "").strip() or None,
                            "site_name": str(spec.get("site_name") or "").strip() or None,
                            "changed": False,
                            "scope_keys": [],
                            "ticket_ids": [],
                            "scope_count": 0,
                        },
                    )
                    site_entry["scope_count"] += 1
                    site_entry["scope_keys"].append(str(spec["scope_key"]))
                    site_entry["ticket_ids"].append(str(spec["ticket_id"]))
                    site_entry["changed"] = bool(site_entry["changed"]) or bool(
                        snapshot_changed
                        or bridge_counts["created"] > 0
                        or (persisted_previous_ticket_state or previous_ticket_state) != final_ticket_state
                    )

            for site_entry in site_notification_specs.values():
                if not bool(site_entry.get("changed")):
                    continue
                recipient_user_ids = _resolve_sentrix_hq_notification_user_ids(
                    conn,
                    tenant_id=tenant_id,
                    site_id=str(site_entry.get("site_id") or "").strip(),
                )
                payload = {
                    "type": "SENTRIX_SUPPORT_ROSTER_UPDATED",
                    "upload_batch_id": str(batch_id),
                    "month": month_key,
                    "site_id": str(site_entry.get("site_id") or "").strip() or None,
                    "site_code": str(site_entry.get("site_code") or "").strip() or None,
                    "site_name": str(site_entry.get("site_name") or "").strip() or None,
                    "scope_count": max(int(site_entry.get("scope_count") or 0), 0),
                    "ticket_ids": sorted(set(site_entry.get("ticket_ids") or [])),
                    "scope_keys": sorted(set(site_entry.get("scope_keys") or [])),
                }
                dedupe_key = (
                    f"sentrix-support-roster:{batch_id}:"
                    f"{str(site_entry.get('site_code') or site_entry.get('site_id') or '').strip()}"
                )
                notification_audit_id = _insert_sentrix_hq_notification_audit(
                    cur,
                    tenant_id=tenant_id,
                    batch_id=batch_id,
                    site_id=str(site_entry.get("site_id") or "").strip(),
                    site_code=str(site_entry.get("site_code") or "").strip(),
                    message=SENTRIX_HQ_ROSTER_NOTIFICATION_MESSAGE,
                    dedupe_key=dedupe_key,
                    recipient_user_ids=recipient_user_ids,
                    payload=payload,
                )
                inserted_count = 0
                for recipient_user_id in recipient_user_ids:
                    cur.execute(
                        """
                        INSERT INTO in_app_notifications (
                            id, tenant_id, user_id, site_id, category, message, dedupe_key,
                            payload_json, created_at
                        )
                        VALUES (
                            arls_random_uuid(), %s, %s, %s, %s, %s, %s,
                            %s::jsonb, timezone('utc', now())
                        )
                        ON CONFLICT DO NOTHING
                        """,
                        (
                            tenant_id,
                            recipient_user_id,
                            str(site_entry.get("site_id") or "").strip() or None,
                            "info",
                            SENTRIX_HQ_ROSTER_NOTIFICATION_MESSAGE,
                            f"{dedupe_key}:{recipient_user_id}",
                            json.dumps(payload, ensure_ascii=False, default=str),
                        ),
                    )
                    inserted_count += max(int(cur.rowcount or 0), 0)
                notifications_created += inserted_count
                notification_sites += 1
                notification_jobs.append(
                    {
                        "audit_id": notification_audit_id,
                        "site_code": str(site_entry.get("site_code") or "").strip() or None,
                        "recipient_user_ids": recipient_user_ids,
                        "payload": payload,
                    }
                )

            result = SupportRosterHqApplyOut(
                batch_id=batch_id,
                applied=True,
                blocked=False,
                blocked_reasons=[],
                issue_count=max(int(batch.get("issue_count") or 0), 0),
                assignments_created=assignments_created,
                assignments_removed=assignments_removed,
                tickets_updated=tickets_updated,
                tickets_auto_approved=tickets_auto_approved,
                tickets_pending=tickets_pending,
                snapshots_created=snapshots_created,
                notifications_created=notifications_created,
                notification_sites=notification_sites,
                push_sent=0,
                push_failed=0,
                bridge_actions_created=bridge_actions_created,
                bridge_upserts=bridge_upserts,
                bridge_retracts=bridge_retracts,
                bridge_processed=0,
                bridge_failed=0,
                arls_materialized_created=0,
                arls_materialized_updated=0,
                arls_materialized_linked=0,
                arls_materialized_retracted=0,
                arls_materialized_noop=0,
                applied_scope_count=len(scope_apply_specs),
                failed_scope_count=0,
                audit_timestamp=datetime.now(timezone.utc),
                scope_results=scope_results,
            )
            _write_sentrix_hq_roster_batch_apply_audit(
                cur,
                batch=batch,
                batch_id=batch_id,
                status="applied",
                user=user,
                result_json=result.model_dump(mode="json"),
            )
        conn.commit()

        if bridge_actions_created > 0:
            bridge_process_result = _process_sentrix_support_arls_bridge_actions(
                conn,
                tenant_id=tenant_id,
                batch_id=str(batch_id),
            )
            bridge_processed = max(int(bridge_process_result.get("processed") or 0), 0)
            bridge_failed = max(int(bridge_process_result.get("failed") or 0), 0)
            arls_materialized_created = max(int(bridge_process_result.get("materialized_created") or 0), 0)
            arls_materialized_updated = max(int(bridge_process_result.get("materialized_updated") or 0), 0)
            arls_materialized_linked = max(int(bridge_process_result.get("materialized_linked") or 0), 0)
            arls_materialized_retracted = max(int(bridge_process_result.get("materialized_retracted") or 0), 0)
            arls_materialized_noop = max(int(bridge_process_result.get("materialized_noop") or 0), 0)
            if bridge_failed > 0:
                partial_failures.append(f"ARLS bridge {bridge_failed}건 실패")
            elif list(bridge_process_result.get("failures") or []):
                partial_failures.append("ARLS bridge 후처리 일부 실패")
            result = result.model_copy(
                update={
                    "bridge_processed": bridge_processed,
                    "bridge_failed": bridge_failed,
                    "arls_materialized_created": arls_materialized_created,
                    "arls_materialized_updated": arls_materialized_updated,
                    "arls_materialized_linked": arls_materialized_linked,
                    "arls_materialized_retracted": arls_materialized_retracted,
                    "arls_materialized_noop": arls_materialized_noop,
                }
            )

        for job in notification_jobs:
            audit_id = job["audit_id"]
            recipient_user_ids = list(job.get("recipient_user_ids") or [])
            try:
                push_result = send_push_notification_to_users(
                    conn,
                    tenant_id=tenant_id,
                    user_ids=recipient_user_ids,
                    title=SENTRIX_HQ_ROSTER_NOTIFICATION_MESSAGE,
                    body=SENTRIX_HQ_ROSTER_NOTIFICATION_MESSAGE,
                    data=dict(job.get("payload") or {}),
                )
                push_sent += max(int(push_result.get("sent_count") or 0), 0)
                push_failed += max(int(push_result.get("failed_count") or 0), 0)
                if max(int(push_result.get("failed_count") or 0), 0) > 0:
                    partial_failures.append(
                        f"{str(job.get('site_code') or '').strip() or 'site'} push 일부 실패"
                    )
                _update_sentrix_hq_notification_audit_after_push(
                    conn,
                    audit_id=audit_id,
                    push_result=push_result,
                )
            except Exception as exc:
                push_failed += len(recipient_user_ids)
                partial_failures.append(
                    f"{str(job.get('site_code') or '').strip() or 'site'} push 전송 실패"
                )
                _update_sentrix_hq_notification_audit_after_push(
                    conn,
                    audit_id=audit_id,
                    push_result={
                        "target_count": len(recipient_user_ids),
                        "sent_count": 0,
                        "failed_count": len(recipient_user_ids),
                    },
                    error_text=str(exc),
                )

        if notification_jobs or bridge_actions_created > 0 or partial_failures:
            result = result.model_copy(
                update={
                    "push_sent": push_sent,
                    "push_failed": push_failed,
                }
            )
            with conn.cursor() as cur:
                _write_sentrix_hq_roster_batch_apply_audit(
                    cur,
                    batch=batch,
                    batch_id=batch_id,
                    status="applied",
                    user=user,
                    result_json=result.model_dump(mode="json"),
                    error_text="; ".join(dict.fromkeys(partial_failures))[:500] or None,
                )
            conn.commit()
        return result
    except Exception as exc:
        conn.rollback()
        failed_result = SupportRosterHqApplyOut(
            batch_id=batch_id,
            applied=False,
            blocked=False,
            blocked_reasons=[],
            issue_count=max(int(batch.get("issue_count") or 0), 0),
            assignments_created=0,
            assignments_removed=0,
            tickets_updated=0,
            tickets_auto_approved=0,
            tickets_pending=0,
            applied_scope_count=0,
            failed_scope_count=len(scope_payloads),
            audit_timestamp=datetime.now(timezone.utc),
            scope_results=[],
        )
        with conn.cursor() as cur:
            _write_sentrix_hq_roster_batch_apply_audit(
                cur,
                batch=batch,
                batch_id=batch_id,
                status="failed",
                user=user,
                result_json=failed_result.model_dump(mode="json"),
                error_text=str(exc),
            )
        conn.commit()
        raise HTTPException(status_code=500, detail="support roster apply failed") from exc


def _get_support_roundtrip_source(conn, *, tenant_id: str, site_id: str, month_key: str) -> dict[str, Any] | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT srs.id,
                   srs.tenant_id,
                   srs.site_id,
                   srs.site_code,
                   srs.month_key,
                   srs.source_batch_id,
                   srs.source_revision,
                   srs.source_filename,
                   srs.source_uploaded_by,
                   srs.source_uploaded_role,
                   srs.source_uploaded_at,
                   srs.state,
                   srs.hq_merge_available,
                   srs.hq_merge_stale,
                   srs.conflict_required,
                   srs.final_download_enabled,
                   srs.latest_hq_batch_id,
                   srs.latest_hq_revision,
                   srs.latest_merged_revision,
                   srs.created_at,
                   srs.updated_at,
                   u.username AS source_uploaded_by_username
            FROM schedule_support_roundtrip_sources srs
            LEFT JOIN arls_users u ON u.id = srs.source_uploaded_by
            WHERE srs.tenant_id = %s
              AND srs.site_id = %s
              AND srs.month_key = %s
            LIMIT 1
            """,
            (tenant_id, site_id, month_key),
        )
        return cur.fetchone()


def _list_support_roundtrip_assignments(
    conn,
    *,
    source_id: str,
    source_revision: str,
) -> list[dict[str, Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id,
                   source_id,
                   source_revision,
                   tenant_id,
                   site_id,
                   site_code,
                   work_date,
                   support_period,
                   slot_index,
                   worker_type,
                   worker_name,
                   employee_id,
                   employee_code,
                   employee_name,
                   is_internal,
                   internal_shift_type,
                   internal_template_id,
                   internal_shift_start_time,
                   internal_shift_end_time,
                   internal_paid_hours,
                   source_batch_id,
                   created_by,
                   created_at,
                   updated_at
            FROM schedule_support_roundtrip_assignments
            WHERE source_id = %s
              AND source_revision = %s
            ORDER BY work_date ASC, support_period ASC, slot_index ASC, worker_name ASC
            """,
            (source_id, source_revision),
        )
        return [dict(row) for row in cur.fetchall()]


def _register_support_roundtrip_source_after_import(
    conn,
    *,
    batch_id: str,
    target_tenant: dict,
    site_row: dict,
    month_key: str,
    user: dict,
) -> dict[str, Any]:
    export_revision = _build_schedule_export_revision(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        site_code=str(site_row["site_code"]),
        month_key=month_key,
    )
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT filename
            FROM schedule_import_batches
            WHERE id = %s
              AND tenant_id = %s
            LIMIT 1
            """,
            (batch_id, target_tenant["id"]),
        )
        batch = cur.fetchone() or {}
        current = _get_support_roundtrip_source(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_id=str(site_row["id"]),
            month_key=month_key,
        )
        stale_hq = bool(current and (current.get("hq_merge_available") or current.get("latest_hq_batch_id")))
        if current:
            next_state = "hq_merge_stale" if stale_hq else "waiting_for_hq_merge"
            cur.execute(
                """
                UPDATE schedule_support_roundtrip_sources
                SET source_batch_id = %s,
                    source_revision = %s,
                    source_filename = %s,
                    source_uploaded_by = %s,
                    source_uploaded_role = %s,
                    source_uploaded_at = timezone('utc', now()),
                    state = %s,
                    hq_merge_available = FALSE,
                    hq_merge_stale = %s,
                    conflict_required = FALSE,
                    final_download_enabled = FALSE,
                    latest_hq_batch_id = CASE WHEN %s THEN latest_hq_batch_id ELSE NULL END,
                    latest_hq_revision = CASE WHEN %s THEN latest_hq_revision ELSE NULL END,
                    latest_merged_revision = CASE WHEN %s THEN latest_merged_revision ELSE NULL END,
                    updated_at = timezone('utc', now())
                WHERE id = %s
                RETURNING *
                """,
                (
                    batch_id,
                    export_revision,
                    str(batch.get("filename") or "").strip() or None,
                    user["id"],
                    _support_roundtrip_normalize_role(user),
                    next_state,
                    stale_hq,
                    stale_hq,
                    stale_hq,
                    stale_hq,
                    current["id"],
                ),
            )
            return cur.fetchone()
        cur.execute(
            """
            INSERT INTO schedule_support_roundtrip_sources (
                id, tenant_id, site_id, site_code, month_key, source_batch_id,
                source_revision, source_filename, source_uploaded_by, source_uploaded_role,
                source_uploaded_at, state, hq_merge_available, hq_merge_stale,
                conflict_required, final_download_enabled, created_at, updated_at
            )
            VALUES (
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                timezone('utc', now()), 'waiting_for_hq_merge', FALSE, FALSE,
                FALSE, FALSE, timezone('utc', now()), timezone('utc', now())
            )
            RETURNING *
            """,
            (
                uuid.uuid4(),
                target_tenant["id"],
                site_row["id"],
                str(site_row["site_code"]),
                month_key,
                batch_id,
                export_revision,
                str(batch.get("filename") or "").strip() or None,
                user["id"],
                _support_roundtrip_normalize_role(user),
            ),
        )
        return cur.fetchone()


def _build_support_roundtrip_status_payload(
    conn,
    *,
    source_row: dict | None,
    tenant_id: str,
    site_id: str,
    site_code: str,
    month_key: str,
) -> SupportRoundtripStatusOut:
    if not source_row:
        return SupportRoundtripStatusOut(
            site_code=site_code,
            month=month_key,
            source_state="source_missing",
            blocked_reasons=["Supervisor 기준 소스 월간 파일이 아직 업로드/반영되지 않았습니다."],
        )
    latest_hq_uploaded_at = None
    latest_hq_uploaded_by = None
    latest_hq_filename = None
    if source_row.get("latest_hq_batch_id"):
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT b.filename, b.created_at, u.username
                FROM schedule_support_roundtrip_batches b
                LEFT JOIN arls_users u ON u.id = b.uploaded_by
                WHERE b.id = %s
                LIMIT 1
                """,
                (source_row["latest_hq_batch_id"],),
            )
            latest = cur.fetchone()
        if latest:
            latest_hq_filename = latest.get("filename")
            latest_hq_uploaded_at = latest.get("created_at")
            latest_hq_uploaded_by = latest.get("username")
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM schedule_support_roundtrip_assignments
            WHERE source_id = %s
              AND source_revision = %s
            """,
            (source_row["id"], source_row["source_revision"]),
        )
        assignment_count_row = cur.fetchone()
        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM schedule_support_roundtrip_batches
            WHERE source_id = %s
              AND status = 'blocked'
            """,
            (source_row["id"],),
        )
        conflict_row = cur.fetchone()
    blocked_reasons: list[str] = []
    if not bool(source_row.get("hq_merge_available")):
        blocked_reasons.append("HQ 지원근무 병합이 아직 없어 최종본 다운로드가 비활성화됩니다.")
    if bool(source_row.get("hq_merge_stale")):
        blocked_reasons.append("현재 HQ 병합본은 최신 Supervisor 소스 리비전에 대해 stale 상태입니다.")
    return SupportRoundtripStatusOut(
        site_code=site_code,
        month=month_key,
        source_state=str(source_row.get("state") or "waiting_for_hq_merge"),
        source_revision=str(source_row.get("source_revision") or "").strip() or None,
        source_uploaded_at=source_row.get("source_uploaded_at"),
        source_uploaded_by=str(source_row.get("source_uploaded_by_username") or source_row.get("source_uploaded_by") or "").strip() or None,
        source_filename=str(source_row.get("source_filename") or "").strip() or None,
        hq_merge_available=bool(source_row.get("hq_merge_available")),
        hq_merge_stale=bool(source_row.get("hq_merge_stale")),
        final_download_enabled=bool(source_row.get("final_download_enabled")),
        latest_hq_uploaded_at=latest_hq_uploaded_at,
        latest_hq_uploaded_by=latest_hq_uploaded_by,
        latest_hq_filename=latest_hq_filename,
        latest_hq_revision=str(source_row.get("latest_hq_revision") or "").strip() or None,
        latest_merged_revision=str(source_row.get("latest_merged_revision") or "").strip() or None,
        support_assignment_count=int((assignment_count_row or {}).get("cnt") or 0),
        conflict_count=int((conflict_row or {}).get("cnt") or 0),
        blocked_reasons=blocked_reasons,
    )


def _support_assignment_display_value(row: dict[str, Any], *, include_internal: bool) -> str:
    worker_type = str(row.get("worker_type") or "").strip().upper()
    worker_name = str(row.get("worker_name") or row.get("employee_name") or row.get("name") or "").strip()
    affiliation = str(row.get("affiliation") or "").strip()
    if bool(row.get("is_internal")) or worker_type == "INTERNAL":
        return worker_name if include_internal else ""
    if worker_type == "UNAVAILABLE":
        return f"지원불가 {worker_name}".strip()
    if affiliation:
        return f"{affiliation} {worker_name}".strip()
    if worker_type in {"F", "BK"} and " " in worker_name:
        return worker_name
    if worker_type:
        return f"{worker_type} {worker_name}".strip()
    return worker_name


def _fallback_support_assignments_from_export_ctx(
    export_ctx: dict[str, Any],
    *,
    include_internal: bool = True,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    day_slot_index: dict[str, int] = {}
    night_slot_index: dict[str, int] = {}
    for row in export_ctx.get("support_rows") or []:
        work_date = row.get("work_date")
        if not isinstance(work_date, date):
            continue
        date_key = work_date.isoformat()
        slot_index = day_slot_index.get(date_key, 0) + 1
        day_slot_index[date_key] = slot_index
        rows.append(
            {
                "work_date": work_date,
                "support_period": "day",
                "slot_index": slot_index,
                "worker_type": str(row.get("worker_type") or "").strip() or "INTERNAL",
                "worker_name": str(row.get("name") or row.get("employee_name") or "").strip(),
                "employee_id": row.get("employee_id"),
                "employee_code": row.get("employee_code"),
                "employee_name": row.get("employee_name"),
                "is_internal": str(row.get("worker_type") or "").strip().upper() == "INTERNAL",
                "internal_shift_type": "day",
                "internal_template_id": None,
                "internal_shift_start_time": None,
                "internal_shift_end_time": None,
                "internal_paid_hours": None,
            }
        )
    if include_internal:
        for row in export_ctx.get("employee_overnight_rows") or []:
            work_date = row.get("work_date")
            if not isinstance(work_date, date):
                continue
            date_key = work_date.isoformat()
            slot_index = night_slot_index.get(date_key, 0) + 1
            night_slot_index[date_key] = slot_index
            rows.append(
                {
                    "work_date": work_date,
                    "support_period": "night",
                    "slot_index": slot_index,
                    "worker_type": "INTERNAL",
                    "worker_name": str(row.get("employee_name") or row.get("employee_code") or "").strip(),
                    "employee_id": row.get("employee_id"),
                    "employee_code": row.get("employee_code"),
                    "employee_name": row.get("employee_name"),
                    "is_internal": True,
                    "internal_shift_type": "night",
                    "internal_template_id": None,
                    "internal_shift_start_time": None,
                    "internal_shift_end_time": None,
                    "internal_paid_hours": row.get("hours"),
                }
            )
    return rows


def _copy_support_row_style(sheet, *, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col_idx in range(1, ARLS_SUMMARY_END_COL + 1):
        source_cell = sheet.cell(row=source_row, column=col_idx)
        target_cell = sheet.cell(row=target_row, column=col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def _locate_support_section_rows(sheet) -> dict[str, Any]:
    start_scan_row = max(_find_template_data_start_row(sheet), 1)
    weekly_start = None
    night_start = None
    weekly_rows: list[int] = []
    night_rows: list[int] = []
    need_rows: list[int] = []
    weekly_count_row = None
    night_count_row = None
    vendor_rows: list[int] = []
    work_note_rows: list[int] = []
    for row_idx in range(start_scan_row, min(sheet.max_row, 320) + 1):
        left = sheet.cell(row=row_idx, column=2).value
        right = sheet.cell(row=row_idx, column=3).value
        right_text = str(right or "").strip().replace("\n", " ")
        if _label_contains_any(left, ARLS_ADDITIONAL_DAY_KEYWORDS) and right_text.startswith("근무자"):
            weekly_start = row_idx
        if _label_contains_any(left, ARLS_ADDITIONAL_NIGHT_KEYWORDS) and right_text.startswith("근무자"):
            night_start = row_idx
        if _label_contains_any(right, ARLS_NEED_COUNT_KEYWORDS):
            need_rows.append(row_idx)
        if _normalize_template_label_token(left) == _normalize_template_label_token("주간 추가 근무자 수"):
            weekly_count_row = row_idx
        if _normalize_template_label_token(left) == _normalize_template_label_token("야간 근무자 총 수"):
            night_count_row = row_idx
        if _label_contains_any(right, ARLS_VENDOR_COUNT_KEYWORDS):
            vendor_rows.append(row_idx)
        if _label_contains_any(right, ARLS_WORK_NOTE_KEYWORDS):
            work_note_rows.append(row_idx)
    if weekly_start:
        row_idx = weekly_start
        while row_idx <= sheet.max_row:
            right = str(sheet.cell(row=row_idx, column=3).value or "").strip().replace("\n", " ")
            if not _is_worker_slot_label(right):
                break
            weekly_rows.append(row_idx)
            row_idx += 1
    if night_start:
        row_idx = night_start
        while row_idx <= sheet.max_row:
            right = str(sheet.cell(row=row_idx, column=3).value or "").strip().replace("\n", " ")
            if not _is_worker_slot_label(right):
                break
            night_rows.append(row_idx)
            row_idx += 1
    weekly_end = weekly_rows[-1] if weekly_rows else weekly_start
    night_end = night_rows[-1] if night_rows else night_start
    day_need_row = None
    night_need_row = None
    day_vendor_count_row = None
    night_vendor_count_row = None
    work_note_row = None
    if weekly_end:
        for row_idx in need_rows:
            if row_idx > weekly_end and (not night_start or row_idx < night_start):
                day_need_row = row_idx
                break
        for row_idx in vendor_rows:
            if row_idx > weekly_end and (not night_start or row_idx < night_start):
                day_vendor_count_row = row_idx
                break
    if night_end:
        for row_idx in need_rows:
            if row_idx > night_end:
                night_need_row = row_idx
                break
        for row_idx in vendor_rows:
            if row_idx > night_end:
                night_vendor_count_row = row_idx
                break
        for row_idx in work_note_rows:
            if row_idx > night_end:
                work_note_row = row_idx
                break
    return {
        "weekly_rows": weekly_rows,
        "night_rows": night_rows,
        "day_need_row": day_need_row,
        "night_need_row": night_need_row,
        "need_row": day_need_row,
        "weekly_count_row": weekly_count_row,
        "night_count_row": night_count_row,
        "day_vendor_count_row": day_vendor_count_row,
        "night_vendor_count_row": night_vendor_count_row,
        "vendor_count_row": day_vendor_count_row,
        "work_note_row": work_note_row,
    }


def _ensure_support_slot_rows(sheet, *, section: str, required_count: int) -> list[int]:
    rows_meta = _locate_support_section_rows(sheet)
    key = "weekly_rows" if section == "day" else "night_rows"
    rows = list(rows_meta.get(key) or [])
    if not rows:
        return rows
    current_count = len(rows)
    if required_count <= current_count:
        for idx, row_idx in enumerate(rows, start=1):
            sheet.cell(row=row_idx, column=3, value=f"근무자 {idx}")
        return rows
    insert_at = rows[-1] + 1
    template_row = rows[-1]
    to_add = required_count - current_count
    sheet.insert_rows(insert_at, amount=to_add)
    for offset in range(to_add):
        target_row = insert_at + offset
        _copy_support_row_style(sheet, source_row=template_row, target_row=target_row)
    rows = [*rows, *[insert_at + offset for offset in range(to_add)]]
    for idx, row_idx in enumerate(rows, start=1):
        if idx == 1:
            if section == "day":
                sheet.cell(row=row_idx, column=2, value=ARLS_SUPPORT_WEEKLY_LABEL)
            else:
                sheet.cell(row=row_idx, column=2, value=ARLS_SUPPORT_NIGHT_LABEL)
        else:
            sheet.cell(row=row_idx, column=2, value=None)
        sheet.cell(row=row_idx, column=3, value=f"근무자 {idx}")
    return rows


def _populate_support_assignment_sections(
    sheet,
    *,
    day_keys: list[str],
    assignment_rows: list[dict[str, Any]],
    include_internal: bool,
) -> None:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in assignment_rows:
        date_value = row.get("work_date")
        if not isinstance(date_value, date):
            continue
        if not include_internal and bool(row.get("is_internal")):
            continue
        display_value = _support_assignment_display_value(row, include_internal=include_internal)
        if not display_value:
            continue
        key = (date_value.isoformat(), str(row.get("support_period") or "day").strip() or "day")
        grouped.setdefault(key, []).append(dict(row))
    max_weekly = max(
        ARLS_SUPPORT_WEEKLY_MAX_SLOTS,
        max((len(grouped.get((date_key, "day"), [])) for date_key in day_keys), default=0),
    )
    max_night = max(
        ARLS_SUPPORT_NIGHT_MAX_SLOTS,
        max((len(grouped.get((date_key, "night"), [])) for date_key in day_keys), default=0),
    )
    weekly_rows = _ensure_support_slot_rows(sheet, section="day", required_count=max_weekly)
    night_rows = _ensure_support_slot_rows(sheet, section="night", required_count=max_night)
    rows_meta = _locate_support_section_rows(sheet)
    for row_idx in [*weekly_rows, *night_rows]:
        for idx, _date_key in enumerate(day_keys):
            col_idx = ARLS_DATE_START_COL + idx
            sheet.cell(row=row_idx, column=col_idx, value=None)
    for idx, date_key in enumerate(day_keys):
        col_idx = ARLS_DATE_START_COL + idx
        weekly_group = sorted(
            grouped.get((date_key, "day"), []),
            key=lambda item: (int(item.get("slot_index") or 0), str(item.get("worker_name") or "")),
        )
        night_group = sorted(
            grouped.get((date_key, "night"), []),
            key=lambda item: (int(item.get("slot_index") or 0), str(item.get("worker_name") or "")),
        )
        for slot_idx, row in enumerate(weekly_group):
            if slot_idx >= len(weekly_rows):
                break
            sheet.cell(
                row=weekly_rows[slot_idx],
                column=col_idx,
                value=_support_assignment_display_value(row, include_internal=include_internal) or None,
            )
        for slot_idx, row in enumerate(night_group):
            if slot_idx >= len(night_rows):
                break
            sheet.cell(
                row=night_rows[slot_idx],
                column=col_idx,
                value=_support_assignment_display_value(row, include_internal=include_internal) or None,
            )
        if rows_meta.get("weekly_count_row"):
            sheet.cell(row=rows_meta["weekly_count_row"], column=col_idx, value=len(weekly_group) or 0)
        if rows_meta.get("night_count_row"):
            sheet.cell(row=rows_meta["night_count_row"], column=col_idx, value=len(night_group) or 0)


def _build_support_only_workbook(
    *,
    export_ctx: dict[str, Any],
    target_tenant: dict,
    site_row: dict,
    month_key: str,
    source_revision: str,
    active_assignments: list[dict[str, Any]],
) -> Workbook:
    workbook = export_ctx["workbook"]
    visible_sheet = workbook[ARLS_SHEET_NAME]
    visible_sheet.title = _build_support_form_sheet_name(
        str(site_row.get("site_name") or "").strip(),
        str(site_row.get("site_code") or "").strip(),
    )
    hidden_sheet_names = [name for name in workbook.sheetnames if name not in {visible_sheet.title, ARLS_METADATA_SHEET_NAME}]
    for name in hidden_sheet_names:
        if name in workbook.sheetnames:
            workbook.remove(workbook[name])
    if ARLS_METADATA_SHEET_NAME in workbook.sheetnames:
        workbook.remove(workbook[ARLS_METADATA_SHEET_NAME])
    data_start_row = _find_template_data_start_row(visible_sheet)
    summary_start_row = _find_template_summary_start_row(visible_sheet, fallback=data_start_row + 42)
    for row_idx in range(data_start_row, summary_start_row):
        visible_sheet.row_dimensions[row_idx].hidden = True
    day_keys = _month_day_keys(*_month_bounds(month_key))
    if active_assignments:
        support_rows = active_assignments
    else:
        support_rows = _fallback_support_assignments_from_export_ctx(export_ctx, include_internal=True)
    _populate_support_assignment_sections(
        visible_sheet,
        day_keys=day_keys,
        assignment_rows=support_rows,
        include_internal=True,
    )
    _write_support_roundtrip_metadata_sheet(
        workbook,
        tenant_code=str(target_tenant.get("tenant_code") or "").strip(),
        site_code=str(site_row.get("site_code") or "").strip(),
        site_name=str(site_row.get("site_name") or "").strip(),
        month_key=month_key,
        source_revision=source_revision,
        template_version=str(export_ctx.get("template_version") or ARLS_EXPORT_TEMPLATE_VERSION),
        support_form_version=ARLS_SUPPORT_FORM_VERSION,
    )
    return workbook


def _parse_support_form_lines(value: object) -> list[str]:
    text = _normalize_workbook_display_value(value)
    if not text:
        return []
    return [line.strip() for line in re.split(r"[\r\n]+", text) if line and line.strip()]


def _split_support_roundtrip_external_label(raw_entry: str) -> tuple[str | None, str]:
    normalized = str(raw_entry or "").strip()
    if not normalized:
        return None, ""
    parts = normalized.split(None, 1)
    if len(parts) < 2:
        return None, normalized
    return parts[0].strip() or None, parts[1].strip()


def _parse_support_form_sheet(sheet) -> dict[str, Any]:
    date_columns, month_ctx = _extract_arls_date_columns(sheet)
    if not date_columns:
        raise HTTPException(status_code=400, detail="지원근무 워크북 날짜 헤더를 읽지 못했습니다.")
    rows_meta = _locate_support_section_rows(sheet)
    parsed_entries: list[dict[str, Any]] = []
    validation_errors: list[str] = []
    occupied_slots: set[tuple[str, str, int]] = set()
    for support_period, slot_rows in (("day", rows_meta.get("weekly_rows") or []), ("night", rows_meta.get("night_rows") or [])):
        for row_slot_index, row_idx in enumerate(slot_rows, start=1):
            for col_idx, schedule_date in date_columns.items():
                raw_value = _normalize_workbook_display_value(sheet.cell(row=row_idx, column=col_idx).value)
                lines = _parse_support_form_lines(raw_value)
                if not lines:
                    continue
                for line_offset, line in enumerate(lines):
                    slot_index = row_slot_index + line_offset
                    slot_key = (schedule_date.isoformat(), support_period, slot_index)
                    if slot_key in occupied_slots:
                        validation_errors.append(
                            f"{sheet.title} {schedule_date.isoformat()} {support_period} slot {slot_index}: line-break/row overlap"
                        )
                        continue
                    occupied_slots.add(slot_key)
                    parsed_entries.append(
                        {
                            "row_no": row_idx,
                            "schedule_date": schedule_date,
                            "support_period": support_period,
                            "slot_index": slot_index,
                            "raw_entry": line,
                            "workbook_value": raw_value,
                            "section_label": ARLS_SUPPORT_WEEKLY_LABEL if support_period == "day" else ARLS_SUPPORT_NIGHT_LABEL,
                        }
                    )
    need_cells: list[dict[str, Any]] = []
    need_row = rows_meta.get("need_row")
    if need_row:
        for col_idx, schedule_date in date_columns.items():
            raw_value = _normalize_workbook_display_value(sheet.cell(row=need_row, column=col_idx).value)
            if not raw_value:
                continue
            need_cells.append(
                {
                    "row_no": need_row,
                    "schedule_date": schedule_date,
                    "work_value": raw_value,
                    "source_block": "daytime_need",
                    "section_label": "필요인원 수",
                }
            )
    return {
        "month_ctx": month_ctx,
        "support_entries": parsed_entries,
        "need_cells": need_cells,
        "validation_errors": validation_errors,
    }


def _resolve_internal_support_template(
    *,
    template_map: dict[str, dict[str, Any]],
    support_period: str,
) -> dict[str, Any] | None:
    duty_type = "night" if support_period == "night" else "day"
    return dict(template_map.get(duty_type) or {}) or None


def _build_support_roundtrip_preview_result(
    conn,
    *,
    workbook: Workbook,
    target_tenant: dict,
    site_row: dict,
    selected_month: str,
    user: dict,
    filename: str,
) -> dict[str, Any]:
    source_row = _get_support_roundtrip_source(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=selected_month,
    )
    blocked_reasons: list[str] = []
    if not source_row:
        blocked_reasons.append("활성 Supervisor 소스 월간 파일이 없어 HQ 지원근무 병합을 진행할 수 없습니다.")
    export_ctx = _collect_monthly_export_context(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=selected_month,
        user=user,
    )
    metadata = _read_support_roundtrip_metadata(workbook)
    if source_row:
        blocked_reasons.extend(
            _validate_support_roundtrip_metadata(
                metadata,
                expected_tenant_code=str(target_tenant.get("tenant_code") or "").strip(),
                expected_site_code=str(site_row.get("site_code") or "").strip(),
                expected_month=selected_month,
                expected_source_revision=str(source_row.get("source_revision") or "").strip(),
            )
        )
    visible_sheet_names = [
        name for name in workbook.sheetnames
        if name not in {ARLS_SUPPORT_METADATA_SHEET_NAME}
        and workbook[name].sheet_state == "visible"
    ]
    expected_sheet_name = _build_support_form_sheet_name(
        str(site_row.get("site_name") or "").strip(),
        str(site_row.get("site_code") or "").strip(),
    )
    parsed = None
    if not visible_sheet_names:
        blocked_reasons.append("지원근무 입력 시트를 찾지 못했습니다.")
    else:
        target_sheet_name = visible_sheet_names[0]
        if target_sheet_name != expected_sheet_name:
            blocked_reasons.append("업로드한 지원근무 파일의 시트명이 현재 현장명과 일치하지 않습니다.")
        parsed = _parse_support_form_sheet(workbook[target_sheet_name])
        blocked_reasons.extend(parsed.get("validation_errors") or [])

    active_assignments = _list_support_roundtrip_assignments(
        conn,
        source_id=str(source_row["id"]),
        source_revision=str(source_row["source_revision"]),
    ) if source_row else []
    current_assignment_index: dict[tuple[str, str, int], dict[str, Any]] = {}
    for item in active_assignments:
        work_date = item.get("work_date")
        if not isinstance(work_date, date):
            continue
        current_assignment_index[(work_date.isoformat(), str(item.get("support_period") or "day"), int(item.get("slot_index") or 1))] = dict(item)

    template_map = _fetch_default_schedule_template_map(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
    )
    current_body_index = _build_visible_value_index(export_ctx["parsed_sheet"].get("body_cells") or [])
    current_need_index = _build_visible_value_index(export_ctx["parsed_sheet"].get("need_cells") or [], include_employee=False)

    resolved_rows: list[dict[str, Any]] = []
    diff_counts: Counter[str] = Counter()
    error_counts: Counter[str] = Counter()

    if parsed:
        for item in parsed.get("support_entries") or []:
            schedule_date = item["schedule_date"]
            date_key = schedule_date.isoformat()
            support_period = str(item.get("support_period") or "day")
            slot_index = int(item.get("slot_index") or 1)
            current_row = current_assignment_index.get((date_key, support_period, slot_index)) or {}
            current_value = _support_assignment_display_value(current_row, include_internal=True)
            raw_entry = str(item.get("raw_entry") or "").strip()
            validation_code = None
            validation_error = None
            resolved_worker_type = None
            resolved_worker_name = None
            resolved_affiliation = None
            employee_row = None
            apply_action = "none"
            diff_category = "unchanged"
            is_blocking = False
            is_protected = False
            protected_reason = None
            internal_template = None

            try:
                worker_type, worker_name = parse_support_entry_text(raw_entry)
                resolved_worker_type = worker_type
                resolved_worker_name = worker_name
            except HTTPException as exc:
                validation_code = "invalid_support_entry"
                validation_error = str(exc.detail or "지원근무자 형식을 해석할 수 없습니다.")
                is_blocking = True
                worker_type = ""
                worker_name = ""

            if not validation_code and resolved_worker_type == "INTERNAL":
                employee_row = _resolve_employee_by_full_name(
                    conn,
                    tenant_id=str(target_tenant["id"]),
                    site_id=str(site_row["id"]),
                    full_name=resolved_worker_name,
                )
                if not employee_row:
                    resolved_worker_type = "F"
                    employee_row = None
                    resolved_affiliation, resolved_worker_name = _split_support_roundtrip_external_label(raw_entry)
                else:
                    internal_template = _resolve_internal_support_template(
                        template_map=template_map,
                        support_period=support_period,
                    )
                    if not internal_template:
                        validation_code = "internal_template_missing"
                        validation_error = "내부 지원근무에 사용할 기본 템플릿이 없습니다."
                        is_blocking = True
                    else:
                        duty_type = "night" if support_period == "night" else "day"
                        target_key = (_normalize_name_token(employee_row.get("full_name")), duty_type, date_key)
                        current_body = current_body_index.get(target_key) or {}
                        target_value = _format_export_hours_value(
                            internal_template.get("paid_hours")
                            if internal_template.get("paid_hours") is not None
                            else _resolve_export_row_hours(
                                {
                                    "paid_hours": internal_template.get("paid_hours"),
                                    "shift_start_time": internal_template.get("start_time"),
                                    "shift_end_time": internal_template.get("end_time"),
                                    "shift_type": "night" if support_period == "night" else "day",
                                }
                            )
                        )
                        current_body_value = str(current_body.get("work_value") or "").strip()
                        if current_body_value and current_body_value != target_value:
                            validation_code = "internal_schedule_conflict"
                            validation_error = "현재 직원 일정 값과 충돌하여 자동 반영할 수 없습니다."
                            is_blocking = True
                            diff_category = "conflict"
                            apply_action = "none"

            if not validation_code:
                if not current_value:
                    diff_category = "create"
                    apply_action = "upsert_support_assignment"
                elif current_value == raw_entry or current_value == _support_assignment_display_value(current_row, include_internal=True):
                    diff_category = "unchanged"
                else:
                    diff_category = "update"
                    apply_action = "upsert_support_assignment"
            else:
                diff_category = "conflict" if is_blocking else "invalid"
                apply_action = "none"
                error_counts[validation_code] += 1

            diff_counts[diff_category] += 1
            resolved_rows.append(
                {
                    "row_no": int(item.get("row_no") or 0),
                    "site_code": str(site_row.get("site_code") or "").strip(),
                    "schedule_date": date_key,
                    "support_period": support_period,
                    "slot_index": slot_index,
                    "section_label": item.get("section_label"),
                    "workbook_value": raw_entry,
                    "current_value": current_value or None,
                    "resolved_worker_type": resolved_worker_type,
                    "resolved_worker_name": resolved_worker_name,
                    "resolved_affiliation": resolved_affiliation,
                    "employee_id": str((employee_row or {}).get("id") or "").strip() or None,
                    "employee_code": str((employee_row or {}).get("employee_code") or "").strip() or None,
                    "employee_name": str((employee_row or {}).get("full_name") or "").strip() or None,
                    "diff_category": diff_category,
                    "apply_action": apply_action,
                    "validation_code": validation_code,
                    "validation_error": validation_error,
                    "is_blocking": is_blocking,
                    "is_protected": is_protected,
                    "protected_reason": protected_reason,
                    "internal_template_id": str((internal_template or {}).get("id") or "").strip() or None,
                    "internal_shift_type": "night" if support_period == "night" else "day",
                    "internal_shift_start_time": _normalize_time_text((internal_template or {}).get("start_time")),
                    "internal_shift_end_time": _normalize_time_text((internal_template or {}).get("end_time")),
                    "internal_paid_hours": (internal_template or {}).get("paid_hours"),
                }
            )

        for row in parsed.get("need_cells") or []:
            schedule_date = row.get("schedule_date")
            if not isinstance(schedule_date, date):
                continue
            date_key = schedule_date.isoformat()
            workbook_value = str(row.get("work_value") or "").strip()
            current_row = current_need_index.get(("daytime_need", date_key)) or {}
            current_value = str(current_row.get("work_value") or "").strip()
            required_count, parsed_raw_text = _parse_daytime_need_value(workbook_value)
            validation_code = None
            validation_error = None
            apply_action = "none"
            diff_category = "unchanged"
            is_blocking = False
            if required_count is None:
                validation_code = "invalid_value"
                validation_error = "필요인원 수를 해석할 수 없습니다."
                is_blocking = True
            else:
                if workbook_value == current_value:
                    diff_category = "unchanged"
                else:
                    diff_category = "update" if current_value else "create"
                    apply_action = "upsert_need_count"
            if validation_code:
                error_counts[validation_code] += 1
                diff_counts["conflict"] += 1
            else:
                diff_counts[diff_category] += 1
            resolved_rows.append(
                {
                    "row_no": int(row.get("row_no") or 0),
                    "site_code": str(site_row.get("site_code") or "").strip(),
                    "schedule_date": date_key,
                    "support_period": "day",
                    "slot_index": 0,
                    "section_label": "필요인원 수",
                    "workbook_value": workbook_value,
                    "current_value": current_value or None,
                    "resolved_worker_type": None,
                    "resolved_worker_name": None,
                    "employee_code": None,
                    "employee_name": None,
                    "diff_category": diff_category if not validation_code else "conflict",
                    "apply_action": apply_action,
                    "validation_code": validation_code,
                    "validation_error": validation_error,
                    "is_blocking": is_blocking,
                    "is_protected": False,
                    "protected_reason": None,
                    "source_block": "daytime_need",
                    "need_raw_text": parsed_raw_text,
                    "need_required_count": required_count,
                }
            )

    can_apply = not blocked_reasons and any(
        str(row.get("apply_action") or "").strip() in {"upsert_support_assignment", "upsert_need_count"} and not bool(row.get("is_blocking"))
        for row in resolved_rows
    )
    return {
        "resolved_rows": resolved_rows,
        "preview_rows": resolved_rows[:IMPORT_PREVIEW_LIMIT],
        "diff_counts": dict(diff_counts),
        "error_counts": dict(error_counts),
        "blocked_reasons": list(dict.fromkeys(reason for reason in blocked_reasons if str(reason).strip())),
        "total_rows": len(resolved_rows),
        "valid_rows": sum(
            1 for row in resolved_rows
            if str(row.get("apply_action") or "").strip() in {"upsert_support_assignment", "upsert_need_count"} and not bool(row.get("is_blocking"))
        ),
        "invalid_rows": sum(1 for row in resolved_rows if bool(row.get("is_blocking"))),
        "metadata": {
            "tenant_code": metadata.get("tenant_code"),
            "site_code": metadata.get("site_code"),
            "site_name": metadata.get("site_name"),
            "month": metadata.get("month"),
            "source_revision": metadata.get("source_revision"),
            "current_source_revision": str((source_row or {}).get("source_revision") or "").strip() or None,
            "template_version": metadata.get("template_version"),
            "support_form_version": metadata.get("support_form_version"),
            "extracted_at_kst": metadata.get("extracted_at_kst"),
            "is_stale": "stale_source_revision" in blocked_reasons,
        },
        "can_apply": can_apply,
        "source_row": source_row,
    }


def _build_support_roundtrip_revision(payload: dict[str, Any]) -> str:
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()[:16]


def _persist_support_roundtrip_preview_batch(
    conn,
    *,
    source_row: dict,
    target_tenant: dict,
    site_row: dict,
    month_key: str,
    preview: dict[str, Any],
    filename: str,
    user: dict,
) -> tuple[uuid.UUID, str]:
    batch_id = uuid.uuid4()
    batch_revision = _build_support_roundtrip_revision(
        {
            "source_revision": str(source_row.get("source_revision") or "").strip(),
            "site_code": str(site_row.get("site_code") or "").strip(),
            "month": month_key,
            "filename": filename,
            "rows": preview.get("resolved_rows") or [],
        }
    )
    resolved_rows = list(preview.get("resolved_rows") or [])
    metadata = preview.get("metadata") or {}
    blocked_reasons = list(dict.fromkeys(str(item).strip() for item in (preview.get("blocked_reasons") or []) if str(item).strip()))
    diff_counts = dict(preview.get("diff_counts") or {})
    status = "blocked" if blocked_reasons else "previewed"
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO schedule_support_roundtrip_batches (
                id, source_id, tenant_id, site_id, site_code, month_key, source_revision,
                workbook_kind, filename, uploaded_by, uploaded_role, status,
                template_version, support_form_version, is_stale, is_partial,
                total_rows, meaningful_rows, blocked_reasons_json, diff_counts_json
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s::jsonb, %s::jsonb
            )
            """,
            (
                batch_id,
                source_row["id"],
                target_tenant["id"],
                site_row["id"],
                str(site_row.get("site_code") or "").strip(),
                month_key,
                str(source_row.get("source_revision") or "").strip(),
                "support_roundtrip_hq",
                filename,
                user["id"],
                _support_roundtrip_normalize_role(user),
                status,
                str(metadata.get("template_version") or "").strip() or None,
                str(metadata.get("support_form_version") or "").strip() or None,
                bool(metadata.get("is_stale")),
                True,
                int(preview.get("total_rows") or 0),
                len(resolved_rows),
                json.dumps(blocked_reasons, ensure_ascii=False),
                json.dumps(diff_counts, ensure_ascii=False),
            ),
        )
        for index, row in enumerate(resolved_rows, start=1):
            cur.execute(
                """
                INSERT INTO schedule_support_roundtrip_rows (
                    id, batch_id, row_no, tenant_id, site_id, site_code, schedule_date,
                    support_period, slot_index, source_block, section_label,
                    workbook_value, current_value, resolved_worker_type, resolved_worker_name,
                    employee_id, employee_code, employee_name, apply_action, diff_category,
                    validation_code, validation_error, is_blocking, is_protected,
                    protected_reason, line_count, payload_json
                )
                VALUES (
                    %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s::jsonb
                )
                """,
                (
                    uuid.uuid4(),
                    batch_id,
                    int(row.get("row_no") or index),
                    target_tenant["id"],
                    site_row["id"],
                    str(row.get("site_code") or site_row.get("site_code") or "").strip(),
                    row.get("schedule_date"),
                    str(row.get("support_period") or "day").strip() or "day",
                    int(row.get("slot_index") or 0),
                    str(row.get("source_block") or "").strip() or "support_assignment",
                    str(row.get("section_label") or "").strip() or None,
                    str(row.get("workbook_value") or "").strip() or None,
                    str(row.get("current_value") or "").strip() or None,
                    str(row.get("resolved_worker_type") or "").strip() or None,
                    str(row.get("resolved_worker_name") or "").strip() or None,
                    row.get("employee_id"),
                    str(row.get("employee_code") or "").strip() or None,
                    str(row.get("employee_name") or "").strip() or None,
                    str(row.get("apply_action") or "none").strip() or "none",
                    str(row.get("diff_category") or "unchanged").strip() or "unchanged",
                    str(row.get("validation_code") or "").strip() or None,
                    str(row.get("validation_error") or "").strip() or None,
                    bool(row.get("is_blocking")),
                    bool(row.get("is_protected")),
                    str(row.get("protected_reason") or "").strip() or None,
                    max(1, len(_parse_support_form_lines(row.get("workbook_value")))) if str(row.get("workbook_value") or "").strip() else 0,
                    json.dumps(row, ensure_ascii=False, sort_keys=True, default=str),
                ),
            )
    return batch_id, batch_revision


def _upsert_support_roundtrip_assignment_row(
    cur,
    *,
    source_row: dict,
    payload: dict[str, Any],
    batch_id: str,
    user_id: str,
) -> None:
    worker_name = str(payload.get("resolved_worker_name") or "").strip()
    affiliation = str(payload.get("resolved_affiliation") or "").strip()
    if affiliation and worker_name:
        stored_worker_name = f"{affiliation} {worker_name}".strip()
    else:
        stored_worker_name = worker_name
    cur.execute(
        """
        INSERT INTO schedule_support_roundtrip_assignments (
            id, source_id, source_revision, tenant_id, site_id, site_code,
            work_date, support_period, slot_index, worker_type, worker_name,
            employee_id, employee_code, employee_name, is_internal,
            internal_shift_type, internal_template_id, internal_shift_start_time,
            internal_shift_end_time, internal_paid_hours, source_batch_id, created_by,
            created_at, updated_at
        )
        VALUES (
            %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, %s,
            timezone('utc', now()), timezone('utc', now())
        )
        ON CONFLICT (source_id, source_revision, work_date, support_period, slot_index)
        DO UPDATE SET
            worker_type = EXCLUDED.worker_type,
            worker_name = EXCLUDED.worker_name,
            employee_id = EXCLUDED.employee_id,
            employee_code = EXCLUDED.employee_code,
            employee_name = EXCLUDED.employee_name,
            is_internal = EXCLUDED.is_internal,
            internal_shift_type = EXCLUDED.internal_shift_type,
            internal_template_id = EXCLUDED.internal_template_id,
            internal_shift_start_time = EXCLUDED.internal_shift_start_time,
            internal_shift_end_time = EXCLUDED.internal_shift_end_time,
            internal_paid_hours = EXCLUDED.internal_paid_hours,
            source_batch_id = EXCLUDED.source_batch_id,
            created_by = COALESCE(schedule_support_roundtrip_assignments.created_by, EXCLUDED.created_by),
            updated_at = timezone('utc', now())
        """,
        (
            uuid.uuid4(),
            source_row["id"],
            str(source_row.get("source_revision") or "").strip(),
            source_row["tenant_id"],
            source_row["site_id"],
            str(source_row.get("site_code") or "").strip(),
            payload["schedule_date"],
            str(payload.get("support_period") or "day").strip() or "day",
            int(payload.get("slot_index") or 1),
            str(payload.get("resolved_worker_type") or "").strip(),
            stored_worker_name,
            payload.get("employee_id"),
            str(payload.get("employee_code") or "").strip() or None,
            str(payload.get("employee_name") or "").strip() or None,
            str(payload.get("resolved_worker_type") or "").strip() == "INTERNAL",
            str(payload.get("internal_shift_type") or "").strip() or None,
            payload.get("internal_template_id"),
            _normalize_time_text(payload.get("internal_shift_start_time")),
            _normalize_time_text(payload.get("internal_shift_end_time")),
            payload.get("internal_paid_hours"),
            batch_id,
            user_id,
        ),
    )


def _build_support_roundtrip_employee_row_index(sheet, *, employee_blocks: list[dict[str, Any]]) -> dict[tuple[str, str], int]:
    data_start_row = _find_template_data_start_row(sheet)
    row_index: dict[tuple[str, str], int] = {}
    current_row = data_start_row
    for block in employee_blocks:
        employee_id = str(block.get("employee_id") or "").strip()
        if employee_id:
            row_index[(employee_id, "day")] = current_row
            row_index[(employee_id, "overtime")] = current_row + 1
            row_index[(employee_id, "night")] = current_row + 2
        current_row += 3
    return row_index


def _apply_internal_support_roundtrip_assignments(
    sheet,
    *,
    export_ctx: dict[str, Any],
    assignment_rows: list[dict[str, Any]],
) -> None:
    row_index = _build_support_roundtrip_employee_row_index(sheet, employee_blocks=export_ctx.get("employee_blocks") or [])
    date_columns, _month_ctx = _extract_arls_date_columns(sheet)
    date_to_col = {value.isoformat(): col_idx for col_idx, value in date_columns.items()}
    for row in assignment_rows:
        if not bool(row.get("is_internal")):
            continue
        employee_id = str(row.get("employee_id") or "").strip()
        if not employee_id:
            continue
        work_date = row.get("work_date")
        if not isinstance(work_date, date):
            continue
        date_key = work_date.isoformat()
        col_idx = date_to_col.get(date_key)
        duty_type = "night" if str(row.get("internal_shift_type") or row.get("support_period") or "").strip() == "night" else "day"
        target_row = row_index.get((employee_id, duty_type))
        if not col_idx or not target_row:
            logger.warning("[schedule][support_roundtrip] missing final overlay target employee=%s date=%s duty=%s", employee_id, date_key, duty_type)
            continue
        derived_value = _format_export_hours_value(
            row.get("internal_paid_hours")
            if row.get("internal_paid_hours") is not None
            else _resolve_export_row_hours(
                {
                    "paid_hours": row.get("internal_paid_hours"),
                    "shift_start_time": row.get("internal_shift_start_time"),
                    "shift_end_time": row.get("internal_shift_end_time"),
                    "shift_type": duty_type,
                }
            )
        )
        if not derived_value:
            logger.warning("[schedule][support_roundtrip] missing internal overlay hours employee=%s date=%s duty=%s", employee_id, date_key, duty_type)
            continue
        cell = sheet.cell(row=target_row, column=col_idx)
        current_value = _normalize_workbook_display_value(cell.value)
        if current_value and current_value != derived_value:
            logger.warning(
                "[schedule][support_roundtrip] skip conflicting internal overlay employee=%s date=%s duty=%s current=%s next=%s",
                employee_id,
                date_key,
                duty_type,
                current_value,
                derived_value,
            )
            continue
        cell.value = derived_value


def _build_support_roundtrip_final_workbook(
    conn,
    *,
    source_row: dict,
    target_tenant: dict,
    site_row: dict,
    month_key: str,
    user: dict,
) -> Workbook:
    export_ctx = _collect_monthly_export_context(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=month_key,
        user=user,
    )
    workbook = export_ctx["workbook"]
    visible_sheet = workbook[ARLS_SHEET_NAME]
    assignment_rows = _list_support_roundtrip_assignments(
        conn,
        source_id=str(source_row["id"]),
        source_revision=str(source_row.get("source_revision") or "").strip(),
    )
    day_keys = _month_day_keys(*_month_bounds(month_key))
    _populate_support_assignment_sections(
        visible_sheet,
        day_keys=day_keys,
        assignment_rows=assignment_rows,
        include_internal=False,
    )
    _apply_internal_support_roundtrip_assignments(
        visible_sheet,
        export_ctx=export_ctx,
        assignment_rows=assignment_rows,
    )
    return workbook


def _build_finance_review_filename(*, month_key: str, site_code: str, generated_at: datetime | None = None) -> str:
    generated = generated_at or datetime.now(timezone(timedelta(hours=9)))
    return f"{month_key[:4]}년 {int(month_key[5:7])}월 근무표_{site_code}_1차확인본_{generated.strftime('%y%m%d')}.xlsx"


def _build_finance_final_filename(*, month_key: str, site_code: str, generated_at: datetime | None = None) -> str:
    generated = generated_at or datetime.now(timezone(timedelta(hours=9)))
    return f"{month_key[:4]}년 {int(month_key[5:7])}월 근무표_{site_code}_2차최종_{generated.strftime('%y%m%d')}.xlsx"


def _get_finance_submission_state(conn, *, tenant_id: str, site_id: str, month_key: str) -> dict[str, Any] | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT fss.*,
                   review_user.username AS review_downloaded_by_username,
                   final_user.username AS final_uploaded_by_username
            FROM schedule_finance_submission_states fss
            LEFT JOIN arls_users review_user ON review_user.id = fss.review_downloaded_by
            LEFT JOIN arls_users final_user ON final_user.id = fss.final_uploaded_by
            WHERE fss.tenant_id = %s
              AND fss.site_id = %s
              AND fss.month_key = %s
            LIMIT 1
            """,
            (tenant_id, site_id, month_key),
        )
        return cur.fetchone()


def _ensure_finance_submission_state(
    conn,
    *,
    tenant_id: str,
    site_id: str,
    site_code: str,
    month_key: str,
) -> dict[str, Any]:
    current = _get_finance_submission_state(
        conn,
        tenant_id=tenant_id,
        site_id=site_id,
        month_key=month_key,
    )
    if current:
        return current
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO schedule_finance_submission_states (
                id, tenant_id, site_id, site_code, month_key, state,
                created_at, updated_at
            )
            VALUES (
                %s, %s, %s, %s, %s, 'review_download_ready',
                timezone('utc', now()), timezone('utc', now())
            )
            RETURNING *
            """,
            (uuid.uuid4(), tenant_id, site_id, site_code, month_key),
        )
        return cur.fetchone()


def _sync_finance_submission_state(
    conn,
    *,
    target_tenant: dict,
    site_row: dict,
    month_key: str,
) -> tuple[dict[str, Any], str]:
    current_revision = _build_schedule_export_revision(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        site_code=str(site_row["site_code"]),
        month_key=month_key,
    )
    state_row = _ensure_finance_submission_state(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        site_code=str(site_row["site_code"]),
        month_key=month_key,
    )
    next_state = str(state_row.get("state") or "review_download_ready").strip() or "review_download_ready"
    final_upload_stale = bool(state_row.get("final_upload_stale"))
    final_download_enabled = bool(state_row.get("final_download_enabled"))
    blocked_reasons: list[str] = []

    active_final_revision = str(state_row.get("active_final_revision") or "").strip()
    review_download_revision = str(state_row.get("review_download_revision") or "").strip()

    if active_final_revision and active_final_revision != current_revision:
        next_state = "final_upload_stale"
        final_upload_stale = True
        final_download_enabled = False
        blocked_reasons.append("최종 업로드 이후 상위 월간 truth가 변경되어 현재 최종본이 stale 상태입니다.")
    elif bool(state_row.get("conflict_required")):
        next_state = "conflict_manual_review_required"
        final_download_enabled = False
        blocked_reasons.append("수동 검토가 필요한 충돌이 있어 최종본 다운로드가 차단됩니다.")
    elif active_final_revision and active_final_revision == current_revision:
        next_state = "hq_final_download_ready"
        final_upload_stale = False
        final_download_enabled = True
    elif review_download_revision and review_download_revision == current_revision:
        next_state = "waiting_final_upload"
        final_upload_stale = False
        final_download_enabled = False
    else:
        next_state = "review_download_ready"
        final_upload_stale = False
        final_download_enabled = False
        if review_download_revision and review_download_revision != current_revision:
            blocked_reasons.append("현재 review download는 최신 assembled revision 기준이 아니어서 다시 내려받아야 합니다.")

    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE schedule_finance_submission_states
            SET current_revision = %s,
                state = %s,
                final_upload_stale = %s,
                final_download_enabled = %s,
                updated_at = timezone('utc', now())
            WHERE id = %s
            """,
            (
                current_revision,
                next_state,
                final_upload_stale,
                final_download_enabled,
                state_row["id"],
            ),
        )
    state_row = _get_finance_submission_state(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month_key,
    ) or state_row
    state_row["_derived_blocked_reasons"] = blocked_reasons
    return state_row, current_revision


def _build_finance_submission_status_payload(
    conn,
    *,
    target_tenant: dict,
    site_row: dict,
    month_key: str,
) -> FinanceSubmissionStatusOut:
    state_row, current_revision = _sync_finance_submission_state(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=month_key,
    )
    blocked_reasons = list(dict.fromkeys(str(item).strip() for item in (state_row.get("_derived_blocked_reasons") or []) if str(item).strip()))
    return FinanceSubmissionStatusOut(
        site_code=str(site_row.get("site_code") or "").strip(),
        month=month_key,
        state=str(state_row.get("state") or "review_download_ready").strip() or "review_download_ready",
        current_revision=current_revision,
        review_download_ready=True,
        review_download_revision=str(state_row.get("review_download_revision") or "").strip() or None,
        review_downloaded_at=state_row.get("review_downloaded_at"),
        review_downloaded_by=str(state_row.get("review_downloaded_by_username") or state_row.get("review_downloaded_by") or "").strip() or None,
        review_download_filename=str(state_row.get("review_download_filename") or "").strip() or None,
        final_download_enabled=bool(state_row.get("final_download_enabled")),
        final_upload_stale=bool(state_row.get("final_upload_stale")),
        active_final_revision=str(state_row.get("active_final_revision") or "").strip() or None,
        active_final_source_revision=str(state_row.get("active_final_source_revision") or "").strip() or None,
        active_final_filename=str(state_row.get("active_final_filename") or "").strip() or None,
        final_uploaded_at=state_row.get("final_uploaded_at"),
        final_uploaded_by=str(state_row.get("final_uploaded_by_username") or state_row.get("final_uploaded_by") or "").strip() or None,
        last_event=str(state_row.get("last_event") or "").strip() or None,
        blocked_reasons=blocked_reasons,
    )


def _create_finance_submission_batch(
    conn,
    *,
    submission_row: dict,
    batch_kind: str,
    source_revision: str,
    filename: str,
    actor_id: str,
    actor_role: str,
    artifact_bytes: bytes | None = None,
    import_batch_id: str | None = None,
    status: str = "created",
    is_stale: bool = False,
    total_rows: int = 0,
    valid_rows: int = 0,
    invalid_rows: int = 0,
    blocked_reasons: list[str] | None = None,
    diff_counts: dict[str, int] | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    batch_id = uuid.uuid4()
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO schedule_finance_submission_batches (
                id, submission_id, tenant_id, site_id, site_code, month_key,
                batch_kind, source_revision, filename, artifact_bytes, import_batch_id,
                actor_id, actor_role, status, is_stale, total_rows, valid_rows, invalid_rows,
                blocked_reasons_json, diff_counts_json, metadata_json
            )
            VALUES (
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s,
                %s::jsonb, %s::jsonb, %s::jsonb
            )
            RETURNING *
            """,
            (
                batch_id,
                submission_row["id"],
                submission_row["tenant_id"],
                submission_row["site_id"],
                submission_row["site_code"],
                submission_row["month_key"],
                batch_kind,
                source_revision,
                filename,
                artifact_bytes,
                import_batch_id,
                actor_id,
                actor_role,
                status,
                is_stale,
                total_rows,
                valid_rows,
                invalid_rows,
                json.dumps(blocked_reasons or [], ensure_ascii=False),
                json.dumps(diff_counts or {}, ensure_ascii=False),
                json.dumps(metadata or {}, ensure_ascii=False, default=str),
            ),
        )
        return cur.fetchone()


def _get_finance_submission_batch(conn, *, batch_id: str, tenant_id: str) -> dict[str, Any] | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT b.*, s.state AS submission_state, s.current_revision AS submission_current_revision,
                   s.site_id AS submission_site_id, s.site_code AS submission_site_code, s.month_key AS submission_month_key
            FROM schedule_finance_submission_batches b
            JOIN schedule_finance_submission_states s ON s.id = b.submission_id
            WHERE b.id = %s
              AND b.tenant_id = %s
            LIMIT 1
            """,
            (batch_id, tenant_id),
        )
        return cur.fetchone()


@router.get("/support-roundtrip/status", response_model=SupportRoundtripStatusOut)
def get_support_roundtrip_status(
    month: str = Query(..., description="YYYY-MM"),
    site_code: str = Query(..., min_length=1, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not (_can_use_support_roundtrip_source(user) or _can_use_support_roundtrip_hq(user) or _can_use_support_roundtrip_final_download(user)):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_code).strip(),
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")
    source_row = _get_support_roundtrip_source(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month,
    )
    return _build_support_roundtrip_status_payload(
        conn,
        source_row=source_row,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        site_code=str(site_row.get("site_code") or "").strip(),
        month_key=month,
    )


@router.get("/support-roundtrip/hq-workspace", response_model=SupportRosterHqWorkspaceOut)
def get_support_roundtrip_hq_workspace(
    month: str = Query(..., description="YYYY-MM"),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_hq(user):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    return _build_support_roster_hq_workspace_payload(
        conn,
        tenant_id=str(target_tenant["id"]),
        tenant_code=str(target_tenant["tenant_code"]),
        month_key=month,
    )


@router.get("/support-roundtrip/hq-roster-workbook")
def download_support_roundtrip_hq_roster_workbook(
    month: str = Query(..., description="YYYY-MM"),
    scope: str = Query(default="all", max_length=16),
    site_code: str | None = Query(default=None, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_hq(user):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    workbook, written_sites = _build_support_roster_hq_download_workbook(
        conn,
        target_tenant=target_tenant,
        month_key=month,
        scope=scope,
        selected_site_code=site_code,
        user=user,
    )
    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    now_kst = datetime.now(timezone(timedelta(hours=9)))
    normalized_scope = "site" if str(scope or "").strip().lower() == "site" else "all"
    filename = (
        f"{month[:4]}년 {int(month[5:7])}월 지원근무자 배정 workbook_"
        f"{str(site_code or 'ALL').strip() if normalized_scope == 'site' else 'ALL'}_"
        f"{now_kst.strftime('%y%m%d')}.xlsx"
    )
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
            "X-Sheet-Count": str(len(written_sites)),
        },
    )


@router.post("/support-roundtrip/hq-roster-upload/inspect", response_model=SupportRosterHqUploadInspectOut)
def inspect_support_roundtrip_hq_roster_upload(
    file: UploadFile,
    month: str | None = Form(default=None),
    tenant_code: str | None = Form(default=None),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_hq(user):
        raise HTTPException(status_code=403, detail="forbidden")
    if not month or not str(month).strip():
        raise HTTPException(status_code=400, detail="month is required")
    selected_month = str(month).strip()
    _month_bounds(selected_month)
    raw_bytes = file.file.read()
    try:
        workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=False, data_only=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid support workbook") from exc
    try:
        target_tenant = _resolve_target_tenant(conn, user, tenant_code)
        return _build_support_roster_hq_upload_inspect_result(
            conn,
            workbook=workbook,
            target_tenant=target_tenant,
            selected_month=selected_month,
            filename=file.filename or "support_roster.xlsx",
            user=user,
        )
    finally:
        workbook.close()


@router.post("/support-roundtrip/hq-roster-upload/{batch_id}/apply", response_model=SupportRosterHqApplyOut)
def apply_support_roundtrip_hq_roster_upload(
    batch_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_hq(user):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    batch = _get_sentrix_hq_roster_batch(
        conn,
        batch_id=batch_id,
        tenant_id=str(target_tenant["id"]),
    )
    if not batch:
        raise HTTPException(status_code=404, detail="support hq roster batch not found")
    restored_result = _restore_sentrix_hq_roster_apply_result(batch)
    if restored_result and str(batch.get("status") or "").strip() in {"applied", "blocked"}:
        return restored_result
    return _apply_sentrix_hq_roster_batch(
        conn,
        batch_id=batch_id,
        batch=batch,
        target_tenant=target_tenant,
        user=user,
    )


@router.post("/support-roundtrip/arls-bridge/process")
def process_sentrix_support_arls_bridge_actions(
    batch_id: uuid.UUID | None = Query(default=None),
    include_failed: bool = Query(default=False),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_hq(user):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    result = _process_sentrix_support_arls_bridge_actions(
        conn,
        tenant_id=str(target_tenant["id"]),
        batch_id=str(batch_id) if batch_id else None,
        include_failed=bool(include_failed),
    )
    return {
        "tenant_code": str(target_tenant.get("tenant_code") or "").strip(),
        "batch_id": str(batch_id) if batch_id else None,
        **result,
    }


@router.get("/support-roundtrip/hq-workbook")
def download_support_roundtrip_hq_workbook(
    month: str = Query(..., description="YYYY-MM"),
    site_code: str = Query(..., min_length=1, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_hq(user):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_code).strip(),
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")
    source_row = _get_support_roundtrip_source(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month,
    )
    if not source_row:
        raise HTTPException(status_code=409, detail="support roundtrip source missing")
    export_ctx = _collect_monthly_export_context(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=month,
        user=user,
    )
    active_assignments = _list_support_roundtrip_assignments(
        conn,
        source_id=str(source_row["id"]),
        source_revision=str(source_row.get("source_revision") or "").strip(),
    )
    workbook = _build_support_only_workbook(
        export_ctx=export_ctx,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=month,
        source_revision=str(source_row.get("source_revision") or "").strip(),
        active_assignments=active_assignments,
    )
    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    now_kst = datetime.now(timezone(timedelta(hours=9)))
    filename = f"{month[:4]}년 {int(month[5:7])}월 지원근무자용 스케쥴 제출_{str(site_row['site_code']).strip()}_{now_kst.strftime('%y%m%d')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/support-roundtrip/hq-upload/preview", response_model=SupportRoundtripPreviewOut)
def preview_support_roundtrip_upload(
    file: UploadFile,
    site_code: str | None = Form(default=None),
    month: str | None = Form(default=None),
    tenant_code: str | None = Form(default=None),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_hq(user):
        raise HTTPException(status_code=403, detail="forbidden")
    if not site_code or not str(site_code).strip():
        raise HTTPException(status_code=400, detail="site_code is required")
    if not month or not str(month).strip():
        raise HTTPException(status_code=400, detail="month is required")
    _month_bounds(str(month).strip())
    raw_bytes = file.file.read()
    try:
        workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=False, data_only=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid support workbook") from exc
    try:
        target_tenant = _resolve_target_tenant(conn, user, tenant_code)
        site_row = _resolve_site_context_by_code(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_code=str(site_code).strip(),
        )
        if not site_row:
            raise HTTPException(status_code=404, detail="site not found")
        preview = _build_support_roundtrip_preview_result(
            conn,
            workbook=workbook,
            target_tenant=target_tenant,
            site_row=site_row,
            selected_month=str(month).strip(),
            user=user,
            filename=file.filename or "support_roundtrip.xlsx",
        )
        source_row = preview.get("source_row")
        if not source_row:
            return SupportRoundtripPreviewOut(
                batch_id=uuid.uuid4(),
                total_rows=int(preview.get("total_rows") or 0),
                valid_rows=int(preview.get("valid_rows") or 0),
                invalid_rows=int(preview.get("invalid_rows") or 0),
                preview_rows=[SupportRoundtripPreviewRowOut(**row) for row in (preview.get("preview_rows") or [])],
                diff_counts=dict(preview.get("diff_counts") or {}),
                blocked_reasons=list(preview.get("blocked_reasons") or []),
                metadata=SupportRoundtripPreviewMetadataOut(**(preview.get("metadata") or {})),
                can_apply=False,
            )
        batch_id, _batch_revision = _persist_support_roundtrip_preview_batch(
            conn,
            source_row=source_row,
            target_tenant=target_tenant,
            site_row=site_row,
            month_key=str(month).strip(),
            preview=preview,
            filename=file.filename or "support_roundtrip.xlsx",
            user=user,
        )
        return SupportRoundtripPreviewOut(
            batch_id=batch_id,
            total_rows=int(preview.get("total_rows") or 0),
            valid_rows=int(preview.get("valid_rows") or 0),
            invalid_rows=int(preview.get("invalid_rows") or 0),
            preview_rows=[SupportRoundtripPreviewRowOut(**row) for row in (preview.get("preview_rows") or [])],
            diff_counts=dict(preview.get("diff_counts") or {}),
            blocked_reasons=list(preview.get("blocked_reasons") or []),
            metadata=SupportRoundtripPreviewMetadataOut(**(preview.get("metadata") or {})),
            can_apply=bool(preview.get("can_apply")),
        )
    finally:
        workbook.close()


@router.post("/support-roundtrip/hq-upload/{batch_id}/apply", response_model=SupportRoundtripApplyOut)
def apply_support_roundtrip_upload(
    batch_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_hq(user):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT b.*, s.source_revision AS active_source_revision, s.id AS active_source_id,
                   s.hq_merge_available, s.hq_merge_stale, s.state AS source_state,
                   s.site_id AS active_site_id, s.site_code AS active_site_code, s.month_key AS active_month_key
            FROM schedule_support_roundtrip_batches b
            JOIN schedule_support_roundtrip_sources s ON s.id = b.source_id
            WHERE b.id = %s
              AND b.tenant_id = %s
            LIMIT 1
            """,
            (batch_id, target_tenant["id"]),
        )
        batch = cur.fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="support roundtrip batch not found")
        if str(batch.get("status") or "").strip() == "applied":
            raise HTTPException(status_code=409, detail="support roundtrip batch already applied")
        blocked_reasons = list(dict.fromkeys(str(item).strip() for item in (batch.get("blocked_reasons_json") or []) if str(item).strip()))
        if bool(batch.get("is_stale")) or str(batch.get("source_revision") or "").strip() != str(batch.get("active_source_revision") or "").strip():
            blocked_reasons.append("현재 HQ 업로드는 최신 Supervisor 소스 리비전 기준이 아니어서 적용할 수 없습니다.")
        if blocked_reasons:
            cur.execute(
                """
                UPDATE schedule_support_roundtrip_batches
                SET status = 'blocked'
                WHERE id = %s
                """,
                (batch_id,),
            )
            return SupportRoundtripApplyOut(
                batch_id=batch_id,
                applied=0,
                skipped=0,
                applied_rows=[],
                skipped_rows=[],
                blocked=True,
                blocked_reasons=blocked_reasons,
            )
        cur.execute(
            """
            SELECT payload_json
            FROM schedule_support_roundtrip_rows
            WHERE batch_id = %s
            ORDER BY row_no, support_period, slot_index
            """,
            (batch_id,),
        )
        payload_rows = [dict(row.get("payload_json") or {}) for row in (cur.fetchall() or [])]

        applied_rows: list[SupportRoundtripApplyRowOut] = []
        skipped_rows: list[SupportRoundtripApplyRowOut] = []
        applied = 0
        skipped = 0
        internal_conversion_count = 0
        conflict_count = 0
        ignored_blank_count = 0

        for payload in payload_rows:
            apply_action = str(payload.get("apply_action") or "none").strip()
            workbook_value = str(payload.get("workbook_value") or "").strip()
            validation_error = str(payload.get("validation_error") or payload.get("protected_reason") or "").strip()
            if not workbook_value:
                ignored_blank_count += 1
                continue
            if bool(payload.get("is_blocking")) or apply_action not in {"upsert_support_assignment", "upsert_need_count"}:
                skipped += 1
                if bool(payload.get("is_blocking")):
                    conflict_count += 1
                if len(skipped_rows) < IMPORT_REPORT_LIMIT:
                    skipped_rows.append(
                        SupportRoundtripApplyRowOut(
                            row_no=int(payload.get("row_no") or 0),
                            site_code=str(payload.get("site_code") or "").strip(),
                            schedule_date=str(payload.get("schedule_date") or ""),
                            support_period=str(payload.get("support_period") or "day"),
                            slot_index=int(payload.get("slot_index") or 0),
                            status="skipped",
                            reason=validation_error or "적용 제외",
                            worker_name=str(payload.get("resolved_worker_name") or "").strip() or None,
                            employee_name=str(payload.get("employee_name") or "").strip() or None,
                        )
                    )
                continue
            if apply_action == "upsert_need_count":
                required_count, parsed_text = _parse_daytime_need_value(payload.get("workbook_value"))
                if required_count is None:
                    skipped += 1
                    conflict_count += 1
                    continue
                _upsert_daytime_need_count_row(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    site_id=str(batch.get("active_site_id") or ""),
                    work_date=date.fromisoformat(str(payload.get("schedule_date"))),
                    required_count=max(0, int(required_count)),
                    raw_text=parsed_text or workbook_value or None,
                    updated_by=str(user["id"]),
                )
                applied += 1
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        SupportRoundtripApplyRowOut(
                            row_no=int(payload.get("row_no") or 0),
                            site_code=str(payload.get("site_code") or "").strip(),
                            schedule_date=str(payload.get("schedule_date") or ""),
                            support_period="day",
                            slot_index=0,
                            status="applied",
                            reason="필요인원 수 반영",
                        )
                    )
                continue
            _upsert_support_roundtrip_assignment_row(
                cur,
                source_row={
                    "id": batch["source_id"],
                    "source_revision": batch["source_revision"],
                    "tenant_id": target_tenant["id"],
                    "site_id": batch["active_site_id"],
                    "site_code": batch["active_site_code"],
                },
                payload=payload,
                batch_id=str(batch_id),
                user_id=str(user["id"]),
            )
            worker_type = str(payload.get("resolved_worker_type") or "").strip() or "F"
            worker_name = str(payload.get("resolved_worker_name") or "").strip()
            worker_affiliation = str(payload.get("resolved_affiliation") or "").strip()
            canonical_name = worker_name
            canonical_employee_id = payload.get("employee_id")
            if worker_type == "INTERNAL":
                canonical_name = str(payload.get("employee_name") or worker_name).strip()
            elif not canonical_name:
                canonical_name = workbook_value
            upsert_support_assignment(
                conn,
                tenant_id=str(target_tenant["id"]),
                site_id=str(batch.get("active_site_id") or ""),
                work_date=date.fromisoformat(str(payload.get("schedule_date"))),
                support_period=str(payload.get("support_period") or "day").strip() or "day",
                slot_index=int(payload.get("slot_index") or 1),
                worker_type=worker_type,
                name=canonical_name,
                source="HQ_ROUNDTRIP",
                employee_id=canonical_employee_id,
                affiliation=worker_affiliation or None,
                source_event_uid=str(batch_id),
            )
            applied += 1
            if str(payload.get("resolved_worker_type") or "").strip() == "INTERNAL":
                internal_conversion_count += 1
            if len(applied_rows) < IMPORT_REPORT_LIMIT:
                applied_rows.append(
                    SupportRoundtripApplyRowOut(
                        row_no=int(payload.get("row_no") or 0),
                        site_code=str(payload.get("site_code") or "").strip(),
                        schedule_date=str(payload.get("schedule_date") or ""),
                        support_period=str(payload.get("support_period") or "day"),
                        slot_index=int(payload.get("slot_index") or 0),
                        status="applied",
                        reason="지원근무 병합 반영",
                        worker_name=str(payload.get("resolved_worker_name") or "").strip() or None,
                        employee_name=str(payload.get("employee_name") or "").strip() or None,
                    )
                )

        latest_hq_revision = _build_support_roundtrip_revision(
            {
                "batch_id": str(batch_id),
                "source_revision": str(batch.get("source_revision") or "").strip(),
                "applied": applied,
                "skipped": skipped,
                "internal_conversion_count": internal_conversion_count,
                "conflict_count": conflict_count,
            }
        )
        merged_revision = _build_support_roundtrip_revision(
            {
                "source_revision": str(batch.get("source_revision") or "").strip(),
                "hq_revision": latest_hq_revision,
            }
        ) if applied > 0 and conflict_count == 0 else None
        source_state = "conflict_manual_review_required" if conflict_count > 0 else "hq_merge_available"
        final_download_enabled = applied > 0 and conflict_count == 0
        cur.execute(
            """
            UPDATE schedule_support_roundtrip_batches
            SET status = %s,
                applied_rows = %s,
                ignored_blank_count = %s,
                internal_conversion_count = %s,
                conflict_count = %s,
                completed_at = timezone('utc', now())
            WHERE id = %s
            """,
            (
                "applied" if final_download_enabled else "partial_applied",
                applied,
                ignored_blank_count,
                internal_conversion_count,
                conflict_count,
                batch_id,
            ),
        )
        cur.execute(
            """
            UPDATE schedule_support_roundtrip_sources
            SET state = %s,
                hq_merge_available = %s,
                hq_merge_stale = FALSE,
                conflict_required = %s,
                final_download_enabled = %s,
                latest_hq_batch_id = %s,
                latest_hq_revision = %s,
                latest_merged_revision = %s,
                updated_at = timezone('utc', now())
            WHERE id = %s
            """,
            (
                source_state,
                final_download_enabled,
                conflict_count > 0,
                final_download_enabled,
                batch_id,
                latest_hq_revision,
                merged_revision,
                batch["source_id"],
            ),
        )
    return SupportRoundtripApplyOut(
        batch_id=batch_id,
        applied=applied,
        skipped=skipped,
        applied_rows=applied_rows,
        skipped_rows=skipped_rows,
        blocked=False,
        blocked_reasons=[],
    )


@router.get("/support-roundtrip/final-excel")
def download_support_roundtrip_final_excel(
    month: str = Query(..., description="YYYY-MM"),
    site_code: str = Query(..., min_length=1, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_use_support_roundtrip_final_download(user):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_code).strip(),
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")
    source_row = _get_support_roundtrip_source(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month,
    )
    if not source_row:
        raise HTTPException(status_code=409, detail="support roundtrip source missing")
    if not bool(source_row.get("final_download_enabled")) or not bool(source_row.get("hq_merge_available")) or bool(source_row.get("hq_merge_stale")):
        raise HTTPException(status_code=409, detail="support roundtrip merged final not available")
    workbook = _build_support_roundtrip_final_workbook(
        conn,
        source_row=source_row,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=month,
        user=user,
    )
    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    now_kst = datetime.now(timezone(timedelta(hours=9)))
    filename = f"{month[:4]}년 {int(month[5:7])}월 근무표_{str(site_row['site_code']).strip()}_지원병합_{now_kst.strftime('%y%m%d')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/finance-submission/status", response_model=FinanceSubmissionStatusOut)
def get_finance_submission_status(
    month: str = Query(..., description="YYYY-MM"),
    site_code: str = Query(..., min_length=1, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_view_finance_submission(user):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_code).strip(),
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")
    return _build_finance_submission_status_payload(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=month,
    )


@router.get("/finance-submission/review-excel")
def download_finance_review_excel(
    month: str = Query(..., description="YYYY-MM"),
    site_code: str = Query(..., min_length=1, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_download_finance_review(user):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_code).strip(),
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")
    export_ctx = _collect_monthly_export_context(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=month,
        user=user,
    )
    workbook = export_ctx["workbook"]
    current_revision = str(export_ctx.get("export_revision") or "").strip()
    submission_row = _ensure_finance_submission_state(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        site_code=str(site_row["site_code"]),
        month_key=month,
    )
    filename = _build_finance_review_filename(
        month_key=month,
        site_code=str(site_row["site_code"]).strip(),
    )
    current_final_is_live = bool(submission_row.get("active_final_revision")) and str(submission_row.get("active_final_revision") or "").strip() == current_revision
    stale_final = bool(submission_row.get("active_final_revision")) and str(submission_row.get("active_final_revision") or "").strip() != current_revision
    _create_finance_submission_batch(
        conn,
        submission_row=submission_row,
        batch_kind="review_download",
        source_revision=current_revision,
        filename=filename,
        actor_id=str(user["id"]),
        actor_role=_finance_submission_normalize_role(user),
        status="downloaded",
        metadata={
            "template_version": ARLS_EXPORT_TEMPLATE_VERSION,
            "export_source_version": ARLS_EXPORT_SOURCE_VERSION,
            "export_revision": current_revision,
            "site_code": str(site_row["site_code"]).strip(),
            "month": month,
        },
    )
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE schedule_finance_submission_states
            SET state = %s,
                current_revision = %s,
                review_download_revision = %s,
                review_downloaded_at = timezone('utc', now()),
                review_downloaded_by = %s,
                review_downloaded_role = %s,
                review_download_filename = %s,
                final_upload_stale = %s,
                final_download_enabled = CASE WHEN %s THEN FALSE ELSE final_download_enabled END,
                last_event = 'review_downloaded',
                updated_at = timezone('utc', now())
            WHERE id = %s
            """,
            (
                "final_upload_stale" if stale_final else ("hq_final_download_ready" if current_final_is_live else "waiting_final_upload"),
                current_revision,
                current_revision,
                user["id"],
                _finance_submission_normalize_role(user),
                filename,
                stale_final,
                stale_final,
                submission_row["id"],
            ),
        )
    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/finance-submission/final-upload/preview", response_model=FinanceSubmissionPreviewOut)
def preview_finance_final_upload(
    file: UploadFile,
    site_code: str | None = Form(default=None),
    month: str | None = Form(default=None),
    tenant_code: str | None = Form(default=None),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_upload_finance_final(user):
        raise HTTPException(status_code=403, detail="forbidden")
    if not site_code or not str(site_code).strip():
        raise HTTPException(status_code=400, detail="site_code is required")
    if not month or not str(month).strip():
        raise HTTPException(status_code=400, detail="month is required")
    _month_bounds(str(month).strip())
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_code).strip(),
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")
    submission_row, current_revision = _sync_finance_submission_state(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=str(month).strip(),
    )
    review_download_revision = str(submission_row.get("review_download_revision") or "").strip()
    if not review_download_revision:
        raise HTTPException(status_code=409, detail="finance review download missing")
    raw_bytes = file.file.read()
    file.file = BytesIO(raw_bytes)
    preview_result = preview_import(
        file=file,
        site_code=str(site_code).strip(),
        month=str(month).strip(),
        tenant_code=tenant_code,
        conn=conn,
        user=user,
    )
    blocked_reasons = list(preview_result.blocked_reasons or [])
    diff_counts = dict(preview_result.diff_counts or {})
    preview_rows = list(preview_result.preview_rows or [])
    ignored_protected_count = int(diff_counts.get("ignored_protected") or 0)
    if ignored_protected_count > 0:
        blocked_reasons.append(f"보호 영역 변경 {ignored_protected_count}건은 Finance 최종 업로드로 반영할 수 없습니다.")
    blocking_row_count = sum(1 for row in preview_rows if bool(getattr(row, "is_blocking", False)))
    if blocking_row_count > 0:
        blocked_reasons.append(f"충돌/차단 행 {blocking_row_count}건을 먼저 해결해야 합니다.")
    metadata = preview_result.metadata
    if metadata and bool(metadata.is_stale):
        blocked_reasons.append("업로드한 파일이 현재 ARLS 리비전보다 오래되어 최종 업로드를 진행할 수 없습니다.")
    if review_download_revision != current_revision:
        blocked_reasons.append("현재 HQ 1차 확인본 기준 리비전이 최신이 아닙니다. HQ에서 다시 다운로드한 뒤 최종 업로드를 진행해야 합니다.")
    finance_batch = _create_finance_submission_batch(
        conn,
        submission_row=submission_row,
        batch_kind="final_upload",
        source_revision=str((metadata or {}).export_revision or current_revision),
        filename=file.filename or _build_finance_final_filename(month_key=str(month).strip(), site_code=str(site_row["site_code"]).strip()),
        actor_id=str(user["id"]),
        actor_role=_finance_submission_normalize_role(user),
        artifact_bytes=raw_bytes,
        import_batch_id=str(preview_result.batch_id),
        status="blocked" if blocked_reasons else "previewed",
        is_stale=bool((metadata or {}).is_stale) or review_download_revision != current_revision,
        total_rows=int(preview_result.total_rows or 0),
        valid_rows=int(preview_result.valid_rows or 0),
        invalid_rows=int(preview_result.invalid_rows or 0),
        blocked_reasons=blocked_reasons,
        diff_counts=diff_counts,
        metadata=(metadata.model_dump() if metadata else {}),
    )
    return FinanceSubmissionPreviewOut(
        finance_batch_id=finance_batch["id"],
        import_batch_id=preview_result.batch_id,
        total_rows=preview_result.total_rows,
        valid_rows=preview_result.valid_rows,
        invalid_rows=preview_result.invalid_rows,
        invalid_samples=list(preview_result.invalid_samples or []),
        preview_rows=preview_rows,
        error_counts=dict(preview_result.error_counts or {}),
        diff_counts=diff_counts,
        blocked_reasons=blocked_reasons,
        metadata=metadata,
        can_apply=bool(preview_result.can_apply) and not blocked_reasons,
    )


@router.post("/finance-submission/final-upload/{finance_batch_id}/apply", response_model=ImportApplyOut)
def apply_finance_final_upload(
    finance_batch_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_upload_finance_final(user):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    finance_batch = _get_finance_submission_batch(
        conn,
        batch_id=str(finance_batch_id),
        tenant_id=str(target_tenant["id"]),
    )
    if not finance_batch:
        raise HTTPException(status_code=404, detail="finance submission batch not found")
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(finance_batch.get("site_code") or "").strip(),
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")
    submission_row, current_revision = _sync_finance_submission_state(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=str(finance_batch.get("month_key") or "").strip(),
    )
    blocked_reasons = list(dict.fromkeys(
        str(item).strip()
        for item in (finance_batch.get("blocked_reasons_json") or [])
        if str(item).strip()
    ))
    if str(finance_batch.get("source_revision") or "").strip() != current_revision:
        blocked_reasons.append("최종 업로드 파일이 최신 assembled revision 기준이 아닙니다.")
    if blocked_reasons:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE schedule_finance_submission_batches
                SET status = 'blocked',
                    is_stale = TRUE,
                    blocked_reasons_json = %s::jsonb,
                    completed_at = timezone('utc', now())
                WHERE id = %s
                """,
                (json.dumps(blocked_reasons, ensure_ascii=False), finance_batch_id),
            )
            cur.execute(
                """
                UPDATE schedule_finance_submission_states
                SET state = 'final_upload_stale',
                    final_upload_stale = TRUE,
                    final_download_enabled = FALSE,
                    last_event = 'final_upload_stale',
                    updated_at = timezone('utc', now())
                WHERE id = %s
                """,
                (submission_row["id"],),
            )
        return ImportApplyOut(
            batch_id=uuid.UUID(str(finance_batch.get("import_batch_id"))),
            applied=0,
            skipped=0,
            applied_rows=[],
            skipped_rows=[],
            blocked=True,
            blocked_reasons=blocked_reasons,
        )
    import_batch_id = str(finance_batch.get("import_batch_id") or "").strip()
    if not import_batch_id:
        raise HTTPException(status_code=409, detail="finance import batch missing")
    result = apply_import(
        batch_id=uuid.UUID(import_batch_id),
        tenant_code=tenant_code,
        conn=conn,
        user=user,
    )
    if result.blocked:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE schedule_finance_submission_batches
                SET status = 'blocked',
                    blocked_reasons_json = %s::jsonb,
                    completed_at = timezone('utc', now())
                WHERE id = %s
                """,
                (json.dumps(result.blocked_reasons or [], ensure_ascii=False), finance_batch_id),
            )
            cur.execute(
                """
                UPDATE schedule_finance_submission_states
                SET state = 'conflict_manual_review_required',
                    conflict_required = TRUE,
                    final_download_enabled = FALSE,
                    last_event = 'conflict_manual_review_required',
                    updated_at = timezone('utc', now())
                WHERE id = %s
                """,
                (submission_row["id"],),
            )
        return result
    final_revision = _build_schedule_export_revision(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        site_code=str(site_row["site_code"]),
        month_key=str(finance_batch.get("month_key") or "").strip(),
    )
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE schedule_finance_submission_batches
            SET status = 'applied',
                final_revision = %s,
                applied_rows = %s,
                skipped_rows = %s,
                conflict_count = %s,
                completed_at = timezone('utc', now())
            WHERE id = %s
            """,
            (
                final_revision,
                int(result.applied or 0),
                int(result.skipped or 0),
                len(result.blocked_reasons or []),
                finance_batch_id,
            ),
        )
        cur.execute(
            """
            UPDATE schedule_finance_submission_states
            SET state = 'hq_final_download_ready',
                current_revision = %s,
                active_final_batch_id = %s,
                active_final_revision = %s,
                active_final_source_revision = %s,
                active_final_filename = %s,
                final_uploaded_at = timezone('utc', now()),
                final_uploaded_by = %s,
                final_uploaded_role = %s,
                final_download_enabled = TRUE,
                final_upload_stale = FALSE,
                conflict_required = FALSE,
                last_event = 'final_uploaded',
                updated_at = timezone('utc', now())
            WHERE id = %s
            """,
            (
                final_revision,
                finance_batch_id,
                final_revision,
                str(finance_batch.get("source_revision") or "").strip(),
                str(finance_batch.get("filename") or "").strip() or None,
                user["id"],
                _finance_submission_normalize_role(user),
                submission_row["id"],
            ),
        )
    return result


@router.get("/finance-submission/final-excel")
def download_finance_final_excel(
    month: str = Query(..., description="YYYY-MM"),
    site_code: str = Query(..., min_length=1, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not _can_download_finance_final(user):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=str(site_code).strip(),
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")
    status_payload = _build_finance_submission_status_payload(
        conn,
        target_tenant=target_tenant,
        site_row=site_row,
        month_key=month,
    )
    if not status_payload.final_download_enabled:
        raise HTTPException(status_code=409, detail="finance final workbook not available")
    submission_row = _get_finance_submission_state(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month,
    )
    active_batch_id = str((submission_row or {}).get("active_final_batch_id") or "").strip()
    if not active_batch_id:
        raise HTTPException(status_code=409, detail="finance final artifact missing")
    finance_batch = _get_finance_submission_batch(
        conn,
        batch_id=active_batch_id,
        tenant_id=str(target_tenant["id"]),
    )
    if not finance_batch or finance_batch.get("artifact_bytes") is None:
        raise HTTPException(status_code=409, detail="finance final artifact missing")
    filename = str(finance_batch.get("filename") or "").strip() or _build_finance_final_filename(
        month_key=month,
        site_code=str(site_row["site_code"]).strip(),
    )
    return StreamingResponse(
        BytesIO(bytes(finance_batch["artifact_bytes"])),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )

@router.get("")
def monthly_view_alias(
    month: str = Query(..., description="YYYY-MM"),
    tenant_code: str | None = None,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    # Legacy alias for older clients still calling /api/v1/schedules?month=YYYY-MM.
    return monthly_view(month=month, tenant_code=tenant_code, conn=conn, user=user)


@router.get("/work-templates", response_model=list[ScheduleTemplateOut])
def list_schedule_work_templates(
    tenant_code: str | None = Query(default=None, max_length=64),
    site_code: str | None = Query(default=None, max_length=64),
    include_inactive: bool = Query(default=False),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    scope_site_id = ""
    if site_code:
        site_row = _resolve_site_context_by_code(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_code=str(site_code).strip(),
        )
        if not site_row:
            raise HTTPException(status_code=404, detail="site not found")
        scope_site_id = str(site_row["id"])
    rows = _fetch_schedule_templates(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=scope_site_id or None,
        include_inactive=include_inactive,
    )
    return [_format_schedule_template_row(row) for row in rows]


@router.post("/work-templates", response_model=ScheduleTemplateOut)
def create_schedule_work_template(
    payload: ScheduleTemplateCreate,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    resolved_site_id = str(payload.site_id) if payload.site_id else ""
    if resolved_site_id:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id
                FROM sites
                WHERE tenant_id = %s
                  AND id = %s
                LIMIT 1
                """,
                (target_tenant["id"], resolved_site_id),
            )
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="site not found")

    template_id = uuid.uuid4()
    with conn.cursor() as cur:
        if payload.is_default:
            cur.execute(
                """
                UPDATE schedule_templates
                SET is_default = FALSE,
                    updated_at = timezone('utc', now())
                WHERE tenant_id = %s
                  AND duty_type = %s
                  AND COALESCE(site_id, '00000000-0000-0000-0000-000000000000'::uuid)
                      = COALESCE(%s::uuid, '00000000-0000-0000-0000-000000000000'::uuid)
                """,
                (target_tenant["id"], payload.duty_type, resolved_site_id or None),
            )
        cur.execute(
            """
            INSERT INTO schedule_templates (
                id,
                tenant_id,
                template_name,
                duty_type,
                start_time,
                end_time,
                paid_hours,
                break_minutes,
                site_id,
                is_default,
                is_active,
                created_by,
                created_at,
                updated_at
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                timezone('utc', now()), timezone('utc', now())
            )
            """,
            (
                template_id,
                target_tenant["id"],
                payload.template_name,
                payload.duty_type,
                payload.start_time,
                payload.end_time,
                payload.paid_hours,
                payload.break_minutes,
                resolved_site_id or None,
                payload.is_default,
                payload.is_active,
                user.get("id"),
            ),
        )
        cur.execute(
            """
            SELECT st.id,
                   st.tenant_id,
                   st.template_name,
                   st.duty_type,
                   st.start_time,
                   st.end_time,
                   st.paid_hours,
                   st.break_minutes,
                   st.site_id,
                   s.site_code,
                   s.site_name,
                   st.is_default,
                   st.is_active,
                   st.created_at,
                   st.updated_at
            FROM schedule_templates st
            LEFT JOIN sites s ON s.id = st.site_id
            WHERE st.id = %s
              AND st.tenant_id = %s
            LIMIT 1
            """,
            (template_id, target_tenant["id"]),
        )
        inserted = cur.fetchone()
    if not inserted:
        raise HTTPException(status_code=500, detail="template create failed")
    return _format_schedule_template_row(dict(inserted))


@router.put("/work-templates/{template_id}", response_model=ScheduleTemplateOut)
def update_schedule_work_template(
    template_id: uuid.UUID,
    payload: ScheduleTemplateUpdate,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, tenant_id, template_name, duty_type, start_time, end_time, paid_hours,
                   break_minutes, site_id, is_default, is_active
            FROM schedule_templates
            WHERE id = %s
              AND tenant_id = %s
            LIMIT 1
            """,
            (str(template_id), target_tenant["id"]),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="template not found")
        current = dict(row)

        update_data = payload.model_dump(exclude_unset=True)
        next_template_name = update_data.get("template_name", current.get("template_name"))
        next_duty_type = update_data.get("duty_type", current.get("duty_type"))
        next_start_time = update_data.get("start_time", _normalize_time_text(current.get("start_time")))
        next_end_time = update_data.get("end_time", _normalize_time_text(current.get("end_time")))
        next_paid_hours = update_data.get("paid_hours", current.get("paid_hours"))
        next_break_minutes = update_data.get("break_minutes", current.get("break_minutes"))
        if "site_id" in update_data:
            next_site_id = str(update_data.get("site_id")) if update_data.get("site_id") else None
        else:
            next_site_id = str(current.get("site_id")) if current.get("site_id") else None
        next_is_default = bool(update_data.get("is_default", current.get("is_default")))
        next_is_active = bool(update_data.get("is_active", current.get("is_active")))

        if next_site_id:
            cur.execute(
                """
                SELECT 1
                FROM sites
                WHERE id = %s
                  AND tenant_id = %s
                LIMIT 1
                """,
                (next_site_id, target_tenant["id"]),
            )
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="site not found")

        if next_is_default:
            cur.execute(
                """
                UPDATE schedule_templates
                SET is_default = FALSE,
                    updated_at = timezone('utc', now())
                WHERE tenant_id = %s
                  AND id <> %s
                  AND duty_type = %s
                  AND COALESCE(site_id, '00000000-0000-0000-0000-000000000000'::uuid)
                      = COALESCE(%s::uuid, '00000000-0000-0000-0000-000000000000'::uuid)
                """,
                (target_tenant["id"], str(template_id), next_duty_type, next_site_id),
            )

        cur.execute(
            """
            UPDATE schedule_templates
            SET template_name = %s,
                duty_type = %s,
                start_time = %s,
                end_time = %s,
                paid_hours = %s,
                break_minutes = %s,
                site_id = %s,
                is_default = %s,
                is_active = %s,
                updated_at = timezone('utc', now())
            WHERE id = %s
              AND tenant_id = %s
            """,
            (
                next_template_name,
                next_duty_type,
                next_start_time,
                next_end_time,
                next_paid_hours,
                next_break_minutes,
                next_site_id,
                next_is_default,
                next_is_active,
                str(template_id),
                target_tenant["id"],
            ),
        )
        cur.execute(
            """
            SELECT st.id,
                   st.tenant_id,
                   st.template_name,
                   st.duty_type,
                   st.start_time,
                   st.end_time,
                   st.paid_hours,
                   st.break_minutes,
                   st.site_id,
                   s.site_code,
                   s.site_name,
                   st.is_default,
                   st.is_active,
                   st.created_at,
                   st.updated_at
            FROM schedule_templates st
            LEFT JOIN sites s ON s.id = st.site_id
            WHERE st.id = %s
              AND st.tenant_id = %s
            LIMIT 1
            """,
            (str(template_id), target_tenant["id"]),
        )
        updated = cur.fetchone()
    if not updated:
        raise HTTPException(status_code=500, detail="template update failed")
    return _format_schedule_template_row(dict(updated))


def _create_schedule_rows_with_template(
    conn,
    *,
    target_tenant: dict,
    site_code: str,
    employee_code: str,
    template_id: str,
    schedule_dates: list[date],
) -> dict:
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=site_code,
    )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")

    employee_row = _resolve_employee_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        employee_code=employee_code,
    )
    if not employee_row:
        raise HTTPException(status_code=404, detail="employee not found")
    if str(employee_row.get("site_id") or "") != str(site_row.get("id") or ""):
        raise HTTPException(status_code=400, detail="employee is not assigned to the target site")

    template_row = _fetch_template_by_id_for_scope(
        conn,
        tenant_id=str(target_tenant["id"]),
        template_id=template_id,
    )
    if not template_row or not bool(template_row.get("is_active")):
        raise HTTPException(status_code=404, detail="template not found")
    template_site_id = str(template_row.get("site_id") or "").strip()
    if template_site_id and template_site_id != str(site_row.get("id")):
        raise HTTPException(status_code=400, detail="template scope does not match site")

    shift_type = _resolve_shift_type_from_duty_type(template_row.get("duty_type"))
    canonical_template = _resolve_canonical_schedule_time(
        {
            **dict(template_row),
            "shift_type": shift_type,
            "duty_type": _normalize_schedule_template_duty_type(template_row.get("duty_type")),
            "template_start_time": template_row.get("start_time"),
            "template_end_time": template_row.get("end_time"),
            "template_paid_hours": template_row.get("paid_hours"),
        }
    )
    shift_start_time = _normalize_time_text(canonical_template.get("start_time"))
    shift_end_time = _normalize_time_text(canonical_template.get("end_time"))
    paid_hours = _coerce_float_or_none(canonical_template.get("hours"))
    if paid_hours is None and template_row.get("paid_hours") is not None:
        paid_hours = float(template_row["paid_hours"])

    created = 0
    skipped = 0
    skipped_dates: list[str] = []
    created_rows: list[dict[str, Any]] = []
    affected_site_days: set[tuple[str, str, str]] = set()
    dedup_dates = sorted(set(schedule_dates))

    with conn.cursor() as cur:
        existing_rows_by_date = _load_existing_schedule_rows_for_dates(
            cur,
            tenant_id=str(target_tenant["id"]),
            employee_id=str(employee_row["id"]),
            schedule_dates=dedup_dates,
        )
        for schedule_date in dedup_dates:
            existing_rows = existing_rows_by_date.get(schedule_date.isoformat(), [])
            has_conflict = False
            if existing_rows:
                if not shift_start_time or not shift_end_time:
                    has_conflict = True
                else:
                    has_conflict = any(
                        _schedule_time_ranges_overlap(
                            shift_start_time,
                            shift_end_time,
                            row.get("shift_start_time"),
                            row.get("shift_end_time"),
                        )
                        for row in existing_rows
                    )
            if has_conflict:
                skipped += 1
                skipped_dates.append(schedule_date.isoformat())
                continue

            schedule_id = _insert_monthly_schedule_row(
                cur,
                tenant_id=str(target_tenant["id"]),
                company_id=str(site_row["company_id"]),
                site_id=str(site_row["id"]),
                employee_id=str(employee_row["id"]),
                schedule_date=schedule_date,
                shift_type=shift_type,
                template_id=str(template_row["id"]),
                shift_start_time=shift_start_time,
                shift_end_time=shift_end_time,
                paid_hours=paid_hours,
            )
            created += 1
            created_rows.append(
                {
                    "id": schedule_id,
                    "tenant_code": str(target_tenant.get("tenant_code") or "").strip().lower(),
                    "company_code": str(site_row.get("company_code") or "").strip(),
                    "site_code": str(site_row.get("site_code") or "").strip().upper(),
                    "site_name": str(site_row.get("site_name") or "").strip(),
                    "employee_id": str(employee_row.get("id") or "").strip(),
                    "employee_code": str(employee_row.get("employee_code") or "").strip(),
                    "employee_name": str(employee_row.get("full_name") or "").strip(),
                    "schedule_date": schedule_date.isoformat(),
                    "shift_type": shift_type,
                    "duty_type": _normalize_schedule_template_duty_type(template_row.get("duty_type")),
                    "template_id": str(template_row.get("id") or "").strip() or None,
                    "template_name": str(template_row.get("template_name") or "").strip() or None,
                    "shift_start_time": shift_start_time,
                    "shift_end_time": shift_end_time,
                    "paid_hours": paid_hours,
                }
            )
            affected_site_days.add((str(target_tenant["id"]), str(site_row["id"]), schedule_date.isoformat()))
            existing_rows_by_date.setdefault(schedule_date.isoformat(), []).append(
                {
                    "schedule_date": schedule_date,
                    "shift_start_time": shift_start_time,
                    "shift_end_time": shift_end_time,
                }
            )

    affected_dates_by_site: dict[tuple[str, str], list[date]] = {}
    for tenant_id, affected_site_id, schedule_date_raw in affected_site_days:
        affected_dates_by_site.setdefault((tenant_id, affected_site_id), []).append(date.fromisoformat(schedule_date_raw))
    for (affected_tenant_id, affected_site_id), affected_dates in affected_dates_by_site.items():
        try:
            _refresh_daily_leader_defaults_for_dates(
                conn,
                tenant_id=affected_tenant_id,
                site_id=affected_site_id,
                schedule_dates=affected_dates,
            )
        except Exception:
            logger.exception(
                "[schedule][template-assign] leader refresh failed tenant=%s site=%s dates=%s",
                affected_tenant_id,
                affected_site_id,
                [item.isoformat() for item in affected_dates],
            )

    return {
        "template_id": str(template_row["id"]),
        "template_name": str(template_row.get("template_name") or "").strip(),
        "employee_code": str(employee_row.get("employee_code") or "").strip(),
        "site_code": str(site_row.get("site_code") or "").strip(),
        "created": created,
        "created_rows": created_rows,
        "skipped": skipped,
        "skipped_dates": skipped_dates,
    }


@router.post("/template-assign/single")
def create_single_schedule_by_template(
    payload: ScheduleTemplateSingleCreateIn,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, payload.tenant_code)
    return _create_schedule_rows_with_template(
        conn,
        target_tenant=target_tenant,
        site_code=str(payload.site_code or "").strip(),
        employee_code=str(payload.employee_code or "").strip(),
        template_id=str(payload.template_id),
        schedule_dates=[payload.schedule_date],
    )


@router.post("/template-assign/bulk")
def create_bulk_schedule_by_template(
    payload: ScheduleTemplateBulkCreateIn,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, payload.tenant_code)
    return _create_schedule_rows_with_template(
        conn,
        target_tenant=target_tenant,
        site_code=str(payload.site_code or "").strip(),
        employee_code=str(payload.employee_code or "").strip(),
        template_id=str(payload.template_id),
        schedule_dates=list(payload.schedule_dates),
    )


@router.post("/bulk", response_model=ScheduleBulkCreateOut)
def create_bulk_schedule(
    payload: ScheduleBulkCreateIn,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, payload.tenant_code)
    tenant_id = str(target_tenant["id"])

    site_row = None
    if payload.site_id:
        site_row = _resolve_site_context_by_id(conn, tenant_id=tenant_id, site_id=str(payload.site_id))
    if not site_row and payload.site_code:
        site_row = _resolve_site_context_by_code(
            conn,
            tenant_id=tenant_id,
            site_code=str(payload.site_code or "").strip(),
        )
    if not site_row:
        raise HTTPException(status_code=404, detail="site not found")

    employee_row = None
    if payload.employee_id:
        employee_row = _resolve_employee_by_id(conn, tenant_id=tenant_id, employee_id=str(payload.employee_id))
    if not employee_row and payload.employee_code:
        employee_row = _resolve_employee_by_code(
            conn,
            tenant_id=tenant_id,
            employee_code=str(payload.employee_code or "").strip(),
        )
    if not employee_row:
        raise HTTPException(status_code=404, detail="employee not found")
    if str(employee_row.get("site_id") or "") != str(site_row.get("id") or ""):
        raise HTTPException(status_code=400, detail="employee is not assigned to the target site")

    if payload.template_id:
        result = _create_schedule_rows_with_template(
            conn,
            target_tenant=target_tenant,
            site_code=str(site_row.get("site_code") or "").strip(),
            employee_code=str(employee_row.get("employee_code") or "").strip(),
            template_id=str(payload.template_id),
            schedule_dates=list(payload.dates),
        )
        return {
            "created_count": int(result.get("created") or 0),
            "skipped_duplicates": int(result.get("skipped") or 0),
            "errors": [],
            "created_rows": list(result.get("created_rows") or []),
        }

    normalized_shift_type = _normalize_shift_type(str(payload.shift_type or ""))
    if not normalized_shift_type:
        raise HTTPException(status_code=400, detail="shift_type is required")

    normalized_shift_start_time = _normalize_time_text(payload.shift_start_time)
    normalized_shift_end_time = _normalize_time_text(payload.shift_end_time)
    if payload.shift_start_time and not normalized_shift_start_time:
        raise HTTPException(status_code=400, detail="shift_start_time invalid")
    if payload.shift_end_time and not normalized_shift_end_time:
        raise HTTPException(status_code=400, detail="shift_end_time invalid")

    schedule_note = str(payload.schedule_note or "").strip() or None
    created = 0
    skipped = 0
    created_rows: list[dict[str, Any]] = []
    affected_site_days: set[tuple[str, str, str]] = set()
    dedup_dates = sorted(set(payload.dates))

    with conn.cursor() as cur:
        existing_rows_by_date = _load_existing_schedule_rows_for_dates(
            cur,
            tenant_id=tenant_id,
            employee_id=str(employee_row["id"]),
            schedule_dates=dedup_dates,
        )
        for schedule_date in dedup_dates:
            existing_rows = existing_rows_by_date.get(schedule_date.isoformat(), [])
            has_conflict = False
            if existing_rows:
                if not normalized_shift_start_time or not normalized_shift_end_time:
                    has_conflict = True
                else:
                    has_conflict = any(
                        _schedule_time_ranges_overlap(
                            normalized_shift_start_time,
                            normalized_shift_end_time,
                            row.get("shift_start_time"),
                            row.get("shift_end_time"),
                        )
                        for row in existing_rows
                    )
            if has_conflict:
                skipped += 1
                continue

            schedule_id = _insert_monthly_schedule_row(
                cur,
                tenant_id=tenant_id,
                company_id=str(site_row["company_id"]),
                site_id=str(site_row["id"]),
                employee_id=str(employee_row["id"]),
                schedule_date=schedule_date,
                shift_type=normalized_shift_type,
                template_id=None,
                shift_start_time=normalized_shift_start_time,
                shift_end_time=normalized_shift_end_time,
                paid_hours=None,
                schedule_note=schedule_note,
            )
            created += 1
            created_rows.append(
                {
                    "id": schedule_id,
                    "tenant_code": str(target_tenant.get("tenant_code") or "").strip().lower(),
                    "company_code": str(site_row.get("company_code") or "").strip(),
                    "site_code": str(site_row.get("site_code") or "").strip().upper(),
                    "site_name": str(site_row.get("site_name") or "").strip(),
                    "employee_id": str(employee_row.get("id") or "").strip(),
                    "employee_code": str(employee_row.get("employee_code") or "").strip(),
                    "employee_name": str(employee_row.get("full_name") or "").strip(),
                    "schedule_date": schedule_date.isoformat(),
                    "shift_type": normalized_shift_type,
                    "duty_type": "night" if normalized_shift_type == "night" else ("overtime" if normalized_shift_type == "overtime" else "day"),
                    "template_id": None,
                    "template_name": None,
                    "shift_start_time": normalized_shift_start_time,
                    "shift_end_time": normalized_shift_end_time,
                    "paid_hours": _infer_canonical_shift_hours(normalized_shift_start_time, normalized_shift_end_time),
                    "schedule_note": schedule_note,
                }
            )
            affected_site_days.add((tenant_id, str(site_row["id"]), schedule_date.isoformat()))
            existing_rows_by_date.setdefault(schedule_date.isoformat(), []).append(
                {
                    "schedule_date": schedule_date,
                    "shift_start_time": normalized_shift_start_time,
                    "shift_end_time": normalized_shift_end_time,
                }
            )

    affected_dates_by_site: dict[tuple[str, str], list[date]] = {}
    for affected_tenant_id, affected_site_id, schedule_date_raw in affected_site_days:
        affected_dates_by_site.setdefault((affected_tenant_id, affected_site_id), []).append(date.fromisoformat(schedule_date_raw))
    for (affected_tenant_id, affected_site_id), affected_dates in affected_dates_by_site.items():
        try:
            _refresh_daily_leader_defaults_for_dates(
                conn,
                tenant_id=affected_tenant_id,
                site_id=affected_site_id,
                schedule_dates=affected_dates,
            )
        except Exception:
            logger.exception(
                "[schedule][bulk-create] leader refresh failed tenant=%s site=%s dates=%s",
                affected_tenant_id,
                affected_site_id,
                [item.isoformat() for item in affected_dates],
            )

    return {
        "created_count": created,
        "skipped_duplicates": skipped,
        "errors": [],
        "created_rows": created_rows,
    }


def _normalize_duty_role(value: str | None, user_role: str | None = None) -> str:
    raw = str(value or "").strip().upper()
    if raw in {VICE_SUPERVISOR_DUTY_ROLE, GUARD_DUTY_ROLE, TEAM_MANAGER_DUTY_ROLE}:
        return raw
    if str(user_role or "").strip().lower() == "branch_manager":
        return VICE_SUPERVISOR_DUTY_ROLE
    return GUARD_DUTY_ROLE


def _fetch_schedule_context(conn, schedule_id: uuid.UUID | str, tenant_id):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT ms.id, ms.tenant_id, ms.company_id, ms.site_id, ms.employee_id, ms.schedule_date, ms.shift_type,
                   ms.leader_user_id, s.site_code, e.employee_code, e.full_name AS employee_name
            FROM monthly_schedules ms
            JOIN sites s ON s.id = ms.site_id
            JOIN employees e ON e.id = ms.employee_id
            WHERE ms.id = %s
              AND ms.tenant_id = %s
            LIMIT 1
            """,
            (str(schedule_id), tenant_id),
        )
        return cur.fetchone()


def _normalize_schedule_delete_scope(scope: str | None) -> str:
    value = str(scope or "single").strip().lower()
    if value not in {"single", "future"}:
        raise HTTPException(status_code=400, detail="invalid delete scope")
    return value


def _fetch_schedule_delete_scope_rows(
    conn,
    *,
    tenant_id,
    employee_id,
    schedule_date: date,
    schedule_id: uuid.UUID | str,
    scope: str,
) -> list[dict]:
    normalized_scope = _normalize_schedule_delete_scope(scope)
    with conn.cursor() as cur:
        if normalized_scope == "future":
            cur.execute(
                """
                SELECT ms.id, ms.site_id, ms.schedule_date
                FROM monthly_schedules ms
                WHERE ms.tenant_id = %s
                  AND ms.employee_id = %s
                  AND ms.schedule_date >= %s
                ORDER BY ms.schedule_date, ms.id
                """,
                (tenant_id, employee_id, schedule_date),
            )
        else:
            cur.execute(
                """
                SELECT ms.id, ms.site_id, ms.schedule_date
                FROM monthly_schedules ms
                WHERE ms.tenant_id = %s
                  AND ms.id = %s
                LIMIT 1
                """,
                (tenant_id, str(schedule_id)),
            )
        return [dict(row) for row in cur.fetchall()]


@router.get("/monthly/{schedule_id}/delete-scope")
def get_monthly_schedule_delete_scope_preview(
    schedule_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    context = _fetch_schedule_context(conn, schedule_id, target_tenant["id"])
    if not context:
        raise HTTPException(status_code=404, detail="schedule not found")

    future_rows = _fetch_schedule_delete_scope_rows(
        conn,
        tenant_id=context["tenant_id"],
        employee_id=context["employee_id"],
        schedule_date=context["schedule_date"],
        schedule_id=schedule_id,
        scope="future",
    )
    total_delete_count = len(future_rows)
    additional_future_count = max(total_delete_count - 1, 0)

    return {
        "schedule_id": str(context["id"]),
        "schedule_date": context["schedule_date"].isoformat() if isinstance(context.get("schedule_date"), date) else None,
        "employee_id": str(context["employee_id"]) if context.get("employee_id") else None,
        "employee_code": str(context.get("employee_code") or "").strip() or None,
        "employee_name": str(context.get("employee_name") or "").strip() or None,
        "site_code": str(context.get("site_code") or "").strip() or None,
        "additional_future_count": additional_future_count,
        "total_future_count": total_delete_count,
        "scope_rule": "same_employee_from_selected_date",
    }


def _fetch_leader_candidates_for_site_day(conn, *, tenant_id, site_id, schedule_date: date) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT au.id AS user_id,
                   au.username,
                   au.full_name,
                   e.employee_code,
                   COALESCE(e.duty_role, '') AS duty_role_raw,
                   COALESCE(au.role, '') AS user_role
            FROM monthly_schedules ms
            JOIN employees e ON e.id = ms.employee_id
            JOIN arls_users au ON au.tenant_id = ms.tenant_id
                              AND au.employee_id = ms.employee_id
                              AND au.is_active = TRUE
            WHERE ms.tenant_id = %s
              AND ms.site_id = %s
              AND ms.schedule_date = %s
              AND lower(ms.shift_type) NOT IN ('off', 'holiday')
            GROUP BY au.id, au.username, au.full_name, e.employee_code, e.duty_role, au.role
            ORDER BY e.employee_code, au.username
            """,
            (tenant_id, site_id, schedule_date),
        )
        rows = [dict(row) for row in cur.fetchall()]

    candidates: list[dict] = []
    for row in rows:
        duty_role = _normalize_duty_role(row.get("duty_role_raw"), row.get("user_role"))
        if duty_role == TEAM_MANAGER_DUTY_ROLE:
            continue
        candidates.append(
            {
                "user_id": row["user_id"],
                "username": row["username"],
                "full_name": row["full_name"],
                "employee_code": row["employee_code"],
                "duty_role": duty_role,
            }
        )

    def _priority(item: dict) -> tuple[int, str]:
        duty_role = item.get("duty_role")
        if duty_role == VICE_SUPERVISOR_DUTY_ROLE:
            return (1, str(item.get("employee_code") or ""))
        if duty_role == GUARD_DUTY_ROLE:
            return (2, str(item.get("employee_code") or ""))
        return (9, str(item.get("employee_code") or ""))

    candidates.sort(key=_priority)
    return candidates


def _fetch_leader_candidates_for_site_dates(
    conn,
    *,
    tenant_id,
    site_id,
    schedule_dates: list[date],
) -> dict[str, list[dict]]:
    normalized_dates = sorted({item for item in schedule_dates if isinstance(item, date)})
    if not normalized_dates:
        return {}
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT ms.schedule_date,
                   au.id AS user_id,
                   au.username,
                   au.full_name,
                   e.employee_code,
                   COALESCE(e.duty_role, '') AS duty_role_raw,
                   COALESCE(au.role, '') AS user_role
            FROM monthly_schedules ms
            JOIN employees e ON e.id = ms.employee_id
            JOIN arls_users au ON au.tenant_id = ms.tenant_id
                              AND au.employee_id = ms.employee_id
                              AND au.is_active = TRUE
            WHERE ms.tenant_id = %s
              AND ms.site_id = %s
              AND ms.schedule_date = ANY(%s::date[])
              AND lower(ms.shift_type) NOT IN ('off', 'holiday')
            GROUP BY ms.schedule_date, au.id, au.username, au.full_name, e.employee_code, e.duty_role, au.role
            ORDER BY ms.schedule_date, e.employee_code, au.username
            """,
            (tenant_id, site_id, normalized_dates),
        )
        rows = [dict(row) for row in cur.fetchall()]

    grouped: dict[str, list[dict]] = {item.isoformat(): [] for item in normalized_dates}
    for row in rows:
        schedule_date_raw = row.get("schedule_date")
        schedule_date_key = schedule_date_raw.isoformat() if isinstance(schedule_date_raw, date) else str(schedule_date_raw or "").strip()
        if not schedule_date_key:
            continue
        duty_role = _normalize_duty_role(row.get("duty_role_raw"), row.get("user_role"))
        if duty_role == TEAM_MANAGER_DUTY_ROLE:
            continue
        grouped.setdefault(schedule_date_key, []).append(
            {
                "user_id": row["user_id"],
                "username": row["username"],
                "full_name": row["full_name"],
                "employee_code": row["employee_code"],
                "duty_role": duty_role,
            }
        )

    def _priority(item: dict) -> tuple[int, str]:
        duty_role = item.get("duty_role")
        if duty_role == VICE_SUPERVISOR_DUTY_ROLE:
            return (1, str(item.get("employee_code") or ""))
        if duty_role == GUARD_DUTY_ROLE:
            return (2, str(item.get("employee_code") or ""))
        return (9, str(item.get("employee_code") or ""))

    for schedule_date_key, candidates in grouped.items():
        candidates.sort(key=_priority)
    return grouped


def _recommended_leader_user_id(candidates: list[dict]) -> str | None:
    if not candidates:
        return None
    for duty_role in (VICE_SUPERVISOR_DUTY_ROLE, GUARD_DUTY_ROLE):
        for candidate in candidates:
            if str(candidate.get("duty_role") or "") == duty_role:
                return str(candidate["user_id"])
    return str(candidates[0]["user_id"])


def _refresh_daily_leader_defaults(conn, *, tenant_id, site_id, schedule_date: date) -> str | None:
    candidates = _fetch_leader_candidates_for_site_day(
        conn,
        tenant_id=tenant_id,
        site_id=site_id,
        schedule_date=schedule_date,
    )
    recommended_user_id = _recommended_leader_user_id(candidates)
    candidate_ids = [str(item["user_id"]) for item in candidates]

    with conn.cursor() as cur:
        if not recommended_user_id:
            # No eligible on-duty leader exists. Clear stale leader assignments for the day.
            cur.execute(
                """
                UPDATE monthly_schedules
                SET leader_user_id = NULL
                WHERE tenant_id = %s
                  AND site_id = %s
                  AND schedule_date = %s
                  AND lower(shift_type) NOT IN ('off', 'holiday')
                """,
                (tenant_id, site_id, schedule_date),
            )
            return None

        cur.execute(
            """
            UPDATE monthly_schedules
            SET leader_user_id = %s
            WHERE tenant_id = %s
              AND site_id = %s
              AND schedule_date = %s
              AND lower(shift_type) NOT IN ('off', 'holiday')
              AND (
                    leader_user_id IS NULL
                    OR NOT (leader_user_id = ANY(%s::uuid[]))
              )
            """,
            (recommended_user_id, tenant_id, site_id, schedule_date, candidate_ids),
        )
    return recommended_user_id


def _refresh_daily_leader_defaults_for_dates(
    conn,
    *,
    tenant_id,
    site_id,
    schedule_dates: list[date],
) -> dict[str, str | None]:
    normalized_dates = sorted({item for item in schedule_dates if isinstance(item, date)})
    if not normalized_dates:
        return {}
    candidates_by_date = _fetch_leader_candidates_for_site_dates(
        conn,
        tenant_id=tenant_id,
        site_id=site_id,
        schedule_dates=normalized_dates,
    )
    results: dict[str, str | None] = {}
    with conn.cursor() as cur:
        for schedule_date in normalized_dates:
            date_key = schedule_date.isoformat()
            candidates = candidates_by_date.get(date_key, [])
            recommended_user_id = _recommended_leader_user_id(candidates)
            candidate_ids = [str(item["user_id"]) for item in candidates]
            if not recommended_user_id:
                cur.execute(
                    """
                    UPDATE monthly_schedules
                    SET leader_user_id = NULL
                    WHERE tenant_id = %s
                      AND site_id = %s
                      AND schedule_date = %s
                      AND lower(shift_type) NOT IN ('off', 'holiday')
                    """,
                    (tenant_id, site_id, schedule_date),
                )
                results[date_key] = None
                continue
            cur.execute(
                """
                UPDATE monthly_schedules
                SET leader_user_id = %s
                WHERE tenant_id = %s
                  AND site_id = %s
                  AND schedule_date = %s
                  AND lower(shift_type) NOT IN ('off', 'holiday')
                  AND (
                        leader_user_id IS NULL
                        OR NOT (leader_user_id = ANY(%s::uuid[]))
                  )
                """,
                (recommended_user_id, tenant_id, site_id, schedule_date, candidate_ids),
            )
            results[date_key] = recommended_user_id
    return results


def _assert_valid_leader_for_site_day(
    conn,
    *,
    tenant_id,
    site_id,
    schedule_date: date,
    leader_user_id: str,
) -> None:
    candidates = _fetch_leader_candidates_for_site_day(
        conn,
        tenant_id=tenant_id,
        site_id=site_id,
        schedule_date=schedule_date,
    )
    candidate_ids = {str(item["user_id"]) for item in candidates}
    if str(leader_user_id) not in candidate_ids:
        raise HTTPException(
            status_code=400,
            detail="leader must be an on-duty VICE_SUPERVISOR/GUARD for this site/date",
        )


@router.get("/monthly")
def monthly_view(
    month: str = Query(..., description="YYYY-MM"),
    tenant_code: str | None = None,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    start, end = _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    staff_scope = enforce_staff_site_scope(user)
    staff_site_id = str((staff_scope or {}).get("site_id") or "").strip()
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT t.tenant_code,
                   ms.id,
                   ms.employee_id,
                   ms.site_id,
                   c.company_code,
                   s.site_code,
                   s.site_name,
                   e.employee_code,
                   e.full_name AS employee_name,
                   COALESCE(e.soc_role, '') AS soc_role,
                   COALESCE(e.duty_role, '') AS duty_role,
                   ms.schedule_date,
                   ms.shift_type,
                   ms.template_id,
                   ms.shift_start_time,
                   ms.shift_end_time,
                   ms.paid_hours,
                   ms.source,
                   ms.source_ticket_id,
                   ms.schedule_note,
                   ms.leader_user_id,
                   lu.username AS leader_username,
                   lu.full_name AS leader_full_name,
                   (lower(COALESCE(ms.schedule_note, '')) LIKE '%%closer%%') AS is_closer,
                   st.duty_type,
                   st.start_time AS template_start_time,
                   st.end_time AS template_end_time,
                   st.paid_hours AS template_paid_hours,
                   st.template_name
            FROM monthly_schedules ms
            JOIN tenants t ON t.id = ms.tenant_id
            JOIN companies c ON c.id = ms.company_id
            JOIN sites s ON s.id = ms.site_id
            JOIN employees e ON e.id = ms.employee_id
            LEFT JOIN arls_users lu ON lu.id = ms.leader_user_id
            LEFT JOIN schedule_templates st ON st.id = ms.template_id
            WHERE ms.tenant_id = %s
              AND ms.schedule_date >= %s
              AND ms.schedule_date < %s
              AND (%s = '' OR ms.site_id::text = %s)
            ORDER BY e.employee_code, ms.schedule_date
            """,
            (target_tenant["id"], start, end, staff_site_id, staff_site_id),
        )
        rows = [dict(r) for r in cur.fetchall()]

    for row in rows:
        canonical = _resolve_canonical_schedule_time(row)
        row["shift_start_time"] = canonical.get("start_time")
        row["shift_end_time"] = canonical.get("end_time")
        row["paid_hours"] = canonical.get("hours")
        row["shift_label"] = canonical.get("label")
        row["duty_type"] = canonical.get("duty_type")
    return rows


@router.get("/monthly-lite")
def monthly_view_lite(
    month: str = Query(..., description="YYYY-MM"),
    tenant_code: str | None = None,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    start, end = _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    staff_scope = enforce_staff_site_scope(user)
    staff_site_id = str((staff_scope or {}).get("site_id") or "").strip()
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
                ms.id,
                ms.employee_id,
                e.employee_code,
                e.full_name AS employee_name,
                ms.schedule_date,
                ms.shift_type,
                s.site_code,
                s.site_name,
                (lower(COALESCE(ms.schedule_note, '')) LIKE '%%closer%%') AS is_closer
            FROM monthly_schedules ms
            JOIN employees e ON e.id = ms.employee_id
            JOIN sites s ON s.id = ms.site_id
            WHERE ms.tenant_id = %s
              AND ms.schedule_date >= %s
              AND ms.schedule_date < %s
              AND (%s = '' OR ms.site_id::text = %s)
            ORDER BY e.employee_code, ms.schedule_date
            """,
            (target_tenant["id"], start, end, staff_site_id, staff_site_id),
        )
        rows = [dict(r) for r in cur.fetchall()]

    employee_map: dict[str, dict] = {}
    for row in rows:
        employee_id = str(row.get("employee_id") or "").strip()
        if not employee_id or employee_id in employee_map:
            continue
        employee_map[employee_id] = {
            "id": employee_id,
            "employee_code": str(row.get("employee_code") or "").strip(),
            "employee_name": str(row.get("employee_name") or "").strip(),
        }

    return {
        "month": month,
        "tenant_code": str(target_tenant.get("tenant_code") or "").strip(),
        "employees": list(employee_map.values()),
        "rows": rows,
    }


@router.get("/monthly-board-lite")
def monthly_board_lite(
    month: str = Query(..., description="YYYY-MM"),
    tenant_code: str | None = None,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    total_started_at = time.perf_counter()
    start, end = _month_bounds(month)
    grid_start, grid_end = _calendar_grid_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    staff_scope = enforce_staff_site_scope(user)
    staff_site_id = str((staff_scope or {}).get("site_id") or "").strip()

    db_started_at = time.perf_counter()
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
                ms.id AS schedule_id,
                ms.employee_id,
                e.employee_code,
                e.full_name AS employee_name,
                COALESCE(e.soc_role, '') AS soc_role,
                COALESCE(e.duty_role, '') AS duty_role,
                ms.schedule_date,
                ms.shift_type,
                ms.template_id,
                ms.shift_start_time,
                ms.shift_end_time,
                ms.paid_hours,
                s.site_code,
                s.site_name,
                st.template_name,
                st.duty_type,
                st.start_time AS template_start_time,
                st.end_time AS template_end_time,
                st.paid_hours AS template_paid_hours
            FROM monthly_schedules ms
            JOIN employees e ON e.id = ms.employee_id
            JOIN sites s ON s.id = ms.site_id
            LEFT JOIN schedule_templates st ON st.id = ms.template_id
            WHERE ms.tenant_id = %s
              AND ms.schedule_date >= %s
              AND ms.schedule_date < %s
              AND (%s = '' OR ms.site_id::text = %s)
            ORDER BY ms.schedule_date ASC, s.site_code ASC, e.employee_code ASC
            """,
            (target_tenant["id"], start, end, staff_site_id, staff_site_id),
        )
        rows = [dict(r) for r in cur.fetchall()]
        cur.execute(
            """
            SELECT
                ms.id AS schedule_id,
                ms.employee_id,
                e.employee_code,
                e.full_name AS employee_name,
                COALESCE(e.soc_role, '') AS soc_role,
                COALESCE(e.duty_role, '') AS duty_role,
                ms.schedule_date,
                ms.shift_type,
                ms.template_id,
                ms.shift_start_time,
                ms.shift_end_time,
                ms.paid_hours,
                s.site_code,
                s.site_name,
                st.template_name,
                st.duty_type,
                st.start_time AS template_start_time,
                st.end_time AS template_end_time,
                st.paid_hours AS template_paid_hours
            FROM monthly_schedules ms
            JOIN employees e ON e.id = ms.employee_id
            JOIN sites s ON s.id = ms.site_id
            LEFT JOIN schedule_templates st ON st.id = ms.template_id
            WHERE ms.tenant_id = %s
              AND ms.schedule_date >= %s
              AND ms.schedule_date < %s
              AND (%s = '' OR ms.site_id::text = %s)
            ORDER BY ms.schedule_date ASC, s.site_code ASC, e.employee_code ASC
            """,
            (target_tenant["id"], grid_start, grid_end, staff_site_id, staff_site_id),
        )
        board_rows = [dict(r) for r in cur.fetchall()]
        cur.execute(
            """
            SELECT work_date, SUM(required_count) AS support_request_count
            FROM site_daytime_need_counts
            WHERE tenant_id = %s
              AND work_date >= %s
              AND work_date < %s
              AND (%s = '' OR site_id::text = %s)
            GROUP BY work_date
            ORDER BY work_date ASC
            """,
            (target_tenant["id"], grid_start, grid_end, staff_site_id, staff_site_id),
        )
        support_rows = [dict(r) for r in cur.fetchall()]
    db_ms = (time.perf_counter() - db_started_at) * 1000

    serialize_started_at = time.perf_counter()
    day_map: dict[str, dict] = {
        date_key: {
            "date": date_key,
            "holiday_name": "",
            "items": [],
            "support_request_count": 0,
        }
        for date_key in _month_day_keys(grid_start, grid_end)
    }
    employee_map: dict[str, dict] = {}

    for row in board_rows:
        date_raw = row.get("schedule_date")
        date_key = date_raw.isoformat() if isinstance(date_raw, date) else str(date_raw or "").strip()
        if not date_key or date_key not in day_map:
            continue
        shift_type = _normalize_shift_type(str(row.get("shift_type") or ""))
        status = "non_working" if shift_type in NON_WORKING_SHIFT_TYPES else "scheduled"
        canonical = _resolve_canonical_schedule_time(row)
        item = {
            "schedule_id": str(row.get("schedule_id") or "").strip(),
            "employee_id": str(row.get("employee_id") or "").strip(),
            "employee_code": str(row.get("employee_code") or "").strip(),
            "employee_name": str(row.get("employee_name") or "").strip(),
            "soc_role": str(row.get("soc_role") or "").strip(),
            "duty_role": str(row.get("duty_role") or "").strip(),
            "site_code": str(row.get("site_code") or "").strip(),
            "site_name": str(row.get("site_name") or "").strip(),
            "shift_type": shift_type,
            "shift_label": str(canonical.get("label") or _row_shift_label(row)).strip(),
            "start_time": _format_time_for_response(canonical.get("start_time")),
            "end_time": _format_time_for_response(canonical.get("end_time")),
            "status": status,
            "template_id": str(row.get("template_id") or "").strip(),
            "template_name": str(row.get("template_name") or "").strip(),
            "duty_type": _normalize_schedule_template_duty_type(row.get("duty_type")),
            "paid_hours": float(canonical["hours"]) if canonical.get("hours") is not None else None,
        }
        day_map[date_key]["items"].append(item)
        if not day_map[date_key]["holiday_name"] and shift_type == "holiday":
            day_map[date_key]["holiday_name"] = "공휴일"

    for row in support_rows:
        date_raw = row.get("work_date")
        date_key = date_raw.isoformat() if isinstance(date_raw, date) else str(date_raw or "").strip()
        if not date_key or date_key not in day_map:
            continue
        day_map[date_key]["support_request_count"] = int(row.get("support_request_count") or 0)

    for date_key in list(day_map.keys()):
        day_map[date_key]["items"] = _merge_board_items_for_calendar(day_map[date_key]["items"])

    for row in rows:
        employee_id = str(row.get("employee_id") or "").strip()
        if employee_id and employee_id not in employee_map:
            employee_map[employee_id] = {
                "id": employee_id,
                "employee_code": str(row.get("employee_code") or "").strip(),
                "employee_name": str(row.get("employee_name") or "").strip(),
            }

    days = [day_map[date_key] for date_key in sorted(day_map.keys())]
    payload = {
        "month": month,
        "tenant_code": str(target_tenant.get("tenant_code") or "").strip(),
        "employees": list(employee_map.values()),
        "rows": rows,
        "days": days,
    }
    response_size_bytes = len(json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8"))
    serialize_ms = (time.perf_counter() - serialize_started_at) * 1000
    total_ms = (time.perf_counter() - total_started_at) * 1000
    metrics = {
        "db_ms": round(db_ms, 3),
        "serialize_ms": round(serialize_ms, 3),
        "total_ms": round(total_ms, 3),
        "response_size_bytes": response_size_bytes,
    }
    payload["metrics"] = metrics
    logger.info(
        "[SCHEDULE][MONTHLY_BOARD_LITE] month=%s tenant=%s db_ms=%.3f serialize_ms=%.3f total_ms=%.3f bytes=%s",
        month,
        str(target_tenant.get("tenant_code") or "").strip(),
        metrics["db_ms"],
        metrics["serialize_ms"],
        metrics["total_ms"],
        metrics["response_size_bytes"],
    )
    return payload


@router.get("/soc-leave-logs")
def list_soc_leave_logs(
    tenant_code: str | None = Query(default=None, max_length=64),
    start_date: str | None = Query(default=None, description="YYYY-MM-DD"),
    end_date: str | None = Query(default=None, description="YYYY-MM-DD"),
    limit: int = Query(default=50, ge=1, le=300),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    staff_scope = enforce_staff_site_scope(user)
    staff_site_id = str((staff_scope or {}).get("site_id") or "").strip()

    if start_date and end_date:
        try:
            period_start = datetime.strptime(start_date, "%Y-%m-%d").date()
            period_end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="start_date/end_date must be YYYY-MM-DD") from exc
        if period_end < period_start:
            raise HTTPException(status_code=400, detail="end_date must be greater than or equal to start_date")
    else:
        period_end = datetime.utcnow().date()
        period_start = period_end - timedelta(days=30)

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT ms.id AS schedule_id,
                   t.tenant_code,
                   e.employee_code,
                   e.full_name AS employee_name,
                   s.site_code,
                   s.site_name,
                   ms.schedule_date,
                   ms.shift_type,
                   ms.source,
                   ms.source_ticket_id,
                   ms.schedule_note,
                   COALESCE((to_jsonb(ms)->>'updated_at')::timestamptz, ms.created_at) AS updated_at
            FROM monthly_schedules ms
            JOIN tenants t ON t.id = ms.tenant_id
            JOIN employees e ON e.id = ms.employee_id
            JOIN sites s ON s.id = ms.site_id
            WHERE ms.tenant_id = %s
              AND (%s = '' OR ms.site_id::text = %s)
              AND ms.schedule_date BETWEEN %s AND %s
              AND lower(COALESCE(ms.source, '')) = 'soc'
              AND lower(ms.shift_type) IN ('off', 'holiday')
            ORDER BY ms.schedule_date DESC, e.employee_code ASC
            LIMIT %s
            """,
            (target_tenant["id"], staff_site_id, staff_site_id, period_start, period_end, limit),
        )
        return [dict(row) for row in cur.fetchall()]


@router.get("/overtime-daily")
def list_overtime_daily(
    date: str = Query(..., description="YYYY-MM-DD"),
    tenant_code: str | None = Query(default=None, max_length=64),
    employee_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    try:
        work_date = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    employee_filter = str(employee_code or "").strip()
    if not can_manage_schedule(user["role"]):
        # Employee role can only read own overtime records.
        own_employee_id = user.get("employee_id")
        if not own_employee_id:
            return []
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT employee_code
                FROM employees
                WHERE id = %s
                  AND tenant_id = %s
                LIMIT 1
                """,
                (own_employee_id, target_tenant["id"]),
            )
            own_row = cur.fetchone()
        own_employee_code = str(own_row["employee_code"] if own_row else "").strip()
        if employee_filter and employee_filter != own_employee_code:
            raise HTTPException(status_code=403, detail="employee mismatch")
        employee_filter = own_employee_code

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT so.id,
                   so.work_date,
                   so.employee_id,
                   e.employee_code,
                   e.full_name AS employee_name,
                   s.site_code,
                   s.site_name,
                   so.ticket_id,
                   so.reason,
                   so.source,
                   so.overtime_source,
                   so.overtime_policy,
                   so.approved_minutes,
                   so.raw_minutes_total,
                   so.overtime_hours_step,
                   so.closer_user_id,
                   cu.username AS closer_username,
                   cu.full_name AS closer_full_name,
                   so.updated_at
            FROM soc_overtime_approvals so
            JOIN employees e ON e.id = so.employee_id
            LEFT JOIN sites s ON s.id = so.site_id
            LEFT JOIN arls_users cu ON cu.id = so.closer_user_id
            WHERE so.tenant_id = %s
              AND so.work_date = %s
              AND (%s = '' OR e.employee_code = %s)
            ORDER BY
              e.employee_code ASC,
              CASE WHEN so.overtime_source = 'SOC_TICKET' THEN 0 ELSE 1 END,
              so.updated_at DESC
            """,
            (target_tenant["id"], work_date, employee_filter, employee_filter),
        )
        return [dict(row) for row in cur.fetchall()]


@router.get("/monthly/{schedule_id}/leader-candidates", response_model=ScheduleLeaderCandidatesOut)
def get_schedule_leader_candidates(
    schedule_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    context = _fetch_schedule_context(conn, schedule_id, target_tenant["id"])
    if not context:
        raise HTTPException(status_code=404, detail="schedule not found")

    candidates = _fetch_leader_candidates_for_site_day(
        conn,
        tenant_id=context["tenant_id"],
        site_id=context["site_id"],
        schedule_date=context["schedule_date"],
    )
    recommended = _recommended_leader_user_id(candidates)

    payload_candidates = [
        ScheduleLeaderCandidateOut(
            user_id=item["user_id"],
            username=item["username"],
            full_name=item["full_name"],
            employee_code=item["employee_code"],
            duty_role=item["duty_role"],
            is_recommended=(recommended is not None and str(item["user_id"]) == str(recommended)),
        )
        for item in candidates
    ]

    return ScheduleLeaderCandidatesOut(
        schedule_id=schedule_id,
        site_code=context["site_code"],
        schedule_date=context["schedule_date"],
        current_leader_user_id=context.get("leader_user_id"),
        recommended_leader_user_id=recommended,
        candidates=payload_candidates,
    )


@router.post("/monthly")
def upsert_monthly_rows(
    rows: list[ScheduleCreateRow],
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    created = 0
    skipped = 0
    affected_site_days: set[tuple[str, str, str]] = set()

    with conn.cursor() as cur:
        for row in rows:
            refs = _lookup_refs(
                conn,
                user["tenant_id"],
                user["tenant_code"],
                row.company_code,
                row.site_code,
                row.employee_code,
            )
            if not refs:
                skipped += 1
                continue

            cur.execute(
                """
                SELECT 1
                FROM monthly_schedules
                WHERE tenant_id = %s
                  AND employee_id = %s
                  AND schedule_date = %s
                  AND lower(COALESCE(shift_type, 'day')) = lower(%s)
                """,
                (user["tenant_id"], refs["employee_id"], row.schedule_date, row.shift_type),
            )
            if cur.fetchone():
                skipped += 1
                continue

            cur.execute(
                """
                INSERT INTO monthly_schedules (id, tenant_id, company_id, site_id, employee_id, schedule_date, shift_type, leader_user_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NULL)
                """,
                (
                    uuid.uuid4(),
                    user["tenant_id"],
                    refs["company_id"],
                    refs["site_id"],
                    refs["employee_id"],
                    row.schedule_date,
                    row.shift_type,
                ),
            )
            affected_site_days.add((str(user["tenant_id"]), str(refs["site_id"]), row.schedule_date.isoformat()))
            created += 1

    for tenant_id, site_id, schedule_date_raw in affected_site_days:
        _refresh_daily_leader_defaults(
            conn,
            tenant_id=tenant_id,
            site_id=site_id,
            schedule_date=date.fromisoformat(schedule_date_raw),
        )

    return {"created": created, "skipped": skipped}


@router.get("/export")
def export_monthly_csv(
    month: str | None = Query(None, description="YYYY-MM"),
    start_date: str | None = Query(None, description="YYYY-MM-DD"),
    end_date: str | None = Query(None, description="YYYY-MM-DD"),
    tenant_code: str | None = None,
    company_code: str | None = None,
    site_code: str | None = None,
    file_format: str = Query("xlsx", alias="format"),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    target_tenant = (tenant_code or user["tenant_code"]).strip()
    company_filter = (company_code or "").strip()
    site_filter = (site_code or "").strip()
    staff_scope = enforce_staff_site_scope(user, request_site_code=site_filter)
    if staff_scope:
        site_filter = str(staff_scope.get("site_code") or "").strip()
    normalized_format = (file_format or "xlsx").strip().lower()
    if normalized_format not in IMPORT_FORMATS:
        raise HTTPException(status_code=400, detail="format must be csv or xlsx")

    period_start, period_end_exclusive, period_label = _parse_export_period(month, start_date, end_date)

    with conn.cursor() as cur:
        cur.execute("SELECT id, tenant_code FROM tenants WHERE tenant_code = %s", (target_tenant,))
        tenant = cur.fetchone()
    if not tenant:
        raise HTTPException(status_code=404, detail="tenant not found")
    if tenant["tenant_code"] != user["tenant_code"] and not is_super_admin(user["role"]):
        raise HTTPException(status_code=403, detail="tenant mismatch")

    payload = _fetch_export_rows(
        conn=conn,
        tenant_id=tenant["id"],
        period_start=period_start,
        period_end_exclusive=period_end_exclusive,
        company_code=company_filter,
        site_code=site_filter,
    )

    headers = _schedule_import_headers()
    if normalized_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "schedule_export"
        sheet.append(headers)
        for row in payload:
            sheet.append(
                [
                    row["tenant_code"],
                    row["company_code"],
                    row["site_code"],
                    row["employee_code"],
                    str(row["schedule_date"]),
                    row["shift_type"],
                ]
            )
        out = BytesIO()
        workbook.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=schedule_export_{period_label}.xlsx"},
        )

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    for row in payload:
        writer.writerow(
            [
                row["tenant_code"],
                row["company_code"],
                row["site_code"],
                row["employee_code"],
                str(row["schedule_date"]),
                row["shift_type"],
            ]
        )

    out.seek(0)
    return StreamingResponse(
        out,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=schedule_export_{period_label}.csv"},
    )


@router.get("/export/monthly-excel")
def export_monthly_board_excel(
    month: str = Query(..., description="YYYY-MM"),
    tenant_code: str | None = Query(default=None, max_length=64),
    site_code: str | None = Query(default=None, min_length=1, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    _month_bounds(month)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    normalized_site_code = str(site_code or "").strip()
    export_all_sites = normalized_site_code.lower() == "all"
    if export_all_sites:
        site_rows = _list_site_contexts_for_export(conn, tenant_id=str(target_tenant["id"]))
        if not site_rows:
            raise HTTPException(status_code=404, detail="site not found")
    else:
        site_row = _resolve_site_context_by_code(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_code=normalized_site_code,
        )
        if not site_row:
            raise HTTPException(status_code=404, detail="site not found")
        site_rows = [site_row]

    blocked_support_status = None
    for site_row in site_rows:
        support_source_row = _get_support_roundtrip_source(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_id=str(site_row["id"]),
            month_key=month,
        )
        if support_source_row and not bool(support_source_row.get("final_download_enabled")):
            blocked_support_status = _build_support_roundtrip_status_payload(
                conn,
                source_row=support_source_row,
                tenant_id=str(target_tenant["id"]),
                site_id=str(site_row["id"]),
                site_code=str(site_row.get("site_code") or "").strip(),
                month_key=month,
            )
            break
    if blocked_support_status:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "HQ 지원근무 병합이 아직 최신 리비전에 대해 완료되지 않아 최종본 다운로드가 비활성화됩니다.",
                "support_roundtrip": blocked_support_status.model_dump(mode="json"),
            },
        )

    if not export_all_sites:
        export_ctx = _collect_monthly_export_context(
            conn,
            target_tenant=target_tenant,
            site_row=site_rows[0],
            month_key=month,
            user=user,
        )
        workbook = export_ctx["workbook"]
    else:
        export_contexts = [
            _collect_monthly_export_context(
                conn,
                target_tenant=target_tenant,
                site_row=site_row,
                month_key=month,
                user=user,
                build_workbook=False,
            )
            for site_row in site_rows
        ]
        workbook, template_path = _load_required_arls_month_workbook()
        template_version = _describe_arls_template_version(template_path)
        template_sheet = workbook[ARLS_SHEET_NAME] if ARLS_SHEET_NAME in workbook.sheetnames else workbook.active
        target_sheets = [template_sheet]
        for _ in range(1, len(export_contexts)):
            target_sheets.append(workbook.copy_worksheet(template_sheet))
        used_titles: set[str] = set()
        for index, (site_ctx, target_sheet) in enumerate(zip(export_contexts, target_sheets), start=1):
            desired_title = _normalize_excel_sheet_title(
                site_ctx.get("site_name"),
                fallback=f"{site_ctx.get('site_code') or 'SITE'}_{index}",
            )
            sheet_title = desired_title
            suffix = 2
            while sheet_title in used_titles:
                base_title = desired_title[: max(0, 31 - len(f"_{suffix}"))]
                sheet_title = f"{base_title}_{suffix}"
                suffix += 1
            used_titles.add(sheet_title)
            target_sheet.title = sheet_title
            _build_arls_month_sheet(
                workbook,
                sheet_name=sheet_title,
                month_key=month,
                rows=site_ctx["rows"],
                tenant_code=str(target_tenant.get("tenant_code") or "").strip(),
                site_code=site_ctx["site_code"],
                site_name=site_ctx["site_name"],
                site_address=site_ctx["site_address"],
                support_rows=site_ctx["support_rows"],
                overnight_rows=site_ctx["overnight_rows"],
                employee_overnight_rows=site_ctx["employee_overnight_rows"],
                daytime_need_rows=site_ctx["daytime_need_rows"],
                export_revision=site_ctx["export_revision"],
                template_version=template_version,
                source_version=f"{ARLS_EXPORT_SOURCE_VERSION}:all-sites",
                write_metadata=False,
            )
        _upsert_arls_export_metadata_sheet(
            workbook,
            tenant_code=str(target_tenant.get("tenant_code") or "").strip(),
            site_code="ALL",
            site_name="전체 지점",
            month_key=month,
            export_revision=_build_all_sites_export_revision(export_contexts),
            template_version=template_version,
            source_version=f"{ARLS_EXPORT_SOURCE_VERSION}:all-sites",
            employee_count=sum(len(ctx.get("employee_blocks") or []) for ctx in export_contexts),
            row_count=sum(len(ctx.get("rows") or []) for ctx in export_contexts),
            support_row_count=sum(len(ctx.get("support_rows") or []) for ctx in export_contexts),
            overnight_row_count=sum(len(ctx.get("overnight_rows") or []) for ctx in export_contexts),
            employee_overnight_row_count=sum(len(ctx.get("employee_overnight_rows") or []) for ctx in export_contexts),
        )

    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    try:
        month_year, month_number = [int(part) for part in str(month).split("-", 1)]
    except Exception:
        now_kst = datetime.now(timezone(timedelta(hours=9)))
        month_year = now_kst.year
        month_number = now_kst.month
    else:
        now_kst = datetime.now(timezone(timedelta(hours=9)))
    generated_on = now_kst.strftime("%Y%m%d")
    if export_all_sites:
        tenant_name_segment = str(
            target_tenant.get("tenant_name")
            or target_tenant.get("company_name")
            or target_tenant.get("tenant_code")
            or "TENANT"
        ).strip()
        download_filename = f"HQ_{tenant_name_segment}_월간 근무표_{generated_on}.xlsx"
    else:
        site_segment = str(site_rows[0]["site_code"]).strip()
        download_filename = f"{month_year}년 {month_number}월 근무표_{site_segment}_{generated_on[-6:]}.xlsx"
    encoded_filename = quote(download_filename)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"monthly_schedule.xlsx\"; filename*=UTF-8''{encoded_filename}"
            )
        },
    )


def _resolve_import_slot_from_schedule_row(row: dict) -> str:
    shift_type = _normalize_shift_type(row.get("shift_type"))
    if shift_type in NON_WORKING_SHIFT_TYPES:
        return "overtime"
    if shift_type == "overtime":
        return "overtime"
    duty_type = _normalize_schedule_template_duty_type(row.get("duty_type"))
    if duty_type in {"day", "overtime", "night"}:
        return duty_type
    return "night" if shift_type == "night" else "day"


def _load_existing_schedule_rows_for_import(
    conn,
    *,
    tenant_id: str,
    site_id: str,
    month_key: str,
) -> list[dict]:
    start_date, end_date = _month_bounds(month_key)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT ms.id AS schedule_id,
                   ms.tenant_id,
                   ms.company_id,
                   ms.site_id,
                   ms.employee_id,
                   ms.schedule_date,
                   ms.shift_type,
                   ms.template_id,
                   ms.shift_start_time,
                   ms.shift_end_time,
                   ms.paid_hours,
                   ms.schedule_note,
                   ms.source,
                   ms.source_ticket_id,
                   st.duty_type,
                   st.template_name,
                   e.employee_code,
                   e.full_name AS employee_name
            FROM monthly_schedules ms
            JOIN employees e ON e.id = ms.employee_id
            LEFT JOIN schedule_templates st ON st.id = ms.template_id
            WHERE ms.tenant_id = %s
              AND ms.site_id = %s
              AND ms.schedule_date >= %s
              AND ms.schedule_date < %s
            ORDER BY ms.schedule_date ASC, e.employee_code ASC
            """,
            (tenant_id, site_id, start_date, end_date),
        )
        return [dict(row) for row in cur.fetchall()]


def _build_existing_schedule_row_index(rows: list[dict]) -> dict[tuple[str, str, str], list[dict]]:
    index: dict[tuple[str, str, str], list[dict]] = {}
    for row in rows:
        employee_id = str(row.get("employee_id") or "").strip()
        schedule_date = row.get("schedule_date")
        if not employee_id or not isinstance(schedule_date, date):
            continue
        key = (employee_id, schedule_date.isoformat(), _resolve_import_slot_from_schedule_row(row))
        index.setdefault(key, []).append(dict(row))
    return index


def _build_protected_overnight_key_set(
    *,
    support_rows: list[dict],
    overnight_rows: list[dict],
    employee_overnight_rows: list[dict],
) -> set[tuple[str, str]]:
    protected: set[tuple[str, str]] = set()
    overnight_days = {
        row.get("work_date").isoformat()
        for row in overnight_rows
        if isinstance(row.get("work_date"), date)
    }
    for row in employee_overnight_rows:
        employee_id = str(row.get("employee_id") or "").strip()
        work_date = row.get("work_date")
        if employee_id and isinstance(work_date, date):
            protected.add((employee_id, work_date.isoformat()))
    for row in support_rows:
        employee_id = str(row.get("employee_id") or "").strip()
        work_date = row.get("work_date")
        if not employee_id or not isinstance(work_date, date):
            continue
        date_key = work_date.isoformat()
        if date_key in overnight_days:
            protected.add((employee_id, date_key))
    return protected


def _match_import_employee(employee_index: dict[str, list[dict]], employee_name: str) -> dict | None:
    token = _normalize_name_token(employee_name)
    matches = employee_index.get(token, [])
    if len(matches) != 1:
        return None
    return dict(matches[0])


def _is_monthly_base_schedule_source(source: object) -> bool:
    return str(source or "").strip().lower() in ARLS_MONTHLY_BASE_SOURCE_ALIASES


def _collect_required_mapping_keys(body_cells: list[dict[str, Any]]) -> set[tuple[str, str]]:
    required: set[tuple[str, str]] = set()
    for row in body_cells:
        if str(row.get("parsed_semantic_type") or "").strip() != "numeric_hours":
            continue
        duty_type = _normalize_schedule_template_duty_type(row.get("duty_type"))
        numeric_hours = row.get("numeric_hours")
        key = _row_type_hours_mapping_key(duty_type, numeric_hours)
        if key[0] and key[1]:
            required.add(key)
    return required


def _validate_mapping_profile_requirements(
    *,
    body_cells: list[dict[str, Any]],
    mapping_profile: dict[str, Any] | None,
    mapping_lookup: dict[tuple[str, str], dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    issues: list[dict[str, Any]] = []
    blocked_reasons: list[str] = []
    missing_entries: list[str] = []
    required_keys = sorted(_collect_required_mapping_keys(body_cells))
    if required_keys and not mapping_profile:
        issues.append(_build_import_issue("TEMPLATE_PROFILE_NOT_PREPARED", section="mapping_profile"))
        blocked_reasons.append("근무 템플릿 매핑 프로필이 준비되지 않아 업로드를 진행할 수 없습니다.")
        missing_entries.extend(f"{row_type}:{hours}" for row_type, hours in required_keys)
        return issues, blocked_reasons, missing_entries
    invalid_mapping_entries = [
        item for item in (mapping_profile or {}).get("entries") or []
        if not str(item.get("template_id") or "").strip()
        or not str(item.get("template_name") or "").strip()
        or not bool(item.get("template_is_active", True))
    ]
    if invalid_mapping_entries:
        issues.append(_build_import_issue("CANNOT_RESOLVE_TEMPLATE", section="mapping_profile"))
        blocked_reasons.append("mapping profile에 유효하지 않은 템플릿 연결이 있어 업로드를 진행할 수 없습니다.")
    for row_type, hours_key in required_keys:
        if (row_type, hours_key) not in mapping_lookup:
            missing_entries.append(f"{row_type}:{hours_key}")
    if missing_entries:
        blocked_reasons.append(f"필수 근무 템플릿 매핑 {len(missing_entries)}건이 누락되어 적용할 수 없습니다.")
    return issues, blocked_reasons, missing_entries


def _resolve_import_body_value(
    *,
    templates: list[dict],
    mapping_lookup: dict[tuple[str, str], dict[str, Any]] | None = None,
    duty_type: str,
    workbook_value: str,
) -> tuple[dict[str, Any], str | None, str | None]:
    value = str(workbook_value or "").strip()
    if not value:
        return {
            "work_value": "",
            "shift_type": "",
            "template_id": None,
            "template_name": None,
            "shift_start_time": None,
            "shift_end_time": None,
            "paid_hours": None,
        }, None, None
    if value in ARLS_LEAVE_MARKERS:
        shift_type = "holiday" if value == "공휴일" else "off"
        return {
            "work_value": value,
            "shift_type": shift_type,
            "template_id": None,
            "template_name": None,
            "shift_start_time": None,
            "shift_end_time": None,
            "paid_hours": None,
        }, None, None
    paid_hours = _parse_numeric_hours(value)
    if paid_hours is None:
        return {}, "UNSUPPORTED_CELL_FORMAT", "잘못된 값"
    normalized_duty_type = _normalize_schedule_template_duty_type(duty_type)
    template_row = None
    if mapping_lookup:
        mapping_entry = mapping_lookup.get(_row_type_hours_mapping_key(normalized_duty_type, float(paid_hours)))
        if not mapping_entry:
            return {}, "TEMPLATE_MAPPING_MISSING", "매핑 가능한 근무 템플릿이 없습니다."
        target_template_id = str(mapping_entry.get("template_id") or "").strip()
        template_row = next((row for row in templates if str(row.get("id") or "").strip() == target_template_id), None)
        if not template_row:
            return {}, "CANNOT_RESOLVE_TEMPLATE", "매핑된 근무 템플릿을 찾을 수 없습니다."
    else:
        template_row = _choose_template_by_duty_hours(
            templates,
            duty_type=normalized_duty_type,
            paid_hours=float(paid_hours),
            strict_hours=True,
        )
    if not template_row:
        return {}, "TEMPLATE_MAPPING_MISSING", "매칭 가능한 근무 템플릿이 없습니다."
    return {
        "work_value": _format_export_hours_value(paid_hours),
        "shift_type": _resolve_shift_type_from_duty_type(normalized_duty_type),
        "template_id": str(template_row.get("id") or "").strip() or None,
        "template_name": str(template_row.get("template_name") or "").strip() or None,
        "shift_start_time": _normalize_time_text(template_row.get("start_time")),
        "shift_end_time": _normalize_time_text(template_row.get("end_time")),
        "paid_hours": float(paid_hours),
        "mapping_key": f"{normalized_duty_type}:{_normalize_hours_key(paid_hours)}",
    }, None, None


def _current_cell_has_time_conflict(current_rows: list[dict], *, next_start: str | None, next_end: str | None, current_schedule_id: str | None = None) -> bool:
    comparable_rows = [
        row for row in current_rows
        if str(row.get("schedule_id") or "").strip() != str(current_schedule_id or "").strip()
    ]
    if not comparable_rows:
        return False
    if not next_start or not next_end:
        return True
    for row in comparable_rows:
        row_start_time, row_end_time = _resolve_schedule_row_conflict_range(row)
        if _schedule_time_ranges_overlap(
            next_start,
            next_end,
            row_start_time,
            row_end_time,
        ):
            return True
    return False


def _build_schedule_import_preview_result(
    conn,
    *,
    workbook: Workbook,
    target_tenant: dict,
    scope_site: dict,
    selected_month: str,
    user: dict,
    filename: str,
) -> dict[str, Any]:
    export_ctx = _collect_monthly_export_context(
        conn,
        target_tenant=target_tenant,
        site_row=scope_site,
        month_key=selected_month,
        user=user,
    )
    current_revision = str(export_ctx.get("export_revision") or "").strip()
    workbook_ctx = _detect_arls_import_workbook_context(
        workbook,
        selected_month=selected_month,
        expected_tenant_code=str(target_tenant.get("tenant_code") or "").strip(),
        expected_site_code=str(scope_site.get("site_code") or "").strip(),
        current_revision=current_revision,
    )
    metadata = dict(workbook_ctx.get("metadata") or {})
    parsed_uploaded = dict(workbook_ctx.get("parsed_sheet") or {})

    current_parsed = export_ctx["parsed_sheet"]
    current_body_index = _build_visible_value_index(current_parsed.get("body_cells") or [])
    current_need_index = _build_visible_value_index(current_parsed.get("need_cells") or [], include_employee=False)
    current_support_index = _build_support_value_index(current_parsed.get("support_cells") or [])

    templates = _fetch_schedule_templates(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(scope_site["id"]),
        include_inactive=False,
    )
    mapping_profile = _fetch_active_schedule_import_mapping_profile(
        conn,
        tenant_id=str(target_tenant["id"]),
    )
    mapping_lookup = _build_schedule_import_mapping_lookup(mapping_profile)
    mapping_summary = _build_schedule_import_mapping_summary(mapping_profile)
    employees = _load_site_employees(conn, tenant_id=str(target_tenant["id"]), site_id=str(scope_site["id"]))
    employee_index = _build_employee_name_index(employees)
    existing_schedule_rows = _load_existing_schedule_rows_for_import(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(scope_site["id"]),
        month_key=selected_month,
    )
    existing_schedule_index = _build_existing_schedule_row_index(existing_schedule_rows)
    current_support_request_rows = _read_monthly_support_request_rows_for_export(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(scope_site["id"]),
        month_key=selected_month,
    )
    current_support_request_index = _build_support_request_ticket_index(current_support_request_rows)
    protected_overnight_keys = _build_protected_overnight_key_set(
        support_rows=export_ctx["support_rows"],
        overnight_rows=export_ctx["overnight_rows"],
        employee_overnight_rows=export_ctx["employee_overnight_rows"],
    )

    resolved_rows: list[dict[str, Any]] = []
    support_ticket_rows: list[dict[str, Any]] = []
    diff_counts: Counter[str] = Counter()
    blocked_reasons: list[str] = list(workbook_ctx.get("blocked_reasons") or [])
    issue_records: list[dict[str, Any]] = list(workbook_ctx.get("issues") or [])
    mapping_issues, mapping_blocked_reasons, missing_mapping_entries = _validate_mapping_profile_requirements(
        body_cells=list(parsed_uploaded.get("body_cells") or []),
        mapping_profile=mapping_profile,
        mapping_lookup=mapping_lookup,
    )
    issue_records.extend(mapping_issues)
    for reason in mapping_blocked_reasons:
        _append_blocked_reason(blocked_reasons, reason)
    mapping_summary["missing_required_entries"] = list(missing_mapping_entries)

    def append_row_issue(
        code: str,
        *,
        row_no: int | None = None,
        col_no: int | None = None,
        section: str | None = None,
        message: str | None = None,
    ) -> None:
        issue_records.append(
            _build_import_issue(
                code,
                message=message,
                sheet_name=str((parsed_uploaded.get("body_cells") or [{}])[0].get("source_sheet") or ARLS_SHEET_NAME) if parsed_uploaded.get("body_cells") else ARLS_SHEET_NAME,
                row_no=row_no,
                col_no=col_no,
                section=section,
            )
        )

    for row in parsed_uploaded.get("body_cells") or []:
        employee_name = str(row.get("employee_name") or "").strip()
        duty_type = _normalize_schedule_template_duty_type(row.get("duty_type"))
        schedule_date = row.get("schedule_date")
        workbook_value = str(row.get("work_value") or "").strip()
        if not isinstance(schedule_date, date):
            continue
        current_key = (_normalize_name_token(employee_name), duty_type, schedule_date.isoformat())
        current_row = current_body_index.get(current_key) or {}
        current_value = str(current_row.get("work_value") or "").strip()
        employee_row = None
        validation_code = _normalize_import_issue_code(row.get("issue_code")) or None
        validation_error = str(row.get("issue_message") or "").strip() or None
        apply_action = "none"
        diff_category = "unchanged"
        is_blocking = _issue_is_blocking(validation_code) if validation_code else False
        is_protected = False
        protected_reason = None
        current_schedule_id = None
        template_meta: dict[str, Any] = {
            "template_id": None,
            "template_name": None,
            "shift_start_time": None,
            "shift_end_time": None,
            "paid_hours": None,
            "shift_type": "",
            "work_value": workbook_value,
            "mapping_key": None,
        }

        if not validation_code:
            employee_row, employee_issue_code, employee_issue_message = _resolve_import_employee_match(
                employee_index,
                employee_name=employee_name,
                schedule_date=schedule_date,
            )
            if employee_issue_code:
                validation_code = employee_issue_code
                validation_error = employee_issue_message
        if validation_code and str(validation_code) in {"EMPLOYEE_MATCH_FAILED", "EMPLOYEE_MATCH_AMBIGUOUS"}:
            is_blocking = True
            append_row_issue(
                validation_code,
                row_no=int(row.get("row_no") or 0),
                col_no=int(row.get("col_no") or 0),
                section="base_schedule",
                message=validation_error,
            )
        else:
            parsed_value_meta, value_error_code, value_error_message = _resolve_import_body_value(
                templates=templates,
                mapping_lookup=mapping_lookup,
                duty_type=duty_type,
                workbook_value=workbook_value,
            )
            if not validation_code and value_error_code:
                validation_code = value_error_code
                validation_error = value_error_message
                is_blocking = True
                append_row_issue(
                    validation_code,
                    row_no=int(row.get("row_no") or 0),
                    col_no=int(row.get("col_no") or 0),
                    section="base_schedule",
                    message=validation_error,
                )
            else:
                template_meta = parsed_value_meta

            existing_key = (str((employee_row or {}).get("id") or "").strip(), schedule_date.isoformat(), duty_type)
            existing_rows = existing_schedule_index.get(existing_key) or []
            base_existing_rows = [item for item in existing_rows if _is_monthly_base_schedule_source(item.get("source"))]
            foreign_existing_rows = [item for item in existing_rows if not _is_monthly_base_schedule_source(item.get("source"))]
            current_schedule_id = str(base_existing_rows[0].get("schedule_id") or "").strip() if len(base_existing_rows) == 1 else None

            if not validation_code and duty_type == "night" and existing_key[:2] in protected_overnight_keys:
                if workbook_value != current_value:
                    diff_category = "ignored_protected"
                    apply_action = "none"
                    is_protected = True
                    protected_reason = "승인된 야간/자체근무 truth가 관리하는 값이라 업로드로 덮어쓸 수 없습니다."
                else:
                    diff_category = "unchanged"
            elif not validation_code and foreign_existing_rows and workbook_value != current_value:
                validation_code = "NON_BASE_LINEAGE_CONFLICT"
                validation_error = "같은 슬롯에 다른 lineage의 일정이 있어 base upload로 수정할 수 없습니다."
                is_blocking = True
                append_row_issue(
                    validation_code,
                    row_no=int(row.get("row_no") or 0),
                    col_no=int(row.get("col_no") or 0),
                    section="base_schedule",
                    message=validation_error,
                )
            elif not validation_code and len(base_existing_rows) > 1 and workbook_value != current_value:
                validation_code = "MULTI_ROW_CONFLICT"
                validation_error = "여러 일정이 한 셀에 합산되어 있어 업로드로 자동 수정할 수 없습니다."
                is_blocking = True
                append_row_issue(
                    validation_code,
                    row_no=int(row.get("row_no") or 0),
                    col_no=int(row.get("col_no") or 0),
                    section="base_schedule",
                    message=validation_error,
                )
            elif not validation_code:
                if workbook_value == current_value:
                    diff_category = "unchanged"
                elif not workbook_value and current_value:
                    diff_category = "delete"
                    apply_action = "delete"
                elif workbook_value and not current_value:
                    diff_category = "create"
                    apply_action = "create"
                else:
                    diff_category = "update"
                    apply_action = "update"

                if apply_action in {"create", "update"}:
                    existing_schedule_id = current_schedule_id if apply_action == "update" else None
                    if _current_cell_has_time_conflict(
                        base_existing_rows,
                        next_start=template_meta.get("shift_start_time"),
                        next_end=template_meta.get("shift_end_time"),
                        current_schedule_id=existing_schedule_id,
                    ):
                        validation_code = "TIME_CONFLICT"
                        validation_error = "같은 날짜의 기존 일정과 시간이 겹칩니다."
                        is_blocking = True
                        apply_action = "none"
                        diff_category = "conflict"
                        append_row_issue(
                            validation_code,
                            row_no=int(row.get("row_no") or 0),
                            col_no=int(row.get("col_no") or 0),
                            section="base_schedule",
                            message=validation_error,
                        )

        if validation_code:
            diff_counts["conflict" if is_blocking else "invalid"] += 1
        else:
            diff_counts[diff_category] += 1

        resolved_rows.append(
            {
                "row_no": int(row.get("row_no") or 0),
                "tenant_id": str(target_tenant["id"]),
                "tenant_code": str(target_tenant.get("tenant_code") or "").strip(),
                "company_id": str(scope_site.get("company_id") or "").strip() or None,
                "company_code": str(scope_site.get("company_code") or "").strip(),
                "site_id": str(scope_site["id"]),
                "site_code": str(scope_site["site_code"]),
                "employee_id": str((employee_row or {}).get("id") or "").strip() or None,
                "employee_code": str((employee_row or {}).get("employee_code") or "").strip(),
                "employee_name": employee_name,
                "schedule_date": schedule_date,
                "shift_type": str(template_meta.get("shift_type") or _resolve_shift_type_from_duty_type(duty_type)),
                "duty_type": duty_type,
                "source_sheet": str(row.get("source_sheet") or ARLS_SHEET_NAME),
                "source_col": _excel_col_label(int(row.get("col_no") or 0)),
                "template_id": template_meta.get("template_id"),
                "template_name": template_meta.get("template_name"),
                "work_value": workbook_value or None,
                "current_work_value": current_value or None,
                "parsed_semantic_type": str(row.get("parsed_semantic_type") or ""),
                "mapped_hours": row.get("numeric_hours"),
                "mapping_key": template_meta.get("mapping_key"),
                "shift_start_time": template_meta.get("shift_start_time"),
                "shift_end_time": template_meta.get("shift_end_time"),
                "paid_hours": template_meta.get("paid_hours"),
                "validation_code": validation_code,
                "validation_error": validation_error,
                "is_valid": validation_code is None,
                "is_blocking": is_blocking,
                "diff_category": diff_category,
                "apply_action": apply_action,
                "source_block": "body",
                "section_label": _schedule_template_duty_label(duty_type),
                "is_protected": is_protected,
                "protected_reason": protected_reason,
                "current_schedule_id": current_schedule_id or None,
            }
        )

    for row in parsed_uploaded.get("need_cells") or []:
        schedule_date = row.get("schedule_date")
        if not isinstance(schedule_date, date):
            continue
        workbook_value = str(row.get("work_value") or "").strip()
        current_key = (str(row.get("source_block") or "day_support_required_count"), schedule_date.isoformat())
        current_row = current_need_index.get(current_key) or {}
        current_value = str(current_row.get("work_value") or "").strip()
        required_count = row.get("required_count_numeric")
        parsed_raw_text = workbook_value
        support_shift_kind = "night" if str(row.get("source_block") or "").strip().startswith("night_support") else "day"
        validation_code = _normalize_import_issue_code(row.get("issue_code")) or None
        validation_error = str(row.get("issue_message") or "").strip() or None
        diff_category = "unchanged"
        apply_action = "none"
        is_blocking = _issue_is_blocking(validation_code) if validation_code else False
        if not validation_code:
            if workbook_value == current_value:
                diff_category = "unchanged"
            elif workbook_value or current_value:
                diff_category = "review"
        else:
            diff_category = "conflict"
        diff_counts[diff_category] += 1
        resolved_rows.append(
            {
                "row_no": int(row.get("row_no") or 0),
                "tenant_id": str(target_tenant["id"]),
                "tenant_code": str(target_tenant.get("tenant_code") or "").strip(),
                "company_id": str(scope_site.get("company_id") or "").strip() or None,
                "company_code": str(scope_site.get("company_code") or "").strip(),
                "site_id": str(scope_site["id"]),
                "site_code": str(scope_site["site_code"]),
                "employee_id": None,
                "employee_code": "",
                "employee_name": "",
                "schedule_date": schedule_date,
                "shift_type": support_shift_kind,
                "duty_type": f"{support_shift_kind}_support",
                "source_sheet": str(row.get("source_sheet") or ARLS_SHEET_NAME),
                "source_col": _excel_col_label(int(row.get("col_no") or 0)),
                "template_id": None,
                "template_name": None,
                "work_value": workbook_value or None,
                "current_work_value": current_value or None,
                "parsed_semantic_type": str(row.get("parsed_semantic_type") or ""),
                "mapped_hours": float(required_count) if required_count is not None else None,
                "mapping_key": None,
                "shift_start_time": None,
                "shift_end_time": None,
                "paid_hours": float(required_count) if required_count is not None else None,
                "validation_code": validation_code,
                "validation_error": validation_error,
                "is_valid": validation_code is None,
                "is_blocking": is_blocking,
                "diff_category": diff_category,
                "apply_action": apply_action,
                "source_block": str(row.get("source_block") or "day_support_required_count"),
                "section_label": "필요인원 수",
                "is_protected": False,
                "protected_reason": None,
                "current_schedule_id": None,
                "schedule_note": parsed_raw_text or None,
            }
        )

    for row in parsed_uploaded.get("support_cells") or []:
        schedule_date = row.get("schedule_date")
        if not isinstance(schedule_date, date):
            continue
        workbook_value = str(row.get("work_value") or "").strip()
        source_block = str(row.get("source_block") or "").strip()
        current_key = (
            source_block,
            schedule_date.isoformat(),
            int(row.get("row_no") or 0),
            int(row.get("col_no") or 0),
        )
        current_row = current_support_index.get(current_key) or {}
        current_value = str(current_row.get("work_value") or "").strip()
        validation_code = _normalize_import_issue_code(row.get("issue_code")) or None
        validation_error = str(row.get("issue_message") or "").strip() or None
        is_blocking = _issue_is_blocking(validation_code) if validation_code else False
        diff_category = "unchanged"
        is_protected = False
        protected_reason = None
        if source_block.endswith("_summary_count"):
            validation_code = validation_code or "PROTECTED_FIELD_IGNORED"
            validation_error = validation_error or "요약 행은 검토만 가능하며 직접 반영되지 않습니다."
            is_protected = True
            protected_reason = "요약/관리 영역은 이번 단계에서 직접 반영되지 않습니다."
        elif source_block in {"day_support_external_count", "night_support_purpose", "day_support_worker", "night_support_worker"}:
            protected_reason = "지원 수요/배정 영역은 이번 단계에서 분석 결과로만 제공합니다."
        if workbook_value != current_value:
            if validation_code and is_blocking:
                diff_category = "conflict"
            elif is_protected:
                diff_category = "ignored_protected"
                if validation_code == "PROTECTED_FIELD_IGNORED":
                    append_row_issue(
                        validation_code,
                        row_no=int(row.get("row_no") or 0),
                        col_no=int(row.get("col_no") or 0),
                        section=source_block,
                        message=validation_error,
                    )
            else:
                diff_category = "review"
        elif validation_code and is_blocking and workbook_value:
            diff_category = "conflict"
        diff_counts[diff_category] += 1
        resolved_rows.append(
            {
                "row_no": int(row.get("row_no") or 0),
                "tenant_id": str(target_tenant["id"]),
                "tenant_code": str(target_tenant.get("tenant_code") or "").strip(),
                "company_id": str(scope_site.get("company_id") or "").strip() or None,
                "company_code": str(scope_site.get("company_code") or "").strip(),
                "site_id": str(scope_site["id"]),
                "site_code": str(scope_site["site_code"]),
                "employee_id": None,
                "employee_code": "",
                "employee_name": "",
                "schedule_date": schedule_date,
                "shift_type": "",
                "duty_type": "",
                "source_sheet": str(row.get("source_sheet") or ARLS_SHEET_NAME),
                "source_col": _excel_col_label(int(row.get("col_no") or 0)),
                "template_id": None,
                "template_name": None,
                "work_value": workbook_value or None,
                "current_work_value": current_value or None,
                "parsed_semantic_type": str(row.get("parsed_semantic_type") or ""),
                "mapped_hours": None,
                "mapping_key": None,
                "shift_start_time": None,
                "shift_end_time": None,
                "paid_hours": None,
                "validation_code": validation_code,
                "validation_error": validation_error,
                "is_valid": validation_code is None or not is_blocking,
                "is_blocking": is_blocking,
                "diff_category": diff_category,
                "apply_action": "none",
                "source_block": source_block,
                "section_label": str(row.get("section_label") or "").strip(),
                "is_protected": is_protected,
                "protected_reason": protected_reason,
                "current_schedule_id": None,
            }
        )

    for block in parsed_uploaded.get("support_blocks") or []:
        target_date = block.get("target_date")
        if not isinstance(target_date, date):
            continue
        block_type = str(block.get("block_type") or "").strip()
        shift_kind = "night" if block_type == "night_support" else "day"
        request_count_numeric = block.get("required_count_numeric")
        current_ticket = current_support_request_index.get((target_date.isoformat(), shift_kind)) or {}
        current_request_count = int(current_ticket.get("request_count") or 0) if current_ticket else 0
        current_work_purpose = str(current_ticket.get("work_purpose") or "").strip()
        validation_code = None
        validation_error = None
        is_blocking = False
        diff_category = "unchanged"
        apply_action = "none"
        if request_count_numeric is None:
            validation_code = "SUPPORT_BLOCK_REQUIRED_COUNT_INVALID"
            validation_error = "필요 인원 수를 해석할 수 없습니다."
            is_blocking = True
        else:
            desired_request_count = max(0, int(request_count_numeric))
            desired_work_purpose = str(block.get("purpose_text") or "").strip()
            current_is_active = str(current_ticket.get("status") or SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS).strip() == SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS
            if desired_request_count > 0:
                if not current_ticket:
                    diff_category = "create"
                    apply_action = "upsert_sentrix_ticket"
                elif (
                    not current_is_active
                    or current_request_count != desired_request_count
                    or current_work_purpose != desired_work_purpose
                ):
                    diff_category = "update"
                    apply_action = "upsert_sentrix_ticket"
            elif current_ticket and current_is_active:
                diff_category = "delete"
                apply_action = "retract_sentrix_ticket"
        if validation_code:
            append_row_issue(
                validation_code,
                row_no=int(block.get("required_row_no") or block.get("required_row") or 0),
                col_no=None,
                section=block_type,
                message=validation_error,
            )
            diff_counts["conflict"] += 1
        elif apply_action != "none":
            diff_counts[f"sentrix_{diff_category}"] += 1
        support_ticket_rows.append(
            {
                "row_no": int(block.get("required_row_no") or block.get("required_row") or 0),
                "tenant_id": str(target_tenant["id"]),
                "tenant_code": str(target_tenant.get("tenant_code") or "").strip(),
                "company_id": str(scope_site.get("company_id") or "").strip() or None,
                "company_code": str(scope_site.get("company_code") or "").strip(),
                "site_id": str(scope_site["id"]),
                "site_code": str(scope_site["site_code"]),
                "employee_id": None,
                "employee_code": "",
                "employee_name": "",
                "schedule_date": target_date,
                "shift_type": shift_kind,
                "duty_type": f"{shift_kind}_support",
                "source_sheet": str(block.get("source_sheet") or ARLS_SHEET_NAME),
                "source_col": str(block.get("source_col") or ""),
                "template_id": None,
                "template_name": None,
                "work_value": str(block.get("required_count_raw") or "").strip() or None,
                "current_work_value": str(current_request_count) if current_ticket and current_is_active else None,
                "parsed_semantic_type": "support_demand",
                "mapped_hours": float(request_count_numeric) if request_count_numeric is not None else None,
                "mapping_key": None,
                "shift_start_time": None,
                "shift_end_time": None,
                "paid_hours": None,
                "validation_code": validation_code,
                "validation_error": validation_error,
                "is_valid": validation_code is None,
                "is_blocking": is_blocking,
                "diff_category": diff_category,
                "apply_action": apply_action,
                "source_block": "sentrix_support_ticket",
                "section_label": "Sentrix 주간 지원 요청" if shift_kind == "day" else "Sentrix 야간 지원 요청",
                "is_protected": False,
                "protected_reason": None,
                "current_schedule_id": None,
                "request_count": int(request_count_numeric) if request_count_numeric is not None else None,
                "current_request_count": current_request_count if current_ticket and current_is_active else 0,
                "purpose_text": str(block.get("purpose_text") or "").strip() or None,
                "current_purpose_text": current_work_purpose or None,
                "worker_slot_count": int(block.get("worker_slot_count") or 0),
                "valid_filled_count": int(block.get("valid_filled_count") or 0),
                "invalid_filled_count": int(block.get("invalid_filled_count") or 0),
                "external_count_raw": str(block.get("external_count_raw") or "").strip() or None,
                "external_count_numeric": block.get("external_count_numeric"),
                "detail_json": {
                    "required_count_raw": str(block.get("required_count_raw") or "").strip() or None,
                    "required_count_numeric": int(request_count_numeric) if request_count_numeric is not None else None,
                    "external_count_raw": str(block.get("external_count_raw") or "").strip() or None,
                    "external_count_numeric": block.get("external_count_numeric"),
                    "purpose_text": str(block.get("purpose_text") or "").strip() or None,
                    "worker_slot_count": int(block.get("worker_slot_count") or 0),
                    "valid_filled_count": int(block.get("valid_filled_count") or 0),
                    "invalid_filled_count": int(block.get("invalid_filled_count") or 0),
                    "issues": list(block.get("issues") or []),
                    "source_sheet": str(block.get("source_sheet") or ARLS_SHEET_NAME),
                    "source_col": str(block.get("source_col") or ""),
                    "required_row_no": int(block.get("required_row_no") or 0) or None,
                    "vendor_row_no": int(block.get("vendor_row_no") or 0) or None,
                    "purpose_row_no": int(block.get("purpose_row_no") or 0) or None,
                },
            }
        )

    total_rows = (
        len(parsed_uploaded.get("body_cells") or [])
        + len(parsed_uploaded.get("need_cells") or [])
        + len(parsed_uploaded.get("support_cells") or [])
    )
    preview_rows = [
        row for row in resolved_rows
        if row.get("diff_category") != "unchanged"
        or row.get("validation_code")
        or row.get("is_protected")
        or (
            (
                str(row.get("source_block") or "").startswith("day_support")
                or str(row.get("source_block") or "").startswith("night_support")
            )
            and (
                str(row.get("diff_category") or "").strip() != "unchanged"
                and (
                str(row.get("work_value") or "").strip()
                or str(row.get("current_work_value") or "").strip()
                )
            )
        )
    ]
    if not preview_rows:
        preview_rows = resolved_rows[: min(len(resolved_rows), 20)]

    error_counts, grouped_issues = _summarize_import_issues(issue_records)
    base_applicable_rows = sum(
        1
        for row in resolved_rows
        if str(row.get("source_block") or "") == "body"
        and str(row.get("apply_action") or "").strip() in {"create", "update", "delete"}
        and not bool(row.get("is_blocking"))
    )
    ticket_applicable_rows = sum(
        1
        for row in support_ticket_rows
        if str(row.get("apply_action") or "").strip() in {"upsert_sentrix_ticket", "retract_sentrix_ticket"}
        and not bool(row.get("is_blocking"))
    )
    applicable_rows = base_applicable_rows + ticket_applicable_rows
    blocked_rows = sum(1 for row in resolved_rows if bool(row.get("is_blocking"))) + sum(1 for row in support_ticket_rows if bool(row.get("is_blocking")))
    unchanged_rows = sum(1 for row in resolved_rows if str(row.get("diff_category") or "") == "unchanged")
    warning_rows = sum(
        1
        for row in resolved_rows
        if not bool(row.get("is_blocking"))
        and str(row.get("diff_category") or "") in {"review", "ignored_protected"}
    )

    can_apply = bool(workbook_ctx.get("workbook_valid")) and not blocked_reasons and blocked_rows == 0

    return {
        "resolved_rows": resolved_rows,
        "preview_rows": preview_rows[:IMPORT_PREVIEW_LIMIT],
        "error_counts": error_counts,
        "diff_counts": diff_counts,
        "issues": grouped_issues,
        "blocked_reasons": blocked_reasons,
        "total_rows": total_rows,
        "valid_rows": applicable_rows,
        "invalid_rows": blocked_rows,
        "applicable_rows": applicable_rows,
        "unchanged_rows": unchanged_rows,
        "blocked_rows": blocked_rows,
        "warning_rows": warning_rows,
        "support_ticket_rows": support_ticket_rows,
        "metadata": {
            "tenant_code": metadata.get("tenant_code") or str(target_tenant.get("tenant_code") or "").strip(),
            "site_code": metadata.get("site_code") or str(scope_site.get("site_code") or "").strip(),
            "month": metadata.get("month") or selected_month,
            "template_family": str(workbook_ctx.get("template_family") or ARLS_TEMPLATE_FAMILY_LABEL),
            "export_revision": metadata.get("export_revision"),
            "template_version": metadata.get("template_version"),
            "export_source_version": metadata.get("export_source_version"),
            "current_revision": current_revision,
            "workbook_kind": workbook_ctx.get("workbook_kind"),
            "workbook_valid": bool(workbook_ctx.get("workbook_valid")),
            "revision_status": workbook_ctx.get("revision_status"),
            "is_stale": bool(workbook_ctx.get("is_stale")),
            "mapping_profile": mapping_summary,
        },
        "can_apply": can_apply,
    }


def _persist_schedule_import_preview_batch(
    conn,
    *,
    preview: dict[str, Any],
    target_tenant: dict,
    scope_site: dict,
    month_key: str,
    filename: str,
    user: dict,
) -> tuple[uuid.UUID, list[str]]:
    batch_id = uuid.uuid4()
    resolved_rows = list(preview.get("resolved_rows") or [])
    support_ticket_rows = list(preview.get("support_ticket_rows") or [])
    metadata = dict(preview.get("metadata") or {})
    mapping_profile = dict(metadata.get("mapping_profile") or {})
    blocked_reasons = list(dict.fromkeys(str(item).strip() for item in (preview.get("blocked_reasons") or []) if str(item).strip()))
    issues = list(preview.get("issues") or [])
    status = "blocked" if blocked_reasons or int(preview.get("blocked_rows") or 0) > 0 else "previewed"
    invalid_samples: list[str] = []
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO schedule_import_batches
            (id, tenant_id, created_by, filename, status, total_rows, valid_rows, invalid_rows,
             import_mode, site_id, site_code, month_key, template_version, export_source_version,
             export_revision, current_revision, metadata_error, blocked_reasons_json, diff_counts_json,
             is_stale, metadata_json, issues_json, mapping_profile_id, mapping_profile_name,
             mapping_profile_updated_at, apply_result_json)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s, %s::jsonb, %s::jsonb, %s, %s, %s, '{}'::jsonb)
            """,
            (
                batch_id,
                target_tenant["id"],
                user["id"],
                filename,
                status,
                int(preview.get("total_rows") or 0),
                int(preview.get("applicable_rows") or 0),
                int(preview.get("blocked_rows") or 0),
                "canonical_workbook",
                scope_site["id"],
                str(scope_site.get("site_code") or "").strip(),
                month_key,
                str(metadata.get("template_version") or "").strip() or None,
                str(metadata.get("export_source_version") or "").strip() or None,
                str(metadata.get("export_revision") or "").strip() or None,
                str(metadata.get("current_revision") or "").strip() or None,
                "; ".join(blocked_reasons) or None,
                json.dumps(blocked_reasons, ensure_ascii=False),
                json.dumps(dict(preview.get("diff_counts") or {}), ensure_ascii=False),
                bool(metadata.get("is_stale")),
                json.dumps(metadata, ensure_ascii=False, default=str),
                json.dumps(issues, ensure_ascii=False, default=str),
                mapping_profile.get("profile_id"),
                str(mapping_profile.get("profile_name") or "").strip() or None,
                mapping_profile.get("updated_at"),
            ),
        )
        for row in [*resolved_rows, *support_ticket_rows]:
            validation_code = str(row.get("validation_code") or "").strip() or None
            validation_error = str(row.get("validation_error") or "").strip() or None
            if (
                str(row.get("source_block") or "").strip() != "sentrix_support_ticket"
                and (validation_code or row.get("is_protected"))
                and len(invalid_samples) < 20
            ):
                invalid_samples.append(
                    f"row {int(row.get('row_no') or 0)}: "
                    f"{validation_error or row.get('protected_reason') or _import_diff_status_label(diff_category=row.get('diff_category'), validation_code=validation_code, is_blocking=bool(row.get('is_blocking')))}"
                )
            schedule_date = row.get("schedule_date")
            cur.execute(
                """
                INSERT INTO schedule_import_rows
                (id, batch_id, row_no, tenant_code, company_code, site_code, employee_code,
                 employee_name, schedule_date, shift_type, duty_type, template_id, template_name,
                 work_value, shift_start_time, shift_end_time, paid_hours, validation_code,
                 validation_error, employee_id, company_id, site_id, tenant_id, is_valid,
                 source_block, section_label, current_work_value, diff_category, apply_action,
                 is_blocking, is_protected, protected_reason, current_schedule_id, payload_json)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                """,
                (
                    uuid.uuid4(),
                    batch_id,
                    int(row.get("row_no") or 0),
                    str(row.get("tenant_code") or "").strip(),
                    str(row.get("company_code") or "").strip(),
                    str(row.get("site_code") or "").strip(),
                    str(row.get("employee_code") or "").strip(),
                    str(row.get("employee_name") or "").strip(),
                    schedule_date if isinstance(schedule_date, date) else _parse_date_or_none(str(schedule_date or "")),
                    str(row.get("shift_type") or "").strip(),
                    str(row.get("duty_type") or "").strip(),
                    row.get("template_id"),
                    str(row.get("template_name") or "").strip() or None,
                    str(row.get("work_value") or "").strip() or None,
                    _normalize_time_text(row.get("shift_start_time")),
                    _normalize_time_text(row.get("shift_end_time")),
                    row.get("paid_hours"),
                    validation_code,
                    validation_error,
                    row.get("employee_id"),
                    row.get("company_id"),
                    row.get("site_id"),
                    row.get("tenant_id"),
                    not bool(row.get("is_blocking")),
                    str(row.get("source_block") or "").strip() or None,
                    str(row.get("section_label") or "").strip() or None,
                    str(row.get("current_work_value") or "").strip() or None,
                    str(row.get("diff_category") or "").strip() or None,
                    str(row.get("apply_action") or "").strip() or None,
                    bool(row.get("is_blocking")),
                    bool(row.get("is_protected")),
                    str(row.get("protected_reason") or "").strip() or None,
                    row.get("current_schedule_id"),
                    json.dumps(row, ensure_ascii=False, sort_keys=True, default=str),
                ),
            )
    return batch_id, invalid_samples


@router.post("/import/preview", response_model=ImportPreviewOut)
def preview_import(
    file: UploadFile,
    site_code: str | None = Form(default=None),
    month: str | None = Form(default=None),
    tenant_code: str | None = Form(default=None),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not (can_manage_schedule(user["role"]) or _can_use_support_roundtrip_source(user)):
        raise HTTPException(status_code=403, detail="forbidden")

    raw_bytes = file.file.read()
    detected_filename = str(file.filename or "").lower()
    is_xlsx = detected_filename.endswith(".xlsx")

    if is_xlsx:
        try:
            workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=False, data_only=False)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="invalid import file") from exc
        try:
            if ARLS_SHEET_NAME in workbook.sheetnames:
                effective_site_code = _resolve_scoped_schedule_site_code(user, request_site_code=site_code)
                if not effective_site_code:
                    raise HTTPException(status_code=400, detail="site_code is required for ARLS import")
                if not month or not str(month).strip():
                    raise HTTPException(status_code=400, detail="month is required for ARLS import")
                _month_bounds(str(month).strip())
                target_tenant = _resolve_target_tenant(conn, user, tenant_code)
                scope_site = _resolve_site_context_by_code(
                    conn,
                    tenant_id=str(target_tenant["id"]),
                    site_code=effective_site_code,
                )
                if not scope_site:
                    raise HTTPException(status_code=404, detail="site not found")
                preview = _build_schedule_import_preview_result(
                    conn,
                    workbook=workbook,
                    target_tenant=target_tenant,
                    scope_site=scope_site,
                    selected_month=str(month).strip(),
                    user=user,
                    filename=file.filename or "schedule_import.xlsx",
                )
                batch_id, errors = _persist_schedule_import_preview_batch(
                    conn,
                    preview=preview,
                    target_tenant=target_tenant,
                    scope_site=scope_site,
                    month_key=str(month).strip(),
                    filename=file.filename or "schedule_import.xlsx",
                    user=user,
                )
                total = int(preview["total_rows"])
                valid = int(preview["valid_rows"])
                invalid = int(preview["invalid_rows"])
                preview_rows = [
                    ImportPreviewRowOut(
                        row_no=int(row.get("row_no") or 0),
                        tenant_code=str(row.get("tenant_code") or "").strip(),
                        company_code=str(row.get("company_code") or "").strip(),
                        site_code=str(row.get("site_code") or "").strip(),
                        employee_code=str(row.get("employee_code") or "").strip(),
                        employee_name=str(row.get("employee_name") or "").strip(),
                        schedule_date=(row.get("schedule_date").isoformat() if isinstance(row.get("schedule_date"), date) else str(row.get("schedule_date") or "")),
                        shift_type=str(row.get("shift_type") or "").strip(),
                        duty_type=_schedule_template_duty_label(row.get("duty_type")),
                        source_sheet=str(row.get("source_sheet") or "").strip() or None,
                        source_col=str(row.get("source_col") or "").strip() or None,
                        source_block=str(row.get("source_block") or "").strip() or None,
                        section_label=str(row.get("section_label") or "").strip() or None,
                        template_id=str(row.get("template_id") or "").strip() or None,
                        template_name=str(row.get("template_name") or "").strip() or None,
                        work_value=str(row.get("work_value") or "").strip() or None,
                        current_work_value=str(row.get("current_work_value") or "").strip() or None,
                        parsed_semantic_type=str(row.get("parsed_semantic_type") or "").strip() or None,
                        mapped_hours=float(row["mapped_hours"]) if row.get("mapped_hours") is not None else None,
                        mapping_key=str(row.get("mapping_key") or "").strip() or None,
                        status_label=_import_diff_status_label(
                            diff_category=str(row.get("diff_category") or "").strip(),
                            validation_code=str(row.get("validation_code") or "").strip() or None,
                            is_blocking=bool(row.get("is_blocking")),
                        ),
                        is_valid=not bool(row.get("is_blocking")),
                        is_blocking=bool(row.get("is_blocking")),
                        diff_category=str(row.get("diff_category") or "").strip() or None,
                        apply_action=str(row.get("apply_action") or "").strip() or None,
                        is_protected=bool(row.get("is_protected")),
                        protected_reason=str(row.get("protected_reason") or "").strip() or None,
                        validation_code=str(row.get("validation_code") or "").strip() or None,
                        validation_error=str(row.get("validation_error") or "").strip() or None,
                    )
                    for row in (preview.get("preview_rows") or [])
                ]

                return ImportPreviewOut(
                    batch_id=batch_id,
                    total_rows=total,
                    valid_rows=valid,
                    invalid_rows=invalid,
                    applicable_rows=int(preview.get("applicable_rows") or valid),
                    unchanged_rows=int(preview.get("unchanged_rows") or 0),
                    blocked_rows=int(preview.get("blocked_rows") or invalid),
                    warning_rows=int(preview.get("warning_rows") or 0),
                    invalid_samples=errors,
                    preview_rows=preview_rows,
                    error_counts=dict(preview["error_counts"]),
                    diff_counts=dict(preview["diff_counts"]),
                    issues=list(preview.get("issues") or []),
                    blocked_reasons=list(preview["blocked_reasons"]),
                    metadata=preview["metadata"],
                    can_apply=bool(preview["can_apply"]),
                )
        finally:
            workbook.close()

    # Legacy flat import path (CSV/legacy XLSX header)
    try:
        detected_format, raw_headers, rows = _read_import_rows(file, raw_bytes)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid import file") from exc
    batch_id = uuid.uuid4()
    total = 0
    valid = 0
    invalid = 0
    errors: list[str] = []
    preview_rows: list[ImportPreviewRowOut] = []
    error_counts: Counter[str] = Counter()
    seen_employee_date: set[tuple[str, str, str]] = set()

    required_headers = set(_schedule_import_headers())
    if (
        not raw_headers
        or len(raw_headers) != len(set(raw_headers))
        or set(raw_headers) != required_headers
    ):
        raise HTTPException(status_code=400, detail="invalid import header")

    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO schedule_import_batches (id, tenant_id, created_by, filename) VALUES (%s, %s, %s, %s)",
            (batch_id, user["tenant_id"], user["id"], file.filename or f"schedule_import.{detected_format}"),
        )

        for row_no, row in enumerate(rows, start=1):
            total += 1

            normalized_row = _normalize_import_row(row)
            normalized_tenant_code = normalized_row.get("tenant_code", "")
            company_code = normalized_row.get("company_code", "")
            normalized_site_code = normalized_row.get("site_code", "")
            employee_code = normalized_row.get("employee_code", "")
            shift_type_raw = normalized_row.get("shift_type", "")
            shift_type = _normalize_shift_type(shift_type_raw)
            schedule_date_raw = normalized_row.get("schedule_date", "")

            validation_code = None
            validation_error = None
            schedule_date = None
            if normalized_tenant_code != user["tenant_code"] and not is_super_admin(user["role"]):
                validation_code = "tenant_code_mismatch"
                validation_error = VALIDATION_MESSAGES["tenant_code_mismatch"]
            if not (
                normalized_tenant_code
                and company_code
                and normalized_site_code
                and employee_code
                and schedule_date_raw
                and shift_type_raw
            ):
                validation_code = "required_column_missing"
                validation_error = VALIDATION_MESSAGES["required_column_missing"]
            elif shift_type not in ALLOWED_SHIFT_TYPES:
                validation_code = "invalid_shift_type"
                validation_error = VALIDATION_MESSAGES["invalid_shift_type"]
            else:
                schedule_date = _parse_date_or_none(schedule_date_raw)
                if not schedule_date:
                    validation_code = "invalid_schedule_date"
                    validation_error = VALIDATION_MESSAGES["invalid_schedule_date"]

            refs = None
            if not validation_code:
                refs, validation_code, validation_error = _resolve_import_refs(
                    cur,
                    user,
                    tenant_code=normalized_tenant_code,
                    company_code=company_code,
                    site_code=normalized_site_code,
                    employee_code=employee_code,
                )

            if not validation_code and refs and schedule_date:
                normalized_shift_type = _normalize_shift_type(shift_type or shift_type_raw)
                dedup_key = (
                    str(refs["tenant_id"]),
                    str(refs["employee_id"]),
                    schedule_date.isoformat(),
                    normalized_shift_type,
                )
                if dedup_key in seen_employee_date:
                    validation_code = "time_conflict"
                    validation_error = "파일 내 중복 스케줄입니다."
                else:
                    cur.execute(
                        """
                        SELECT 1
                        FROM monthly_schedules
                        WHERE tenant_id = %s
                          AND employee_id = %s
                          AND schedule_date = %s
                          AND lower(COALESCE(shift_type, 'day')) = lower(%s)
                        """,
                        (refs["tenant_id"], refs["employee_id"], schedule_date, normalized_shift_type),
                    )
                    if cur.fetchone():
                        validation_code = "time_conflict"
                        validation_error = "이미 등록된 스케줄과 충돌합니다."
                    else:
                        seen_employee_date.add(dedup_key)

            cur.execute(
                """
                INSERT INTO schedule_import_rows
                (id, batch_id, row_no, tenant_code, company_code, site_code, employee_code,
                 schedule_date, shift_type, validation_error, employee_id, company_id, site_id,
                 tenant_id, is_valid)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    uuid.uuid4(),
                    batch_id,
                    row_no,
                    normalized_tenant_code,
                    company_code,
                    normalized_site_code,
                    employee_code,
                    schedule_date,
                    shift_type or shift_type_raw,
                    validation_error,
                    refs["employee_id"] if refs else None,
                    refs["company_id"] if refs else None,
                    refs["site_id"] if refs else None,
                    refs["tenant_id"] if refs else user["tenant_id"],
                    validation_code is None,
                ),
            )

            preview_rows.append(
                ImportPreviewRowOut(
                    row_no=row_no,
                    tenant_code=normalized_tenant_code,
                    company_code=company_code,
                    site_code=normalized_site_code,
                    employee_code=employee_code,
                    schedule_date=schedule_date_raw or (schedule_date.isoformat() if schedule_date else None),
                    shift_type=shift_type or shift_type_raw,
                    status_label=_import_status_label(validation_code),
                    is_valid=validation_code is None,
                    validation_code=validation_code,
                    validation_error=validation_error,
                )
            )

            if validation_code:
                invalid += 1
                error_counts[validation_code] += 1
                if len(errors) < 20:
                    errors.append(f"row {row_no}: {validation_error}")
            else:
                valid += 1

        cur.execute(
            "UPDATE schedule_import_batches SET total_rows = %s, valid_rows = %s, invalid_rows = %s WHERE id = %s",
            (total, valid, invalid, batch_id),
        )

    return ImportPreviewOut(
        batch_id=batch_id,
        total_rows=total,
        valid_rows=valid,
        invalid_rows=invalid,
        invalid_samples=errors,
        preview_rows=preview_rows[:IMPORT_PREVIEW_LIMIT],
        error_counts=dict(error_counts),
    )


def _normalize_schedule_import_payload_date(value: object) -> date | None:
    if isinstance(value, date):
        return value
    return _parse_date_or_none(str(value or "").strip())


def _coerce_int_or_none(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _load_schedule_import_payload_rows(conn, *, batch_id: uuid.UUID | str) -> list[dict[str, Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT payload_json
            FROM schedule_import_rows
            WHERE batch_id = %s
            ORDER BY row_no, source_block, schedule_date NULLS LAST
            """,
            (str(batch_id),),
        )
        return [dict(row.get("payload_json") or {}) for row in (cur.fetchall() or [])]


def _load_daytime_need_count_rows_for_apply(
    conn,
    *,
    tenant_id: str,
    site_id: str,
    month_key: str,
) -> list[dict[str, Any]]:
    start_date, end_date = _month_bounds(month_key)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT work_date, required_count, raw_text, source
            FROM site_daytime_need_counts
            WHERE tenant_id = %s
              AND site_id = %s
              AND work_date >= %s
              AND work_date < %s
            ORDER BY work_date ASC
            """,
            (tenant_id, site_id, start_date, end_date),
        )
        return [dict(row) for row in cur.fetchall()]


def _is_daytime_need_base_source(source: object) -> bool:
    return str(source or "").strip().lower() in ARLS_DAYTIME_NEED_SOURCE_ALIASES


def _schedule_import_current_row_matches(current_row: dict[str, Any], desired_row: dict[str, Any]) -> bool:
    return (
        _normalize_shift_type(current_row.get("shift_type")) == _normalize_shift_type(desired_row.get("shift_type"))
        and str(current_row.get("template_id") or "").strip() == str(desired_row.get("template_id") or "").strip()
        and _normalize_time_text(current_row.get("shift_start_time")) == _normalize_time_text(desired_row.get("shift_start_time"))
        and _normalize_time_text(current_row.get("shift_end_time")) == _normalize_time_text(desired_row.get("shift_end_time"))
        and _coerce_float_or_none(current_row.get("paid_hours")) == _coerce_float_or_none(desired_row.get("paid_hours"))
        and str(current_row.get("schedule_note") or "").strip() == str(desired_row.get("schedule_note") or "").strip()
    )


def _build_sentrix_support_request_event_id(
    *,
    batch_id: str,
    work_date: date,
    shift_kind: str,
    event_type: str,
) -> str:
    return f"arls-monthly-base:{batch_id}:{work_date.isoformat()}:{shift_kind}:{event_type}"


def _log_schedule_import_integration_event(
    cur,
    *,
    tenant_id: str,
    site_code: str,
    batch_id: str,
    work_date: date,
    shift_kind: str,
    event_type: str,
    payload: dict[str, Any],
    status: str = "SUCCESS",
    error_message: str | None = None,
) -> None:
    payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    cur.execute(
        """
        INSERT INTO integration_event_log (
            id, source, event_id, event_type, tenant_id, site_id,
            occurred_at, payload_digest, status, error_message, created_at
        )
        VALUES (
            %s, 'ARLS', %s, %s, %s, %s,
            timezone('utc', now()), %s, %s, %s, timezone('utc', now())
        )
        ON CONFLICT (event_id)
        DO UPDATE SET
            payload_digest = EXCLUDED.payload_digest,
            status = EXCLUDED.status,
            error_message = EXCLUDED.error_message,
            created_at = timezone('utc', now())
        """,
        (
            uuid.uuid4(),
            _build_sentrix_support_request_event_id(
                batch_id=batch_id,
                work_date=work_date,
                shift_kind=shift_kind,
                event_type=event_type,
            ),
            event_type,
            tenant_id,
            site_code,
            hashlib.sha256(payload_json.encode("utf-8")).hexdigest()[:16],
            status,
            error_message,
        ),
    )


def _upsert_sentrix_support_request_ticket_row(
    cur,
    *,
    tenant_id: str,
    site_id: str,
    site_code: str,
    month_key: str,
    work_date: date,
    shift_kind: str,
    request_count: int,
    work_purpose: str | None,
    batch_id: str,
    source_revision: str | None,
    detail_json: dict[str, Any],
) -> None:
    cur.execute(
        """
        INSERT INTO sentrix_support_request_tickets (
            id, tenant_id, site_id, site_code, month_key, work_date, shift_kind,
            request_count, work_purpose, status, source_workflow, source_batch_id,
            source_revision, detail_json, retracted_at, created_at, updated_at
        )
        VALUES (
            %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s::jsonb, NULL, timezone('utc', now()), timezone('utc', now())
        )
        ON CONFLICT (tenant_id, site_id, work_date, shift_kind, source_workflow)
        DO UPDATE SET
            site_code = EXCLUDED.site_code,
            month_key = EXCLUDED.month_key,
            request_count = EXCLUDED.request_count,
            work_purpose = EXCLUDED.work_purpose,
            status = EXCLUDED.status,
            source_batch_id = EXCLUDED.source_batch_id,
            source_revision = EXCLUDED.source_revision,
            detail_json = EXCLUDED.detail_json,
            retracted_at = NULL,
            updated_at = timezone('utc', now())
        """,
        (
            uuid.uuid4(),
            tenant_id,
            site_id,
            site_code,
            month_key,
            work_date,
            shift_kind,
            max(0, int(request_count)),
            str(work_purpose or "").strip() or None,
            SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS,
            SENTRIX_SUPPORT_REQUEST_WORKFLOW,
            batch_id,
            source_revision,
            json.dumps(detail_json or {}, ensure_ascii=False, default=str),
        ),
    )


def _retract_sentrix_support_request_ticket_row(
    cur,
    *,
    tenant_id: str,
    site_id: str,
    work_date: date,
    shift_kind: str,
    batch_id: str,
    source_revision: str | None,
) -> None:
    cur.execute(
        """
        UPDATE sentrix_support_request_tickets
        SET status = %s,
            source_batch_id = %s,
            source_revision = %s,
            retracted_at = timezone('utc', now()),
            updated_at = timezone('utc', now())
        WHERE tenant_id = %s
          AND site_id = %s
          AND work_date = %s
          AND shift_kind = %s
          AND source_workflow = %s
        """,
        (
            SENTRIX_SUPPORT_REQUEST_RETRACTED_STATUS,
            batch_id,
            source_revision,
            tenant_id,
            site_id,
            work_date,
            shift_kind,
            SENTRIX_SUPPORT_REQUEST_WORKFLOW,
        ),
    )


def _write_schedule_import_batch_apply_audit(
    cur,
    *,
    batch_id: uuid.UUID | str,
    status: str,
    user: dict,
    result_json: dict[str, Any],
    error_text: str | None = None,
) -> None:
    cur.execute(
        """
        UPDATE schedule_import_batches
        SET status = %s,
            applied_by = %s,
            applied_role = %s,
            completed_at = timezone('utc', now()),
            error_text = %s,
            apply_result_json = %s::jsonb
        WHERE id = %s
        """,
        (
            status,
            user["id"],
            normalize_role(user.get("role")),
            error_text,
            json.dumps(result_json or {}, ensure_ascii=False, default=str),
            str(batch_id),
        ),
    )


def _apply_canonical_schedule_import_batch(
    conn,
    *,
    batch_id: uuid.UUID,
    batch: dict[str, Any],
    target_tenant: dict,
    user: dict,
) -> tuple[ImportApplyOut, set[tuple[str, str, str]]]:
    site_code = str(batch.get("site_code") or "").strip()
    month_key = str(batch.get("month_key") or "").strip()
    site_row = _resolve_site_context_by_code(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_code=site_code,
    )
    blocked_reasons = list(dict.fromkeys(str(item).strip() for item in (batch.get("blocked_reasons_json") or []) if str(item).strip()))
    metadata = dict(batch.get("metadata_json") or {})
    if not site_row:
        blocked_reasons.append("지점 정보를 찾을 수 없어 적용할 수 없습니다.")
    if not metadata.get("workbook_valid"):
        blocked_reasons.append("검증된 workbook metadata가 없어 적용할 수 없습니다.")
    current_revision = _build_schedule_export_revision(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str((site_row or {}).get("id") or batch.get("site_id") or ""),
        site_code=site_code,
        month_key=month_key,
    ) if site_row and month_key else ""
    if current_revision and str(batch.get("current_revision") or "").strip() and current_revision != str(batch.get("current_revision") or "").strip():
        blocked_reasons.append("분석 이후 현재 월간 기준본이 변경되어 다시 분석해야 합니다.")
    batch_mapping_profile_id = str(batch.get("mapping_profile_id") or "").strip()
    if batch_mapping_profile_id:
        current_profile = _fetch_active_schedule_import_mapping_profile(conn, tenant_id=str(target_tenant["id"]))
        if not current_profile or str(current_profile.get("profile_id") or "").strip() != batch_mapping_profile_id:
            blocked_reasons.append("근무 템플릿 매핑 프로필이 변경되어 다시 분석해야 합니다.")
        elif batch.get("mapping_profile_updated_at") and current_profile.get("updated_at") and batch.get("mapping_profile_updated_at") != current_profile.get("updated_at"):
            blocked_reasons.append("근무 템플릿 매핑 프로필이 변경되어 다시 분석해야 합니다.")
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM schedule_import_rows
            WHERE batch_id = %s
              AND is_blocking = TRUE
            """,
            (str(batch_id),),
        )
        row = cur.fetchone()
        if int((row or {}).get("cnt") or 0) > 0:
            blocked_reasons.append("분석 결과에 차단 이슈가 남아 있어 적용할 수 없습니다.")
    payload_rows = _load_schedule_import_payload_rows(conn, batch_id=batch_id)
    body_payloads = [row for row in payload_rows if str(row.get("source_block") or "").strip() == "body"]
    support_ticket_payloads = [row for row in payload_rows if str(row.get("source_block") or "").strip() == "sentrix_support_ticket"]
    if blocked_reasons:
        result = ImportApplyOut(
            batch_id=batch_id,
            upload_batch_id=batch_id,
            applied=0,
            skipped=0,
            applied_rows=[],
            skipped_rows=[],
            blocked=True,
            blocked_reasons=blocked_reasons,
            apply_status="blocked",
            audit_timestamp=datetime.now(timezone.utc),
            blocking_failures=blocked_reasons,
        )
        with conn.cursor() as cur:
            _write_schedule_import_batch_apply_audit(
                cur,
                batch_id=batch_id,
                status="blocked",
                user=user,
                result_json=result.model_dump(mode="json"),
                error_text="; ".join(blocked_reasons[:5]) or None,
            )
        return result, set()

    employees = _load_site_employees(conn, tenant_id=str(target_tenant["id"]), site_id=str(site_row["id"]))
    employee_index = _build_employee_name_index(employees)
    existing_schedule_rows = _load_existing_schedule_rows_for_import(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month_key,
    )
    existing_schedule_index = _build_existing_schedule_row_index(existing_schedule_rows)
    current_need_rows = _load_daytime_need_count_rows_for_apply(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month_key,
    )
    current_need_index = {
        row.get("work_date").isoformat(): dict(row)
        for row in current_need_rows
        if isinstance(row.get("work_date"), date)
    }
    current_ticket_rows = _read_monthly_support_request_rows_for_export(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site_row["id"]),
        month_key=month_key,
    )
    current_ticket_index = _build_support_request_ticket_index(current_ticket_rows)

    desired_base_rows: dict[tuple[str, str, str], dict[str, Any]] = {}
    desired_ticket_rows: dict[tuple[str, str], dict[str, Any]] = {}
    desired_day_need_rows: dict[str, dict[str, Any]] = {}
    blocking_failures: list[str] = []
    skipped_rows: list[ImportApplyRowOut] = []

    def append_blocking_failure(message: str, *, payload: dict[str, Any] | None = None, shift_type: str = "") -> None:
        if message not in blocking_failures:
            blocking_failures.append(message)
        if payload and len(skipped_rows) < IMPORT_REPORT_LIMIT:
            schedule_date = _normalize_schedule_import_payload_date(payload.get("schedule_date"))
            skipped_rows.append(
                ImportApplyRowOut(
                    row_no=int(payload.get("row_no") or 0),
                    employee_code=str(payload.get("employee_code") or "").strip(),
                    site_code=site_code,
                    schedule_date=schedule_date.isoformat() if isinstance(schedule_date, date) else None,
                    shift_type=shift_type or str(payload.get("shift_type") or "").strip(),
                    section_label=str(payload.get("section_label") or "").strip() or None,
                    status="blocked",
                    reason=message,
                )
            )

    for payload in body_payloads:
        if bool(payload.get("is_blocking")):
            append_blocking_failure(str(payload.get("validation_error") or "차단 행이 남아 있어 적용할 수 없습니다.").strip(), payload=payload)
            continue
        schedule_date = _normalize_schedule_import_payload_date(payload.get("schedule_date"))
        duty_type = _normalize_schedule_template_duty_type(payload.get("duty_type"))
        workbook_value = str(payload.get("work_value") or "").strip()
        if not isinstance(schedule_date, date) or duty_type not in {"day", "overtime", "night"}:
            continue
        employee_row, employee_issue_code, employee_issue_message = _resolve_import_employee_match(
            employee_index,
            employee_name=str(payload.get("employee_name") or "").strip(),
            schedule_date=schedule_date,
        )
        if employee_issue_code or not employee_row:
            append_blocking_failure(employee_issue_message or "직원을 매칭할 수 없습니다.", payload=payload, shift_type=str(payload.get("shift_type") or "").strip())
            continue
        if not workbook_value:
            continue
        shift_type = _normalize_shift_type(payload.get("shift_type") or _resolve_shift_type_from_duty_type(duty_type))
        if shift_type not in ALLOWED_SHIFT_TYPES:
            append_blocking_failure("지원하지 않는 shift_type입니다.", payload=payload, shift_type=shift_type)
            continue
        if str(payload.get("parsed_semantic_type") or "").strip() == "numeric_hours" and not str(payload.get("template_id") or "").strip():
            append_blocking_failure("근무 템플릿 매핑이 없어 적용할 수 없습니다.", payload=payload, shift_type=shift_type)
            continue
        key = (str(employee_row.get("id") or "").strip(), schedule_date.isoformat(), duty_type)
        if key in desired_base_rows:
            append_blocking_failure("같은 직원/날짜/근무유형이 업로드 결과에 중복되어 있습니다.", payload=payload, shift_type=shift_type)
            continue
        existing_rows = existing_schedule_index.get(key) or []
        foreign_rows = [item for item in existing_rows if not _is_monthly_base_schedule_source(item.get("source"))]
        if foreign_rows:
            append_blocking_failure("같은 슬롯에 다른 lineage의 일정이 있어 base upload로 덮어쓸 수 없습니다.", payload=payload, shift_type=shift_type)
            continue
        desired_base_rows[key] = {
            "row_no": int(payload.get("row_no") or 0),
            "tenant_id": str(target_tenant["id"]),
            "company_id": str(employee_row.get("company_id") or site_row.get("company_id") or "").strip(),
            "site_id": str(site_row["id"]),
            "employee_id": str(employee_row["id"]),
            "employee_code": str(employee_row.get("employee_code") or "").strip(),
            "employee_name": str(employee_row.get("full_name") or payload.get("employee_name") or "").strip(),
            "schedule_date": schedule_date,
            "shift_type": shift_type,
            "duty_type": duty_type,
            "template_id": str(payload.get("template_id") or "").strip() or None,
            "template_name": str(payload.get("template_name") or "").strip() or None,
            "shift_start_time": _normalize_time_text(payload.get("shift_start_time")),
            "shift_end_time": _normalize_time_text(payload.get("shift_end_time")),
            "paid_hours": _coerce_float_or_none(payload.get("paid_hours")),
            "schedule_note": workbook_value if shift_type in NON_WORKING_SHIFT_TYPES else None,
        }

    for payload in support_ticket_payloads:
        if bool(payload.get("is_blocking")):
            append_blocking_failure(str(payload.get("validation_error") or "지원 요청 차단 이슈가 남아 있습니다.").strip(), payload=payload, shift_type=str(payload.get("shift_type") or "").strip())
            continue
        schedule_date = _normalize_schedule_import_payload_date(payload.get("schedule_date"))
        if not isinstance(schedule_date, date):
            continue
        shift_kind = "night" if str(payload.get("shift_type") or "day").strip().lower() == "night" else "day"
        request_count = _coerce_int_or_none(payload.get("request_count"))
        if request_count is None:
            append_blocking_failure("지원 요청 인원 수를 해석할 수 없습니다.", payload=payload, shift_type=shift_kind)
            continue
        ticket_key = (schedule_date.isoformat(), shift_kind)
        if ticket_key in desired_ticket_rows:
            append_blocking_failure("같은 날짜/주야 지원 요청이 중복되었습니다.", payload=payload, shift_type=shift_kind)
            continue
        detail_json = dict(payload.get("detail_json") or {})
        if request_count > 0:
            desired_ticket_rows[ticket_key] = {
                "row_no": int(payload.get("row_no") or 0),
                "work_date": schedule_date,
                "shift_kind": shift_kind,
                "request_count": request_count,
                "work_purpose": str(payload.get("purpose_text") or "").strip() or None,
                "detail_json": detail_json,
            }
            if shift_kind == "day":
                desired_day_need_rows[schedule_date.isoformat()] = {
                    "work_date": schedule_date,
                    "required_count": request_count,
                    "raw_text": str(detail_json.get("required_count_raw") or payload.get("work_value") or "").strip() or str(request_count),
                }

    base_existing_by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for key, rows in existing_schedule_index.items():
        current_base_rows = [item for item in rows if _is_monthly_base_schedule_source(item.get("source"))]
        if current_base_rows:
            base_existing_by_key[key] = current_base_rows
        if len(current_base_rows) > 1:
            append_blocking_failure("기존 base lineage 일정이 중복되어 자동 반영할 수 없습니다.")

    if blocking_failures:
        result = ImportApplyOut(
            batch_id=batch_id,
            upload_batch_id=batch_id,
            applied=0,
            skipped=len(skipped_rows),
            applied_rows=[],
            skipped_rows=skipped_rows,
            blocked=True,
            blocked_reasons=blocking_failures,
            apply_status="blocked",
            audit_timestamp=datetime.now(timezone.utc),
            failed_items=len(skipped_rows),
            blocking_failures=blocking_failures,
        )
        with conn.cursor() as cur:
            _write_schedule_import_batch_apply_audit(
                cur,
                batch_id=batch_id,
                status="blocked",
                user=user,
                result_json=result.model_dump(mode="json"),
                error_text="; ".join(blocking_failures[:5]) or None,
            )
        return result, set()

    source_revision = str(batch.get("export_revision") or batch.get("current_revision") or "").strip() or None
    base_create_ops: list[dict[str, Any]] = []
    base_update_ops: list[dict[str, Any]] = []
    base_delete_ops: list[dict[str, Any]] = []
    need_upsert_ops: list[dict[str, Any]] = []
    need_delete_ops: list[dict[str, Any]] = []
    ticket_create_ops: list[dict[str, Any]] = []
    ticket_update_ops: list[dict[str, Any]] = []
    ticket_retract_ops: list[dict[str, Any]] = []

    for key, desired in desired_base_rows.items():
        current_rows = base_existing_by_key.get(key) or []
        if not current_rows:
            base_create_ops.append(desired)
            continue
        current_row = current_rows[0]
        if not _schedule_import_current_row_matches(current_row, desired):
            base_update_ops.append({**desired, "schedule_id": str(current_row.get("schedule_id") or "").strip()})

    for key, current_rows in base_existing_by_key.items():
        if key in desired_base_rows:
            continue
        for current_row in current_rows:
            base_delete_ops.append(dict(current_row))

    for date_key, desired in desired_day_need_rows.items():
        current_row = current_need_index.get(date_key)
        if not current_row:
            need_upsert_ops.append(desired)
            continue
        current_required = _coerce_int_or_none(current_row.get("required_count")) or 0
        current_raw = str(current_row.get("raw_text") or "").strip()
        if (
            not _is_daytime_need_base_source(current_row.get("source"))
            and (current_required != int(desired.get("required_count") or 0) or current_raw != str(desired.get("raw_text") or "").strip())
        ):
            append_blocking_failure("기존 필요인원 수가 다른 source에 의해 관리되고 있어 업로드로 변경할 수 없습니다.")
            continue
        if current_required != int(desired.get("required_count") or 0) or current_raw != str(desired.get("raw_text") or "").strip():
            need_upsert_ops.append(desired)

    for date_key, current_row in current_need_index.items():
        if date_key in desired_day_need_rows:
            continue
        if _is_daytime_need_base_source(current_row.get("source")):
            need_delete_ops.append(dict(current_row))

    for key, desired in desired_ticket_rows.items():
        current_ticket = current_ticket_index.get(key) or {}
        current_is_active = str(current_ticket.get("status") or SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS).strip() == SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS
        if not current_ticket:
            ticket_create_ops.append(desired)
        elif (
            not current_is_active
            or _coerce_int_or_none(current_ticket.get("request_count")) != int(desired.get("request_count") or 0)
            or str(current_ticket.get("work_purpose") or "").strip() != str(desired.get("work_purpose") or "").strip()
            or dict(current_ticket.get("detail_json") or {}) != dict(desired.get("detail_json") or {})
        ):
            ticket_update_ops.append(desired)

    for key, current_ticket in current_ticket_index.items():
        current_is_active = str(current_ticket.get("status") or SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS).strip() == SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS
        if current_is_active and key not in desired_ticket_rows:
            ticket_retract_ops.append(dict(current_ticket))

    if blocking_failures:
        result = ImportApplyOut(
            batch_id=batch_id,
            upload_batch_id=batch_id,
            applied=0,
            skipped=len(skipped_rows),
            applied_rows=[],
            skipped_rows=skipped_rows,
            blocked=True,
            blocked_reasons=blocking_failures,
            apply_status="blocked",
            audit_timestamp=datetime.now(timezone.utc),
            failed_items=len(skipped_rows),
            blocking_failures=blocking_failures,
        )
        with conn.cursor() as cur:
            _write_schedule_import_batch_apply_audit(
                cur,
                batch_id=batch_id,
                status="blocked",
                user=user,
                result_json=result.model_dump(mode="json"),
                error_text="; ".join(blocking_failures[:5]) or None,
            )
        return result, set()

    applied_rows: list[ImportApplyRowOut] = []
    affected_site_days: set[tuple[str, str, str]] = set()
    base_created = 0
    base_updated = 0
    base_removed = 0
    sentrix_created = 0
    sentrix_updated = 0
    sentrix_retracted = 0

    try:
        with conn.cursor() as cur:
            for op in base_create_ops:
                _insert_monthly_schedule_row(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    company_id=str(op.get("company_id") or site_row.get("company_id") or ""),
                    site_id=str(site_row["id"]),
                    employee_id=str(op["employee_id"]),
                    schedule_date=op["schedule_date"],
                    shift_type=str(op.get("shift_type") or ""),
                    template_id=op.get("template_id"),
                    shift_start_time=op.get("shift_start_time"),
                    shift_end_time=op.get("shift_end_time"),
                    paid_hours=op.get("paid_hours"),
                    schedule_note=op.get("schedule_note"),
                    source=ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                    source_batch_id=str(batch_id),
                    source_revision=source_revision,
                )
                base_created += 1
                affected_site_days.add((str(target_tenant["id"]), str(site_row["id"]), op["schedule_date"].isoformat()))
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=int(op.get("row_no") or 0),
                            employee_code=str(op.get("employee_code") or ""),
                            site_code=site_code,
                            schedule_date=op["schedule_date"].isoformat(),
                            shift_type=str(op.get("shift_type") or ""),
                            section_label=_schedule_template_duty_label(op.get("duty_type")),
                            status="applied",
                            reason="기본 스케줄 생성",
                        )
                    )
            for op in base_update_ops:
                _update_monthly_schedule_row(
                    cur,
                    schedule_id=str(op.get("schedule_id") or ""),
                    shift_type=str(op.get("shift_type") or ""),
                    template_id=op.get("template_id"),
                    shift_start_time=op.get("shift_start_time"),
                    shift_end_time=op.get("shift_end_time"),
                    paid_hours=op.get("paid_hours"),
                    schedule_note=op.get("schedule_note"),
                    source=ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                    source_batch_id=str(batch_id),
                    source_revision=source_revision,
                )
                base_updated += 1
                affected_site_days.add((str(target_tenant["id"]), str(site_row["id"]), op["schedule_date"].isoformat()))
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=int(op.get("row_no") or 0),
                            employee_code=str(op.get("employee_code") or ""),
                            site_code=site_code,
                            schedule_date=op["schedule_date"].isoformat(),
                            shift_type=str(op.get("shift_type") or ""),
                            section_label=_schedule_template_duty_label(op.get("duty_type")),
                            status="applied",
                            reason="기본 스케줄 수정",
                        )
                    )
            for op in base_delete_ops:
                _delete_monthly_schedule_row(cur, schedule_id=str(op.get("schedule_id") or ""))
                base_removed += 1
                schedule_date = op.get("schedule_date")
                if isinstance(schedule_date, date):
                    affected_site_days.add((str(target_tenant["id"]), str(site_row["id"]), schedule_date.isoformat()))
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=0,
                            employee_code=str(op.get("employee_code") or ""),
                            site_code=site_code,
                            schedule_date=schedule_date.isoformat() if isinstance(schedule_date, date) else None,
                            shift_type=str(op.get("shift_type") or ""),
                            section_label=_schedule_template_duty_label(_resolve_import_slot_from_schedule_row(op)),
                            status="applied",
                            reason="기본 스케줄 삭제",
                        )
                    )

            for op in need_upsert_ops:
                _upsert_daytime_need_count_row(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    site_id=str(site_row["id"]),
                    work_date=op["work_date"],
                    required_count=int(op.get("required_count") or 0),
                    raw_text=str(op.get("raw_text") or "").strip() or None,
                    updated_by=str(user["id"]),
                    source=ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                    source_batch_id=str(batch_id),
                    source_revision=source_revision,
                )
            for op in need_delete_ops:
                work_date = op.get("work_date")
                if isinstance(work_date, date):
                    _delete_daytime_need_count_row(
                        cur,
                        tenant_id=str(target_tenant["id"]),
                        site_id=str(site_row["id"]),
                        work_date=work_date,
                    )

            for op in ticket_create_ops:
                _upsert_sentrix_support_request_ticket_row(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    site_id=str(site_row["id"]),
                    site_code=site_code,
                    month_key=month_key,
                    work_date=op["work_date"],
                    shift_kind=str(op.get("shift_kind") or "day"),
                    request_count=int(op.get("request_count") or 0),
                    work_purpose=op.get("work_purpose"),
                    batch_id=str(batch_id),
                    source_revision=source_revision,
                    detail_json=dict(op.get("detail_json") or {}),
                )
                sentrix_created += 1
                _log_schedule_import_integration_event(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    site_code=site_code,
                    batch_id=str(batch_id),
                    work_date=op["work_date"],
                    shift_kind=str(op.get("shift_kind") or "day"),
                    event_type="SENTRIX_SUPPORT_REQUEST_CREATED",
                    payload=op,
                )
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=int(op.get("row_no") or 0),
                            employee_code="",
                            site_code=site_code,
                            schedule_date=op["work_date"].isoformat(),
                            shift_type=str(op.get("shift_kind") or "day"),
                            section_label="Sentrix 지원 요청",
                            status="applied",
                            reason="Sentrix 지원 요청 생성",
                        )
                    )
            for op in ticket_update_ops:
                _upsert_sentrix_support_request_ticket_row(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    site_id=str(site_row["id"]),
                    site_code=site_code,
                    month_key=month_key,
                    work_date=op["work_date"],
                    shift_kind=str(op.get("shift_kind") or "day"),
                    request_count=int(op.get("request_count") or 0),
                    work_purpose=op.get("work_purpose"),
                    batch_id=str(batch_id),
                    source_revision=source_revision,
                    detail_json=dict(op.get("detail_json") or {}),
                )
                sentrix_updated += 1
                _log_schedule_import_integration_event(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    site_code=site_code,
                    batch_id=str(batch_id),
                    work_date=op["work_date"],
                    shift_kind=str(op.get("shift_kind") or "day"),
                    event_type="SENTRIX_SUPPORT_REQUEST_UPDATED",
                    payload=op,
                )
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=int(op.get("row_no") or 0),
                            employee_code="",
                            site_code=site_code,
                            schedule_date=op["work_date"].isoformat(),
                            shift_type=str(op.get("shift_kind") or "day"),
                            section_label="Sentrix 지원 요청",
                            status="applied",
                            reason="Sentrix 지원 요청 갱신",
                        )
                    )
            for op in ticket_retract_ops:
                work_date = op.get("work_date")
                if not isinstance(work_date, date):
                    continue
                shift_kind = "night" if str(op.get("shift_kind") or "day").strip().lower() == "night" else "day"
                _retract_sentrix_support_request_ticket_row(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    site_id=str(site_row["id"]),
                    work_date=work_date,
                    shift_kind=shift_kind,
                    batch_id=str(batch_id),
                    source_revision=source_revision,
                )
                sentrix_retracted += 1
                _log_schedule_import_integration_event(
                    cur,
                    tenant_id=str(target_tenant["id"]),
                    site_code=site_code,
                    batch_id=str(batch_id),
                    work_date=work_date,
                    shift_kind=shift_kind,
                    event_type="SENTRIX_SUPPORT_REQUEST_RETRACTED",
                    payload=op,
                )
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=0,
                            employee_code="",
                            site_code=site_code,
                            schedule_date=work_date.isoformat(),
                            shift_type=shift_kind,
                            section_label="Sentrix 지원 요청",
                            status="applied",
                            reason="Sentrix 지원 요청 철회",
                        )
                    )

            result = ImportApplyOut(
                batch_id=batch_id,
                upload_batch_id=batch_id,
                applied=base_created + base_updated + base_removed + sentrix_created + sentrix_updated + sentrix_retracted + len(need_upsert_ops) + len(need_delete_ops),
                skipped=0,
                applied_rows=applied_rows,
                skipped_rows=[],
                blocked=False,
                blocked_reasons=[],
                apply_status="applied",
                audit_timestamp=datetime.now(timezone.utc),
                base_schedule_created=base_created,
                base_schedule_updated=base_updated,
                base_schedule_removed=base_removed,
                sentrix_tickets_created=sentrix_created,
                sentrix_tickets_updated=sentrix_updated,
                sentrix_tickets_retracted=sentrix_retracted,
                failed_items=0,
                blocking_failures=[],
                partial_failures=[],
            )
            _write_schedule_import_batch_apply_audit(
                cur,
                batch_id=batch_id,
                status="applied",
                user=user,
                result_json=result.model_dump(mode="json"),
                error_text=None,
            )
        return result, affected_site_days
    except Exception as exc:
        logger.exception("[schedule][import-apply] canonical apply failed batch=%s", batch_id)
        try:
            conn.rollback()
        except Exception:
            logger.exception("[schedule][import-apply] rollback failed batch=%s", batch_id)
        failure_message = f"적용 중 오류가 발생했습니다: {str(exc).strip() or exc.__class__.__name__}"
        result = ImportApplyOut(
            batch_id=batch_id,
            upload_batch_id=batch_id,
            applied=0,
            skipped=0,
            applied_rows=[],
            skipped_rows=[],
            blocked=True,
            blocked_reasons=[failure_message],
            apply_status="failed",
            audit_timestamp=datetime.now(timezone.utc),
            failed_items=1,
            blocking_failures=[],
            partial_failures=[failure_message],
        )
        with conn.cursor() as cur:
            _write_schedule_import_batch_apply_audit(
                cur,
                batch_id=batch_id,
                status="failed",
                user=user,
                result_json=result.model_dump(mode="json"),
                error_text=failure_message,
            )
        return result, set()


@router.post("/import/{batch_id}/apply", response_model=ImportApplyOut)
def apply_import(
    batch_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not (can_manage_schedule(user["role"]) or _can_use_support_roundtrip_source(user)):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT status,
                   import_mode,
                   site_id,
                   site_code,
                   month_key,
                   metadata_error,
                   blocked_reasons_json,
                   is_stale,
                   current_revision,
                   export_revision,
                   metadata_json,
                   issues_json,
                   mapping_profile_id,
                   mapping_profile_name,
                   mapping_profile_updated_at
            FROM schedule_import_batches
            WHERE id = %s AND tenant_id = %s
            """,
            (batch_id, target_tenant["id"]),
        )
        batch = cur.fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="batch not found")
        _resolve_scoped_schedule_site_code(user, request_site_code=str(batch.get("site_code") or "").strip())
        if str(batch["status"]).lower() == "applied":
            raise HTTPException(status_code=409, detail="batch already applied")
        if str(batch.get("import_mode") or "").strip() == "canonical_workbook":
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch=batch,
                target_tenant=target_tenant,
                user=user,
            )
            partial_failures: list[str] = []
            if not result.blocked and str(batch.get("site_code") or "").strip() and str(batch.get("month_key") or "").strip() and _can_use_support_roundtrip_source(user):
                try:
                    site_row = _resolve_site_context_by_code(
                        conn,
                        tenant_id=str(target_tenant["id"]),
                        site_code=str(batch.get("site_code") or "").strip(),
                    )
                    if site_row:
                        _register_support_roundtrip_source_after_import(
                            conn,
                            batch_id=str(batch_id),
                            target_tenant=target_tenant,
                            site_row=site_row,
                            month_key=str(batch.get("month_key") or "").strip(),
                            user=user,
                        )
                except Exception:
                    logger.exception("[schedule][import-apply] support roundtrip source registration failed batch=%s", batch_id)
                    partial_failures.append("support roundtrip source registration failed")
            for tenant_id, site_id, schedule_date_raw in affected_site_days:
                try:
                    _refresh_daily_leader_defaults(
                        conn,
                        tenant_id=tenant_id,
                        site_id=site_id,
                        schedule_date=date.fromisoformat(schedule_date_raw),
                    )
                except Exception:
                    logger.exception("[schedule][import-apply] leader refresh failed tenant=%s site=%s date=%s", tenant_id, site_id, schedule_date_raw)
                    partial_failures.append(f"leader refresh failed:{schedule_date_raw}")
            if partial_failures and not result.blocked:
                result.apply_status = "partial_failed"
                result.partial_failures = list(dict.fromkeys([*result.partial_failures, *partial_failures]))
                with conn.cursor() as cur:
                    _write_schedule_import_batch_apply_audit(
                        cur,
                        batch_id=batch_id,
                        status="partial_failed",
                        user=user,
                        result_json=result.model_dump(mode="json"),
                        error_text="; ".join(result.partial_failures[:5]) or None,
                    )
            return result
        blocked_reasons = list(batch.get("blocked_reasons_json") or [])
        metadata_error = str(batch.get("metadata_error") or "").strip()
        if metadata_error:
            blocked_reasons.append(metadata_error)
        if bool(batch.get("is_stale")):
            blocked_reasons.append("구버전 파일이라 적용할 수 없습니다.")
        if blocked_reasons:
            return ImportApplyOut(
                batch_id=batch_id,
                applied=0,
                skipped=0,
                applied_rows=[],
                skipped_rows=[],
                blocked=True,
                blocked_reasons=list(dict.fromkeys(str(item).strip() for item in blocked_reasons if str(item).strip())),
            )

        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM schedule_import_rows
            WHERE batch_id = %s AND is_blocking = TRUE
            """,
            (batch_id,),
        )
        invalid_count_row = cur.fetchone()
        invalid_count = int(invalid_count_row["cnt"]) if invalid_count_row else 0

        cur.execute(
            """
            SELECT row_no, employee_code, site_code, schedule_date, shift_type, section_label, validation_error
            FROM schedule_import_rows
            WHERE batch_id = %s AND is_blocking = TRUE
            ORDER BY row_no
            LIMIT %s
            """,
            (batch_id, IMPORT_REPORT_LIMIT),
        )
        invalid_rows = cur.fetchall()

        cur.execute(
            """
            SELECT row_no, tenant_id, company_id, site_id, employee_id, employee_code, site_code, schedule_date, shift_type,
                   template_id, duty_type, shift_start_time, shift_end_time, paid_hours, template_name,
                   work_value, current_work_value, apply_action, source_block, section_label,
                   protected_reason, current_schedule_id
            FROM schedule_import_rows
            WHERE batch_id = %s
              AND is_blocking = FALSE
              AND apply_action IN ('create', 'update', 'delete', 'upsert_need_count', 'delete_need_count')
            ORDER BY row_no
            """,
            (batch_id,),
        )
        rows = cur.fetchall()

        applied = 0
        skipped = invalid_count
        applied_rows: list[ImportApplyRowOut] = []
        skipped_rows: list[ImportApplyRowOut] = []
        dedup_applied: set[tuple[str, str, str]] = set()
        affected_site_days: set[tuple[str, str, str]] = set()

        for row in invalid_rows:
            if len(skipped_rows) >= IMPORT_REPORT_LIMIT:
                break
            skipped_rows.append(
                ImportApplyRowOut(
                    row_no=row["row_no"],
                    employee_code=row["employee_code"],
                    site_code=row["site_code"],
                    schedule_date=str(row["schedule_date"]) if row["schedule_date"] else None,
                    shift_type=row["shift_type"],
                    section_label=row.get("section_label"),
                    status="invalid",
                    reason=row["validation_error"] or "미리보기 검증 실패",
                )
            )

        for row in rows:
            shift_type = _normalize_shift_type(row["shift_type"])
            apply_action = str(row.get("apply_action") or "").strip()
            source_block = str(row.get("source_block") or "").strip()
            if source_block == "daytime_need":
                if not (row["tenant_id"] and row["site_id"] and row["schedule_date"]):
                    skipped += 1
                    if len(skipped_rows) < IMPORT_REPORT_LIMIT:
                        skipped_rows.append(
                            ImportApplyRowOut(
                                row_no=row["row_no"],
                                employee_code=row["employee_code"],
                                site_code=row["site_code"],
                                schedule_date=str(row["schedule_date"]) if row["schedule_date"] else None,
                                shift_type="day",
                                section_label=row.get("section_label"),
                                status="skipped",
                                reason="need-count metadata missing",
                            )
                        )
                    continue
                if apply_action == "delete_need_count":
                    _delete_daytime_need_count_row(
                        cur,
                        tenant_id=str(row["tenant_id"]),
                        site_id=str(row["site_id"]),
                        work_date=row["schedule_date"],
                    )
                    applied += 1
                    if len(applied_rows) < IMPORT_REPORT_LIMIT:
                        applied_rows.append(
                            ImportApplyRowOut(
                                row_no=row["row_no"],
                                employee_code="",
                                site_code=row["site_code"],
                                schedule_date=str(row["schedule_date"]),
                                shift_type="day",
                                section_label=row.get("section_label"),
                                status="applied",
                                reason="필요인원 수 삭제",
                            )
                        )
                    continue
                required_count, parsed_text = _parse_daytime_need_value(row.get("work_value"))
                _upsert_daytime_need_count_row(
                    cur,
                    tenant_id=str(row["tenant_id"]),
                    site_id=str(row["site_id"]),
                    work_date=row["schedule_date"],
                    required_count=max(0, int(required_count or 0)),
                    raw_text=parsed_text or str(row.get("work_value") or "").strip() or None,
                    updated_by=str(user["id"]),
                )
                applied += 1
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=row["row_no"],
                            employee_code="",
                            site_code=row["site_code"],
                            schedule_date=str(row["schedule_date"]),
                            shift_type="day",
                            section_label=row.get("section_label"),
                            status="applied",
                            reason="필요인원 수 반영",
                        )
                    )
                continue

            if shift_type not in ALLOWED_SHIFT_TYPES:
                skipped += 1
                if len(skipped_rows) < IMPORT_REPORT_LIMIT:
                    skipped_rows.append(
                        ImportApplyRowOut(
                            row_no=row["row_no"],
                            employee_code=row["employee_code"],
                            site_code=row["site_code"],
                            schedule_date=str(row["schedule_date"]) if row["schedule_date"] else None,
                            shift_type=row["shift_type"],
                            section_label=row.get("section_label"),
                            status="skipped",
                            reason="invalid shift_type",
                        )
                    )
                continue

            if not (row["tenant_id"] and row["company_id"] and row["site_id"] and row["employee_id"] and row["schedule_date"]):
                skipped += 1
                if len(skipped_rows) < IMPORT_REPORT_LIMIT:
                    skipped_rows.append(
                        ImportApplyRowOut(
                            row_no=row["row_no"],
                            employee_code=row["employee_code"],
                            site_code=row["site_code"],
                            schedule_date=str(row["schedule_date"]) if row["schedule_date"] else None,
                            shift_type=shift_type,
                            section_label=row.get("section_label"),
                            status="skipped",
                            reason="lookup metadata missing",
                        )
                    )
                continue

            dedup_key = (str(row["tenant_id"]), str(row["employee_id"]), str(row["schedule_date"]), apply_action)
            if dedup_key in dedup_applied:
                skipped += 1
                if len(skipped_rows) < IMPORT_REPORT_LIMIT:
                    skipped_rows.append(
                        ImportApplyRowOut(
                            row_no=row["row_no"],
                            employee_code=row["employee_code"],
                            site_code=row["site_code"],
                            schedule_date=str(row["schedule_date"]),
                            shift_type=shift_type,
                            section_label=row.get("section_label"),
                            status="skipped",
                            reason="duplicate in batch",
                        )
                    )
                continue
            dedup_applied.add(dedup_key)

            if apply_action == "delete":
                current_schedule_id = str(row.get("current_schedule_id") or "").strip()
                if not current_schedule_id:
                    skipped += 1
                    if len(skipped_rows) < IMPORT_REPORT_LIMIT:
                        skipped_rows.append(
                            ImportApplyRowOut(
                                row_no=row["row_no"],
                                employee_code=row["employee_code"],
                                site_code=row["site_code"],
                                schedule_date=str(row["schedule_date"]),
                                shift_type=shift_type,
                                section_label=row.get("section_label"),
                                status="skipped",
                                reason="delete target missing",
                            )
                        )
                    continue
                _delete_monthly_schedule_row(cur, schedule_id=current_schedule_id)
                affected_site_days.add((str(row["tenant_id"]), str(row["site_id"]), str(row["schedule_date"])))
                applied += 1
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=row["row_no"],
                            employee_code=row["employee_code"],
                            site_code=row["site_code"],
                            schedule_date=str(row["schedule_date"]),
                            shift_type=shift_type,
                            section_label=row.get("section_label"),
                            status="applied",
                            reason="삭제",
                        )
                    )
                continue

            if apply_action == "update":
                current_schedule_id = str(row.get("current_schedule_id") or "").strip()
                if not current_schedule_id:
                    skipped += 1
                    if len(skipped_rows) < IMPORT_REPORT_LIMIT:
                        skipped_rows.append(
                            ImportApplyRowOut(
                                row_no=row["row_no"],
                                employee_code=row["employee_code"],
                                site_code=row["site_code"],
                                schedule_date=str(row["schedule_date"]),
                                shift_type=shift_type,
                                section_label=row.get("section_label"),
                                status="skipped",
                                reason="update target missing",
                            )
                        )
                    continue
                _update_monthly_schedule_row(
                    cur,
                    schedule_id=current_schedule_id,
                    shift_type=shift_type,
                    template_id=str(row["template_id"]) if row.get("template_id") else None,
                    shift_start_time=_normalize_time_text(row.get("shift_start_time")),
                    shift_end_time=_normalize_time_text(row.get("shift_end_time")),
                    paid_hours=float(row["paid_hours"]) if row.get("paid_hours") is not None else None,
                    schedule_note=str(row.get("work_value") or "").strip() if shift_type in NON_WORKING_SHIFT_TYPES else None,
                )
                affected_site_days.add((str(row["tenant_id"]), str(row["site_id"]), str(row["schedule_date"])))
                applied += 1
                if len(applied_rows) < IMPORT_REPORT_LIMIT:
                    applied_rows.append(
                        ImportApplyRowOut(
                            row_no=row["row_no"],
                            employee_code=row["employee_code"],
                            site_code=row["site_code"],
                            schedule_date=str(row["schedule_date"]),
                            shift_type=shift_type,
                            section_label=row.get("section_label"),
                            status="applied",
                            reason=str(row.get("template_name") or row.get("work_value") or "updated"),
                        )
                    )
                continue

            cur.execute(
                """
                SELECT 1 FROM monthly_schedules
                WHERE tenant_id = %s AND employee_id = %s AND schedule_date = %s
                  AND lower(COALESCE(shift_type, 'day')) = lower(%s)
                  AND (
                    (template_id IS NOT DISTINCT FROM %s)
                    OR COALESCE(template_id::text, '') = ''
                  )
                LIMIT 1
                """,
                (
                    row["tenant_id"],
                    row["employee_id"],
                    row["schedule_date"],
                    shift_type,
                    row.get("template_id"),
                ),
            )
            if cur.fetchone():
                skipped += 1
                if len(skipped_rows) < IMPORT_REPORT_LIMIT:
                    skipped_rows.append(
                        ImportApplyRowOut(
                            row_no=row["row_no"],
                            employee_code=row["employee_code"],
                            site_code=row["site_code"],
                            schedule_date=str(row["schedule_date"]),
                            shift_type=shift_type,
                            section_label=row.get("section_label"),
                            status="skipped",
                            reason="already exists",
                        )
                    )
                continue

            _insert_monthly_schedule_row(
                cur,
                tenant_id=str(row["tenant_id"]),
                company_id=str(row["company_id"]),
                site_id=str(row["site_id"]),
                employee_id=str(row["employee_id"]),
                schedule_date=row["schedule_date"],
                shift_type=shift_type,
                template_id=str(row["template_id"]) if row.get("template_id") else None,
                shift_start_time=_normalize_time_text(row.get("shift_start_time")),
                shift_end_time=_normalize_time_text(row.get("shift_end_time")),
                paid_hours=float(row["paid_hours"]) if row.get("paid_hours") is not None else None,
                schedule_note=str(row.get("work_value") or "").strip() if shift_type in NON_WORKING_SHIFT_TYPES else None,
            )
            affected_site_days.add((str(row["tenant_id"]), str(row["site_id"]), str(row["schedule_date"])))
            applied += 1
            if len(applied_rows) < IMPORT_REPORT_LIMIT:
                applied_rows.append(
                    ImportApplyRowOut(
                        row_no=row["row_no"],
                        employee_code=row["employee_code"],
                        site_code=row["site_code"],
                        schedule_date=str(row["schedule_date"]),
                        shift_type=shift_type,
                        section_label=row.get("section_label"),
                        status="applied",
                        reason=str(row.get("template_name") or "applied"),
                    )
                )

        cur.execute(
            "UPDATE schedule_import_batches SET status = 'applied', completed_at = timezone('utc', now()) WHERE id = %s",
            (batch_id,),
        )

    if (
        str(batch.get("import_mode") or "").strip() == "canonical_workbook"
        and str(batch.get("site_code") or "").strip()
        and str(batch.get("month_key") or "").strip()
        and _can_use_support_roundtrip_source(user)
    ):
        site_row = _resolve_site_context_by_code(
            conn,
            tenant_id=str(target_tenant["id"]),
            site_code=str(batch.get("site_code") or "").strip(),
        )
        if site_row:
            _register_support_roundtrip_source_after_import(
                conn,
                batch_id=str(batch_id),
                target_tenant=target_tenant,
                site_row=site_row,
                month_key=str(batch.get("month_key") or "").strip(),
                user=user,
            )

    for tenant_id, site_id, schedule_date_raw in affected_site_days:
        _refresh_daily_leader_defaults(
            conn,
            tenant_id=tenant_id,
            site_id=site_id,
            schedule_date=date.fromisoformat(schedule_date_raw),
        )

    return ImportApplyOut(
        batch_id=batch_id,
        applied=applied,
        skipped=skipped,
        applied_rows=applied_rows,
        skipped_rows=skipped_rows,
        blocked=False,
        blocked_reasons=[],
    )


@router.put("/monthly/{schedule_id}")
def update_monthly_schedule(
    schedule_id: uuid.UUID,
    payload: ScheduleUpdate,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    context = _fetch_schedule_context(conn, schedule_id, target_tenant["id"])
    if not context:
        raise HTTPException(status_code=404, detail="schedule not found")

    leader_field_provided = "leader_user_id" in payload.model_fields_set
    next_shift_type = _normalize_shift_type(payload.shift_type)
    if next_shift_type not in ALLOWED_SHIFT_TYPES:
        raise HTTPException(status_code=400, detail="shift_type invalid")

    next_leader_user_id = context.get("leader_user_id")
    if leader_field_provided:
        next_leader_user_id = payload.leader_user_id

    if next_shift_type in NON_WORKING_SHIFT_TYPES:
        next_leader_user_id = None
    elif next_leader_user_id:
        _assert_valid_leader_for_site_day(
            conn,
            tenant_id=context["tenant_id"],
            site_id=context["site_id"],
            schedule_date=context["schedule_date"],
            leader_user_id=str(next_leader_user_id),
        )

    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE monthly_schedules
            SET shift_type = %s,
                leader_user_id = %s
            WHERE id = %s AND tenant_id = %s
            """,
            (next_shift_type, next_leader_user_id, str(schedule_id), target_tenant["id"]),
        )

    recommended_user_id = _refresh_daily_leader_defaults(
        conn,
        tenant_id=context["tenant_id"],
        site_id=context["site_id"],
        schedule_date=context["schedule_date"],
    )

    return {
        "id": str(schedule_id),
        "shift_type": next_shift_type,
        "leader_user_id": str(next_leader_user_id) if next_leader_user_id else None,
        "recommended_leader_user_id": recommended_user_id,
    }


@router.post("/monthly/{schedule_id}/closer")
def set_monthly_schedule_closer(
    schedule_id: uuid.UUID,
    payload: ScheduleCloserUpdate,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    context = _fetch_schedule_context(conn, schedule_id, target_tenant["id"])
    if not context:
        raise HTTPException(status_code=404, detail="schedule not found")

    if payload.enabled and str(context.get("shift_type") or "").strip().lower() in NON_WORKING_SHIFT_TYPES:
        raise HTTPException(status_code=400, detail="closer must be an on-duty schedule row")

    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE monthly_schedules
            SET schedule_note = NULLIF(
                btrim(
                    regexp_replace(
                        regexp_replace(COALESCE(schedule_note, ''), '(\\[closer\\]|\\bcloser\\b)', '', 'gi'),
                        '\\s{2,}',
                        ' ',
                        'g'
                    )
                ),
                ''
            )
            WHERE tenant_id = %s
              AND site_id = %s
              AND schedule_date = %s
            """,
            (context["tenant_id"], context["site_id"], context["schedule_date"]),
        )

        if payload.enabled:
            cur.execute(
                """
                UPDATE monthly_schedules
                SET schedule_note = CASE
                    WHEN lower(COALESCE(schedule_note, '')) LIKE '%%closer%%' THEN schedule_note
                    WHEN COALESCE(schedule_note, '') = '' THEN '[closer]'
                    ELSE btrim(schedule_note || ' [closer]')
                END
                WHERE id = %s
                  AND tenant_id = %s
                  AND lower(shift_type) NOT IN ('off', 'holiday')
                RETURNING employee_id
                """,
                (str(schedule_id), target_tenant["id"]),
            )
            marked = cur.fetchone()
            if not marked:
                raise HTTPException(status_code=400, detail="failed to assign closer to schedule row")

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT e.employee_code,
                   e.full_name AS employee_name,
                   au.id AS closer_user_id,
                   au.username AS closer_username,
                   au.full_name AS closer_full_name
            FROM monthly_schedules ms
            JOIN employees e ON e.id = ms.employee_id
            LEFT JOIN arls_users au
                   ON au.tenant_id = ms.tenant_id
                  AND au.employee_id = ms.employee_id
                  AND au.is_active = TRUE
            WHERE ms.tenant_id = %s
              AND ms.site_id = %s
              AND ms.schedule_date = %s
              AND lower(COALESCE(ms.schedule_note, '')) LIKE '%%closer%%'
            ORDER BY e.employee_code
            LIMIT 1
            """,
            (context["tenant_id"], context["site_id"], context["schedule_date"]),
        )
        closer = cur.fetchone()

    return {
        "schedule_id": str(schedule_id),
        "enabled": bool(payload.enabled),
        "site_code": context["site_code"],
        "schedule_date": context["schedule_date"].isoformat(),
        "closer_employee_code": str(closer["employee_code"]) if closer else None,
        "closer_employee_name": str(closer["employee_name"]) if closer else None,
        "closer_user_id": str(closer["closer_user_id"]) if closer and closer.get("closer_user_id") else None,
        "closer_username": str(closer["closer_username"]) if closer and closer.get("closer_username") else None,
        "closer_full_name": str(closer["closer_full_name"]) if closer and closer.get("closer_full_name") else None,
    }


@router.delete("/monthly/{schedule_id}")
def delete_monthly_schedule(
    schedule_id: uuid.UUID,
    scope: str = Query(default="single", max_length=16),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    context = _fetch_schedule_context(conn, schedule_id, target_tenant["id"])
    if not context:
        raise HTTPException(status_code=404, detail="schedule not found")
    normalized_scope = _normalize_schedule_delete_scope(scope)

    target_rows = _fetch_schedule_delete_scope_rows(
        conn,
        tenant_id=context["tenant_id"],
        employee_id=context["employee_id"],
        schedule_date=context["schedule_date"],
        schedule_id=schedule_id,
        scope=normalized_scope,
    )
    if not target_rows:
        raise HTTPException(status_code=404, detail="schedule not found")

    with conn.cursor() as cur:
        if normalized_scope == "future":
            cur.execute(
                """
                DELETE FROM monthly_schedules
                WHERE tenant_id = %s
                  AND employee_id = %s
                  AND schedule_date >= %s
                """,
                (target_tenant["id"], context["employee_id"], context["schedule_date"]),
            )
        else:
            cur.execute(
                "DELETE FROM monthly_schedules WHERE id = %s AND tenant_id = %s",
                (str(schedule_id), target_tenant["id"]),
            )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="schedule not found")

    schedule_dates_by_site: dict[str, list[date]] = {}
    for row in target_rows:
        site_id = str(row.get("site_id") or "").strip()
        schedule_date = row.get("schedule_date")
        if not site_id or not isinstance(schedule_date, date):
            continue
        schedule_dates_by_site.setdefault(site_id, []).append(schedule_date)

    for site_id, schedule_dates in schedule_dates_by_site.items():
        _refresh_daily_leader_defaults_for_dates(
            conn,
            tenant_id=context["tenant_id"],
            site_id=site_id,
            schedule_dates=schedule_dates,
        )

    return {
        "deleted": True,
        "id": str(schedule_id),
        "scope": normalized_scope,
        "deleted_count": len(target_rows),
    }


def _parse_ymd_or_400(value: str, *, field_name: str = "date") -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} must be YYYY-MM-DD") from exc


def _resolve_site_or_404(conn, *, tenant_id, site_code: str):
    site = resolve_site(conn, tenant_id=tenant_id, site_code=site_code)
    if not site:
        raise HTTPException(status_code=404, detail="site not found")
    return site


def _normalize_support_status_shift_kind(value: object) -> str:
    normalized = str(value or "").strip().lower()
    return "night" if normalized == "night" else "day"


def _support_status_source_label(value: object) -> str:
    normalized = str(value or "").strip()
    upper = normalized.upper()
    if not normalized:
        return "-"
    if normalized == SENTRIX_SUPPORT_REQUEST_WORKFLOW:
        return "ARLS 원본 요청"
    if upper == "HQ_ROUNDTRIP":
        return "HQ 병합"
    if upper == "MANUAL":
        return "수동"
    if upper == "SENTRIX_TICKET":
        return "Sentrix 티켓"
    if upper == "ARLS_MONTHLY_BASE_UPLOAD":
        return "ARLS 원본"
    return normalized


def _support_status_worker_display_value(row: dict[str, Any]) -> str:
    worker_type = str(row.get("worker_type") or "").strip().upper()
    worker_name = str(row.get("employee_name") or row.get("name") or "").strip()
    affiliation = str(row.get("affiliation") or "").strip()
    if worker_type == "INTERNAL":
        return f"자체 {worker_name}".strip() if worker_name else "자체"
    if worker_type == "UNAVAILABLE":
        return f"지원불가 {worker_name}".strip() if worker_name else "지원불가"
    if affiliation and worker_name:
        return f"{affiliation} {worker_name}".strip()
    if worker_type in {"F", "BK"} and worker_name:
        return f"{worker_type} {worker_name}".strip()
    return worker_name or "-"


def _serialize_support_status_assignment_row(row: dict[str, Any]) -> SupportStatusAssignmentOut:
    worker_name = str(row.get("employee_name") or row.get("name") or "").strip()
    return SupportStatusAssignmentOut(
        id=row["id"],
        slot_index=int(row.get("slot_index") or 1),
        worker_type=str(row.get("worker_type") or "").strip().upper(),
        employee_id=row.get("employee_id"),
        employee_code=str(row.get("employee_code") or "").strip() or None,
        employee_name=str(row.get("employee_name") or "").strip() or None,
        worker_name=worker_name or "-",
        display_value=_support_status_worker_display_value(row),
        affiliation=str(row.get("affiliation") or "").strip() or None,
        source=str(row.get("source") or "").strip() or "-",
        source_ticket_id=row.get("source_ticket_id"),
        source_event_uid=str(row.get("source_event_uid") or "").strip() or None,
        created_at=row["created_at"],
        updated_at=row.get("updated_at"),
    )


def _load_support_status_workspace_rows(
    conn,
    *,
    tenant_id: str,
    tenant_code: str,
    month_key: str,
    site_id: str | None = None,
) -> list[SupportStatusWorkspaceRowOut]:
    start_date, end_date = _month_bounds(month_key)
    ticket_clauses = [
        "sr.tenant_id = %s",
        "sr.work_date >= %s",
        "sr.work_date < %s",
    ]
    ticket_params: list[Any] = [tenant_id, start_date, end_date]
    assignment_clauses = [
        "sa.tenant_id = %s",
        "sa.work_date >= %s",
        "sa.work_date < %s",
    ]
    assignment_params: list[Any] = [tenant_id, start_date, end_date]
    if site_id:
        ticket_clauses.append("sr.site_id = %s")
        ticket_params.append(site_id)
        assignment_clauses.append("sa.site_id = %s")
        assignment_params.append(site_id)

    with conn.cursor() as cur:
        cur.execute(
            f"""
            SELECT sr.site_id,
                   s.site_code,
                   s.site_name,
                   sr.work_date,
                   sr.shift_kind,
                   sr.request_count,
                   sr.work_purpose,
                   sr.status,
                   sr.source_workflow,
                   sr.source_batch_id,
                   sr.source_revision,
                   sr.updated_at
            FROM sentrix_support_request_tickets sr
            JOIN sites s ON s.id = sr.site_id
            WHERE {' AND '.join(ticket_clauses)}
            ORDER BY sr.work_date DESC, s.site_code ASC, sr.shift_kind ASC
            """,
            tuple(ticket_params),
        )
        ticket_rows = [dict(row) for row in (cur.fetchall() or [])]

        cur.execute(
            f"""
            SELECT sa.id,
                   sa.site_id,
                   s.site_code,
                   s.site_name,
                   sa.work_date,
                   sa.support_period,
                   sa.slot_index,
                   sa.worker_type,
                   sa.employee_id,
                   e.employee_code,
                   e.full_name AS employee_name,
                   sa.name,
                   sa.affiliation,
                   sa.source,
                   sa.source_ticket_id,
                   sa.source_event_uid,
                   sa.created_at,
                   sa.updated_at
            FROM support_assignment sa
            JOIN sites s ON s.id = sa.site_id
            LEFT JOIN employees e ON e.id = sa.employee_id
            WHERE {' AND '.join(assignment_clauses)}
            ORDER BY sa.work_date DESC, s.site_code ASC, sa.support_period ASC, sa.slot_index ASC
            """,
            tuple(assignment_params),
        )
        assignment_rows = [dict(row) for row in (cur.fetchall() or [])]

    bucket_index: dict[tuple[str, str, str], dict[str, Any]] = {}

    def ensure_bucket(*, site_code: str, site_name: str, work_date: date, shift_kind: str) -> dict[str, Any]:
        key = (site_code, work_date.isoformat(), shift_kind)
        bucket = bucket_index.get(key)
        if bucket:
            return bucket
        bucket = {
            "row_key": f"{site_code}:{work_date.isoformat()}:{shift_kind}",
            "tenant_code": tenant_code,
            "site_code": site_code,
            "site_name": site_name or site_code,
            "work_date": work_date,
            "shift_kind": shift_kind,
            "request_count": 0,
            "assigned_count": 0,
            "filled_count": 0,
            "request_status": None,
            "work_purpose": None,
            "source_workflow": None,
            "source_batch_id": None,
            "source_revision": None,
            "source_labels": [],
            "worker_display_values": [],
            "assignments": [],
            "has_request_ticket": False,
            "updated_at": None,
        }
        bucket_index[key] = bucket
        return bucket

    for ticket in ticket_rows:
        site_code = str(ticket.get("site_code") or "").strip().upper()
        if not site_code:
            continue
        shift_kind = _normalize_support_status_shift_kind(ticket.get("shift_kind"))
        work_date = ticket.get("work_date")
        if not isinstance(work_date, date):
            continue
        bucket = ensure_bucket(
            site_code=site_code,
            site_name=str(ticket.get("site_name") or site_code).strip(),
            work_date=work_date,
            shift_kind=shift_kind,
        )
        bucket["request_count"] = max(0, int(ticket.get("request_count") or 0))
        bucket["request_status"] = _extract_sentrix_ticket_hq_roster_status(ticket) or str(ticket.get("status") or "").strip() or None
        bucket["work_purpose"] = str(ticket.get("work_purpose") or "").strip() or None
        bucket["source_workflow"] = str(ticket.get("source_workflow") or "").strip() or None
        bucket["source_batch_id"] = ticket.get("source_batch_id")
        bucket["source_revision"] = str(ticket.get("source_revision") or "").strip() or None
        bucket["has_request_ticket"] = True
        source_label = _support_status_source_label(ticket.get("source_workflow"))
        if source_label not in bucket["source_labels"]:
            bucket["source_labels"].append(source_label)
        ticket_updated_at = ticket.get("updated_at")
        if ticket_updated_at and (
            not bucket["updated_at"] or ticket_updated_at > bucket["updated_at"]
        ):
            bucket["updated_at"] = ticket_updated_at

    for assignment in assignment_rows:
        site_code = str(assignment.get("site_code") or "").strip().upper()
        if not site_code:
            continue
        shift_kind = _normalize_support_status_shift_kind(assignment.get("support_period"))
        work_date = assignment.get("work_date")
        if not isinstance(work_date, date):
            continue
        bucket = ensure_bucket(
            site_code=site_code,
            site_name=str(assignment.get("site_name") or site_code).strip(),
            work_date=work_date,
            shift_kind=shift_kind,
        )
        display_value = _support_status_worker_display_value(assignment)
        if display_value:
            bucket["worker_display_values"].append(display_value)
        bucket["assignments"].append(_serialize_support_status_assignment_row(assignment))
        bucket["assigned_count"] = len(bucket["assignments"])
        bucket["filled_count"] = bucket["assigned_count"]
        source_label = _support_status_source_label(assignment.get("source"))
        if source_label not in bucket["source_labels"]:
            bucket["source_labels"].append(source_label)
        assignment_updated_at = assignment.get("updated_at") or assignment.get("created_at")
        if assignment_updated_at and (
            not bucket["updated_at"] or assignment_updated_at > bucket["updated_at"]
        ):
            bucket["updated_at"] = assignment_updated_at

    rows = [
        SupportStatusWorkspaceRowOut(
            row_key=str(bucket["row_key"]),
            tenant_code=str(bucket["tenant_code"]),
            site_code=str(bucket["site_code"]),
            site_name=str(bucket.get("site_name") or "").strip() or None,
            work_date=bucket["work_date"],
            shift_kind=str(bucket["shift_kind"]),
            request_count=max(0, int(bucket.get("request_count") or 0)),
            assigned_count=max(0, int(bucket.get("assigned_count") or 0)),
            filled_count=max(0, int(bucket.get("filled_count") or 0)),
            request_status=str(bucket.get("request_status") or "").strip() or None,
            work_purpose=str(bucket.get("work_purpose") or "").strip() or None,
            source_workflow=str(bucket.get("source_workflow") or "").strip() or None,
            source_batch_id=bucket.get("source_batch_id"),
            source_revision=str(bucket.get("source_revision") or "").strip() or None,
            source_labels=list(bucket.get("source_labels") or []),
            worker_display_values=list(bucket.get("worker_display_values") or []),
            assignments=list(bucket.get("assignments") or []),
            has_request_ticket=bool(bucket.get("has_request_ticket")),
            updated_at=bucket.get("updated_at"),
        )
        for bucket in bucket_index.values()
    ]

    return sorted(
        rows,
        key=lambda row: (
            -row.work_date.toordinal(),
            str(row.site_code or "").upper(),
            0 if str(row.shift_kind or "").strip().lower() == "day" else 1,
        ),
    )


@router.get("/support-status-workspace", response_model=SupportStatusWorkspaceOut)
def get_support_status_workspace(
    month: str | None = Query(default=None, description="YYYY-MM"),
    site_code: str | None = Query(default=None, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    month_key = str(month or datetime.utcnow().strftime("%Y-%m")).strip()
    _month_bounds(month_key)
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=site_code) if site_code else None
    rows = _load_support_status_workspace_rows(
        conn,
        tenant_id=str(target_tenant["id"]),
        tenant_code=str(target_tenant["tenant_code"]),
        month_key=month_key,
        site_id=str(site["id"]) if site else None,
    )
    return SupportStatusWorkspaceOut(
        tenant_code=str(target_tenant["tenant_code"]),
        month=month_key,
        total_count=len(rows),
        day_count=sum(1 for row in rows if str(row.shift_kind or "").strip().lower() == "day"),
        night_count=sum(1 for row in rows if str(row.shift_kind or "").strip().lower() == "night"),
        rows=rows,
        generated_at=datetime.utcnow(),
    )


@router.get("/site-shift-policy", response_model=SiteShiftPolicyOut)
def get_site_shift_policy(
    site_code: str = Query(..., min_length=1, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=site_code)
    policy = get_or_create_site_shift_policy(conn, tenant_id=target_tenant["id"], site_id=site["id"])
    return SiteShiftPolicyOut(
        tenant_code=target_tenant["tenant_code"],
        site_code=site["site_code"],
        weekday_headcount=int(policy["weekday_headcount"]),
        weekend_headcount=int(policy["weekend_headcount"]),
        updated_at=policy["updated_at"],
    )


@router.put("/site-shift-policy", response_model=SiteShiftPolicyOut)
def put_site_shift_policy(
    payload: SiteShiftPolicyUpdate,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")

    target_tenant = _resolve_target_tenant(conn, user, payload.tenant_code)
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=payload.site_code)
    policy = upsert_site_shift_policy(
        conn,
        tenant_id=target_tenant["id"],
        site_id=site["id"],
        weekday_headcount=payload.weekday_headcount,
        weekend_headcount=payload.weekend_headcount,
    )
    return SiteShiftPolicyOut(
        tenant_code=target_tenant["tenant_code"],
        site_code=site["site_code"],
        weekday_headcount=int(policy["weekday_headcount"]),
        weekend_headcount=int(policy["weekend_headcount"]),
        updated_at=policy["updated_at"],
    )


@router.get("/apple-daytime-shift", response_model=AppleDaytimeShiftOut)
def get_apple_daytime_shift(
    date: str = Query(..., description="YYYY-MM-DD"),
    site_code: str = Query(..., min_length=1, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if normalize_role(user.get("role")) not in {"dev", "branch_manager", "employee"}:
        raise HTTPException(status_code=403, detail="forbidden")

    work_date = _parse_ymd_or_400(date)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=site_code)
    policy = get_or_create_site_shift_policy(conn, tenant_id=target_tenant["id"], site_id=site["id"])
    generated = generate_apple_daytime_shift(
        work_date=work_date,
        weekday_headcount=int(policy["weekday_headcount"]),
        weekend_headcount=int(policy["weekend_headcount"]),
    )
    return AppleDaytimeShiftOut(
        tenant_code=target_tenant["tenant_code"],
        site_code=site["site_code"],
        work_date=generated["work_date"],
        is_weekend=bool(generated["is_weekend"]),
        total_headcount=int(generated["total_headcount"]),
        supervisor_count=int(generated["supervisor_count"]),
        guard_count=int(generated["guard_count"]),
        supervisor_time=generated["supervisor_time"],
        guard_time=generated["guard_time"],
        supervisor_hours=float(generated["supervisor_hours"]),
        guard_hours=float(generated["guard_hours"]),
    )


@router.post("/apple-overtime", response_model=AppleOvertimeOut)
def create_apple_daytime_overtime(
    payload: AppleOvertimeCreate,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    target_tenant = _resolve_target_tenant(conn, user, payload.tenant_code)
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=payload.site_code)

    actor_role = normalize_role(user.get("role"))
    if actor_role not in {"dev", "branch_manager", "employee"}:
        raise HTTPException(status_code=403, detail="forbidden")

    leader_user_id = payload.leader_user_id or user["id"]
    if actor_role == "employee" and str(leader_user_id) != str(user["id"]):
        raise HTTPException(status_code=403, detail="employee can only submit own leader record")

    row = create_apple_overtime_log(
        conn,
        tenant_id=target_tenant["id"],
        site_id=site["id"],
        work_date=payload.work_date,
        leader_user_id=leader_user_id,
        reason=payload.reason,
    )

    rows = list_apple_overtime_logs(
        conn,
        tenant_id=target_tenant["id"],
        work_date=payload.work_date,
        site_id=site["id"],
    )
    hydrated = next((item for item in rows if str(item["id"]) == str(row["id"])), None) or row
    return AppleOvertimeOut(
        id=hydrated["id"],
        tenant_code=target_tenant["tenant_code"],
        site_code=hydrated.get("site_code") or site["site_code"],
        work_date=hydrated["work_date"],
        leader_user_id=hydrated["leader_user_id"],
        leader_username=hydrated.get("leader_username"),
        leader_full_name=hydrated.get("leader_full_name"),
        reason=hydrated["reason"],
        hours=float(hydrated["hours"]),
        source=hydrated.get("source") or "APPLE_DAYTIME_OT",
        created_at=hydrated["created_at"],
    )


@router.get("/apple-overtime", response_model=list[AppleOvertimeOut])
def get_apple_daytime_overtime(
    date: str | None = Query(default=None, description="YYYY-MM-DD"),
    site_code: str | None = Query(default=None, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    work_date = _parse_ymd_or_400(date) if date else None
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=site_code) if site_code else None
    rows = list_apple_overtime_logs(
        conn,
        tenant_id=target_tenant["id"],
        work_date=work_date,
        site_id=site["id"] if site else None,
    )
    return [
        AppleOvertimeOut(
            id=row["id"],
            tenant_code=row["tenant_code"],
            site_code=row["site_code"],
            work_date=row["work_date"],
            leader_user_id=row["leader_user_id"],
            leader_username=row.get("leader_username"),
            leader_full_name=row.get("leader_full_name"),
            reason=row["reason"],
            hours=float(row["hours"]),
            source=row.get("source") or "APPLE_DAYTIME_OT",
            created_at=row["created_at"],
        )
        for row in rows
    ]


@router.post("/late-shifts", response_model=LateShiftOut)
def post_late_shift(
    payload: LateShiftCreate,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, payload.tenant_code)
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=payload.site_code)
    employee = resolve_employee(
        conn,
        tenant_id=target_tenant["id"],
        employee_code=payload.employee_code,
        site_id=site["id"],
    )
    if not employee:
        raise HTTPException(status_code=404, detail="employee not found")

    row = create_late_shift_log(
        conn,
        tenant_id=target_tenant["id"],
        site_id=site["id"],
        work_date=payload.work_date,
        employee_id=employee["id"],
        minutes_late=payload.minutes_late,
        note=payload.note,
    )
    return LateShiftOut(
        id=row["id"],
        tenant_code=target_tenant["tenant_code"],
        site_code=site["site_code"],
        work_date=row["work_date"],
        employee_id=employee["id"],
        employee_code=employee["employee_code"],
        employee_name=employee.get("full_name"),
        minutes_late=int(row["minutes_late"]),
        note=row.get("note"),
        created_at=row["created_at"],
    )


@router.get("/late-shifts", response_model=list[LateShiftOut])
def get_late_shifts(
    date: str | None = Query(default=None, description="YYYY-MM-DD"),
    site_code: str | None = Query(default=None, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    work_date = _parse_ymd_or_400(date) if date else None
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=site_code) if site_code else None
    rows = list_late_shift_logs(
        conn,
        tenant_id=target_tenant["id"],
        work_date=work_date,
        site_id=site["id"] if site else None,
    )
    if not can_manage_schedule(user["role"]) and user.get("employee_id"):
        own_id = str(user.get("employee_id"))
        rows = [row for row in rows if str(row.get("employee_id")) == own_id]
    return [
        LateShiftOut(
            id=row["id"],
            tenant_code=row["tenant_code"],
            site_code=row["site_code"],
            work_date=row["work_date"],
            employee_id=row["employee_id"],
            employee_code=row["employee_code"],
            employee_name=row.get("employee_name"),
            minutes_late=int(row["minutes_late"]),
            note=row.get("note"),
            created_at=row["created_at"],
        )
        for row in rows
    ]


@router.delete("/late-shifts/{late_shift_id}")
def remove_late_shift(
    late_shift_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    deleted = delete_late_shift_log(conn, tenant_id=target_tenant["id"], row_id=late_shift_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="late shift log not found")
    return {"deleted": True, "id": str(late_shift_id)}


@router.post("/daily-events", response_model=DailyEventOut)
def post_daily_event(
    payload: DailyEventCreate,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, payload.tenant_code)
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=payload.site_code)
    row = create_daily_event_log(
        conn,
        tenant_id=target_tenant["id"],
        site_id=site["id"],
        work_date=payload.work_date,
        event_type=payload.type,
        description=payload.description,
    )
    return DailyEventOut(
        id=row["id"],
        tenant_code=target_tenant["tenant_code"],
        site_code=site["site_code"],
        work_date=row["work_date"],
        type=row["type"],
        description=row["description"],
        created_at=row["created_at"],
    )


@router.get("/daily-events", response_model=list[DailyEventOut])
def get_daily_events(
    date: str | None = Query(default=None, description="YYYY-MM-DD"),
    site_code: str | None = Query(default=None, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    work_date = _parse_ymd_or_400(date) if date else None
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=site_code) if site_code else None
    rows = list_daily_event_logs(
        conn,
        tenant_id=target_tenant["id"],
        work_date=work_date,
        site_id=site["id"] if site else None,
    )
    return [
        DailyEventOut(
            id=row["id"],
            tenant_code=row["tenant_code"],
            site_code=row["site_code"],
            work_date=row["work_date"],
            type=row["type"],
            description=row["description"],
            created_at=row["created_at"],
        )
        for row in rows
    ]


@router.delete("/daily-events/{event_id}")
def remove_daily_event(
    event_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    deleted = delete_daily_event_log(conn, tenant_id=target_tenant["id"], row_id=event_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="daily event not found")
    return {"deleted": True, "id": str(event_id)}


@router.post("/support-assignments", response_model=SupportAssignmentOut)
def post_support_assignment(
    payload: SupportAssignmentCreate,
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, payload.tenant_code)
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=payload.site_code)
    support_source_row = _get_support_roundtrip_source(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(site["id"]),
        month_key=f"{payload.work_date.year:04d}-{payload.work_date.month:02d}",
    )
    if support_source_row:
        raise HTTPException(
            status_code=409,
            detail="support roundtrip source is active for this site/month; direct support-assignment writes are gated",
        )
    employee_id = None
    if payload.employee_code:
        employee = resolve_employee(
            conn,
            tenant_id=target_tenant["id"],
            employee_code=payload.employee_code,
            site_id=site["id"],
        )
        if not employee:
            raise HTTPException(status_code=404, detail="employee not found")
        employee_id = employee["id"]

    row, _created = upsert_support_assignment(
        conn,
        tenant_id=target_tenant["id"],
        site_id=site["id"],
        work_date=payload.work_date,
        worker_type=payload.worker_type,
        name=payload.name,
        support_period=payload.support_period,
        slot_index=payload.slot_index,
        source=payload.source,
        employee_id=employee_id,
        affiliation=payload.affiliation,
    )
    if not row:
        raise HTTPException(status_code=500, detail="failed to save support assignment")
    return SupportAssignmentOut(
        id=row["id"],
        tenant_code=row["tenant_code"],
        site_code=row["site_code"],
        work_date=row["work_date"],
        support_period=row.get("support_period") or "day",
        slot_index=int(row.get("slot_index") or 1),
        worker_type=row["worker_type"],
        employee_id=row.get("employee_id"),
        employee_code=row.get("employee_code"),
        employee_name=row.get("employee_name"),
        name=row["name"],
        affiliation=row.get("affiliation"),
        source=row["source"],
        source_ticket_id=row.get("source_ticket_id"),
        source_event_uid=row.get("source_event_uid"),
        created_at=row["created_at"],
        updated_at=row.get("updated_at"),
    )


@router.get("/support-assignments", response_model=list[SupportAssignmentOut])
def get_support_assignments(
    date: str | None = Query(default=None, description="YYYY-MM-DD"),
    site_code: str | None = Query(default=None, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    work_date = _parse_ymd_or_400(date) if date else None
    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=site_code) if site_code else None
    rows = list_support_assignments(
        conn,
        tenant_id=target_tenant["id"],
        work_date=work_date,
        site_id=site["id"] if site else None,
    )
    return [
        SupportAssignmentOut(
            id=row["id"],
            tenant_code=row["tenant_code"],
            site_code=row["site_code"],
            work_date=row["work_date"],
            support_period=row.get("support_period") or "day",
            slot_index=int(row.get("slot_index") or 1),
            worker_type=row["worker_type"],
            employee_id=row.get("employee_id"),
            employee_code=row.get("employee_code"),
            employee_name=row.get("employee_name"),
            name=row["name"],
            affiliation=row.get("affiliation"),
            source=row["source"],
            source_ticket_id=row.get("source_ticket_id"),
            source_event_uid=row.get("source_event_uid"),
            created_at=row["created_at"],
            updated_at=row.get("updated_at"),
        )
        for row in rows
    ]


@router.delete("/support-assignments/{assignment_id}")
def remove_support_assignment(
    assignment_id: uuid.UUID,
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    if not can_manage_schedule(user["role"]):
        raise HTTPException(status_code=403, detail="forbidden")
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT site_id, work_date
            FROM support_assignment
            WHERE id = %s
              AND tenant_id = %s
            LIMIT 1
            """,
            (assignment_id, target_tenant["id"]),
        )
        existing_row = cur.fetchone()
    if not existing_row:
        raise HTTPException(status_code=404, detail="support assignment not found")
    support_source_row = _get_support_roundtrip_source(
        conn,
        tenant_id=str(target_tenant["id"]),
        site_id=str(existing_row["site_id"]),
        month_key=f"{existing_row['work_date'].year:04d}-{existing_row['work_date'].month:02d}",
    )
    if support_source_row:
        raise HTTPException(
            status_code=409,
            detail="support roundtrip source is active for this site/month; direct support-assignment deletes are gated",
        )
    deleted = delete_support_assignment(conn, tenant_id=target_tenant["id"], row_id=assignment_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="support assignment not found")
    return {"deleted": True, "id": str(assignment_id)}


@router.get("/duty-log", response_model=DutyLogOut)
def get_duty_log(
    month: str = Query(..., description="YYYY-MM"),
    employee_code: str | None = Query(default=None, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    actor_role = normalize_role(user.get("role"))
    if actor_role == "employee":
        if not user.get("employee_id"):
            raise HTTPException(status_code=400, detail="employee account is not linked")
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT employee_code
                FROM employees
                WHERE id = %s
                  AND tenant_id = %s
                LIMIT 1
                """,
                (user["employee_id"], target_tenant["id"]),
            )
            own = cur.fetchone()
        if not own:
            raise HTTPException(status_code=404, detail="employee not found")
        target_employee_code = own["employee_code"]
    else:
        target_employee_code = (employee_code or "").strip()
        if not target_employee_code:
            raise HTTPException(status_code=400, detail="employee_code is required")

    employee = resolve_employee(
        conn,
        tenant_id=target_tenant["id"],
        employee_code=target_employee_code,
    )
    if not employee:
        raise HTTPException(status_code=404, detail="employee not found")

    rows = build_duty_log(
        conn,
        tenant_id=target_tenant["id"],
        employee_id=employee["id"],
        month=month,
    )
    return DutyLogOut(
        tenant_code=target_tenant["tenant_code"],
        employee_code=employee["employee_code"],
        month=month,
        rows=[
            DutyLogRowOut(
                work_date=row["work_date"],
                mark=row["mark"],
                shift_type=row.get("shift_type"),
                leave_type=row.get("leave_type"),
                source=row.get("source"),
            )
            for row in rows
        ],
    )


@router.get("/date-details")
def get_schedule_date_details(
    date: str = Query(..., description="YYYY-MM-DD"),
    site_code: str | None = Query(default=None, max_length=64),
    tenant_code: str | None = Query(default=None, max_length=64),
    conn=Depends(get_db_conn),
    user=Depends(get_current_user),
):
    work_date = _parse_ymd_or_400(date)
    target_tenant = _resolve_target_tenant(conn, user, tenant_code)
    staff_scope = enforce_staff_site_scope(user, request_site_code=site_code)
    effective_site_code = str(site_code or "").strip()
    if staff_scope:
        effective_site_code = str(staff_scope.get("site_code") or "").strip()

    site = _resolve_site_or_404(conn, tenant_id=target_tenant["id"], site_code=effective_site_code) if effective_site_code else None
    site_id = site["id"] if site else None

    apple_ot_rows = list_apple_overtime_logs(
        conn,
        tenant_id=target_tenant["id"],
        work_date=work_date,
        site_id=site_id,
    )
    late_rows = list_late_shift_logs(
        conn,
        tenant_id=target_tenant["id"],
        work_date=work_date,
        site_id=site_id,
    )
    support_rows = list_support_assignments(
        conn,
        tenant_id=target_tenant["id"],
        work_date=work_date,
        site_id=site_id,
    )
    event_rows = list_daily_event_logs(
        conn,
        tenant_id=target_tenant["id"],
        work_date=work_date,
        site_id=site_id,
    )

    daytime_payload: dict[str, Any] | None = None
    policy_payload: dict[str, Any] | None = None
    if site:
        policy = get_or_create_site_shift_policy(conn, tenant_id=target_tenant["id"], site_id=site["id"])
        generated = generate_apple_daytime_shift(
            work_date=work_date,
            weekday_headcount=int(policy["weekday_headcount"]),
            weekend_headcount=int(policy["weekend_headcount"]),
        )
        policy_payload = {
            "tenant_code": target_tenant["tenant_code"],
            "site_code": site["site_code"],
            "weekday_headcount": int(policy["weekday_headcount"]),
            "weekend_headcount": int(policy["weekend_headcount"]),
            "updated_at": policy["updated_at"],
        }
        daytime_payload = {
            "tenant_code": target_tenant["tenant_code"],
            "site_code": site["site_code"],
            **generated,
        }

    return {
        "tenant_code": target_tenant["tenant_code"],
        "work_date": work_date.isoformat(),
        "site_code": site["site_code"] if site else None,
        "site_policy": policy_payload,
        "daytime_shift": daytime_payload,
        "apple_overtime": [dict(row) for row in apple_ot_rows],
        "late_shift": [dict(row) for row in late_rows],
        "support_assignment": [dict(row) for row in support_rows],
        "daily_events": [dict(row) for row in event_rows],
    }
