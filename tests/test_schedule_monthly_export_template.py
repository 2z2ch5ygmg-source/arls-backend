from __future__ import annotations

from datetime import date
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.routers.v1.schedules import (
    _build_arls_month_sheet,
    _merge_board_items_for_calendar,
    _current_cell_has_time_conflict,
    _extract_arls_date_columns,
    _locate_support_section_rows,
    _resolve_canonical_schedule_time,
    _resolve_schedule_row_conflict_range,
    _resolve_export_overnight_value,
    _schedule_time_ranges_overlap,
)


TEMPLATE_PATH = Path("/Users/seoseong-won/Documents/rg-arls-dev/backend/app/templates/monthly_schedule_template.xlsx")


class MonthlyScheduleTemplateExportTests(unittest.TestCase):
    def test_monthly_board_items_merge_day_and_night_for_same_employee_site(self):
        merged = _merge_board_items_for_calendar(
            [
                {
                    "schedule_id": "day-1",
                    "employee_id": "emp-1",
                    "employee_code": "R692-2",
                    "employee_name": "조태환",
                    "site_code": "R692",
                    "site_name": "Apple_가로수길",
                    "shift_type": "day",
                    "start_time": "10:00:00",
                    "end_time": "22:00:00",
                    "status": "scheduled",
                },
                {
                    "schedule_id": "night-1",
                    "employee_id": "emp-1",
                    "employee_code": "R692-2",
                    "employee_name": "조태환",
                    "site_code": "R692",
                    "site_name": "Apple_가로수길",
                    "shift_type": "night",
                    "start_time": "22:00:00",
                    "end_time": "08:00:00",
                    "status": "scheduled",
                },
            ]
        )

        self.assertEqual(len(merged), 1)
        self.assertEqual(merged[0]["display_variant"], "combined")
        self.assertEqual(merged[0]["display_shift_types"], ["day", "night"])
        self.assertEqual(merged[0]["start_time"], "10:00:00")
        self.assertEqual(merged[0]["end_time"], "08:00:00")

    def test_canonical_schedule_time_repairs_malformed_day_ranges_by_role_default(self):
        supervisor = _resolve_canonical_schedule_time(
            {
                "shift_type": "day",
                "duty_type": "day",
                "shift_start_time": "08:00:00",
                "shift_end_time": "06:00:00",
                "template_start_time": "08:00:00",
                "template_end_time": "06:00:00",
                "soc_role": "Supervisor",
            }
        )
        officer = _resolve_canonical_schedule_time(
            {
                "shift_type": "day",
                "duty_type": "day",
                "shift_start_time": "10:00:00",
                "shift_end_time": "10:00:00",
                "template_start_time": "10:00:00",
                "template_end_time": "10:00:00",
                "soc_role": "Officer",
            }
        )

        self.assertEqual((supervisor["start_time"], supervisor["end_time"], supervisor["hours"]), ("08:00:00", "18:00:00", 10.0))
        self.assertEqual((officer["start_time"], officer["end_time"], officer["hours"]), ("10:00:00", "22:00:00", 12.0))
        self.assertEqual(supervisor["source"], "role_default")
        self.assertEqual(officer["source"], "role_default")

    def test_canonical_schedule_time_prefers_template_for_invalid_day_range(self):
        canonical = _resolve_canonical_schedule_time(
            {
                "shift_type": "day",
                "duty_type": "day",
                "shift_start_time": "10:00:00",
                "shift_end_time": "10:00:00",
                "paid_hours": 12,
                "template_start_time": "10:00:00",
                "template_end_time": "22:00:00",
                "template_paid_hours": 12,
            }
        )

        self.assertEqual(canonical["start_time"], "10:00:00")
        self.assertEqual(canonical["end_time"], "22:00:00")
        self.assertEqual(canonical["label"], "10:00-22:00")
        self.assertEqual(canonical["hours"], 12)
        self.assertIn(canonical["source"], {"template", "row_hours"})

    def test_canonical_schedule_time_repairs_start_only_night_range(self):
        canonical = _resolve_canonical_schedule_time(
            {
                "shift_type": "night",
                "duty_type": "night",
                "template_start_time": "22:00:00",
                "template_end_time": None,
                "template_paid_hours": None,
            }
        )

        self.assertEqual(canonical["start_time"], "22:00:00")
        self.assertEqual(canonical["end_time"], "08:00:00")
        self.assertEqual(canonical["hours"], 10.0)
        self.assertEqual(canonical["label"], "22:00-08:00")

    def test_existing_schedule_conflict_range_uses_template_when_row_time_is_malformed(self):
        start_time, end_time = _resolve_schedule_row_conflict_range(
            {
                "shift_type": "day",
                "duty_type": "day",
                "shift_start_time": "10:00:00",
                "shift_end_time": "10:00:00",
                "template_start_time": "10:00:00",
                "template_end_time": "22:00:00",
                "template_paid_hours": 12,
                "soc_role": "Officer",
            }
        )

        self.assertEqual(start_time, "10:00:00")
        self.assertEqual(end_time, "22:00:00")

    def test_current_cell_conflict_allows_same_day_day_and_night_when_times_do_not_overlap(self):
        has_conflict = _current_cell_has_time_conflict(
            [
                {
                    "schedule_id": "day-1",
                    "shift_type": "day",
                    "duty_type": "day",
                    "shift_start_time": "10:00:00",
                    "shift_end_time": "10:00:00",
                    "template_start_time": "10:00:00",
                    "template_end_time": "22:00:00",
                    "template_paid_hours": 12,
                    "soc_role": "Officer",
                }
            ],
            next_start="22:00:00",
            next_end="08:00:00",
        )

        self.assertFalse(has_conflict)

    def test_schedule_time_ranges_treat_adjacent_day_and_night_as_non_overlap(self):
        self.assertFalse(
            _schedule_time_ranges_overlap(
                "10:00:00",
                "22:00:00",
                "22:00:00",
                "08:00:00",
            )
        )

    def test_overnight_duration_resolution_handles_cross_midnight_and_24(self):
        self.assertEqual(_resolve_export_overnight_value(None, "22:00-08:00"), "10")
        self.assertEqual(_resolve_export_overnight_value(None, "18:00-09:00"), "15")
        self.assertEqual(_resolve_export_overnight_value(None, "22:00-24:00"), "2")
        self.assertEqual(_resolve_export_overnight_value(None, "24:00-08:00"), "8")

    def test_monthly_schedule_template_build_handles_merged_cells(self):
        workbook = load_workbook(TEMPLATE_PATH)

        _build_arls_month_sheet(
            workbook,
            month_key="2026-03",
            rows=[
                {
                    "employee_id": "emp-2",
                    "employee_code": "R692-2",
                    "employee_name": "부관리자",
                    "sequence_no": 2,
                    "schedule_date": "2026-03-01",
                    "duty_type": "day",
                    "shift_type": "day",
                    "paid_hours": None,
                    "shift_start_time": "08:00:00",
                    "shift_end_time": "18:00:00",
                    "template_start_time": "08:00:00",
                    "template_end_time": "18:00:00",
                    "soc_role": "Vice_Supervisor",
                },
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "관리자",
                    "sequence_no": 1,
                    "schedule_date": "2026-03-01",
                    "duty_type": "day",
                    "shift_type": "day",
                    "paid_hours": None,
                    "shift_start_time": "10:00:00",
                    "shift_end_time": "10:00:00",
                    "template_start_time": "10:00:00",
                    "template_end_time": "22:00:00",
                    "soc_role": "Supervisor",
                    "schedule_note": "테스트 메모",
                },
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "관리자",
                    "sequence_no": 1,
                    "schedule_date": "2026-03-02",
                    "duty_type": "night",
                    "shift_type": "night",
                    "paid_hours": None,
                    "shift_start_time": "22:00:00",
                    "shift_end_time": "08:00:00",
                    "template_start_time": "22:00:00",
                    "template_end_time": "08:00:00",
                    "soc_role": "Supervisor",
                },
                {
                    "employee_id": "emp-3",
                    "employee_code": "R692-3",
                    "employee_name": "경비원",
                    "sequence_no": 3,
                    "schedule_date": "2026-03-01",
                    "duty_type": "day",
                    "shift_type": "day",
                    "paid_hours": None,
                    "shift_start_time": "10:00:00",
                    "shift_end_time": "10:00:00",
                    "template_start_time": "10:00:00",
                    "template_end_time": "22:00:00",
                    "soc_role": "Officer",
                },
            ],
            tenant_code="srs_korea",
            site_code="R692",
            site_name="Apple_가로수길",
            site_address="서울시 강남구",
            export_revision="test-revision-1",
            template_version="test-template-v1",
        )

        sheet = workbook["본사 스케쥴 양식"]
        self.assertEqual(sheet["B1"].value, "3월")
        self.assertEqual(sheet["B2"].value, "2026년 3월")
        self.assertIn("Apple_가로수길", str(sheet["B3"].value or ""))
        self.assertEqual(sheet["B5"].value, "관리자")
        self.assertEqual(sheet["C5"].value, "주간근무")
        self.assertEqual(sheet["C6"].value, "초과근무")
        self.assertEqual(sheet["C7"].value, "야간근무")
        self.assertEqual(sheet["D5"].value, 12)
        self.assertEqual(sheet["E7"].value, 10)
        self.assertEqual(sheet["AK5"].value, "* 테스트 메모")
        self.assertEqual(sheet["B8"].value, "부관리자")
        self.assertEqual(sheet["D8"].value, 10)
        self.assertEqual(sheet["B11"].value, "경비원")
        self.assertEqual(sheet["D11"].value, 12)
        self.assertTrue(str(sheet["D5"].fill.fgColor.rgb or "").endswith("C2D6EC"))
        self.assertTrue(str(sheet["E7"].fill.fgColor.rgb or "").endswith("F0C9AF"))
        self.assertEqual(sheet["D5"].number_format, "0")
        self.assertEqual(sheet["E7"].number_format, "0")
        self.assertIn("_ARLS_EXPORT_META", workbook.sheetnames)
        meta_sheet = workbook["_ARLS_EXPORT_META"]
        self.assertEqual(meta_sheet.sheet_state, "hidden")
        self.assertEqual(meta_sheet["A1"].value, "tenant_code")
        self.assertEqual(meta_sheet["B1"].value, "srs_korea")
        self.assertEqual(meta_sheet["A6"].value, "export_revision")
        self.assertEqual(meta_sheet["B6"].value, "test-revision-1")
        self.assertEqual(meta_sheet["A7"].value, "template_version")
        self.assertEqual(meta_sheet["B7"].value, "test-template-v1")

    def test_monthly_schedule_template_build_ignores_orphan_employee_overnight_rows(self):
        workbook = load_workbook(TEMPLATE_PATH)

        _build_arls_month_sheet(
            workbook,
            month_key="2026-03",
            rows=[
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "야간근무자",
                    "sequence_no": 1,
                    "schedule_date": "2026-03-01",
                    "duty_type": "day",
                    "shift_type": "day",
                    "paid_hours": 12,
                    "shift_start_time": "10:00:00",
                    "shift_end_time": "22:00:00",
                    "soc_role": "Officer",
                }
            ],
            tenant_code="srs_korea",
            site_code="R692",
            site_name="Apple_가로수길",
            employee_overnight_rows=[
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "야간근무자",
                    "sequence_no": 1,
                    "soc_role": "Officer",
                    "work_date": date(2026, 3, 1),
                    "hours": None,
                    "time_range": "22:00-08:00",
                    "site_code": "R692",
                }
            ],
            export_revision="test-revision-2",
        )

        sheet = workbook["본사 스케쥴 양식"]
        self.assertEqual(sheet["B5"].value, "야간근무자")
        self.assertEqual(sheet["D5"].value, 12)
        self.assertTrue(str(sheet["D5"].fill.fgColor.rgb or "").endswith("C2D6EC"))
        self.assertEqual(sheet["D5"].number_format, "0")
        self.assertIsNone(sheet["D7"].value)

    def test_monthly_schedule_template_build_writes_external_support_and_internal_overlay(self):
        workbook = load_workbook(TEMPLATE_PATH)

        _build_arls_month_sheet(
            workbook,
            month_key="2026-03",
            rows=[
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "관리자",
                    "sequence_no": 1,
                    "schedule_date": "2026-03-01",
                    "duty_type": "day",
                    "shift_type": "day",
                    "paid_hours": 12,
                    "shift_start_time": "10:00:00",
                    "shift_end_time": "22:00:00",
                    "soc_role": "Supervisor",
                }
            ],
            tenant_code="srs_korea",
            site_code="R692",
            site_name="Apple_가로수길",
            support_rows=[
                {
                    "work_date": date(2026, 3, 2),
                    "support_period": "day",
                    "slot_index": 1,
                    "worker_type": "F",
                    "name": "최유진",
                    "affiliation": "Apple_가로수길",
                    "employee_name": None,
                    "employee_id": None,
                    "soc_role": "",
                    "duty_role": "",
                    "is_internal": False,
                },
                {
                    "work_date": date(2026, 3, 2),
                    "support_period": "day",
                    "slot_index": 2,
                    "worker_type": "INTERNAL",
                    "name": "관리자",
                    "employee_name": "관리자",
                    "employee_code": "R692-1",
                    "employee_id": "emp-1",
                    "soc_role": "Supervisor",
                    "duty_role": "",
                    "is_internal": True,
                },
            ],
            export_revision="test-revision-3",
        )

        sheet = workbook["본사 스케쥴 양식"]
        date_columns, _ = _extract_arls_date_columns(sheet)
        day_col = next(col for col, value in date_columns.items() if value.isoformat() == "2026-03-02")
        rows_meta = _locate_support_section_rows(sheet)

        self.assertEqual(sheet.cell(row=5, column=day_col).value, 10)
        self.assertEqual(sheet.cell(row=rows_meta["weekly_rows"][0], column=day_col).value, "Apple_가로수길 최유진")

    def test_monthly_schedule_template_build_styles_annual_leave_and_employment_markers(self):
        workbook = load_workbook(TEMPLATE_PATH)

        _build_arls_month_sheet(
            workbook,
            month_key="2026-03",
            rows=[
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "입사자",
                    "sequence_no": 1,
                    "schedule_date": "2026-03-01",
                    "duty_type": "day",
                    "shift_type": "day",
                    "paid_hours": 10,
                    "shift_start_time": "08:00:00",
                    "shift_end_time": "18:00:00",
                    "hire_date": date(2026, 3, 1),
                    "soc_role": "Officer",
                },
                {
                    "employee_id": "emp-2",
                    "employee_code": "R692-2",
                    "employee_name": "퇴사자",
                    "sequence_no": 2,
                    "schedule_date": "2026-03-02",
                    "duty_type": "night",
                    "shift_type": "night",
                    "paid_hours": None,
                    "shift_start_time": "18:00:00",
                    "shift_end_time": "09:00:00",
                    "leave_date": date(2026, 3, 2),
                    "soc_role": "Officer",
                },
                {
                    "employee_id": "emp-3",
                    "employee_code": "R692-3",
                    "employee_name": "연차자",
                    "sequence_no": 3,
                    "schedule_date": "2026-03-03",
                    "duty_type": "day",
                    "shift_type": "off",
                    "schedule_note": "연차",
                    "soc_role": "Officer",
                },
            ],
            tenant_code="srs_korea",
            site_code="R692",
            site_name="Apple_가로수길",
            export_revision="test-revision-4",
        )

        sheet = workbook["본사 스케쥴 양식"]
        self.assertEqual(sheet["D7"].value, "입사")
        self.assertTrue(str(sheet["D7"].fill.fgColor.rgb or "").endswith("4FAD5B"))
        self.assertTrue(str(sheet["D7"].font.color.value or "").upper().endswith("000000"))

        self.assertEqual(sheet["E10"].value, "퇴사")
        self.assertTrue(str(sheet["E10"].fill.fgColor.rgb or "").endswith("EA3323"))
        self.assertTrue(str(sheet["E10"].font.color.value or "").upper().endswith("FFFFFF"))

        self.assertEqual(sheet["F12"].value, "연차")
        self.assertTrue(str(sheet["F12"].fill.fgColor.rgb or "").endswith("4FAEEA"))


if __name__ == "__main__":
    unittest.main()
