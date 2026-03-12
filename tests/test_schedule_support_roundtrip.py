from __future__ import annotations

from io import BytesIO
from datetime import date
import unittest
from pathlib import Path
from unittest.mock import patch
import uuid

from openpyxl import Workbook, load_workbook

from app.routers.v1.schedules import (
    ARLS_EXPORT_TEMPLATE_VERSION,
    ARLS_SHEET_NAME,
    ARLS_SUPPORT_METADATA_SHEET_NAME,
    GUARD_DAY_SHIFT_HOURS,
    SENTRIX_HQ_ROSTER_ASSIGNMENT_SOURCE,
    _apply_internal_support_roundtrip_assignments,
    _build_export_employee_blocks,
    _build_sentrix_hq_bridge_candidates,
    _build_arls_month_sheet,
    _build_employee_name_index,
    _build_support_roster_hq_upload_inspect_result,
    _build_sentrix_hq_snapshot_signature,
    _build_sentrix_support_roster_apply_result_from_handoff,
    _build_sentrix_support_roster_handoff_failure_result,
    _build_sentrix_support_roster_handoff_payload,
    _clone_support_hq_sheet_to_workbook,
    _build_support_only_workbook,
    _build_support_roundtrip_employee_row_index,
    _extract_sentrix_ticket_hq_roster_status,
    _extract_arls_date_columns,
    _locate_support_section_rows,
    _normalize_sentrix_hq_roster_final_state,
    _parse_sentrix_hq_worker_cell,
    _read_support_roundtrip_metadata,
    _resolve_sentrix_support_materialized_shift_defaults,
)


TEMPLATE_PATH = Path("/Users/mark/Desktop/rg-arls-dev/app/templates/monthly_schedule_template.xlsx")


class ScheduleSupportRoundtripTests(unittest.TestCase):
    def _build_export_ctx(self):
        workbook = load_workbook(TEMPLATE_PATH)
        rows = [
            {
                "employee_id": "emp-1",
                "employee_code": "R692-1",
                "employee_name": "이보한",
                "sequence_no": 1,
                "schedule_date": "2026-03-01",
                "duty_type": "day",
                "shift_type": "day",
                "paid_hours": 12,
                "shift_start_time": "10:00:00",
                "shift_end_time": "22:00:00",
                "soc_role": "Supervisor",
            }
        ]
        _build_arls_month_sheet(
            workbook,
            month_key="2026-03",
            rows=rows,
            tenant_code="srs_korea",
            site_code="R692",
            site_name="Apple_가로수길",
            site_address="서울시 강남구",
            daytime_need_rows=[],
            export_revision="rev-support-1",
            template_version=ARLS_EXPORT_TEMPLATE_VERSION,
            source_version="phase1",
        )
        return {
            "workbook": workbook,
            "template_version": ARLS_EXPORT_TEMPLATE_VERSION,
            "employee_blocks": [
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "이보한",
                }
            ],
        }

    def test_build_support_only_workbook_preserves_site_sheet_and_hidden_metadata(self):
        export_ctx = self._build_export_ctx()
        workbook = _build_support_only_workbook(
            export_ctx=export_ctx,
            target_tenant={"tenant_code": "srs_korea"},
            site_row={"site_code": "R692", "site_name": "Apple_가로수길"},
            month_key="2026-03",
            source_revision="src-rev-001",
            active_assignments=[
                {
                    "work_date": date(2026, 3, 1),
                    "support_period": "day",
                    "slot_index": 1,
                    "worker_type": "EXTERNAL",
                    "worker_name": "BK 박준연",
                    "is_internal": False,
                }
            ],
        )

        self.assertIn("Apple_가로수길", workbook.sheetnames)
        self.assertIn(ARLS_SUPPORT_METADATA_SHEET_NAME, workbook.sheetnames)
        meta = _read_support_roundtrip_metadata(workbook)
        self.assertEqual(meta.get("tenant_code"), "srs_korea")
        self.assertEqual(meta.get("site_code"), "R692")
        self.assertEqual(meta.get("month"), "2026-03")
        self.assertEqual(meta.get("source_revision"), "src-rev-001")

    def test_internal_support_overlay_writes_employee_day_and_night_rows(self):
        export_ctx = self._build_export_ctx()
        workbook = export_ctx["workbook"]
        sheet = workbook[ARLS_SHEET_NAME]
        date_columns, _ = _extract_arls_date_columns(sheet)
        day_col = next(col for col, value in date_columns.items() if value.isoformat() == "2026-03-01")
        row_index = _build_support_roundtrip_employee_row_index(sheet, employee_blocks=export_ctx["employee_blocks"])

        _apply_internal_support_roundtrip_assignments(
            sheet,
            export_ctx=export_ctx,
            assignment_rows=[
                {
                    "is_internal": True,
                    "employee_id": "emp-1",
                    "work_date": date(2026, 3, 1),
                    "support_period": "day",
                    "internal_shift_type": "day",
                    "internal_paid_hours": 12,
                },
                {
                    "is_internal": True,
                    "employee_id": "emp-1",
                    "work_date": date(2026, 3, 1),
                    "support_period": "night",
                    "internal_shift_type": "night",
                    "internal_shift_start_time": "22:00:00",
                    "internal_shift_end_time": "08:00:00",
                },
            ],
        )

        self.assertEqual(sheet.cell(row=row_index[("emp-1", "day")], column=day_col).value, "12")
        self.assertEqual(sheet.cell(row=row_index[("emp-1", "night")], column=day_col).value, "10")

    def test_support_only_workbook_keeps_external_display_label_and_extends_slots(self):
        export_ctx = self._build_export_ctx()
        workbook = _build_support_only_workbook(
            export_ctx=export_ctx,
            target_tenant={"tenant_code": "srs_korea"},
            site_row={"site_code": "R692", "site_name": "Apple_가로수길"},
            month_key="2026-03",
            source_revision="src-rev-002",
            active_assignments=[
                {
                    "work_date": date(2026, 3, 1),
                    "support_period": "day",
                    "slot_index": index + 1,
                    "worker_type": "F",
                    "worker_name": "Apple_가로수길 박준연" if index == 0 else f"협력사{index} 지원자{index}",
                    "is_internal": False,
                }
                for index in range(7)
            ],
        )

        sheet = workbook["Apple_가로수길"]
        date_columns, _ = _extract_arls_date_columns(sheet)
        day_col = next(col for col, value in date_columns.items() if value.isoformat() == "2026-03-01")
        rows_meta = _locate_support_section_rows(sheet)

        self.assertGreaterEqual(len(rows_meta["weekly_rows"]), 7)
        self.assertEqual(sheet.cell(row=rows_meta["weekly_rows"][0], column=day_col).value, "Apple_가로수길 박준연")

    def test_support_only_workbook_can_export_blank_worker_slots_without_prefill(self):
        export_ctx = self._build_export_ctx()
        workbook = _build_support_only_workbook(
            export_ctx=export_ctx,
            target_tenant={"tenant_code": "srs_korea"},
            site_row={"site_code": "R692", "site_name": "Apple_가로수길"},
            month_key="2026-03",
            source_revision="src-rev-blank",
            active_assignments=[],
            include_existing_assignments=False,
        )

        sheet = workbook["Apple_가로수길"]
        date_columns, _ = _extract_arls_date_columns(sheet)
        day_col = next(col for col, value in date_columns.items() if value.isoformat() == "2026-03-01")
        rows_meta = _locate_support_section_rows(sheet)

        self.assertIsNone(sheet.cell(row=rows_meta["weekly_rows"][0], column=day_col).value)
        self.assertIsNone(sheet.cell(row=rows_meta["night_rows"][0], column=day_col).value)
        self.assertEqual(sheet.cell(row=rows_meta["weekly_count_row"], column=day_col).value, 0)

    def test_hq_roster_inspect_ignores_ticket_scopes_not_signaled_by_workbook(self):
        export_ctx = self._build_export_ctx()
        workbook = _build_support_only_workbook(
            export_ctx=export_ctx,
            target_tenant={"tenant_code": "srs_korea"},
            site_row={"site_code": "R692", "site_name": "Apple_가로수길"},
            month_key="2026-03",
            source_revision="src-rev-inspect",
            active_assignments=[],
            include_existing_assignments=False,
        )

        sheet = workbook["Apple_가로수길"]
        date_columns, _ = _extract_arls_date_columns(sheet)
        rows_meta = _locate_support_section_rows(sheet)
        day_col = next(col for col, value in date_columns.items() if value.isoformat() == "2026-03-01")
        blank_ticket_col = next(col for col, value in date_columns.items() if value.isoformat() == "2026-03-02")
        night_col = next(col for col, value in date_columns.items() if value.isoformat() == "2026-03-03")

        sheet.cell(row=rows_meta["day_need_row"], column=day_col).value = "섭외 2인 요청"
        sheet.cell(row=rows_meta["day_need_row"], column=blank_ticket_col).value = None
        sheet.cell(row=rows_meta["night_need_row"], column=night_col).value = "섭외 1인 요청"
        sheet.cell(row=rows_meta["work_note_row"], column=night_col).value = "Project"

        ticket_day = uuid.uuid4()
        ticket_blank = uuid.uuid4()
        ticket_night = uuid.uuid4()

        with patch("app.routers.v1.schedules._list_support_roundtrip_workspace_sites", return_value=[{
            "site_id": "site-1",
            "site_code": "R692",
            "site_name": "Apple_가로수길",
        }]), patch("app.routers.v1.schedules._build_support_roster_hq_source_map", return_value={
            "R692": {"source_revision": "src-rev-inspect"}
        }), patch("app.routers.v1.schedules._load_sentrix_support_ticket_scope_map", return_value={
            ("R692", "2026-03-01", "day"): {"id": ticket_day, "request_count": 2},
            ("R692", "2026-03-02", "day"): {"id": ticket_blank, "request_count": 1},
            ("R692", "2026-03-03", "night"): {"id": ticket_night, "request_count": 1},
        }), patch("app.routers.v1.schedules._load_site_employees", return_value=[]), patch(
            "app.routers.v1.schedules._persist_sentrix_hq_roster_preview_batch",
            return_value=uuid.uuid4(),
        ):
            preview = _build_support_roster_hq_upload_inspect_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                selected_month="2026-03",
                filename="병합T1.xlsx",
                user={"id": "user-1"},
            )

        self.assertEqual(preview.total_scope_count, 2)
        self.assertEqual(preview.summary["scope_total"], 2)
        self.assertEqual(preview.summary["approval_pending"], 2)
        self.assertEqual(
            {item.scope_key for item in preview.scope_summaries},
            {
                "R692:2026-03-01:day",
                "R692:2026-03-03:night",
            },
        )
        self.assertNotIn(
            "R692:2026-03-02:day",
            {
                f"{item.site_code}:{item.work_date.isoformat()}:{item.shift_kind}"
                for item in preview.review_rows
                if item.work_date and item.shift_kind and item.site_code
            },
        )

    def test_clone_support_hq_sheet_to_workbook_copies_dimension_style_without_error(self):
        source_workbook = Workbook()
        source_sheet = source_workbook.active
        source_sheet.title = "Apple_가로수길"
        source_sheet["A1"] = "지원근무"
        source_sheet.column_dimensions["A"].width = 24
        source_sheet.row_dimensions[1].height = 28
        source_sheet["A1"].number_format = "@"

        target_workbook = Workbook()
        target_workbook.remove(target_workbook.active)

        _clone_support_hq_sheet_to_workbook(
            source_sheet,
            target_workbook=target_workbook,
            title="Apple_가로수길",
        )

        cloned_sheet = target_workbook["Apple_가로수길"]
        self.assertEqual(cloned_sheet["A1"].value, "지원근무")
        self.assertEqual(cloned_sheet.column_dimensions["A"].width, 24)
        self.assertEqual(cloned_sheet.row_dimensions[1].height, 28)

        out = BytesIO()
        target_workbook.save(out)
        out.seek(0)
        reloaded = load_workbook(out)
        self.assertEqual(reloaded["Apple_가로수길"]["A1"].value, "지원근무")

    def test_parse_sentrix_hq_worker_cell_normalizes_external_and_internal(self):
        employee_index = _build_employee_name_index(
            [
                {
                    "id": "emp-1",
                    "employee_code": "R692-1",
                    "full_name": "조태환",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                }
            ]
        )

        external = _parse_sentrix_hq_worker_cell(
            "bk   박준연",
            schedule_date=date(2026, 3, 1),
            employee_index=employee_index,
        )
        self.assertTrue(external["is_valid"])
        self.assertEqual(external["affiliation"], "BK")
        self.assertEqual(external["name"], "박준연")
        self.assertEqual(external["display_value"], "BK 박준연")
        self.assertEqual(external["worker_type"], "BK")

        internal = _parse_sentrix_hq_worker_cell(
            "자체 조태환",
            schedule_date=date(2026, 3, 1),
            employee_index=employee_index,
        )
        self.assertTrue(internal["is_valid"])
        self.assertTrue(internal["self_staff"])
        self.assertEqual(internal["employee_id"], "emp-1")
        self.assertEqual(internal["worker_type"], "INTERNAL")

    def test_parse_sentrix_hq_worker_cell_rejects_multi_person_and_invalid_self(self):
        employee_index = _build_employee_name_index(
            [
                {
                    "id": "emp-1",
                    "employee_code": "R692-1",
                    "full_name": "조태환",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                },
                {
                    "id": "emp-2",
                    "employee_code": "R692-2",
                    "full_name": "조태환",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                },
            ]
        )

        multi = _parse_sentrix_hq_worker_cell(
            "BK 박준연 / BK 김하늘",
            schedule_date=date(2026, 3, 1),
            employee_index=employee_index,
        )
        self.assertFalse(multi["is_valid"])
        self.assertEqual(multi["issue_code"], "MULTI_PERSON_CELL")

        invalid_self = _parse_sentrix_hq_worker_cell(
            "자체 bk 박준연",
            schedule_date=date(2026, 3, 1),
            employee_index=employee_index,
        )
        self.assertFalse(invalid_self["is_valid"])
        self.assertEqual(invalid_self["issue_code"], "SELF_STAFF_FORMAT_INVALID")

        ambiguous_self = _parse_sentrix_hq_worker_cell(
            "자체 조태환",
            schedule_date=date(2026, 3, 1),
            employee_index=employee_index,
        )
        self.assertFalse(ambiguous_self["is_valid"])
        self.assertEqual(ambiguous_self["issue_code"], "SELF_STAFF_EMPLOYEE_AMBIGUOUS")

    def test_build_sentrix_support_roster_handoff_payload_preserves_scope_lineage(self):
        batch_id = uuid.uuid4()
        scope_apply_specs = [
            {
                "scope_key": "R692:2026-03-01:day",
                "sheet_name": "Apple_가로수길",
                "site_code": "R692",
                "site_name": "Apple_가로수길",
                "work_date": date(2026, 3, 1),
                "shift_kind": "day",
                "ticket": {"status": "pending", "payload_json": {}},
                "ticket_id": uuid.uuid4(),
                "request_count": 2,
                "valid_filled_count": 2,
                "invalid_filled_count": 0,
                "target_status": "approved",
                "scope_payload": {
                    "site_id": "site-r692",
                    "workbook_required_count": 2,
                    "workbook_required_raw": "섭외 2인 요청",
                    "external_count_raw": "0",
                    "purpose_text": "셀 값 Project",
                    "matched_ticket": True,
                },
                "worker_payloads": [
                    {
                        "slot_index": 1,
                        "raw_cell_text": "BK 박준연",
                        "parsed_display_value": "BK 박준연",
                        "affiliation": "BK",
                        "worker_name": "박준연",
                        "worker_type": "BK",
                        "self_staff": False,
                        "countable": True,
                        "issue_code": "",
                        "sheet_name": "Apple_가로수길",
                        "source_row": 55,
                        "source_col": 4,
                        "source_cell_ref": "D55",
                    },
                    {
                        "slot_index": 2,
                        "raw_cell_text": "자체 조태환",
                        "parsed_display_value": "자체 조태환",
                        "affiliation": "",
                        "worker_name": "조태환",
                        "worker_type": "INTERNAL",
                        "self_staff": True,
                        "countable": True,
                        "issue_code": "",
                        "employee_id": "emp-1",
                        "employee_code": "R692-1",
                        "employee_name": "조태환",
                        "sheet_name": "Apple_가로수길",
                        "source_row": 56,
                        "source_col": 4,
                        "source_cell_ref": "D56",
                    },
                ],
            }
        ]

        payload = _build_sentrix_support_roster_handoff_payload(
            batch_id=batch_id,
            batch={
                "month_key": "2026-03",
                "selected_site_code": "R692",
                "bundle_revision": "rev-001",
                "download_scope": "site",
                "upload_meta_json": {
                    "workbook_family": "support_hq_assignment",
                    "template_version": "2026.03",
                },
            },
            target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
            scope_apply_specs=scope_apply_specs,
        )

        self.assertEqual(payload["artifact_id"], "sentrix-hq:SRS_KOREA:2026-03:R692:rev-001")
        self.assertEqual(payload["source_upload_batch_id"], str(batch_id))
        self.assertEqual(payload["affected_scope_count"], 1)
        self.assertEqual(payload["affected_site_codes"], ["R692"])
        self.assertEqual(payload["affected_dates"], ["2026-03-01"])
        self.assertNotIn("current_ticket_hint", payload["scopes"][0])
        self.assertNotIn("site_id", payload["scopes"][0])
        self.assertEqual(payload["scopes"][0]["worker_entries"][0]["source_cell_ref"], "D55")
        self.assertEqual(payload["scopes"][0]["worker_entries"][1]["canonical_employee_id_hint"], "emp-1")

    def test_build_sentrix_support_roster_handoff_failure_result_is_retryable(self):
        batch_id = uuid.uuid4()
        result = _build_sentrix_support_roster_handoff_failure_result(
            batch_id=batch_id,
            batch={"month_key": "2026-03"},
            issue_count=0,
            artifact_id="sentrix-hq:SRS_KOREA:2026-03:R692:rev-001",
            scope_apply_specs=[
                {
                    "scope_key": "R692:2026-03-01:day",
                    "sheet_name": "Apple_가로수길",
                    "site_code": "R692",
                    "site_name": "Apple_가로수길",
                    "work_date": date(2026, 3, 1),
                    "shift_kind": "day",
                    "request_count": 2,
                    "valid_filled_count": 2,
                    "target_status": "approved",
                    "ticket": {"status": "pending"},
                    "valid_worker_payloads": [{"slot_index": 1}, {"slot_index": 2}],
                }
            ],
            error_message="Sentrix handoff timeout",
        )

        self.assertFalse(result.applied)
        self.assertTrue(result.partial_success)
        self.assertEqual(result.handoff_status, "failed")
        self.assertEqual(result.retry_token, str(batch_id))
        self.assertEqual(result.handoff_failed_count, 1)
        self.assertEqual(result.scope_results[0].handoff_message, "Sentrix handoff timeout")

    def test_build_sentrix_support_roster_apply_result_from_handoff_maps_scope_statuses(self):
        batch_id = uuid.uuid4()
        ticket_id = uuid.uuid4()
        scope_specs = [
            {
                "scope_key": "R692:2026-03-01:day",
                "sheet_name": "Apple_가로수길",
                "site_code": "R692",
                "site_name": "Apple_가로수길",
                "work_date": date(2026, 3, 1),
                "shift_kind": "day",
                "request_count": 2,
                "valid_filled_count": 2,
                "target_status": "approved",
                "ticket": {"status": "pending"},
                "valid_worker_payloads": [{"slot_index": 1}, {"slot_index": 2}],
            }
        ]
        result = _build_sentrix_support_roster_apply_result_from_handoff(
            batch_id=batch_id,
            batch={"month_key": "2026-03"},
            issue_count=0,
            handoff_payload={
                "artifact_id": "sentrix-hq:SRS_KOREA:2026-03:R692:rev-001",
                "affected_scope_count": 1,
                "affected_site_codes": ["R692"],
                "affected_dates": ["2026-03-01"],
            },
            handoff_response={
                "applied": False,
                "partial_success": True,
                "handoff_status": "partial",
                "handoff_message": "일부 scope만 Sentrix에 반영되었습니다.",
                "artifact_id": "sentrix-hq:SRS_KOREA:2026-03:R692:rev-001",
                "retry_token": str(batch_id),
                "handoff_success_count": 1,
                "handoff_failed_count": 1,
                "affected_scope_count": 1,
                "affected_site_codes": ["R692"],
                "affected_dates": ["2026-03-01"],
                "applied_scope_count": 1,
                "failed_scope_count": 1,
                "updated_scope_count": 1,
                "tickets_updated": 1,
                "tickets_auto_approved": 1,
                "tickets_pending": 0,
                "scope_results": [
                    {
                        "scope_key": "R692:2026-03-01:day",
                        "handoff_status": "success",
                        "handoff_message": "Sentrix ticket updated",
                        "sentrix_ticket_id": str(ticket_id),
                    }
                ],
            },
            scope_apply_specs=scope_specs,
        )

        self.assertFalse(result.applied)
        self.assertTrue(result.partial_success)
        self.assertEqual(result.handoff_status, "partial")
        self.assertEqual(result.handoff_success_count, 1)
        self.assertEqual(result.handoff_failed_count, 1)
        self.assertEqual(result.scope_results[0].handoff_status, "success")
        self.assertEqual(result.scope_results[0].sentrix_ticket_id, str(ticket_id))

    def test_extract_sentrix_ticket_hq_roster_status_prefers_detail_json(self):
        self.assertEqual(
            _extract_sentrix_ticket_hq_roster_status(
                {
                    "detail_json": {
                        "hq_roster": {
                            "status": "auto_approved",
                        }
                    }
                }
            ),
            "auto_approved",
        )
        self.assertEqual(
            _extract_sentrix_ticket_hq_roster_status(
                {
                    "detail_json": {
                        "hq_roster": {
                            "status": "approval_pending",
                        }
                    }
                }
            ),
            "approval_pending",
        )
        self.assertIsNone(_extract_sentrix_ticket_hq_roster_status({"detail_json": {}}))

    def test_normalize_sentrix_hq_roster_final_state_maps_auto_approved_to_approved(self):
        self.assertEqual(_normalize_sentrix_hq_roster_final_state("auto_approved"), "approved")
        self.assertEqual(_normalize_sentrix_hq_roster_final_state("approval_pending"), "approval_pending")
        self.assertEqual(_normalize_sentrix_hq_roster_final_state("APPROVED"), "approved")

    def test_build_sentrix_hq_bridge_candidates_keeps_unique_self_staff_employee(self):
        candidates = _build_sentrix_hq_bridge_candidates(
            [
                {
                    "slot_index": 1,
                    "self_staff": True,
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "조태환",
                    "display_value": "자체 조태환",
                    "validity_state": "valid",
                },
                {
                    "slot_index": 2,
                    "self_staff": True,
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "조태환",
                    "display_value": "자체 조태환",
                    "validity_state": "valid",
                },
                {
                    "slot_index": 3,
                    "self_staff": False,
                    "employee_id": None,
                    "display_value": "BK 박준연",
                    "validity_state": "valid",
                },
            ]
        )
        self.assertEqual(list(candidates.keys()), ["emp-1"])
        self.assertEqual(candidates["emp-1"]["employee_name"], "조태환")

    def test_sentrix_hq_snapshot_signature_changes_when_ticket_state_changes(self):
        base_entries = [
            {
                "slot_index": 1,
                "display_value": "자체 조태환",
                "self_staff": True,
                "employee_id": "emp-1",
                "employee_code": "R692-1",
                "employee_name": "조태환",
                "worker_type": "INTERNAL",
                "validity_state": "valid",
            }
        ]
        approved_signature = _build_sentrix_hq_snapshot_signature(
            entries=base_entries,
            request_count=1,
            valid_filled_count=1,
            invalid_filled_count=0,
            ticket_state="approved",
        )
        pending_signature = _build_sentrix_hq_snapshot_signature(
            entries=base_entries,
            request_count=1,
            valid_filled_count=1,
            invalid_filled_count=0,
            ticket_state="approval_pending",
        )
        self.assertNotEqual(approved_signature, pending_signature)

    def test_build_export_employee_blocks_skips_hq_roundtrip_internal_overlay(self):
        rows = []
        blocks = _build_export_employee_blocks(
            rows,
            support_rows=[
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "조태환",
                    "work_date": date(2026, 3, 1),
                    "support_period": "day",
                    "is_internal": True,
                    "source": SENTRIX_HQ_ROSTER_ASSIGNMENT_SOURCE,
                    "soc_role": "guard",
                    "duty_role": "GUARD",
                }
            ],
        )
        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0]["day"], {})

    def test_resolve_sentrix_support_materialized_shift_defaults_uses_day_and_night_rules(self):
        guard_employee = {
            "soc_role": "guard",
            "duty_role": "GUARD",
        }
        start_time, end_time, paid_hours = _resolve_sentrix_support_materialized_shift_defaults(
            guard_employee,
            shift_kind="day",
        )
        self.assertEqual(start_time, "10:00:00")
        self.assertEqual(end_time, "22:00:00")
        self.assertEqual(paid_hours, GUARD_DAY_SHIFT_HOURS)

        night_start, night_end, night_hours = _resolve_sentrix_support_materialized_shift_defaults(
            guard_employee,
            shift_kind="night",
        )
        self.assertEqual(night_start, "22:00:00")
        self.assertEqual(night_end, "08:00:00")
        self.assertEqual(night_hours, 10.0)


if __name__ == "__main__":
    unittest.main()
