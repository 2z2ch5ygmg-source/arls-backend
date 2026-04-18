from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
import unittest
from pathlib import Path
import uuid
from unittest.mock import patch

from openpyxl import load_workbook

from app.routers.v1.schedules import (
    ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
    ARLS_SUPPORT_UPLOAD_INTERNAL_SOURCE,
    ARLS_EXPORT_SOURCE_VERSION,
    ARLS_EXPORT_TEMPLATE_VERSION,
    ARLS_SHEET_NAME,
    SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS,
    _apply_canonical_schedule_import_batch,
    _build_support_request_rows_from_import_payloads,
    _classify_import_preview_visibility,
    _build_arls_month_sheet,
    _build_employee_name_index,
    _build_import_current_body_index_from_existing_schedule_rows,
    _build_schedule_import_preview_result,
    _filter_support_assignment_rows_for_active_scopes,
    _format_schedule_import_mapping_requirement_label,
    _build_schedule_import_mapping_lookup,
    _collect_monthly_export_context,
    _extract_arls_date_columns,
    _load_canonical_schedule_import_apply_payload_rows,
    _leader_candidate_role_label,
    _leader_candidate_role_priority,
    _locate_support_section_rows,
    _normalize_name_token,
    _select_arls_import_visible_sheet_name,
    _parse_arls_canonical_import_sheet,
    _parse_daytime_need_value,
    _parse_support_worker_cell,
    _read_monthly_support_request_rows_for_export,
    _read_arls_export_metadata,
    _resolve_import_support_worker_match,
    _resolve_leader_candidate_role_key,
    _resolve_import_body_value,
    _resolve_schedule_display_meta,
    _resolve_shift_type_from_duty_type,
    preview_import,
    _validate_mapping_profile_requirements,
    _validate_support_shift_worker,
    _validate_arls_import_metadata,
)


TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "app" / "templates" / "monthly_schedule_template.xlsx"


class _CursorContext:
    def __init__(self, responses=None):
        self._responses = list(responses or [])
        self.executed: list[tuple[str, object]] = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False

    def execute(self, sql, params=None):
        self.executed.append((str(sql), params))

    def executemany(self, sql, params_seq):
        self.executed.append((str(sql), list(params_seq)))

    def fetchone(self):
        if not self._responses:
            return None
        return self._responses.pop(0)

    def fetchall(self):
        rows = list(self._responses)
        self._responses.clear()
        return rows


class _ConnectionStub:
    def __init__(self, cursor_responses=None):
        self._cursor_responses = list(cursor_responses or [])
        self.rollbacks = 0

    def cursor(self):
        responses = self._cursor_responses.pop(0) if self._cursor_responses else []
        return _CursorContext(responses)

    def rollback(self):
        self.rollbacks += 1


class _UploadFileStub:
    def __init__(self, *, filename: str, raw_bytes: bytes, content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
        self.filename = filename
        self.file = BytesIO(raw_bytes)
        self.content_type = content_type


class MonthlyScheduleCanonicalImportTests(unittest.TestCase):
    def _build_sample_workbook(self):
        workbook = load_workbook(TEMPLATE_PATH)
        _build_arls_month_sheet(
            workbook,
            month_key="2026-03",
            rows=[
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "이보한",
                    "sequence_no": 1,
                    "schedule_date": "2026-03-01",
                    "duty_type": "day",
                    "shift_type": "day",
                    "paid_hours": 12,
                    "shift_start_time": "10:00:00",
                    "shift_end_time": "22:00:00",
                    "soc_role": "Supervisor",
                },
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "이보한",
                    "sequence_no": 1,
                    "schedule_date": "2026-03-01",
                    "duty_type": "night",
                    "shift_type": "night",
                    "paid_hours": 10,
                    "shift_start_time": "22:00:00",
                    "shift_end_time": "08:00:00",
                    "soc_role": "Supervisor",
                },
                {
                    "employee_id": "emp-2",
                    "employee_code": "R692-2",
                    "employee_name": "서성원",
                    "sequence_no": 2,
                    "schedule_date": "2026-03-02",
                    "duty_type": "day",
                    "shift_type": "day",
                    "paid_hours": 10,
                    "shift_start_time": "08:00:00",
                    "shift_end_time": "18:00:00",
                    "soc_role": "Officer",
                },
            ],
            tenant_code="srs_korea",
            site_code="R692",
            site_name="Apple_가로수길",
            site_address="서울시 강남구",
            daytime_need_rows=[
                {
                    "work_date": date(2026, 3, 1),
                    "required_count": 4,
                    "raw_text": "4",
                }
            ],
            export_revision="rev-20260309",
            template_version=ARLS_EXPORT_TEMPLATE_VERSION,
            source_version=ARLS_EXPORT_SOURCE_VERSION,
        )
        return workbook

    def test_select_arls_import_visible_sheet_name_accepts_site_named_sheet(self):
        workbook = self._build_sample_workbook()
        workbook[ARLS_SHEET_NAME].title = "Apple_가로수길"
        try:
            self.assertEqual(_select_arls_import_visible_sheet_name(workbook), "Apple_가로수길")
        finally:
            workbook.close()

    def test_preview_import_accepts_site_named_arls_workbook(self):
        workbook = self._build_sample_workbook()
        workbook[ARLS_SHEET_NAME].title = "Apple_가로수길"
        raw_bytes = BytesIO()
        workbook.save(raw_bytes)
        workbook.close()
        upload = _UploadFileStub(
            filename="가로수길 3월 테스트 스케쥴.xlsx",
            raw_bytes=raw_bytes.getvalue(),
        )
        preview_payload = {
            "total_rows": 1,
            "valid_rows": 1,
            "invalid_rows": 0,
            "applicable_rows": 1,
            "unchanged_rows": 0,
            "blocked_rows": 0,
            "warning_rows": 0,
            "preview_rows": [],
            "error_counts": {},
            "diff_counts": {},
            "issues": [],
            "blocked_reasons": [],
            "metadata": {},
            "can_apply": True,
        }
        with patch("app.routers.v1.schedules._resolve_scoped_schedule_site_code", return_value="R692"), patch(
            "app.routers.v1.schedules._resolve_target_tenant",
            return_value={"id": "tenant-1", "tenant_code": "srs_korea"},
        ), patch(
            "app.routers.v1.schedules._resolve_site_context_by_code",
            return_value={"id": "site-1", "site_code": "R692"},
        ), patch(
            "app.routers.v1.schedules._build_schedule_import_preview_result",
            return_value=preview_payload,
        ), patch(
            "app.routers.v1.schedules._persist_schedule_import_preview_batch",
            return_value=(str(uuid.uuid4()), []),
        ):
            result = preview_import(
                file=upload,
                site_code="R692",
                month="2026-03",
                tenant_code="srs_korea",
                conn=_ConnectionStub(),
                user={"role": "developer", "tenant_id": "tenant-1", "id": "user-1"},
            )
        self.assertEqual(result.total_rows, 1)

    def _build_preview_employee_rows(self):
        return [
            {
                "id": "emp-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "R692-1",
                "full_name": "이보한",
                "hire_date": None,
                "leave_date": None,
            },
            {
                "id": "emp-2",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "R692-2",
                "full_name": "서성원",
                "hire_date": None,
                "leave_date": None,
            },
        ]

    def _build_preview_template_rows(self):
        return [
            {
                "id": "tpl-day-12",
                "template_name": "주간 12시간",
                "duty_type": "day",
                "start_time": "10:00:00",
                "end_time": "22:00:00",
                "paid_hours": 12,
            },
            {
                "id": "tpl-day-10",
                "template_name": "주간 10시간",
                "duty_type": "day",
                "start_time": "08:00:00",
                "end_time": "18:00:00",
                "paid_hours": 10,
            },
            {
                "id": "tpl-night-10",
                "template_name": "야간 10시간",
                "duty_type": "night",
                "start_time": "22:00:00",
                "end_time": "08:00:00",
                "paid_hours": 10,
            },
        ]

    def _build_preview_mapping_profile(self):
        return {
            "entries": [
                {
                    "row_type": "day",
                    "numeric_hours": 12,
                    "template_id": "tpl-day-12",
                    "template_name": "주간 12시간",
                    "template_is_active": True,
                },
                {
                    "row_type": "day",
                    "numeric_hours": 10,
                    "template_id": "tpl-day-10",
                    "template_name": "주간 10시간",
                    "template_is_active": True,
                },
                {
                    "row_type": "night",
                    "numeric_hours": 10,
                    "template_id": "tpl-night-10",
                    "template_name": "야간 10시간",
                    "template_is_active": True,
                },
            ]
        }

    def _build_existing_schedule_rows_for_preview(self):
        return [
            {
                "schedule_id": "sch-day-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 1),
                "shift_type": "day",
                "template_id": "tpl-day-12",
                "shift_start_time": "10:00:00",
                "shift_end_time": "22:00:00",
                "paid_hours": 12,
                "schedule_note": None,
                "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                "source_ticket_id": None,
                "duty_type": "day",
                "template_name": "주간 12시간",
                "employee_code": "R692-1",
                "employee_name": "이보한",
            },
            {
                "schedule_id": "sch-night-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 1),
                "shift_type": "night",
                "template_id": "tpl-night-10",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                "source_ticket_id": None,
                "duty_type": "night",
                "template_name": "야간 10시간",
                "employee_code": "R692-1",
                "employee_name": "이보한",
            },
            {
                "schedule_id": "sch-day-2",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-2",
                "schedule_date": date(2026, 3, 2),
                "shift_type": "day",
                "template_id": "tpl-day-10",
                "shift_start_time": "08:00:00",
                "shift_end_time": "18:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                "source_ticket_id": None,
                "duty_type": "day",
                "template_name": "주간 10시간",
                "employee_code": "R692-2",
                "employee_name": "서성원",
            },
        ]

    def test_metadata_validation_accepts_matching_export(self):
        workbook = self._build_sample_workbook()
        metadata = _read_arls_export_metadata(workbook)
        errors = _validate_arls_import_metadata(
            metadata,
            expected_tenant_code="srs_korea",
            expected_site_code="R692",
            expected_month="2026-03",
        )
        self.assertEqual(errors, [])

    def test_metadata_validation_rejects_mismatch(self):
        workbook = self._build_sample_workbook()
        metadata = _read_arls_export_metadata(workbook)
        errors = _validate_arls_import_metadata(
            metadata,
            expected_tenant_code="srs_korea",
            expected_site_code="R999",
            expected_month="2026-03",
        )
        self.assertIn("metadata_mismatch:site_code", errors)

    def test_parse_canonical_sheet_reads_body_need_and_protected_sections(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        rows_meta = _locate_support_section_rows(sheet)
        sheet["D6"] = "연차"
        sheet.cell(row=rows_meta["day_need_row"], column=4, value="5")
        sheet.cell(row=rows_meta["night_need_row"], column=4, value="3")
        sheet.cell(row=rows_meta["night_rows"][0], column=4, value="홍길동")
        sheet.cell(row=rows_meta["work_note_row"], column=4, value="정기 점검")

        parsed = _parse_arls_canonical_import_sheet(sheet)
        body = parsed["body_cells"]
        need_cells = parsed["need_cells"]
        support_cells = parsed["support_cells"]

        leave_cell = next(
            row for row in body
            if row["employee_name"] == "이보한"
            and row["duty_type"] == "overtime"
            and row["schedule_date"].isoformat() == "2026-03-01"
        )
        need_cell = next(row for row in need_cells if row["schedule_date"].isoformat() == "2026-03-01")
        protected_cell = next(
            row for row in support_cells
            if row["schedule_date"].isoformat() == "2026-03-01"
            and row["section_label"] == "야간 지원 근무자"
        )
        support_block = next(
            row for row in parsed["support_blocks"]
            if row["target_date"].isoformat() == "2026-03-01"
            and row["block_type"] == "night_support"
        )
        night_need_cell = next(
            row for row in need_cells
            if row["schedule_date"].isoformat() == "2026-03-01"
            and row["source_block"] == "night_support_required_count"
        )

        self.assertEqual(leave_cell["work_value"], "연차")
        self.assertEqual(need_cell["work_value"], "5")
        self.assertEqual(night_need_cell["work_value"], "3")
        self.assertEqual(protected_cell["work_value"], "홍길동")
        self.assertEqual(support_block["valid_filled_count"], 1)
        self.assertEqual(support_block["invalid_filled_count"], 0)
        self.assertEqual(support_block["required_count_numeric"], 3)
        self.assertEqual(support_block["purpose_text"], "정기 점검")

    def test_extract_date_columns_prefers_actual_header_dates_over_stale_month_label(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        sheet.cell(row=2, column=2, value="2026년 3월")
        start_day = datetime(2026, 2, 1)
        for offset in range(31):
            sheet.cell(row=2, column=4 + offset, value=start_day + timedelta(days=offset))

        date_columns, month_ctx = _extract_arls_date_columns(sheet)

        self.assertEqual(month_ctx, (2026, 2))
        self.assertEqual(date_columns[4].isoformat(), "2026-02-01")
        self.assertEqual(date_columns[5].isoformat(), "2026-02-02")

    def test_parse_daytime_need_value_supports_numeric_and_text(self):
        self.assertEqual(_parse_daytime_need_value("4"), (4, "4"))
        self.assertEqual(_parse_daytime_need_value("5인"), (5, "5인"))
        self.assertEqual(_parse_daytime_need_value("섭외 2인 요청"), (2, "섭외 2인 요청"))
        self.assertEqual(_parse_daytime_need_value("-"), (0, "-"))
        self.assertEqual(_parse_daytime_need_value("미정"), (None, "미정"))

    def test_resolve_import_body_value_uses_leave_marker_convention(self):
        resolved, code, message = _resolve_import_body_value(
            templates=[],
            duty_type="overtime",
            workbook_value="연차",
        )
        self.assertIsNone(code)
        self.assertIsNone(message)
        self.assertEqual(resolved["shift_type"], "off")
        self.assertEqual(resolved["work_value"], "연차")

    def test_resolve_import_body_value_matches_template_for_numeric_hours(self):
        templates = [
            {
                "id": "tpl-day-12",
                "template_name": "주간 12시간",
                "duty_type": "day",
                "start_time": "10:00:00",
                "end_time": "22:00:00",
                "paid_hours": 12,
            },
            {
                "id": "tpl-night-10",
                "template_name": "야간 10시간",
                "duty_type": "night",
                "start_time": "22:00:00",
                "end_time": "08:00:00",
                "paid_hours": 10,
            },
        ]
        mapping_profile = {
            "entries": [
                {
                    "row_type": "night",
                    "numeric_hours": 10,
                    "template_id": "tpl-night-10",
                    "template_name": "야간 10시간",
                    "template_is_active": True,
                }
            ]
        }
        resolved, code, message = _resolve_import_body_value(
            templates=templates,
            mapping_lookup=_build_schedule_import_mapping_lookup(mapping_profile),
            duty_type="night",
            workbook_value="10",
        )
        self.assertIsNone(code)
        self.assertIsNone(message)
        self.assertEqual(resolved["template_id"], "tpl-night-10")
        self.assertEqual(resolved["shift_type"], "night")

    def test_preview_visibility_marks_base_review_rows_as_actionable(self):
        visibility_class, actionable, protected_info_only = _classify_import_preview_visibility({
            "source_block": "body",
            "parsed_semantic_type": "numeric_hours",
            "diff_category": "review",
            "is_blocking": False,
            "is_protected": False,
            "is_valid": True,
        })
        self.assertEqual(visibility_class, "review_actionable")
        self.assertTrue(actionable)
        self.assertFalse(protected_info_only)

    def test_preview_visibility_hides_support_metadata_rows_from_default_preview(self):
        visibility_class, actionable, protected_info_only = _classify_import_preview_visibility({
            "source_block": "night_support_purpose",
            "parsed_semantic_type": "protected_support_purpose",
            "diff_category": "ignored_protected",
            "is_blocking": False,
            "is_protected": True,
            "is_valid": True,
        })
        self.assertEqual(visibility_class, "protected_info_only")
        self.assertFalse(actionable)
        self.assertTrue(protected_info_only)

    def test_preview_visibility_keeps_support_worker_apply_rows_actionable(self):
        visibility_class, actionable, protected_info_only = _classify_import_preview_visibility({
            "source_block": "night_support_worker",
            "parsed_semantic_type": "support_worker",
            "diff_category": "create",
            "decision_stage": "apply",
            "is_blocking": False,
            "is_protected": False,
            "is_valid": True,
        })
        self.assertEqual(visibility_class, "apply_actionable")
        self.assertTrue(actionable)
        self.assertFalse(protected_info_only)

    def test_preview_visibility_keeps_blocked_support_rows_actionable(self):
        visibility_class, actionable, protected_info_only = _classify_import_preview_visibility({
            "source_block": "night_support_required_count",
            "parsed_semantic_type": "support_demand",
            "diff_category": "invalid",
            "is_blocking": True,
            "is_protected": False,
            "is_valid": False,
        })
        self.assertEqual(visibility_class, "blocked")
        self.assertTrue(actionable)
        self.assertFalse(protected_info_only)

    def test_resolve_import_body_value_requires_mapping_when_lookup_is_supplied(self):
        templates = [
            {
                "id": "tpl-day-12",
                "template_name": "주간 12시간",
                "duty_type": "day",
                "start_time": "10:00:00",
                "end_time": "22:00:00",
                "paid_hours": 12,
            }
        ]
        resolved, code, message = _resolve_import_body_value(
            templates=templates,
            mapping_lookup={},
            duty_type="day",
            workbook_value="12",
        )
        self.assertEqual(resolved, {})
        self.assertEqual(code, "TEMPLATE_MAPPING_MISSING")
        self.assertEqual(message, "매핑 가능한 근무 템플릿이 없습니다.")

    def test_validate_mapping_profile_requirements_requires_active_profile_for_numeric_hours(self):
        workbook = self._build_sample_workbook()
        parsed = _parse_arls_canonical_import_sheet(workbook[ARLS_SHEET_NAME])
        issues, blocked_reasons, missing_entries = _validate_mapping_profile_requirements(
            body_cells=parsed["body_cells"],
            mapping_profile=None,
            mapping_lookup={},
        )
        self.assertTrue(issues)
        self.assertTrue(blocked_reasons)
        self.assertTrue(missing_entries)
        self.assertIn("주간근무 10시간", " ".join(blocked_reasons))

    def test_format_schedule_import_mapping_requirement_label_is_operator_friendly(self):
        self.assertEqual(
            _format_schedule_import_mapping_requirement_label("day", 10),
            "주간근무 10시간",
        )
        self.assertEqual(
            _format_schedule_import_mapping_requirement_label("night", "10.00"),
            "야간근무 10시간",
        )

    def test_parse_support_worker_cell_rejects_multi_person_input(self):
        parsed = _parse_support_worker_cell("BK 박준연 / BK 김민수")
        self.assertEqual(parsed["issue_code"], "MULTI_PERSON_CELL")
        self.assertTrue(parsed["is_filled"])

    def test_parse_support_worker_cell_marks_plain_and_self_prefixed_names_as_internal_match_targets(self):
        plain = _parse_support_worker_cell("민경민")
        prefixed = _parse_support_worker_cell("자체 민경민")
        external = _parse_support_worker_cell("BK 박준연")
        affiliated_external = _parse_support_worker_cell("ks 민경민")

        self.assertEqual(plain["worker_type"], "INTERNAL")
        self.assertEqual(plain["worker_name"], "민경민")
        self.assertTrue(plain["requires_employee_match"])
        self.assertEqual(prefixed["worker_type"], "INTERNAL")
        self.assertEqual(prefixed["worker_name"], "민경민")
        self.assertTrue(prefixed["requires_employee_match"])
        self.assertTrue(prefixed["self_staff"])
        self.assertEqual(external["worker_type"], "BK")
        self.assertEqual(external["worker_name"], "박준연")
        self.assertFalse(external["requires_employee_match"])
        self.assertEqual(affiliated_external["worker_type"], "F")
        self.assertEqual(affiliated_external["worker_name"], "민경민")
        self.assertEqual(affiliated_external["affiliation"], "KS")
        self.assertFalse(affiliated_external["requires_employee_match"])

    def test_resolve_import_support_worker_match_accepts_existing_self_prefixed_employee(self):
        employee_index = _build_employee_name_index(
            [
                {
                    "id": "emp-1",
                    "employee_code": "R692-1",
                    "full_name": "민경민",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                }
            ]
        )

        employee_row, issue_code, issue_message = _resolve_import_support_worker_match(
            parsed_worker=_parse_support_worker_cell("자체 민경민"),
            employee_index=employee_index,
            schedule_date=date(2026, 3, 5),
        )

        self.assertIsNotNone(employee_row)
        self.assertEqual(employee_row["full_name"], "민경민")
        self.assertIsNone(issue_code)
        self.assertIsNone(issue_message)

    def test_resolve_import_support_worker_match_blocks_missing_prefixed_employee(self):
        employee_index = _build_employee_name_index(
            [
                {
                    "id": "emp-1",
                    "employee_code": "R692-1",
                    "full_name": "민경민",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                }
            ]
        )

        employee_row, issue_code, issue_message = _resolve_import_support_worker_match(
            parsed_worker=_parse_support_worker_cell("자체 조정현"),
            employee_index=employee_index,
            schedule_date=date(2026, 3, 5),
        )

        self.assertIsNone(employee_row)
        self.assertEqual(issue_code, "EMPLOYEE_MATCH_FAILED")
        self.assertEqual(issue_message, "해당 스토어 직원 명단에 없는 자체 지원근무자입니다: 조정현")

    def test_resolve_import_support_worker_match_accepts_plain_existing_employee_name(self):
        employee_index = _build_employee_name_index(
            [
                {
                    "id": "emp-1",
                    "employee_code": "R692-1",
                    "full_name": "민경민",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                }
            ]
        )

        employee_row, issue_code, issue_message = _resolve_import_support_worker_match(
            parsed_worker=_parse_support_worker_cell("민경민"),
            employee_index=employee_index,
            schedule_date=date(2026, 3, 5),
        )

        self.assertIsNotNone(employee_row)
        self.assertEqual(employee_row["employee_code"], "R692-1")
        self.assertIsNone(issue_code)
        self.assertIsNone(issue_message)

    def test_parse_support_worker_cell_blocks_blank_internal_prefix_without_name(self):
        parsed = _parse_support_worker_cell("자체")

        self.assertEqual(parsed["issue_code"], "WORKER_CELL_INVALID")
        self.assertTrue(parsed["is_filled"])

    def test_parse_support_worker_cell_treats_zero_as_blank(self):
        parsed = _parse_support_worker_cell("0")

        self.assertEqual(parsed["semantic_type"], "blank")
        self.assertFalse(parsed["is_filled"])
        self.assertIsNone(parsed["issue_code"])

    def test_validate_support_shift_worker_applies_same_store_internal_candidate(self):
        employee_index = _build_employee_name_index(
            [
                {
                    "id": "emp-1",
                    "tenant_id": "tenant-1",
                    "company_id": "company-1",
                    "site_id": "site-1",
                    "employee_code": "R692-1",
                    "full_name": "민경민",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                }
            ]
        )

        result = _validate_support_shift_worker(
            _ConnectionStub(),
            tenant_id="tenant-1",
            site_id="site-1",
            parsed_worker=_parse_support_worker_cell("자체 민경민"),
            employee_index=employee_index,
            schedule_date=date(2026, 3, 5),
        )

        self.assertEqual(result["decision_stage"], "apply")
        self.assertEqual(result["support_origin_type"], "same_store_internal_candidate")
        self.assertIsNone(result["validation_code"])
        self.assertEqual(result["employee_row"]["full_name"], "민경민")

    def test_validate_support_shift_worker_blocks_missing_same_store_internal_candidate(self):
        employee_index = _build_employee_name_index(
            [
                {
                    "id": "emp-1",
                    "tenant_id": "tenant-1",
                    "company_id": "company-1",
                    "site_id": "site-1",
                    "employee_code": "R692-1",
                    "full_name": "민경민",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                }
            ]
        )

        result = _validate_support_shift_worker(
            _ConnectionStub(cursor_responses=[[]]),
            tenant_id="tenant-1",
            site_id="site-1",
            parsed_worker=_parse_support_worker_cell("자체 조정현"),
            employee_index=employee_index,
            schedule_date=date(2026, 3, 5),
        )

        self.assertEqual(result["decision_stage"], "block")
        self.assertEqual(result["validation_code"], "EMPLOYEE_MATCH_FAILED")
        self.assertEqual(result["validation_error"], "해당 스토어 직원 명단에 없는 자체 지원근무자입니다: 조정현")

    def test_validate_support_shift_worker_marks_same_store_duplicate_name_for_review(self):
        employee_index = _build_employee_name_index(
            [
                {
                    "id": "emp-1",
                    "tenant_id": "tenant-1",
                    "company_id": "company-1",
                    "site_id": "site-1",
                    "employee_code": "R692-1",
                    "full_name": "민경민",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                },
                {
                    "id": "emp-2",
                    "tenant_id": "tenant-1",
                    "company_id": "company-1",
                    "site_id": "site-1",
                    "employee_code": "R692-2",
                    "full_name": "민경민",
                    "hire_date": date(2025, 1, 1),
                    "leave_date": None,
                },
            ]
        )

        result = _validate_support_shift_worker(
            _ConnectionStub(),
            tenant_id="tenant-1",
            site_id="site-1",
            parsed_worker=_parse_support_worker_cell("자체 민경민"),
            employee_index=employee_index,
            schedule_date=date(2026, 3, 5),
        )

        self.assertEqual(result["decision_stage"], "review")
        self.assertEqual(result["validation_code"], "SUPPORT_INTERNAL_MATCH_AMBIGUOUS_REVIEW")

    def test_validate_support_shift_worker_marks_other_site_candidate_for_review(self):
        employee_index = _build_employee_name_index([])
        tenant_rows = [
            {
                "id": "emp-9",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-9",
                "employee_code": "R999-1",
                "full_name": "민경민",
                "hire_date": date(2025, 1, 1),
                "leave_date": None,
            }
        ]

        result = _validate_support_shift_worker(
            _ConnectionStub(cursor_responses=[tenant_rows]),
            tenant_id="tenant-1",
            site_id="site-1",
            parsed_worker=_parse_support_worker_cell("민경민"),
            employee_index=employee_index,
            schedule_date=date(2026, 3, 5),
        )

        self.assertEqual(result["decision_stage"], "review")
        self.assertEqual(result["validation_code"], "SUPPORT_OTHER_SITE_MATCH_REVIEW")

    def test_validate_support_shift_worker_applies_affiliated_external_without_store_match(self):
        result = _validate_support_shift_worker(
            _ConnectionStub(),
            tenant_id="tenant-1",
            site_id="site-1",
            parsed_worker=_parse_support_worker_cell("rk 김민지"),
            employee_index=_build_employee_name_index([]),
            schedule_date=date(2026, 3, 5),
        )

        self.assertEqual(result["decision_stage"], "apply")
        self.assertEqual(result["support_origin_type"], "f_support_worker")
        self.assertIsNone(result["validation_code"])
        self.assertIsNone(result["employee_row"])

    def test_locate_support_section_rows_separates_day_and_night_meta_rows(self):
        workbook = self._build_sample_workbook()
        rows_meta = _locate_support_section_rows(workbook[ARLS_SHEET_NAME])
        self.assertEqual(rows_meta["day_vendor_count_row"], 54)
        self.assertEqual(rows_meta["day_need_row"], 55)
        self.assertEqual(rows_meta["night_vendor_count_row"], 64)
        self.assertEqual(rows_meta["night_need_row"], 65)
        self.assertEqual(rows_meta["work_note_row"], 66)

    def test_resolve_shift_type_from_duty_type_keeps_overtime_distinct(self):
        self.assertEqual(_resolve_shift_type_from_duty_type("overtime"), "overtime")

    def test_build_support_request_rows_from_import_payloads_keeps_meaningful_day_and_night_scopes(self):
        rows = _build_support_request_rows_from_import_payloads(
            [
                {
                    "source_block": "sentrix_support_ticket",
                    "schedule_date": date(2026, 3, 2),
                    "shift_type": "day",
                    "request_count": 2,
                    "work_value": "섭외 2인 요청",
                    "purpose_text": "",
                    "detail_json": {"external_count_raw": "1"},
                },
                {
                    "source_block": "sentrix_support_ticket",
                    "schedule_date": date(2026, 3, 2),
                    "shift_type": "night",
                    "request_count": 3,
                    "work_value": "섭외 3인 요청",
                    "purpose_text": "프로젝트",
                    "detail_json": {"required_row_no": 65},
                },
                {
                    "source_block": "sentrix_support_ticket",
                    "schedule_date": date(2026, 3, 3),
                    "shift_type": "day",
                    "request_count": 0,
                    "work_value": "",
                    "detail_json": {},
                },
                {
                    "source_block": "sentrix_support_ticket",
                    "schedule_date": date(2026, 3, 4),
                    "shift_type": "day",
                    "request_count": 1,
                    "is_blocking": True,
                    "work_value": "1",
                    "detail_json": {},
                },
            ]
        )

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["shift_kind"], "day")
        self.assertEqual(rows[0]["request_count"], 2)
        self.assertEqual(rows[0]["detail_json"]["required_count_raw"], "섭외 2인 요청")
        self.assertEqual(rows[1]["shift_kind"], "night")
        self.assertEqual(rows[1]["work_purpose"], "프로젝트")

    @patch("app.routers.v1.schedules._read_live_sentrix_support_request_rows", return_value=[])
    @patch("app.routers.v1.schedules._load_schedule_import_payload_rows")
    def test_read_monthly_support_request_rows_for_export_prefers_source_batch_payloads(
        self,
        mock_load_payload_rows,
        _mock_live_rows,
    ):
        mock_load_payload_rows.return_value = [
            {
                "source_block": "sentrix_support_ticket",
                "schedule_date": date(2026, 3, 15),
                "shift_type": "night",
                "request_count": 2,
                "work_value": "섭외 2인 요청",
                "purpose_text": "야간 작업",
                "detail_json": {"external_count_raw": "0"},
            }
        ]

        rows = _read_monthly_support_request_rows_for_export(
            None,
            tenant_id="tenant-1",
            site_id="site-1",
            month_key="2026-03",
            source_batch_id="batch-123",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["work_date"].isoformat(), "2026-03-15")
        self.assertEqual(rows[0]["shift_kind"], "night")
        self.assertEqual(rows[0]["request_count"], 2)
        self.assertEqual(rows[0]["work_purpose"], "야간 작업")

    def test_schedule_display_meta_distinguishes_off_holiday_and_annual_leave(self):
        self.assertEqual(
            _resolve_schedule_display_meta({"shift_type": "off", "schedule_note": "연차"})["type"],
            "annual_leave",
        )
        self.assertEqual(
            _resolve_schedule_display_meta({"shift_type": "off", "schedule_note": "반차"})["type"],
            "half_leave",
        )
        self.assertEqual(
            _resolve_schedule_display_meta({"shift_type": "off", "schedule_note": ""})["label"],
            "휴무",
        )
        self.assertEqual(
            _resolve_schedule_display_meta({"shift_type": "holiday", "schedule_note": None})["label"],
            "공휴일",
        )

    def test_leader_role_resolution_preserves_admin_labels(self):
        self.assertEqual(_resolve_leader_candidate_role_key("GUARD", "hq_admin"), "HQ_ADMIN")
        self.assertEqual(_resolve_leader_candidate_role_key("TEAM_MANAGER", "supervisor"), "SUPERVISOR")
        self.assertEqual(_resolve_leader_candidate_role_key("GUARD", "vice_supervisor"), "VICE_SUPERVISOR")
        self.assertEqual(_resolve_leader_candidate_role_key("GUARD", "officer"), "GUARD")
        self.assertEqual(_leader_candidate_role_label("HQ_ADMIN"), "HQ Admin")
        self.assertEqual(_leader_candidate_role_label("SUPERVISOR"), "Supervisor")
        self.assertEqual(_leader_candidate_role_label("VICE_SUPERVISOR"), "Vice Supervisor")
        self.assertEqual(_leader_candidate_role_priority("HQ_ADMIN"), 0)
        self.assertEqual(_leader_candidate_role_priority("GUARD"), 3)

    def test_parse_canonical_sheet_ignores_placeholder_zero_employee_rows(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        sheet["B5"] = "0"
        sheet["B8"] = 0

        parsed = _parse_arls_canonical_import_sheet(sheet)

        self.assertEqual(parsed["body_cells"], [])
        self.assertFalse(any(row.get("employee_name") in {"0", 0} for row in parsed["body_cells"]))

    def test_support_section_label_normalization_accepts_newlines_and_spacing(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        rows_meta = _locate_support_section_rows(sheet)
        sheet.cell(row=rows_meta["weekly_rows"][0], column=2, value="주간\n추가 근무자")
        sheet.cell(row=rows_meta["night_rows"][0], column=2, value="야간\n추가 근무자")
        sheet.cell(row=rows_meta["day_vendor_count_row"], column=3, value="외부인원 \n투입 수")
        sheet.cell(row=rows_meta["work_note_row"], column=3, value="작업 내용")
        sheet.cell(row=rows_meta["work_note_row"], column=4, value="정기 점검")

        relocated = _locate_support_section_rows(sheet)
        parsed = _parse_arls_canonical_import_sheet(sheet)
        night_block = next(
            row for row in parsed["support_blocks"]
            if row["target_date"].isoformat() == "2026-03-01"
            and row["block_type"] == "night_support"
        )

        self.assertEqual(relocated["weekly_rows"], rows_meta["weekly_rows"])
        self.assertEqual(relocated["night_rows"], rows_meta["night_rows"])
        self.assertEqual(relocated["day_vendor_count_row"], rows_meta["day_vendor_count_row"])
        self.assertEqual(relocated["work_note_row"], rows_meta["work_note_row"])
        self.assertEqual(night_block["purpose_text"], "정기 점검")

    def test_blank_support_required_count_with_no_demand_is_non_blocking(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        rows_meta = _locate_support_section_rows(sheet)
        sheet.cell(row=rows_meta["day_need_row"], column=4).value = None
        sheet.cell(row=rows_meta["day_vendor_count_row"], column=4).value = None

        parsed = _parse_arls_canonical_import_sheet(sheet)
        block = next(
            row for row in parsed["support_blocks"]
            if row["target_date"].isoformat() == "2026-03-01"
            and row["block_type"] == "day_support"
        )
        need_cell = next(
            row for row in parsed["need_cells"]
            if row["schedule_date"].isoformat() == "2026-03-01"
            and row["source_block"] == "day_support_required_count"
        )

        self.assertEqual(block["required_count_state"], "no_demand")
        self.assertEqual(block["required_count_numeric"], 0)
        self.assertEqual(need_cell["parsed_semantic_type"], "no_demand")
        self.assertIsNone(need_cell["issue_code"])
        self.assertNotIn("SUPPORT_BLOCK_REQUIRED_COUNT_INVALID", block["issues"])

    def test_blank_support_required_count_with_meaningful_payload_is_blocking(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        rows_meta = _locate_support_section_rows(sheet)
        sheet.cell(row=rows_meta["day_need_row"], column=4).value = None
        sheet.cell(row=rows_meta["day_vendor_count_row"], column=4, value=2)

        parsed = _parse_arls_canonical_import_sheet(sheet)
        block = next(
            row for row in parsed["support_blocks"]
            if row["target_date"].isoformat() == "2026-03-01"
            and row["block_type"] == "day_support"
        )
        need_cell = next(
            row for row in parsed["need_cells"]
            if row["schedule_date"].isoformat() == "2026-03-01"
            and row["source_block"] == "day_support_required_count"
        )

        self.assertEqual(block["required_count_state"], "invalid_blank")
        self.assertIsNone(block["required_count_numeric"])
        self.assertEqual(need_cell["issue_code"], "SUPPORT_BLOCK_REQUIRED_COUNT_INVALID")
        self.assertIn("SUPPORT_BLOCK_REQUIRED_COUNT_INVALID", block["issues"])

    def test_build_import_current_body_index_uses_base_rows_only(self):
        index = _build_import_current_body_index_from_existing_schedule_rows(
            [
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "이보한",
                    "schedule_date": date(2026, 3, 1),
                    "shift_type": "day",
                    "duty_type": "day",
                    "shift_start_time": "08:00:00",
                    "shift_end_time": "18:00:00",
                    "paid_hours": 10,
                    "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                },
                {
                    "employee_id": "emp-1",
                    "employee_code": "R692-1",
                    "employee_name": "이보한",
                    "schedule_date": date(2026, 3, 1),
                    "shift_type": "day",
                    "duty_type": "day",
                    "shift_start_time": "10:00:00",
                    "shift_end_time": "22:00:00",
                    "paid_hours": 12,
                    "source": "manual_override",
                },
            ]
        )

        current_row = index[("이보한", "day", "2026-03-01")]
        self.assertEqual(current_row["work_value"], "10")

    @patch("app.routers.v1.schedules._build_schedule_export_revision", return_value="rev-preview")
    @patch("app.routers.v1.schedules._read_monthly_support_request_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._read_monthly_daytime_need_rows_for_export")
    @patch("app.routers.v1.schedules._read_monthly_employee_overnight_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._read_monthly_overnight_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._read_monthly_support_assignment_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._build_export_rows_from_board_payload", return_value=[])
    @patch("app.routers.v1.schedules.monthly_board_lite", return_value={})
    @patch("app.routers.v1.schedules._read_monthly_board_rows_for_export", return_value=[])
    def test_collect_monthly_export_context_allows_empty_employee_blocks_for_preview(
        self,
        _mock_board_rows,
        _mock_board_payload,
        _mock_export_rows,
        _mock_support_rows,
        _mock_overnight_rows,
        _mock_employee_overnight_rows,
        mock_daytime_need_rows,
        _mock_support_request_rows,
        _mock_export_revision,
    ):
        mock_daytime_need_rows.return_value = [
            {
                "work_date": date(2026, 3, 1),
                "required_count": 2,
                "raw_text": "2",
            }
        ]

        export_ctx = _collect_monthly_export_context(
            None,
            target_tenant={"id": "tenant-1", "tenant_code": "TENANT"},
            site_row={"id": "site-1", "site_code": "R692", "site_name": "Apple_가로수길", "address": "서울시 강남구"},
            month_key="2026-03",
            user={},
            allow_empty_employee_blocks=True,
        )

        self.assertEqual(export_ctx["employee_blocks"], [])
        self.assertEqual(export_ctx["export_revision"], "rev-preview")
        need_cell = next(
            row for row in export_ctx["parsed_sheet"]["need_cells"]
            if row["source_block"] == "day_support_required_count"
            and row["schedule_date"].isoformat() == "2026-03-01"
        )
        self.assertEqual(need_cell["work_value"], "2")

    @patch("app.routers.v1.schedules._build_schedule_export_revision", return_value="rev-empty-preview")
    @patch("app.routers.v1.schedules._read_monthly_support_request_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._read_monthly_daytime_need_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._read_monthly_employee_overnight_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._read_monthly_overnight_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._read_monthly_support_assignment_rows_for_export", return_value=[])
    @patch("app.routers.v1.schedules._build_export_rows_from_board_payload", return_value=[])
    @patch("app.routers.v1.schedules.monthly_board_lite", return_value={})
    @patch("app.routers.v1.schedules._read_monthly_board_rows_for_export", return_value=[])
    def test_collect_monthly_export_context_returns_empty_current_context_when_no_export_data_for_preview(
        self,
        _mock_board_rows,
        _mock_board_payload,
        _mock_export_rows,
        _mock_support_rows,
        _mock_overnight_rows,
        _mock_employee_overnight_rows,
        _mock_daytime_need_rows,
        _mock_support_request_rows,
        _mock_export_revision,
    ):
        export_ctx = _collect_monthly_export_context(
            None,
            target_tenant={"id": "tenant-1", "tenant_code": "TENANT"},
            site_row={"id": "site-2", "site_code": "M001", "site_name": "명동", "address": "서울시 중구"},
            month_key="2026-03",
            user={},
            allow_empty_employee_blocks=True,
        )

        self.assertEqual(export_ctx["rows"], [])
        self.assertEqual(export_ctx["employee_blocks"], [])
        self.assertEqual(export_ctx["export_revision"], "rev-empty-preview")
        self.assertEqual(export_ctx["parsed_sheet"]["body_cells"], [])
        self.assertEqual(export_ctx["parsed_sheet"]["need_cells"], [])
        self.assertEqual(export_ctx["parsed_sheet"]["support_cells"], [])

    def test_filter_support_assignment_rows_for_active_scopes_drops_stale_rows(self):
        support_rows = [
            {
                "work_date": date(2026, 3, 3),
                "support_period": "night",
                "slot_index": 1,
                "worker_type": "F",
                "name": "형준",
                "affiliation": "F",
            },
            {
                "work_date": date(2026, 3, 3),
                "support_period": "day",
                "slot_index": 1,
                "worker_type": "BK",
                "name": "강상모",
                "affiliation": "BK",
            },
        ]
        support_request_rows = [
            {
                "work_date": date(2026, 3, 3),
                "shift_kind": "day",
                "request_count": 1,
                "status": SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS,
            }
        ]

        filtered = _filter_support_assignment_rows_for_active_scopes(
            support_rows,
            support_request_rows,
        )

        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0]["support_period"], "day")
        self.assertEqual(filtered[0]["name"], "강상모")

    def test_build_schedule_import_preview_result_keeps_full_snapshot_rows_for_safe_reupload(self):
        workbook = self._build_sample_workbook()
        parsed_sheet = _parse_arls_canonical_import_sheet(workbook[ARLS_SHEET_NAME])
        export_ctx = {
            "parsed_sheet": parsed_sheet,
            "export_revision": "rev-20260309",
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }
        support_request_rows = [
            {
                "work_date": date(2026, 3, 1),
                "shift_kind": "day",
                "request_count": 4,
                "work_purpose": None,
                "status": SENTRIX_SUPPORT_REQUEST_ACTIVE_STATUS,
                "detail_json": {
                    "required_count_raw": "4",
                    "required_count_numeric": 4,
                },
            }
        ]

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=self._build_existing_schedule_rows_for_preview(),
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=support_request_rows,
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        unchanged_body_row = next(
            row
            for row in preview["resolved_rows"]
            if row["source_block"] == "body"
            and row["employee_name"] == "이보한"
            and row["schedule_date"] == date(2026, 3, 1)
            and row["duty_type"] == "day"
        )
        unchanged_support_scope = next(
            row
            for row in preview["support_ticket_rows"]
            if row["schedule_date"] == date(2026, 3, 1)
            and row["shift_type"] == "day"
        )

        self.assertEqual(unchanged_body_row["diff_category"], "unchanged")
        self.assertEqual(unchanged_support_scope["diff_category"], "unchanged")
        self.assertEqual(unchanged_support_scope["request_count"], 4)
        self.assertFalse(
            any(
                row["source_block"] == "body"
                and row["employee_name"] == "이보한"
                and row["schedule_date"] == date(2026, 3, 1)
                for row in preview["preview_rows"]
            )
        )
        self.assertTrue(preview["can_apply"])
        workbook.close()

    def test_build_schedule_import_preview_result_keeps_support_apply_rows_non_blocking(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        sheet.cell(row=48, column=8).value = "ks 민경민"
        sheet.cell(row=55, column=8).value = "1"
        sheet.cell(row=58, column=8).value = "자체 이보한"
        sheet.cell(row=65, column=8).value = "1"
        sheet.cell(row=66, column=8).value = "LiDAR 작업"

        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {"body_cells": [], "need_cells": [], "support_cells": []},
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        external_support_row = next(
            row
            for row in preview["resolved_rows"]
            if row["source_block"] == "day_support_worker"
            and row["work_value"] == "ks 민경민"
        )
        self_staff_support_row = next(
            row
            for row in preview["resolved_rows"]
            if row["source_block"] == "night_support_worker"
            and row["work_value"] == "자체 이보한"
        )

        self.assertEqual(external_support_row["decision_stage"], "apply")
        self.assertFalse(external_support_row["is_blocking"])
        self.assertIsNone(external_support_row["validation_code"])
        self.assertEqual(self_staff_support_row["decision_stage"], "apply")
        self.assertFalse(self_staff_support_row["is_blocking"])
        self.assertIsNone(self_staff_support_row["validation_code"])
        workbook.close()

    def test_build_schedule_import_preview_result_exposes_blocking_support_ticket_rows(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        rows_meta = _locate_support_section_rows(sheet)
        sheet.cell(row=rows_meta["day_need_row"], column=4).value = None
        sheet.cell(row=rows_meta["day_vendor_count_row"], column=4, value=2)

        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {"body_cells": [], "need_cells": [], "support_cells": []},
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        blocking_ticket = next(
            row
            for row in preview["preview_rows"]
            if row["source_block"] == "sentrix_support_ticket"
            and row["schedule_date"] == date(2026, 3, 1)
            and row["shift_type"] == "day"
        )

        self.assertGreater(preview["blocked_rows"], 0)
        self.assertTrue(blocking_ticket["is_blocking"])
        self.assertEqual(
            blocking_ticket["validation_code"],
            "SUPPORT_BLOCK_REQUIRED_COUNT_INVALID",
        )
        self.assertEqual(blocking_ticket["source_col"], "D")
        workbook.close()

    def test_build_schedule_import_preview_result_hides_internal_support_row_when_same_slot_is_already_covered_by_body_schedule(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        sheet.cell(row=58, column=4).value = "자체 이보한"
        sheet.cell(row=65, column=4).value = "섭외 1인 요청"
        sheet.cell(row=66, column=4).value = "매장 지원"

        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {"body_cells": [], "need_cells": [], "support_cells": []},
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        self.assertFalse(
            any(
                row["source_block"] == "night_support_worker"
                and row["employee_name"] == "이보한"
                and row["schedule_date"] == date(2026, 3, 1)
                for row in preview["preview_rows"]
            )
        )
        self.assertEqual(preview["blocked_rows"], 0)
        workbook.close()

    def test_build_schedule_import_preview_result_ignores_blank_base_cell_for_foreign_lineage_slot(self):
        workbook = self._build_sample_workbook()
        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {
                "body_cells": [
                    {
                        "employee_name": "이보한",
                        "duty_type": "night",
                        "schedule_date": date(2026, 3, 5),
                        "work_value": "10",
                    }
                ],
                "need_cells": [],
                "support_cells": [],
            },
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }
        foreign_existing_rows = [
            {
                "schedule_id": "support-night-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "night",
                "template_id": "tpl-night-10",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_SUPPORT_UPLOAD_INTERNAL_SOURCE,
                "source_ticket_id": None,
                "duty_type": "night",
                "template_name": "야간 10시간",
                "employee_code": "R692-1",
                "employee_name": "이보한",
            }
        ]

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=foreign_existing_rows,
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        self.assertFalse(
            any(
                row["source_block"] == "body"
                and row["employee_name"] == "이보한"
                and row["schedule_date"] == date(2026, 3, 5)
                and row["duty_type"] == "night"
                for row in preview["preview_rows"]
            )
        )
        self.assertTrue(preview["can_apply"])
        self.assertEqual(preview["blocked_rows"], 0)
        workbook.close()

    def test_build_schedule_import_preview_result_ignores_stale_internal_support_rows_without_active_request(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        sheet.cell(row=58, column=6, value="자체 서성원")
        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {"body_cells": [], "need_cells": [], "support_cells": []},
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }
        stale_existing_rows = [
            {
                "schedule_id": "support-day-stale-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-2",
                "schedule_date": date(2026, 3, 3),
                "shift_type": "day",
                "template_id": "tpl-day-stale",
                "shift_start_time": "12:00:00",
                "shift_end_time": "23:00:00",
                "paid_hours": 11,
                "schedule_note": None,
                "source": ARLS_SUPPORT_UPLOAD_INTERNAL_SOURCE,
                "source_ticket_id": None,
                "duty_type": "day",
                "template_name": "주간 11시간",
                "employee_code": "R692-2",
                "employee_name": "서성원",
            }
        ]

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=stale_existing_rows,
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        target_row = next(
            row
            for row in preview["preview_rows"]
            if row["source_block"] == "night_support_worker"
            and row["schedule_date"] == date(2026, 3, 3)
            and row["work_value"] == "자체 서성원"
        )
        self.assertIsNone(target_row["validation_code"])
        self.assertFalse(target_row["is_blocking"])
        self.assertNotEqual(target_row["validation_error"], "같은 날짜의 기존 일정과 시간이 겹칩니다.")
        workbook.close()

    def test_build_schedule_import_preview_result_allows_matching_foreign_lineage_value_as_noop(self):
        workbook = self._build_sample_workbook()
        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {"body_cells": [], "need_cells": [], "support_cells": []},
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }
        foreign_existing_rows = [
            {
                "schedule_id": "support-night-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 6),
                "shift_type": "night",
                "template_id": "tpl-night-10",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_SUPPORT_UPLOAD_INTERNAL_SOURCE,
                "source_ticket_id": None,
                "duty_type": "night",
                "template_name": "야간 10시간",
                "employee_code": "R692-1",
                "employee_name": "이보한",
            }
        ]

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=foreign_existing_rows,
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        self.assertFalse(
            any(
                row["source_block"] == "body"
                and row["employee_name"] == "이보한"
                and row["schedule_date"] == date(2026, 3, 6)
                and row["duty_type"] == "night"
                for row in preview["preview_rows"]
            )
        )
        self.assertEqual(preview["blocked_rows"], 0)
        workbook.close()

    def test_build_schedule_import_preview_result_links_existing_base_shift_when_internal_template_hours_are_null(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        sheet.cell(row=58, column=4, value="자체 이보한")
        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {"body_cells": [], "need_cells": [], "support_cells": []},
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }
        template_rows = self._build_preview_template_rows()
        for row in template_rows:
            if row["id"] == "tpl-night-10":
                row["paid_hours"] = None

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=template_rows,
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=self._build_existing_schedule_rows_for_preview(),
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        self.assertFalse(
            any(
                row["source_block"] == "night_support_worker"
                and row["schedule_date"] == date(2026, 3, 1)
                and row["work_value"] == "자체 이보한"
                and row.get("validation_code") == "TIME_CONFLICT"
                for row in preview["preview_rows"]
            )
        )
        workbook.close()

    def test_build_schedule_import_preview_result_matches_existing_support_worker_by_value_not_row_position(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        rows_meta = _locate_support_section_rows(sheet)
        target_col = 6  # 2026-03-03
        sheet.cell(row=rows_meta["night_rows"][0], column=target_col, value="자체 서성원")
        sheet.cell(row=rows_meta["night_rows"][1], column=target_col, value="BK 김진수")

        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {
                "body_cells": [],
                "need_cells": [],
                "support_cells": [
                    {
                        "row_no": rows_meta["night_rows"][0],
                        "col_no": target_col,
                        "source_sheet": sheet.title,
                        "schedule_date": date(2026, 3, 3),
                        "work_value": "BK 김진수",
                        "source_block": "night_support_worker",
                        "section_label": "야간 지원 근무자",
                        "parsed_semantic_type": "worker",
                    }
                ],
            },
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        create_row = next(
            row
            for row in preview["preview_rows"]
            if row["source_block"] == "night_support_worker"
            and row["schedule_date"] == date(2026, 3, 3)
            and row["work_value"] == "자체 서성원"
        )
        self.assertEqual(create_row["diff_category"], "create")
        self.assertIsNone(create_row["current_work_value"])
        self.assertFalse(
            any(
                row["source_block"] == "night_support_worker"
                and row["schedule_date"] == date(2026, 3, 3)
                and row["work_value"] == "BK 김진수"
                for row in preview["preview_rows"]
            )
        )
        workbook.close()

    def test_build_schedule_import_preview_result_treats_same_night_time_as_existing_even_when_template_differs(self):
        workbook = self._build_sample_workbook()
        sheet = workbook[ARLS_SHEET_NAME]
        rows_meta = _locate_support_section_rows(sheet)
        target_col = 6  # 2026-03-03
        sheet.cell(row=rows_meta["night_rows"][0], column=target_col, value="자체 서성원")

        export_ctx = {
            "export_revision": "rev-20260309",
            "parsed_sheet": {"body_cells": [], "need_cells": [], "support_cells": []},
            "support_rows": [],
            "overnight_rows": [],
            "employee_overnight_rows": [],
        }
        existing_rows = [
            {
                "schedule_id": "soc-night-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-2",
                "schedule_date": date(2026, 3, 3),
                "shift_type": "night",
                "template_id": "tpl-night-soc",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": "SOC",
                "source_ticket_id": "ticket-1",
                "duty_type": "night",
                "template_name": "SOC 야간 10시간",
                "employee_code": "R692-2",
                "employee_name": "서성원",
            }
        ]

        with patch("app.routers.v1.schedules._collect_monthly_export_context", return_value=export_ctx), patch(
            "app.routers.v1.schedules._fetch_schedule_templates",
            return_value=self._build_preview_template_rows(),
        ), patch(
            "app.routers.v1.schedules._fetch_active_schedule_import_mapping_profile",
            return_value=self._build_preview_mapping_profile(),
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=self._build_preview_employee_rows(),
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=existing_rows,
        ), patch(
            "app.routers.v1.schedules._read_monthly_support_request_rows_for_export",
            return_value=[],
        ):
            preview = _build_schedule_import_preview_result(
                None,
                workbook=workbook,
                target_tenant={"id": "tenant-1", "tenant_code": "srs_korea"},
                scope_site={"id": "site-1", "site_code": "R692", "company_id": "company-1", "company_code": "APPLE"},
                selected_month="2026-03",
                user={"id": "user-1", "role": "developer"},
                filename="preview.xlsx",
                file_sha256="sha-preview",
            )

        self.assertFalse(
            any(
                row["source_block"] == "night_support_worker"
                and row["schedule_date"] == date(2026, 3, 3)
                and row["work_value"] == "자체 서성원"
                and row.get("validation_code") == "TIME_CONFLICT"
                for row in preview["preview_rows"]
            )
        )
        workbook.close()

    def test_load_canonical_schedule_import_apply_payload_rows_rebuilds_from_raw_workbook(self):
        workbook = self._build_sample_workbook()
        rebuilt_rows = {
            "resolved_rows": [
                {"source_block": "body", "row_no": 5},
                {"source_block": "body", "row_no": 8},
            ],
            "support_ticket_rows": [
                {"source_block": "sentrix_support_ticket", "row_no": 55},
            ],
        }

        with patch(
            "app.routers.v1.schedules._load_schedule_import_payload_rows",
            return_value=[{"source_block": "body", "row_no": 999}],
        ), patch(
            "app.routers.v1.schedules._load_schedule_import_batch_raw_workbook",
            return_value={
                "filename": "raw.xlsx",
                "sha256": "sha-raw",
                "workbook": workbook,
            },
        ), patch(
            "app.routers.v1.schedules._build_schedule_import_preview_result",
            return_value=rebuilt_rows,
        ):
            payload_rows = _load_canonical_schedule_import_apply_payload_rows(
                None,
                batch_id="batch-1",
                batch={"month_key": "2026-03", "metadata_json": {"workbook_valid": True}},
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                site_row={"id": "site-1", "site_code": "R692"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertEqual([row["row_no"] for row in payload_rows], [5, 8, 55])

    def test_apply_canonical_schedule_import_batch_reupload_replaces_only_target_scope(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-000000000101")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        existing_rows = [
            {
                "schedule_id": "sch-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 1),
                "shift_type": "day",
                "template_id": "tpl-day-12",
                "shift_start_time": "10:00:00",
                "shift_end_time": "22:00:00",
                "paid_hours": 12,
                "schedule_note": None,
                "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "duty_type": "day",
            },
            {
                "schedule_id": "sch-2",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-2",
                "schedule_date": date(2026, 3, 2),
                "shift_type": "day",
                "template_id": "tpl-day-10",
                "shift_start_time": "08:00:00",
                "shift_end_time": "18:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                "employee_code": "M001-2",
                "employee_name": "서성원",
                "duty_type": "day",
            },
            {
                "schedule_id": "sch-3",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-3",
                "schedule_date": date(2026, 3, 3),
                "shift_type": "day",
                "template_id": "tpl-day-10",
                "shift_start_time": "08:00:00",
                "shift_end_time": "18:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                "employee_code": "M001-3",
                "employee_name": "이보한",
                "duty_type": "day",
            },
        ]
        payload_rows = [
            {
                "row_no": 5,
                "source_block": "body",
                "employee_name": "민경민",
                "employee_code": "M001-1",
                "schedule_date": date(2026, 3, 1),
                "duty_type": "day",
                "shift_type": "day",
                "work_value": "12",
                "parsed_semantic_type": "numeric_hours",
                "template_id": "tpl-day-12",
                "template_name": "주간 12시간",
                "shift_start_time": "10:00:00",
                "shift_end_time": "22:00:00",
                "paid_hours": 12,
                "is_blocking": False,
                "section_label": "주간근무",
            },
            {
                "row_no": 8,
                "source_block": "body",
                "employee_name": "서성원",
                "employee_code": "M001-2",
                "schedule_date": date(2026, 3, 2),
                "duty_type": "day",
                "shift_type": "day",
                "work_value": "12",
                "parsed_semantic_type": "numeric_hours",
                "template_id": "tpl-day-12",
                "template_name": "주간 12시간",
                "shift_start_time": "10:00:00",
                "shift_end_time": "22:00:00",
                "paid_hours": 12,
                "is_blocking": False,
                "section_label": "주간근무",
            },
        ]
        employees = [
            {
                "id": "emp-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "M001-1",
                "full_name": "민경민",
                "hire_date": None,
                "leave_date": None,
            },
            {
                "id": "emp-2",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "M001-2",
                "full_name": "서성원",
                "hire_date": None,
                "leave_date": None,
            },
            {
                "id": "emp-3",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "M001-3",
                "full_name": "이보한",
                "hire_date": None,
                "leave_date": None,
            },
        ]
        updated_ids: list[str] = []
        deleted_ids: list[str] = []

        def _record_update(_cur, *, schedule_id, **_kwargs):
            updated_ids.append(schedule_id)

        def _record_delete(_cur, *, schedule_id):
            deleted_ids.append(schedule_id)

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=payload_rows,
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=employees,
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=existing_rows,
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._ensure_monthly_schedule_shift_type_constraint",
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
        ) as insert_row, patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
            side_effect=_record_update,
        ), patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
            side_effect=_record_delete,
        ), patch(
            "app.routers.v1.schedules._upsert_daytime_need_count_row",
        ), patch(
            "app.routers.v1.schedules._delete_daytime_need_count_row",
        ), patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertFalse(result.blocked)
        self.assertEqual(result.base_schedule_created, 0)
        self.assertEqual(result.base_schedule_updated, 1)
        self.assertEqual(result.base_schedule_removed, 1)
        self.assertEqual(updated_ids, ["sch-2"])
        self.assertEqual(deleted_ids, ["sch-3"])
        insert_row.assert_not_called()
        self.assertIn(("tenant-1", "site-1", "2026-03-02"), affected_site_days)
        self.assertIn(("tenant-1", "site-1", "2026-03-03"), affected_site_days)
        self.assertNotIn(("tenant-1", "site-1", "2026-03-01"), affected_site_days)

    def test_apply_canonical_schedule_import_batch_keeps_existing_rows_when_validation_blocks(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-000000000102")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=[],
        ) as load_payload_rows, patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
        ) as insert_row, patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
        ) as delete_row, patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": ["검증 실패로 재업로드가 차단되었습니다."],
                    "current_revision": "",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertTrue(result.blocked)
        self.assertIn("검증 실패로 재업로드가 차단되었습니다.", result.blocked_reasons)
        self.assertEqual(affected_site_days, set())
        load_payload_rows.assert_called_once()
        insert_row.assert_not_called()
        update_row.assert_not_called()
        delete_row.assert_not_called()

    def test_apply_canonical_schedule_import_batch_materializes_internal_support_shift_without_overwriting_day_shift(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-000000000103")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        existing_rows = [
            {
                "schedule_id": "sch-day-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "day",
                "template_id": "tpl-day-10",
                "shift_start_time": "08:00:00",
                "shift_end_time": "18:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "duty_type": "day",
            }
        ]
        payload_rows = [
            {
                "row_no": 5,
                "source_block": "body",
                "employee_name": "민경민",
                "employee_code": "M001-1",
                "schedule_date": date(2026, 3, 5),
                "duty_type": "day",
                "shift_type": "day",
                "work_value": "10",
                "parsed_semantic_type": "numeric_hours",
                "template_id": "tpl-day-10",
                "template_name": "주간 10시간",
                "shift_start_time": "08:00:00",
                "shift_end_time": "18:00:00",
                "paid_hours": 10,
                "is_blocking": False,
                "section_label": "주간근무",
            },
            {
                "row_no": 55,
                "source_block": "night_support_worker",
                "employee_id": "emp-1",
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "night",
                "decision_stage": "apply",
                "is_blocking": False,
                "internal_template_id": "tpl-night-10",
                "internal_shift_type": "night",
                "internal_shift_start_time": "22:00:00",
                "internal_shift_end_time": "08:00:00",
                "internal_paid_hours": 10,
                "section_label": "야간 지원근무자",
            },
        ]
        employees = [
            {
                "id": "emp-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "M001-1",
                "full_name": "민경민",
                "hire_date": None,
                "leave_date": None,
            }
        ]
        inserted_rows: list[dict] = []
        deleted_ids: list[str] = []

        def _record_insert(_cur, **kwargs):
            inserted_rows.append(dict(kwargs))
            return "new-support-row"

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=payload_rows,
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=employees,
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=existing_rows,
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._fetch_default_schedule_template_map",
            return_value={"night": {"id": "tpl-night-10", "template_name": "야간 10시간", "start_time": "22:00:00", "end_time": "08:00:00", "paid_hours": 10}},
        ), patch(
            "app.routers.v1.schedules._ensure_monthly_schedule_shift_type_constraint",
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
            side_effect=_record_insert,
        ), patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
            side_effect=lambda _cur, *, schedule_id: deleted_ids.append(schedule_id),
        ), patch(
            "app.routers.v1.schedules._upsert_daytime_need_count_row",
        ), patch(
            "app.routers.v1.schedules._delete_daytime_need_count_row",
        ), patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertFalse(result.blocked)
        self.assertEqual(len(inserted_rows), 1)
        self.assertEqual(inserted_rows[0]["source"], ARLS_SUPPORT_UPLOAD_INTERNAL_SOURCE)
        self.assertEqual(inserted_rows[0]["source_action"], "night_support")
        self.assertEqual(inserted_rows[0]["shift_type"], "night")
        self.assertEqual(inserted_rows[0]["shift_start_time"], "22:00:00")
        self.assertEqual(inserted_rows[0]["shift_end_time"], "08:00:00")
        self.assertEqual(deleted_ids, [])
        update_row.assert_not_called()
        self.assertIn(("tenant-1", "site-1", "2026-03-05"), affected_site_days)

    def test_apply_canonical_schedule_import_batch_skips_internal_support_insert_when_same_slot_is_already_created_by_body_row(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-000000000103")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        payload_rows = [
            {
                "row_no": 7,
                "source_block": "body",
                "employee_name": "민경민",
                "employee_code": "M001-1",
                "schedule_date": date(2026, 3, 3),
                "duty_type": "night",
                "shift_type": "night",
                "work_value": "10",
                "parsed_semantic_type": "numeric_hours",
                "template_id": "tpl-night-10",
                "template_name": "야간 10시간",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "is_blocking": False,
                "section_label": "야간근무",
            },
            {
                "row_no": 58,
                "source_block": "night_support_worker",
                "employee_id": "emp-1",
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "schedule_date": date(2026, 3, 3),
                "shift_type": "night",
                "decision_stage": "apply",
                "is_blocking": False,
                "internal_template_id": "tpl-night-10",
                "internal_shift_type": "night",
                "internal_shift_start_time": "22:00:00",
                "internal_shift_end_time": "08:00:00",
                "internal_paid_hours": 10,
                "section_label": "야간 지원근무자",
            },
        ]
        employees = [
            {
                "id": "emp-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "M001-1",
                "full_name": "민경민",
                "hire_date": None,
                "leave_date": None,
            }
        ]
        inserted_rows: list[dict] = []

        def _record_insert(_cur, **kwargs):
            inserted_rows.append(dict(kwargs))
            return f"new-{len(inserted_rows)}"

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=payload_rows,
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=employees,
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._fetch_default_schedule_template_map",
            return_value={"night": {"id": "tpl-night-10", "template_name": "야간 10시간", "start_time": "22:00:00", "end_time": "08:00:00", "paid_hours": 10}},
        ), patch(
            "app.routers.v1.schedules._ensure_monthly_schedule_shift_type_constraint",
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
            side_effect=_record_insert,
        ), patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
        ) as delete_row, patch(
            "app.routers.v1.schedules._upsert_daytime_need_count_row",
        ), patch(
            "app.routers.v1.schedules._delete_daytime_need_count_row",
        ), patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertFalse(result.blocked)
        self.assertEqual(len(inserted_rows), 1)
        self.assertEqual(inserted_rows[0]["source"], ARLS_MONTHLY_BASE_UPLOAD_SOURCE)
        self.assertEqual(inserted_rows[0]["shift_type"], "night")
        self.assertEqual(result.applied, 1)
        update_row.assert_not_called()
        delete_row.assert_not_called()
        self.assertIn(("tenant-1", "site-1", "2026-03-03"), affected_site_days)

    def test_apply_canonical_schedule_import_batch_keeps_existing_matching_shift_when_internal_support_row_targets_same_slot(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-000000000104")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        existing_rows = [
            {
                "schedule_id": "sch-day-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "day",
                "template_id": "tpl-day-10",
                "shift_start_time": "08:00:00",
                "shift_end_time": "18:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_MONTHLY_BASE_UPLOAD_SOURCE,
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "duty_type": "day",
            }
        ]
        payload_rows = [
            {
                "row_no": 5,
                "source_block": "body",
                "employee_name": "민경민",
                "employee_code": "M001-1",
                "schedule_date": date(2026, 3, 5),
                "duty_type": "day",
                "shift_type": "day",
                "work_value": "10",
                "parsed_semantic_type": "numeric_hours",
                "template_id": "tpl-day-10",
                "template_name": "주간 10시간",
                "shift_start_time": "08:00:00",
                "shift_end_time": "18:00:00",
                "paid_hours": 10,
                "is_blocking": False,
                "section_label": "주간근무",
            },
            {
                "row_no": 55,
                "source_block": "day_support_worker",
                "employee_id": "emp-1",
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "day",
                "decision_stage": "apply",
                "is_blocking": False,
                "internal_template_id": "tpl-day-10",
                "internal_shift_type": "day",
                "internal_shift_start_time": "08:00:00",
                "internal_shift_end_time": "18:00:00",
                "internal_paid_hours": None,
                "section_label": "주간 지원근무자",
            },
        ]
        employees = [
            {
                "id": "emp-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "M001-1",
                "full_name": "민경민",
                "hire_date": None,
                "leave_date": None,
            }
        ]

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=payload_rows,
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=employees,
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=existing_rows,
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._fetch_default_schedule_template_map",
            return_value={"day": {"id": "tpl-day-10", "template_name": "주간 10시간", "start_time": "08:00:00", "end_time": "18:00:00", "paid_hours": None}},
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
        ) as insert_row, patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
        ) as delete_row, patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, _affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertFalse(result.blocked)
        self.assertEqual(result.blocked_reasons, [])
        insert_row.assert_not_called()
        update_row.assert_not_called()
        delete_row.assert_not_called()

    def test_apply_canonical_schedule_import_batch_skips_external_support_worker_rows(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-000000000104")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        payload_rows = [
            {
                "row_no": 48,
                "source_block": "day_support_worker",
                "employee_id": None,
                "employee_code": "",
                "employee_name": "민경민",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "day",
                "decision_stage": "apply",
                "is_blocking": False,
                "support_origin_type": "f_support_worker",
                "worker_type": "F",
                "work_value": "ks 민경민",
                "section_label": "주간 지원근무자",
            },
        ]

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=payload_rows,
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._fetch_default_schedule_template_map",
            return_value={},
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
        ) as insert_row, patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
        ) as delete_row, patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertFalse(result.blocked)
        self.assertEqual(result.blocked_reasons, [])
        self.assertEqual(affected_site_days, set())
        insert_row.assert_not_called()
        update_row.assert_not_called()
        delete_row.assert_not_called()

    def test_apply_canonical_schedule_import_batch_deletes_stale_internal_support_rows(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-000000000104")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        existing_rows = [
            {
                "schedule_id": "support-night-stale-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-2",
                "schedule_date": date(2026, 3, 3),
                "shift_type": "night",
                "template_id": "tpl-night-10",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_SUPPORT_UPLOAD_INTERNAL_SOURCE,
                "employee_code": "M001-2",
                "employee_name": "서성원",
                "duty_type": "night",
            }
        ]
        deleted_ids: list[str] = []

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=existing_rows,
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._fetch_default_schedule_template_map",
            return_value={},
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
        ) as insert_row, patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
            side_effect=lambda _cur, *, schedule_id: deleted_ids.append(schedule_id),
        ), patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertFalse(result.blocked)
        self.assertEqual(deleted_ids, ["support-night-stale-1"])
        self.assertEqual(result.applied, 1)
        self.assertIn(("tenant-1", "site-1", "2026-03-03"), affected_site_days)
        insert_row.assert_not_called()
        update_row.assert_not_called()

    def test_apply_canonical_schedule_import_batch_ignores_blank_body_row_when_only_foreign_lineage_exists(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-00000000010a")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        payload_rows = [
            {
                "row_no": 7,
                "source_block": "body",
                "employee_name": "민경민",
                "employee_code": "M001-1",
                "schedule_date": date(2026, 3, 5),
                "duty_type": "night",
                "shift_type": "night",
                "work_value": "",
                "parsed_semantic_type": "blank",
                "template_id": None,
                "template_name": None,
                "shift_start_time": None,
                "shift_end_time": None,
                "paid_hours": None,
                "is_blocking": False,
                "section_label": "야간근무",
            },
        ]
        employees = [
            {
                "id": "emp-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "M001-1",
                "full_name": "민경민",
                "hire_date": None,
                "leave_date": None,
            }
        ]
        existing_rows = [
            {
                "schedule_id": "support-night-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "night",
                "template_id": "tpl-night-10",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_SUPPORT_UPLOAD_INTERNAL_SOURCE,
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "duty_type": "night",
            }
        ]

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=payload_rows,
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=employees,
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=existing_rows,
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
        ) as insert_row, patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
        ) as delete_row, patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertFalse(result.blocked)
        self.assertEqual(result.blocked_reasons, [])
        self.assertEqual(affected_site_days, set())
        insert_row.assert_not_called()
        update_row.assert_not_called()
        delete_row.assert_not_called()

    def test_apply_canonical_schedule_import_batch_ignores_matching_foreign_lineage_body_value(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-00000000010b")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        payload_rows = [
            {
                "row_no": 7,
                "source_block": "body",
                "employee_name": "민경민",
                "employee_code": "M001-1",
                "schedule_date": date(2026, 3, 6),
                "duty_type": "night",
                "shift_type": "night",
                "work_value": "10",
                "current_work_value": "10",
                "parsed_semantic_type": "numeric_hours",
                "template_id": "tpl-night-10",
                "template_name": "야간 10시간",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "is_blocking": False,
                "section_label": "야간근무",
            },
        ]
        employees = [
            {
                "id": "emp-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_code": "M001-1",
                "full_name": "민경민",
                "hire_date": None,
                "leave_date": None,
            }
        ]
        existing_rows = [
            {
                "schedule_id": "support-night-1",
                "tenant_id": "tenant-1",
                "company_id": "company-1",
                "site_id": "site-1",
                "employee_id": "emp-1",
                "schedule_date": date(2026, 3, 6),
                "shift_type": "night",
                "template_id": "tpl-night-10",
                "shift_start_time": "22:00:00",
                "shift_end_time": "08:00:00",
                "paid_hours": 10,
                "schedule_note": None,
                "source": ARLS_SUPPORT_UPLOAD_INTERNAL_SOURCE,
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "duty_type": "night",
            }
        ]

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=payload_rows,
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=employees,
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=existing_rows,
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
        ) as insert_row, patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
        ) as delete_row, patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertFalse(result.blocked)
        self.assertEqual(result.blocked_reasons, [])
        self.assertEqual(affected_site_days, set())
        insert_row.assert_not_called()
        update_row.assert_not_called()
        delete_row.assert_not_called()

    def test_apply_canonical_schedule_import_batch_blocks_duplicate_internal_support_rows_same_shift(self):
        batch_id = uuid.UUID("00000000-0000-0000-0000-000000000105")
        conn = _ConnectionStub(cursor_responses=[[{"cnt": 0}], []])
        payload_rows = [
            {
                "row_no": 55,
                "source_block": "night_support_worker",
                "employee_id": "emp-1",
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "night",
                "decision_stage": "apply",
                "is_blocking": False,
                "internal_template_id": "tpl-night-10",
                "internal_shift_type": "night",
                "internal_shift_start_time": "22:00:00",
                "internal_shift_end_time": "08:00:00",
                "internal_paid_hours": 10,
                "section_label": "야간 지원근무자",
            },
            {
                "row_no": 56,
                "source_block": "night_support_worker",
                "employee_id": "emp-1",
                "employee_code": "M001-1",
                "employee_name": "민경민",
                "schedule_date": date(2026, 3, 5),
                "shift_type": "night",
                "decision_stage": "apply",
                "is_blocking": False,
                "internal_template_id": "tpl-night-10",
                "internal_shift_type": "night",
                "internal_shift_start_time": "22:00:00",
                "internal_shift_end_time": "08:00:00",
                "internal_paid_hours": 10,
                "section_label": "야간 지원근무자",
            },
        ]

        with patch("app.routers.v1.schedules._resolve_site_context_by_code", return_value={"id": "site-1", "site_code": "M001", "company_id": "company-1"}), patch(
            "app.routers.v1.schedules._load_canonical_schedule_import_apply_payload_rows",
            return_value=payload_rows,
        ), patch(
            "app.routers.v1.schedules._load_site_employees",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._load_existing_schedule_rows_for_import",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._load_daytime_need_count_rows_for_apply",
            return_value=[],
        ), patch(
            "app.routers.v1.schedules._fetch_default_schedule_template_map",
            return_value={"night": {"id": "tpl-night-10", "template_name": "야간 10시간", "start_time": "22:00:00", "end_time": "08:00:00", "paid_hours": 10}},
        ), patch(
            "app.routers.v1.schedules._insert_monthly_schedule_row",
        ) as insert_row, patch(
            "app.routers.v1.schedules._update_monthly_schedule_row",
        ) as update_row, patch(
            "app.routers.v1.schedules._delete_monthly_schedule_row",
        ) as delete_row, patch(
            "app.routers.v1.schedules._write_schedule_import_batch_apply_audit",
        ):
            result, _affected_site_days = _apply_canonical_schedule_import_batch(
                conn,
                batch_id=batch_id,
                batch={
                    "site_code": "M001",
                    "month_key": "2026-03",
                    "metadata_json": {"workbook_valid": True},
                    "blocked_reasons_json": [],
                    "current_revision": "",
                    "export_revision": "rev-1",
                },
                target_tenant={"id": "tenant-1", "tenant_code": "SRS_KOR"},
                user={"id": "user-1", "role": "developer"},
            )

        self.assertTrue(result.blocked)
        self.assertIn("같은 직원/날짜/지원근무유형이 업로드 결과에 중복되어 있습니다.", result.blocked_reasons)
        insert_row.assert_not_called()
        update_row.assert_not_called()
        delete_row.assert_not_called()

    def test_normalize_name_token_handles_unicode_and_zero_width_variants(self):
        import unicodedata

        base_name = "서성원"
        nfd_name = unicodedata.normalize("NFD", base_name)
        hidden_name = "서\u200b성원"

        self.assertEqual(_normalize_name_token(base_name), _normalize_name_token(nfd_name))
        self.assertEqual(_normalize_name_token(base_name), _normalize_name_token(hidden_name))


if __name__ == "__main__":
    unittest.main()
