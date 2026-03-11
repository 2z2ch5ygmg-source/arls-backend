#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


ARLS_SHEET_NAME = "본사 스케쥴 양식"
ARLS_DATE_START_COL = 4
ARLS_DATE_END_COL = 34
ARLS_SUMMARY_END_COL = 37
ARLS_METADATA_SHEETS = {"_ARLS_EXPORT_META", "_ARLS_SUPPORT_META"}
ARLS_DATA_LABELS = {"주간근무", "초과근무", "야간근무"}
ARLS_SUMMARY_LABEL_KEYWORDS = (
    "주간 근무자(직원) 수",
    "주간 추가 근무자 수",
    "주간 출근자 총 수",
    "야간 근무자 총 수",
    "주간 지원 근무자",
    "야간 지원 근무자",
    "필요인원 수",
    "외부인원 투입 수",
    "작업 내용",
    "작업 목적",
)


def _normalize_label(value: object) -> str:
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def _set_if_writable(sheet, *, row: int, column: int, value: Any) -> bool:
    cell = sheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def _find_data_start_row(sheet) -> int:
    for row_idx in range(1, sheet.max_row + 1):
        if _normalize_label(sheet.cell(row=row_idx, column=3).value) in ARLS_DATA_LABELS:
            return row_idx
    return 5


def _find_summary_start_row(sheet, *, fallback: int) -> int:
    for row_idx in range(fallback, sheet.max_row + 1):
        left = _normalize_label(sheet.cell(row=row_idx, column=2).value)
        right = _normalize_label(sheet.cell(row=row_idx, column=3).value)
        combined = f"{left} {right}".strip()
        if any(keyword in combined for keyword in ARLS_SUMMARY_LABEL_KEYWORDS):
            return row_idx
    return fallback


def _clear_hidden_sheet(sheet) -> None:
    for row in range(2, min(sheet.max_row, 64) + 1):
        for col in range(2, min(sheet.max_column, 26) + 1):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None


def prepare_template(source: Path, dest: Path) -> None:
    workbook = load_workbook(source)
    sheet = workbook[ARLS_SHEET_NAME]

    data_start_row = _find_data_start_row(sheet)
    summary_start_row = _find_summary_start_row(sheet, fallback=data_start_row + 42)

    for row_idx in range(data_start_row, summary_start_row):
        group_offset = (row_idx - data_start_row) % 3
        for col_idx in range(1, ARLS_SUMMARY_END_COL + 1):
            if col_idx == 3:
                continue
            if group_offset != 0 and col_idx in {1, 2, 36, 37}:
                continue
            _set_if_writable(sheet, row=row_idx, column=col_idx, value=None)
        _set_if_writable(
            sheet,
            row=row_idx,
            column=3,
            value=("주간근무", "초과근무", "야간근무")[group_offset],
        )
        if group_offset == 0:
            _set_if_writable(sheet, row=row_idx, column=36, value=None)
            _set_if_writable(sheet, row=row_idx, column=37, value=None)

    for col_idx in range(ARLS_DATE_START_COL, ARLS_DATE_END_COL + 1):
        _set_if_writable(sheet, row=2, column=col_idx, value=None)
        _set_if_writable(sheet, row=3, column=col_idx, value=None)
        _set_if_writable(sheet, row=4, column=col_idx, value=None)

    for row_idx in range(summary_start_row, min(sheet.max_row, 84) + 1):
        for col_idx in range(1, ARLS_SUMMARY_END_COL + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            if col_idx >= ARLS_DATE_START_COL or col_idx in {1, 35, 36, 37}:
                cell.value = None

    sheet["B1"] = None
    sheet["B2"] = None
    sheet["B3"] = None

    hidden_sheet_name = "출동.잔업 초과수당(2)"
    if hidden_sheet_name in workbook.sheetnames:
        _clear_hidden_sheet(workbook[hidden_sheet_name])

    for meta_sheet in ARLS_METADATA_SHEETS:
        if meta_sheet in workbook.sheetnames:
            workbook.remove(workbook[meta_sheet])

    dest.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(dest)


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare canonical blank ARLS monthly schedule template from a sample workbook.")
    parser.add_argument("source", help="Source workbook path")
    parser.add_argument("dest", help="Destination workbook path")
    args = parser.parse_args()
    prepare_template(Path(args.source).expanduser(), Path(args.dest).expanduser())


if __name__ == "__main__":
    main()
